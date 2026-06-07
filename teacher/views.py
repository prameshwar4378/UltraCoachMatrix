from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from institute_admin.forms import get_or_create_academic_year
from institute_admin.models import Batch, Course, Subject
from student_parent.models import StudentAcademicSession, StudentEnrollment
from super_admin.decorators import teacher_required
from .forms import ExamQuestionForm, ExamQuestionOptionFormSet, TeacherExamForm, TeacherHomeworkForm
from .models import (
    Attendance,
    Exam,
    ExamAttempt,
    ExamAttemptUpload,
    ExamQuestion,
    ExamQuestionAttempt,
    ExamQuestionOption,
    ExamResult,
    Homework,
)


BULK_QUESTION_HEADERS = [
    "Question Text",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Correct Answer",
    "Marks",
]
BULK_ANSWER_TO_ORDER = {"A": 1, "B": 2, "C": 3, "D": 4}


def close_teacher_popup_response(fallback_url="/teacher/exams/"):
    return HttpResponse(
        f"""
        <script>
            if (window.opener) {{
                window.opener.location.reload();
                window.close();
            }} else {{
                window.location.href = "{fallback_url}";
            }}
        </script>
        """
    )


def teacher_selected_academic_year(request):
    profile = getattr(request.user, "profile", None)
    if not profile or not profile.institute_id:
        return None
    year_id = request.session.get("academic_year_id")
    if year_id:
        year = profile.institute.academic_years.filter(pk=year_id, is_active=True).first()
        if year:
            return year
    year = get_or_create_academic_year(profile.institute)
    request.session["academic_year_id"] = year.pk
    return year


def teacher_batches(request):
    profile = getattr(request.user, "profile", None)
    if not profile or not profile.institute_id:
        return Batch.objects.none()
    batches = Batch.objects.filter(
        institute=profile.institute,
        teachers=request.user,
        is_active=True,
    ).prefetch_related("courses")
    selected_year = teacher_selected_academic_year(request)
    if selected_year:
        batches = batches.filter(academic_year=selected_year)
    return batches


def teacher_students_for_batches(batches):
    return (
        StudentAcademicSession.objects.filter(
            enrollments__batch__in=batches,
            enrollments__status=StudentEnrollment.Status.ACTIVE,
            status=StudentAcademicSession.Status.ACTIVE,
            student__is_active=True,
        )
        .select_related("student", "student__user")
        .distinct()
        .order_by("admission_number", "student__user__first_name", "student__user__username")
    )


def get_teacher_batch_course_data(batches):
    return {
        str(batch.pk): [
            {
                "id": str(course.pk),
                "name": course.name,
                "fee": str(course.fee_amount),
            }
            for course in batch.courses.all()
        ]
        for batch in batches.prefetch_related("courses")
    }


@teacher_required
def dashboard(request):
    batches = teacher_batches(request)
    today = date.today()
    week_end = today + timedelta(days=7)
    students_qs = teacher_students_for_batches(batches)
    homework_qs = Homework.objects.filter(batch__in=batches)
    exam_qs = Exam.objects.filter(batch__in=batches)
    today_attendance = Attendance.objects.filter(batch__in=batches, date=today)
    present_today = today_attendance.filter(status=Attendance.Status.PRESENT).count()
    absent_today = today_attendance.filter(status=Attendance.Status.ABSENT).count()
    late_today = today_attendance.filter(status=Attendance.Status.LATE).count()
    attendance_marked_count = today_attendance.count()
    attendance_rate = round((present_today / attendance_marked_count) * 100, 1) if attendance_marked_count else 0
    assigned_batches = (
        batches.annotate(
            active_students=Count(
                "enrollments",
                filter=Q(
                    enrollments__status=StudentEnrollment.Status.ACTIVE,
                    enrollments__academic_session__status=StudentAcademicSession.Status.ACTIVE,
                    enrollments__student__is_active=True,
                ),
                distinct=True,
            ),
            course_total=Count("courses", distinct=True),
        )
        .order_by("name")
    )
    upcoming_exams = exam_qs.filter(exam_date__gte=today).select_related("batch", "subject").order_by("exam_date")[:6]
    due_homework = (
        homework_qs.filter(due_date__gte=today)
        .select_related("batch", "subject", "course")
        .order_by("due_date", "-created_at")[:6]
    )
    context = {
        "batch_count": batches.count(),
        "student_count": students_qs.count(),
        "homework_count": homework_qs.count(),
        "exam_count": exam_qs.count(),
        "today_attendance_count": attendance_marked_count,
        "present_today": present_today,
        "absent_today": absent_today,
        "late_today": late_today,
        "attendance_rate": attendance_rate,
        "open_homework_count": homework_qs.filter(Q(due_date__isnull=True) | Q(due_date__gte=today)).count(),
        "due_soon_homework_count": homework_qs.filter(due_date__range=(today, week_end)).count(),
        "overdue_homework_count": homework_qs.filter(due_date__lt=today).count(),
        "published_exam_count": exam_qs.filter(is_published=True).count(),
        "draft_exam_count": exam_qs.filter(is_published=False).count(),
        "upcoming_exam_count": exam_qs.filter(exam_date__gte=today).count(),
        "submitted_attempt_count": ExamAttempt.objects.filter(exam__in=exam_qs, submitted_at__isnull=False).count(),
        "result_count": ExamAttempt.objects.filter(exam__in=exam_qs, submitted_at__isnull=False).count(),
        "assigned_batches": assigned_batches,
        "recent_homework": homework_qs.select_related("batch", "subject", "course").order_by("-created_at")[:6],
        "due_homework": due_homework,
        "recent_exams": upcoming_exams,
        "recent_results": ExamAttempt.objects.filter(
            exam__in=exam_qs,
            submitted_at__isnull=False,
        ).select_related("exam", "exam__batch", "student", "student__user").order_by("-submitted_at")[:6],
        "today": today,
        "week_end": week_end,
        "selected_academic_year": teacher_selected_academic_year(request),
        "teacher_name": request.user.get_full_name() or request.user.username,
    }
    return render(request, "teacher/dashboard.html", context)


@teacher_required
def attendance(request):
    batch_queryset = teacher_batches(request)
    selected_year = teacher_selected_academic_year(request)
    selected_date = parse_date(request.GET.get("date", "")) or date.today()
    batch_id = request.GET.get("batch", "").strip()
    selected_batch = batch_queryset.filter(pk=batch_id).first() if batch_id else batch_queryset.first()

    student_sessions = StudentAcademicSession.objects.none()
    export_student_sessions = teacher_students_for_batches(batch_queryset)
    attendance_map = {}

    if selected_batch:
        student_sessions = teacher_students_for_batches(batch_queryset.filter(pk=selected_batch.pk))
        existing_attendance = Attendance.objects.filter(
            batch=selected_batch,
            date=selected_date,
            academic_session__in=student_sessions,
        ).select_related("academic_session", "student", "marked_by")
        attendance_map = {record.academic_session_id: record for record in existing_attendance}

    if request.method == "POST" and selected_batch:
        posted_student_ids = request.POST.getlist("student_ids")
        saved_count = 0
        for student_id in posted_student_ids:
            status = request.POST.get(f"status_{student_id}", Attendance.Status.PRESENT)
            note = request.POST.get(f"note_{student_id}", "").strip()
            if status not in Attendance.Status.values:
                status = Attendance.Status.PRESENT
            student_session = student_sessions.filter(student_id=student_id).first()
            if not student_session:
                continue
            Attendance.objects.update_or_create(
                academic_session=student_session,
                batch=selected_batch,
                date=selected_date,
                defaults={
                    "student": student_session.student,
                    "status": status,
                    "note": note,
                    "marked_by": request.user,
                },
            )
            saved_count += 1
        messages.success(request, f"Attendance saved for {saved_count} student(s).")
        return redirect(f"{reverse('teacher:attendance')}?batch={selected_batch.pk}&date={selected_date.isoformat()}")

    selected_date_records = Attendance.objects.filter(batch__in=batch_queryset, date=selected_date)
    all_records = Attendance.objects.filter(batch__in=batch_queryset)
    if selected_year:
        selected_date_records = selected_date_records.filter(academic_session__academic_year=selected_year)
        all_records = all_records.filter(academic_session__academic_year=selected_year, batch__academic_year=selected_year)
    if selected_batch:
        selected_date_records = selected_date_records.filter(batch=selected_batch)

    total_today = selected_date_records.count()
    present_today = selected_date_records.filter(status=Attendance.Status.PRESENT).count()
    absent_today = selected_date_records.filter(status=Attendance.Status.ABSENT).count()
    late_today = selected_date_records.filter(status=Attendance.Status.LATE).count()
    rate_today = round((present_today / total_today) * 100, 1) if total_today else 0

    rows = []
    for student_session in student_sessions:
        record = attendance_map.get(student_session.pk)
        rows.append(
            {
                "student": student_session.student,
                "student_session": student_session,
                "record": record,
                "status": record.status if record else Attendance.Status.PRESENT,
                "note": record.note if record else "",
            }
        )

    context = {
        "batches": batch_queryset,
        "selected_batch": selected_batch,
        "selected_date": selected_date,
        "attendance_rows": rows,
        "export_students": export_student_sessions,
        "status_choices": Attendance.Status.choices,
        "present_value": Attendance.Status.PRESENT,
        "absent_value": Attendance.Status.ABSENT,
        "late_value": Attendance.Status.LATE,
        "total_students": student_sessions.count() if selected_batch else 0,
        "marked_count": len(attendance_map),
        "total_today": total_today,
        "present_today": present_today,
        "absent_today": absent_today,
        "late_today": late_today,
        "rate_today": rate_today,
        "recent_attendance": all_records.select_related("student", "student__user", "batch", "marked_by")[:8],
    }
    return render(request, "teacher/attendance.html", context)


def get_teacher_attendance_export_queryset(request):
    batches = teacher_batches(request)
    records = Attendance.objects.select_related(
        "academic_session", "student", "student__user", "batch", "marked_by"
    ).filter(batch__in=batches)
    selected_year = teacher_selected_academic_year(request)
    if selected_year:
        records = records.filter(academic_session__academic_year=selected_year)

    batch_id = request.GET.get("batch", "").strip()
    student_id = request.GET.get("student", "").strip()
    status = request.GET.get("status", "").strip()
    date_from = parse_date(request.GET.get("date_from", ""))
    date_to = parse_date(request.GET.get("date_to", ""))

    if batch_id:
        records = records.filter(batch_id=batch_id, batch__in=batches)
    if student_id:
        records = records.filter(student_id=student_id)
    if status in Attendance.Status.values:
        records = records.filter(status=status)
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)

    return records.order_by("-date", "batch__name", "academic_session__admission_number")


def attendance_export_filename(file_format):
    stamp = timezone.now().strftime("%Y%m%d_%H%M")
    extension = "xlsx" if file_format == "excel" else "pdf"
    return f"attendance_report_{stamp}.{extension}"


def attendance_status_counts(records):
    total = len(records)
    present = sum(1 for record in records if record.status == Attendance.Status.PRESENT)
    absent = sum(1 for record in records if record.status == Attendance.Status.ABSENT)
    late = sum(1 for record in records if record.status == Attendance.Status.LATE)
    rate = round((present / total) * 100, 1) if total else 0
    return total, present, absent, late, rate


def build_attendance_excel(records, institute, filters, include_notes=True):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    primary = "2563EB"
    dark = "0F172A"
    muted = "64748B"
    light_blue = "DBEAFE"
    light_green = "DCFCE7"
    light_red = "FEE2E2"
    light_amber = "FEF3C7"
    border_color = "CBD5E1"

    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    columns = ["Date", "Admission No", "Student Name", "Batch", "Status", "Marked By"]
    if include_notes:
        columns.append("Note")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1, value="Attendance Report")
    title_cell.font = Font(size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=primary)
    title_cell.alignment = center
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    institute_name = getattr(institute, "name", "All Institutes") if institute else "All Institutes"
    sheet.cell(row=2, column=1, value=f"{institute_name} | Generated {timezone.now().strftime('%d-%m-%Y %I:%M %p')}").font = Font(size=10, color=muted)
    sheet.cell(row=2, column=1).alignment = center

    total, present, absent, late, rate = attendance_status_counts(records)
    summary = [
        ("Total", total, light_blue),
        ("Present", present, light_green),
        ("Absent", absent, light_red),
        ("Late", late, light_amber),
        ("Rate", f"{rate}%", light_blue),
    ]
    col = 1
    for label, value, fill in summary:
        sheet.cell(row=4, column=col, value=label).font = Font(bold=True, color=dark)
        sheet.cell(row=4, column=col).fill = PatternFill("solid", fgColor=fill)
        sheet.cell(row=4, column=col).alignment = center
        sheet.cell(row=4, column=col).border = border
        sheet.cell(row=5, column=col, value=value).font = Font(size=14, bold=True, color=dark)
        sheet.cell(row=5, column=col).alignment = center
        sheet.cell(row=5, column=col).border = border
        col += 1

    filter_text = " | ".join(part for part in filters if part)
    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(columns))
    sheet.cell(row=7, column=1, value=f"Filters: {filter_text or 'All records'}").font = Font(italic=True, color=muted)

    header_row = 9
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border

    status_fills = {
        Attendance.Status.PRESENT: light_green,
        Attendance.Status.ABSENT: light_red,
        Attendance.Status.LATE: light_amber,
    }
    for row_index, record in enumerate(records, start=header_row + 1):
        values = [
            record.date.strftime("%d-%m-%Y"),
            record.academic_session.admission_number,
            record.student.user.get_full_name() or record.student.user.username,
            record.batch.name,
            record.get_status_display(),
            record.marked_by.get_full_name() if record.marked_by else "Not set",
        ]
        if include_notes:
            values.append(record.note or "")
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = left if col_index in (3, 4, len(values)) else center
            if col_index == 5:
                cell.fill = PatternFill("solid", fgColor=status_fills.get(record.status, "F8FAFC"))
                cell.font = Font(bold=True, color=dark)

    widths = [14, 16, 28, 24, 14, 24, 34]
    for index, width in enumerate(widths[: len(columns)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:{get_column_letter(len(columns))}{max(10, header_row + len(records))}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def pdf_escape(value):
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(x, y, text, size=9, font="F1", color="0 0 0"):
    return f"{color} rg BT /{font} {size} Tf {x} {y} Td ({pdf_escape(text)}) Tj ET\n"


def pdf_rect(x, y, w, h, color):
    return f"{color} rg {x} {y} {w} {h} re f\n"


def build_pdf_document(page_streams, width=842, height=595):
    objects = ["<< /Type /Catalog /Pages 2 0 R >>"]
    kids = []
    current_id = 3
    for stream in page_streams:
        page_id = current_id
        content_id = current_id + 1
        kids.append(f"{page_id} 0 R")
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> /F2 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >> >> >> /Contents {content_id} 0 R >>")
        encoded = stream.encode("latin-1", errors="replace")
        objects.append(f"<< /Length {len(encoded)} >>\nstream\n{stream}endstream")
        current_id += 2
    objects.insert(1, f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>")

    output = "%PDF-1.4\n"
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output.encode("latin-1", errors="replace")))
        output += f"{index} 0 obj\n{obj}\nendobj\n"
    xref_offset = len(output.encode("latin-1", errors="replace"))
    output += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for offset in offsets[1:]:
        output += f"{offset:010d} 00000 n \n"
    output += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
    return output.encode("latin-1", errors="replace")


def build_attendance_pdf(records, institute, filters, include_notes=True, title="Attendance Report"):
    width, height = 842, 595
    rows_per_page = 18 if include_notes else 22
    pages = []
    total, present, absent, late, rate = attendance_status_counts(records)
    institute_name = getattr(institute, "name", "All Institutes") if institute else "All Institutes"
    chunks = [records[index:index + rows_per_page] for index in range(0, len(records), rows_per_page)] or [[]]

    for page_number, chunk in enumerate(chunks, start=1):
        stream = ""
        stream += pdf_rect(0, 545, width, 50, "0.145 0.388 0.922")
        stream += pdf_text(34, 570, title, 18, "F2", "1 1 1")
        stream += pdf_text(34, 552, f"{institute_name} | Generated {timezone.now().strftime('%d-%m-%Y %I:%M %p')}", 8, "F1", "1 1 1")
        stream += pdf_text(700, 552, f"Page {page_number} of {len(chunks)}", 8, "F1", "1 1 1")

        summary = [("Total", total), ("Present", present), ("Absent", absent), ("Late", late), ("Rate", f"{rate}%")]
        x = 34
        for label, value in summary:
            stream += pdf_rect(x, 500, 128, 32, "0.93 0.96 1")
            stream += pdf_text(x + 8, 519, label, 7, "F1", "0.39 0.45 0.55")
            stream += pdf_text(x + 8, 506, value, 12, "F2", "0.06 0.09 0.16")
            x += 138

        filter_text = " | ".join(part for part in filters if part) or "All records"
        stream += pdf_text(34, 480, f"Filters: {filter_text[:135]}", 8, "F1", "0.39 0.45 0.55")

        headers = ["Date", "Adm No", "Student", "Batch", "Status", "Marked By"]
        widths = [68, 76, 168, 142, 76, 126]
        if include_notes:
            headers.append("Note")
            widths.append(118)
        start_x = 34
        y = 452
        stream += pdf_rect(start_x, y - 6, sum(widths), 22, "0.06 0.09 0.16")
        x = start_x
        for header, col_width in zip(headers, widths):
            stream += pdf_text(x + 4, y, header, 8, "F2", "1 1 1")
            x += col_width
        y -= 24

        for record in chunk:
            status_color = "0.86 0.99 0.91" if record.status == Attendance.Status.PRESENT else "1 0.89 0.89" if record.status == Attendance.Status.ABSENT else "1 0.95 0.78"
            stream += pdf_rect(start_x, y - 5, sum(widths), 20, "0.98 0.99 1")
            row_values = [
                record.date.strftime("%d-%m-%Y"),
                record.academic_session.admission_number,
                (record.student.user.get_full_name() or record.student.user.username)[:28],
                record.batch.name[:22],
                record.get_status_display(),
                (record.marked_by.get_full_name() if record.marked_by else "Not set")[:20],
            ]
            if include_notes:
                row_values.append((record.note or "")[:24])
            x = start_x
            for index, (value, col_width) in enumerate(zip(row_values, widths)):
                if index == 4:
                    stream += pdf_rect(x + 2, y - 3, col_width - 4, 15, status_color)
                    stream += pdf_text(x + 5, y, value, 7, "F2", "0.06 0.09 0.16")
                else:
                    stream += pdf_text(x + 4, y, value, 7, "F1", "0.06 0.09 0.16")
                x += col_width
            y -= 21

        stream += pdf_text(34, 28, "UltraCoachMatrix attendance export", 8, "F1", "0.39 0.45 0.55")
        pages.append(stream)

    return build_pdf_document(pages, width=width, height=height)


@teacher_required
def attendance_export(request):
    profile = getattr(request.user, "profile", None)
    institute = profile.institute if profile and profile.institute_id else None
    file_format = request.GET.get("format", "excel").strip()
    include_notes = request.GET.get("include_notes", "1") == "1"
    records = list(get_teacher_attendance_export_queryset(request))
    batches = teacher_batches(request)

    batch_label = "All Batches"
    batch_id = request.GET.get("batch", "").strip()
    if batch_id:
        batch = batches.filter(pk=batch_id).first()
        batch_label = batch.name if batch else batch_label
    student_label = "All Students"
    student_id = request.GET.get("student", "").strip()
    if student_id:
        student_session = teacher_students_for_batches(batches).filter(student_id=student_id).first()
        if student_session:
            student_label = str(student_session)
    status_value = request.GET.get("status", "").strip()
    status_label = dict(Attendance.Status.choices).get(status_value, "All Status")
    filters = [
        f"Batch: {batch_label}",
        f"Student: {student_label}",
        f"Status: {status_label}",
        f"From: {request.GET.get('date_from') or 'Any'}",
        f"To: {request.GET.get('date_to') or 'Any'}",
    ]

    if file_format == "pdf":
        content = build_attendance_pdf(records, institute, filters, include_notes=include_notes)
        response = HttpResponse(content, content_type="application/pdf")
    else:
        content = build_attendance_excel(records, institute, filters, include_notes=include_notes)
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file_format = "excel"
    response["Content-Disposition"] = f'attachment; filename="{attendance_export_filename(file_format)}"'
    return response


@teacher_required
def homework(request):
    batches = teacher_batches(request)
    homework_qs = (
        Homework.objects.filter(batch__in=batches)
        .select_related("batch", "subject", "course", "created_by")
        .prefetch_related("attachments")
    )
    search_query = request.GET.get("search", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    subject_filter = request.GET.get("subject", "").strip()
    course_filter = request.GET.get("course", "").strip()

    if search_query:
        homework_qs = homework_qs.filter(
            Q(title__icontains=search_query)
            | Q(instructions__icontains=search_query)
            | Q(batch__name__icontains=search_query)
            | Q(subject__name__icontains=search_query)
            | Q(course__name__icontains=search_query)
        )
    if batch_filter:
        homework_qs = homework_qs.filter(batch_id=batch_filter)
    if subject_filter:
        homework_qs = homework_qs.filter(subject_id=subject_filter)
    if course_filter:
        homework_qs = homework_qs.filter(course_id=course_filter)

    courses = Course.objects.filter(batches__in=batches, is_active=True).distinct()
    subjects = Subject.objects.filter(
        institute_id__in=batches.values_list("institute_id", flat=True).distinct(),
        academic_year_id__in=batches.values_list("academic_year_id", flat=True).distinct(),
        is_active=True,
    ).distinct()
    selected_year = teacher_selected_academic_year(request)
    if selected_year:
        courses = courses.filter(academic_year=selected_year)
        subjects = subjects.filter(academic_year=selected_year)
    base_queryset = Homework.objects.filter(batch__in=batches)
    context = {
        "homework_list": homework_qs,
        "batches": batches,
        "subjects": subjects,
        "courses": courses,
        "search_query": search_query,
        "batch_filter": batch_filter,
        "subject_filter": subject_filter,
        "course_filter": course_filter,
        "today": date.today(),
        "total_homework": base_queryset.count(),
        "due_homework": base_queryset.filter(due_date__gte=date.today()).count(),
        "expired_homework": base_queryset.filter(due_date__lt=date.today()).count(),
        "subject_count": subjects.count(),
    }
    return render(request, "teacher/homework.html", context)


@teacher_required
def homework_create(request):
    batches = teacher_batches(request)
    form = TeacherHomeworkForm(request.POST or None, request.FILES or None, batches=batches)
    if request.method == "POST" and form.is_valid():
        homework = form.save(commit=False)
        homework.created_by = request.user
        homework.save()
        form.save_attachments(homework)
        messages.success(request, "Homework created successfully.")
        return close_teacher_popup_response("/teacher/homework/")
    return render(
        request,
        "teacher/homework_form.html",
        {
            "form": form,
            "title": "Create Homework",
            "subtitle": "Assign work to a batch with separate subject and course details.",
            "button_text": "Save Homework",
            "batch_course_data": get_teacher_batch_course_data(batches),
            "attachments": [],
        },
    )


@teacher_required
def homework_update(request, pk):
    batches = teacher_batches(request)
    homework_item = get_object_or_404(
        Homework.objects.filter(batch__in=batches).select_related("batch", "subject", "course").prefetch_related("attachments"),
        pk=pk,
    )
    form = TeacherHomeworkForm(
        request.POST or None,
        request.FILES or None,
        batches=batches,
        instance=homework_item,
    )
    if request.method == "POST" and form.is_valid():
        homework_item = form.save()
        form.save_attachments(homework_item)
        messages.success(request, "Homework updated successfully.")
        return close_teacher_popup_response("/teacher/homework/")
    return render(
        request,
        "teacher/homework_form.html",
        {
            "form": form,
            "title": "Edit Homework",
            "subtitle": "Update batch, subject, course, instructions and due date.",
            "button_text": "Update Homework",
            "batch_course_data": get_teacher_batch_course_data(batches),
            "attachments": homework_item.attachments.all(),
        },
    )


@teacher_required
def homework_delete(request, pk):
    homework_item = get_object_or_404(Homework.objects.filter(batch__in=teacher_batches(request)), pk=pk)
    if request.method == "POST":
        homework_item.delete()
        messages.success(request, "Homework deleted successfully.")
    return redirect("teacher:homework")


@teacher_required
def exams(request):
    batches = teacher_batches(request)
    selected_year = teacher_selected_academic_year(request)
    exam_qs = (
        Exam.objects.filter(batch__in=batches)
        .select_related("batch", "academic_year", "subject")
        .annotate(question_count=Count("questions", distinct=True), marks_from_questions=Sum("questions__marks"))
    )
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    if search_query:
        exam_qs = exam_qs.filter(Q(title__icontains=search_query) | Q(batch__name__icontains=search_query) | Q(subject__name__icontains=search_query))
    if status_filter == "published":
        exam_qs = exam_qs.filter(is_published=True)
    elif status_filter == "draft":
        exam_qs = exam_qs.filter(is_published=False)
    if batch_filter:
        exam_qs = exam_qs.filter(batch_id=batch_filter)
    base_qs = Exam.objects.filter(batch__in=batches)
    return render(
        request,
        "teacher/exams.html",
        {
            "exams": exam_qs,
            "batches": batches,
            "selected_academic_year": selected_year,
            "search_query": search_query,
            "status_filter": status_filter,
            "batch_filter": batch_filter,
            "total_exams": base_qs.count(),
            "published_exams": base_qs.filter(is_published=True).count(),
            "draft_exams": base_qs.filter(is_published=False).count(),
            "total_attempts": ExamAttempt.objects.filter(exam__in=base_qs, submitted_at__isnull=False).count(),
        },
    )


@teacher_required
def exam_create(request):
    batches = teacher_batches(request)
    form = TeacherExamForm(request.POST or None, batches=batches)
    if request.method == "POST" and form.is_valid():
        exam = form.save(commit=False)
        exam.academic_year = exam.batch.academic_year
        exam.created_by = request.user
        exam.total_marks = 0
        exam.save()
        messages.success(request, "Exam created successfully.")
        return close_teacher_popup_response()
    return render(
        request,
        "teacher/exam_form.html",
        {
            "form": form,
            "title": "Create Exam",
            "subtitle": "Set exam details for the selected academic year.",
            "button_text": "Save Exam",
        },
    )


@teacher_required
def exam_update(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    form = TeacherExamForm(request.POST or None, instance=exam, batches=teacher_batches(request))
    if request.method == "POST" and form.is_valid():
        exam = form.save(commit=False)
        exam.academic_year = exam.batch.academic_year
        exam.save()
        messages.success(request, "Exam updated successfully.")
        return close_teacher_popup_response()
    return render(
        request,
        "teacher/exam_form.html",
        {
            "form": form,
            "title": "Edit Exam",
            "subtitle": "Update exam schedule, instructions and publish settings.",
            "button_text": "Update Exam",
        },
    )


def teacher_exam_queryset(request):
    return Exam.objects.filter(batch__in=teacher_batches(request)).select_related("batch", "academic_year", "subject")


def sync_exam_total_marks(exam):
    total = exam.questions.aggregate(total=Sum("marks")).get("total") or 0
    Exam.objects.filter(pk=exam.pk).update(total_marks=total)
    exam.total_marks = total


def recalculate_exam_attempt(attempt):
    questions = list(attempt.exam.questions.prefetch_related("options"))
    answer_map = {
        row.question_id: row
        for row in attempt.question_attempts.select_related("selected_option")
    }
    score = 0
    correct_count = 0
    wrong_count = 0
    unattempted_count = 0
    total_marks = sum(question.marks for question in questions)
    for question in questions:
        answer = answer_map.get(question.pk)
        selected_option = answer.selected_option if answer else None
        if selected_option and selected_option.question_id != question.pk:
            selected_option = None
        is_correct = bool(selected_option and selected_option.is_correct)
        marks_awarded = question.marks if is_correct else 0
        if selected_option is None:
            unattempted_count += 1
        elif is_correct:
            correct_count += 1
            score += question.marks
        else:
            wrong_count += 1
        ExamQuestionAttempt.objects.update_or_create(
            attempt=attempt,
            question=question,
            defaults={
                "selected_option": selected_option,
                "is_correct": is_correct,
                "marks_awarded": marks_awarded,
            },
        )
    attempt.score = score
    attempt.total_marks = total_marks
    attempt.correct_count = correct_count
    attempt.wrong_count = wrong_count
    attempt.unattempted_count = unattempted_count
    attempt.save(
        update_fields=[
            "score",
            "total_marks",
            "correct_count",
            "wrong_count",
            "unattempted_count",
        ]
    )
    if attempt.is_submitted:
        ExamResult.objects.update_or_create(
            exam=attempt.exam,
            student=attempt.student,
            defaults={
                "marks_obtained": score,
                "remark": "Updated by teacher from exam submission management.",
            },
        )
    return attempt


def option_formset_has_one_correct(formset):
    option_rows = [
        data
        for data in formset.cleaned_data
        if data and not data.get("DELETE") and data.get("text")
    ]
    return len(option_rows) == 4 and sum(1 for data in option_rows if data.get("is_correct")) == 1


@teacher_required
def exam_questions(request, pk):
    exam = get_object_or_404(
        teacher_exam_queryset(request).prefetch_related("questions", "questions__options"),
        pk=pk,
    )
    return render(request, "teacher/exam_questions.html", {"exam": exam})


@teacher_required
def exam_question_create(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    question = ExamQuestion(exam=exam, order=exam.questions.count() + 1)
    form = ExamQuestionForm(request.POST or None, request.FILES or None, instance=question)
    formset = ExamQuestionOptionFormSet(request.POST or None, instance=question, prefix="options")
    if request.method == "POST" and form.is_valid() and formset.is_valid():
        if not option_formset_has_one_correct(formset):
            messages.error(request, "Add exactly four options and select exactly one correct option.")
        else:
            question = form.save(commit=False)
            question.exam = exam
            question.order = exam.questions.count() + 1
            question.save()
            formset.instance = question
            options = formset.save(commit=False)
            for index, option in enumerate(options, start=1):
                option.question = question
                option.order = index
                option.save()
            for deleted in formset.deleted_objects:
                deleted.delete()
            sync_exam_total_marks(exam)
            messages.success(request, "Question added successfully.")
            return close_teacher_popup_response(reverse("teacher:exam_questions", args=[exam.pk]))
    return render(
        request,
        "teacher/exam_question_form.html",
        {"exam": exam, "form": form, "formset": formset, "title": "Add Question"},
    )


@teacher_required
def exam_question_update(request, exam_pk, question_pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=exam_pk)
    question = get_object_or_404(ExamQuestion.objects.filter(exam=exam), pk=question_pk)
    form = ExamQuestionForm(request.POST or None, request.FILES or None, instance=question)
    formset = ExamQuestionOptionFormSet(request.POST or None, instance=question, prefix="options")
    if request.method == "POST" and form.is_valid() and formset.is_valid():
        if not option_formset_has_one_correct(formset):
            messages.error(request, "Add exactly four options and select exactly one correct option.")
        else:
            question = form.save()
            options = formset.save(commit=False)
            for deleted in formset.deleted_objects:
                deleted.delete()
            for index, option in enumerate(options, start=1):
                option.question = question
                option.order = index
                option.save()
            sync_exam_total_marks(exam)
            messages.success(request, "Question updated successfully.")
            return close_teacher_popup_response(reverse("teacher:exam_questions", args=[exam.pk]))
    return render(
        request,
        "teacher/exam_question_form.html",
        {"exam": exam, "form": form, "formset": formset, "title": "Edit Question"},
    )


def normalize_bulk_cell(value):
    if value is None:
        return ""
    return str(value).strip()


@teacher_required
def exam_question_import_template(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Questions"
    worksheet.append(BULK_QUESTION_HEADERS)

    sample_rows = [
        [
            "A body is moving with uniform velocity. Its acceleration is?",
            "Zero",
            "Positive",
            "Negative",
            "Variable",
            "A",
            1,
        ],
        [
            "If x + 5 = 12, then x equals?",
            "5",
            "6",
            "7",
            "8",
            "C",
            1,
        ],
        [
            "Which option best represents the SI unit of force?",
            "Joule",
            "Newton",
            "Watt",
            "Pascal",
            "B",
            1,
        ],
    ]
    for row in sample_rows:
        worksheet.append(row)

    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_index, width in enumerate([48, 28, 28, 28, 28, 18, 12], start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    answer_validation = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
    answer_validation.error = "Select only A, B, C, or D as the correct answer."
    answer_validation.errorTitle = "Invalid answer"
    answer_validation.prompt = "Choose the correct option."
    answer_validation.promptTitle = "Correct Answer"
    worksheet.add_data_validation(answer_validation)
    answer_validation.add("F2:F500")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"bulk-question-template-{exam.pk}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@teacher_required
def exam_question_bulk_import(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    if request.method != "POST":
        return redirect("teacher:exam_questions", pk=exam.pk)

    upload = request.FILES.get("question_file")
    if not upload:
        messages.error(request, "Please select the completed question template.")
        return redirect("teacher:exam_questions", pk=exam.pk)
    if not upload.name.lower().endswith(".xlsx"):
        messages.error(request, "Please upload the Excel template in .xlsx format.")
        return redirect("teacher:exam_questions", pk=exam.pk)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded Excel file. Please download the template and try again.")
        return redirect("teacher:exam_questions", pk=exam.pk)

    worksheet = workbook.active
    imported_rows = []
    errors = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [normalize_bulk_cell(value) for value in row[:7]]
        values += [""] * (7 - len(values))
        question_text, option_a, option_b, option_c, option_d, correct_answer, marks_value = values

        if not any(values):
            continue

        row_errors = []
        if not question_text:
            row_errors.append("question text is required")
        options = [option_a, option_b, option_c, option_d]
        if any(not option for option in options):
            row_errors.append("all four options are required")
        correct_answer = correct_answer.upper()
        if correct_answer not in BULK_ANSWER_TO_ORDER:
            row_errors.append("correct answer must be A, B, C, or D")
        try:
            marks = int(float(marks_value or 1))
        except (TypeError, ValueError):
            marks = 0
        if marks < 1:
            row_errors.append("marks must be 1 or higher")

        if row_errors:
            errors.append(f"Row {row_number}: {', '.join(row_errors)}.")
            continue

        imported_rows.append(
            {
                "text": question_text,
                "options": options,
                "correct_order": BULK_ANSWER_TO_ORDER[correct_answer],
                "marks": marks,
            }
        )

    if errors:
        messages.error(request, "Bulk import failed. Fix these rows and upload again: " + " ".join(errors[:8]))
        if len(errors) > 8:
            messages.error(request, f"{len(errors) - 8} more row(s) also need correction.")
        return redirect("teacher:exam_questions", pk=exam.pk)
    if not imported_rows:
        messages.error(request, "No question rows found in the uploaded template.")
        return redirect("teacher:exam_questions", pk=exam.pk)

    start_order = exam.questions.count() + 1
    with transaction.atomic():
        for offset, row in enumerate(imported_rows):
            question = ExamQuestion.objects.create(
                exam=exam,
                text=row["text"],
                marks=row["marks"],
                order=start_order + offset,
            )
            ExamQuestionOption.objects.bulk_create(
                [
                    ExamQuestionOption(
                        question=question,
                        text=option_text,
                        order=option_order,
                        is_correct=option_order == row["correct_order"],
                    )
                    for option_order, option_text in enumerate(row["options"], start=1)
                ]
            )
        sync_exam_total_marks(exam)

    messages.success(request, f"{len(imported_rows)} question(s) imported successfully.")
    return redirect("teacher:exam_questions", pk=exam.pk)


@teacher_required
def exam_submissions(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    enrollments = (
        StudentEnrollment.objects.filter(
            batch=exam.batch,
            academic_session__academic_year=exam.academic_year,
            academic_session__status=StudentAcademicSession.Status.ACTIVE,
            status=StudentEnrollment.Status.ACTIVE,
            student__is_active=True,
        )
        .select_related("academic_session", "student", "student__user")
        .order_by("academic_session__admission_number", "student__user__first_name", "student__user__username")
    )
    attempts = list(
        ExamAttempt.objects.filter(exam=exam)
        .select_related("student", "student__user", "academic_session")
        .prefetch_related("uploads")
    )
    attempts_by_session = {attempt.academic_session_id: attempt for attempt in attempts}
    rows = []
    seen_session_ids = set()

    for enrollment in enrollments:
        attempt = attempts_by_session.get(enrollment.academic_session_id)
        seen_session_ids.add(enrollment.academic_session_id)
        rows.append(
            {
                "academic_session": enrollment.academic_session,
                "student": enrollment.student,
                "attempt": attempt,
                "upload_count": len(attempt.uploads.all()) if attempt else 0,
                "status": "Attempted" if attempt and attempt.is_submitted else "In Progress" if attempt else "Not Attempted",
            }
        )

    for attempt in attempts:
        if attempt.academic_session_id in seen_session_ids:
            continue
        rows.append(
            {
                "academic_session": attempt.academic_session,
                "student": attempt.student,
                "attempt": attempt,
                "upload_count": len(attempt.uploads.all()),
                "status": "Attempted" if attempt.is_submitted else "In Progress",
            }
        )

    attempted_count = sum(1 for row in rows if row["attempt"] and row["attempt"].is_submitted)
    in_progress_count = sum(1 for row in rows if row["attempt"] and not row["attempt"].is_submitted)
    not_attempted_count = sum(1 for row in rows if not row["attempt"])
    return render(
        request,
        "teacher/exam_submissions.html",
        {
            "exam": exam,
            "rows": rows,
            "total_students": len(rows),
            "attempted_count": attempted_count,
            "in_progress_count": in_progress_count,
            "not_attempted_count": not_attempted_count,
        },
    )


@teacher_required
def exam_attempt_manage(request, exam_pk, attempt_pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=exam_pk)
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related("student", "student__user", "academic_session", "exam")
        .prefetch_related("question_attempts", "uploads", "activities"),
        pk=attempt_pk,
        exam=exam,
    )
    questions = list(exam.questions.prefetch_related("options"))

    if request.method == "POST":
        action = request.POST.get("action", "save_answers")
        with transaction.atomic():
            for question in questions:
                selected_option = None
                option_id = request.POST.get(f"answer_{question.pk}")
                if option_id:
                    selected_option = question.options.filter(pk=option_id).first()
                ExamQuestionAttempt.objects.update_or_create(
                    attempt=attempt,
                    question=question,
                    defaults={"selected_option": selected_option},
                )
            if action == "force_submit" and not attempt.is_submitted:
                attempt.submitted_at = timezone.now()
                attempt.save(update_fields=["submitted_at"])
            recalculate_exam_attempt(attempt)
        if action == "force_submit":
            messages.success(request, "Attempt answers saved and marked as submitted.")
        else:
            messages.success(request, "Student answers updated and score recalculated.")
        return redirect("teacher:exam_attempt_manage", exam_pk=exam.pk, attempt_pk=attempt.pk)

    recalculate_exam_attempt(attempt)
    attempts_by_question = {
        answer.question_id: answer
        for answer in attempt.question_attempts.select_related("selected_option")
    }
    question_rows = [
        {
            "question": question,
            "answer": attempts_by_question.get(question.pk),
            "uploads": [upload for upload in attempt.uploads.all() if upload.question_id == question.pk],
        }
        for question in questions
    ]
    unlinked_uploads = [upload for upload in attempt.uploads.all() if not upload.question_id]
    return render(
        request,
        "teacher/exam_attempt_manage.html",
        {
            "exam": exam,
            "attempt": attempt,
            "question_rows": question_rows,
            "unlinked_uploads": unlinked_uploads,
        },
    )


@teacher_required
@require_POST
def exam_attempt_reset(request, exam_pk, attempt_pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=exam_pk)
    attempt = get_object_or_404(ExamAttempt.objects.select_related("student"), pk=attempt_pk, exam=exam)
    student_name = attempt.student.user.get_full_name() or attempt.student.user.username
    ExamResult.objects.filter(exam=exam, student=attempt.student).delete()
    attempt.delete()
    messages.success(request, f"Attempt reset for {student_name}. The student can attend this exam again.")
    return redirect("teacher:exam_submissions", pk=exam.pk)


@teacher_required
@require_POST
def exam_publish(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    if exam.is_published:
        messages.info(request, "Exam is already published to students.")
    else:
        exam.is_published = True
        exam.save(update_fields=["is_published"])
        messages.success(request, "Exam published successfully. Students can now view and attempt this exam.")
    return redirect("teacher:exam_submissions", pk=exam.pk)


@teacher_required
@require_POST
def exam_toggle_result_publish(request, pk):
    exam = get_object_or_404(teacher_exam_queryset(request), pk=pk)
    action = request.POST.get("action")
    if action == "publish":
        exam.show_result_after_submit = True
        exam.save(update_fields=["show_result_after_submit"])
        messages.success(request, "Exam results published successfully. Students can now view their scores.")
    elif action == "hide":
        exam.show_result_after_submit = False
        exam.save(update_fields=["show_result_after_submit"])
        messages.success(request, "Exam results hidden successfully. Students will not see scores until you publish again.")
    else:
        messages.error(request, "Invalid result visibility action.")
    return redirect("teacher:exam_submissions", pk=exam.pk)


@teacher_required
def results(request):
    return render(request, "teacher/results.html", get_teacher_result_report(request))


def teacher_attempt_percentage(attempt):
    if not attempt.total_marks:
        return Decimal("0.00")
    return (attempt.score / attempt.total_marks * Decimal("100")).quantize(Decimal("0.01"))


def teacher_attempt_ranks(attempts):
    ranks = {}
    grouped = {}
    for attempt in attempts:
        grouped.setdefault(attempt.exam_id, []).append(attempt)
    for exam_attempts in grouped.values():
        previous_score = None
        rank = 0
        for position, attempt in enumerate(
            sorted(exam_attempts, key=lambda item: (-item.score, item.submitted_at, item.pk)),
            start=1,
        ):
            if previous_score is None or attempt.score != previous_score:
                rank = position
                previous_score = attempt.score
            ranks[attempt.pk] = rank
    return ranks


def get_teacher_result_report(request):
    profile = request.user.profile
    academic_year = teacher_selected_academic_year(request)
    batches = teacher_batches(request).order_by("name")
    exams = Exam.objects.filter(batch__in=batches).select_related(
        "batch", "course", "subject", "academic_year"
    )
    all_attempts = list(
        ExamAttempt.objects.filter(exam__in=exams, submitted_at__isnull=False)
        .select_related(
            "exam",
            "exam__batch",
            "exam__course",
            "exam__subject",
            "academic_session",
            "student",
            "student__user",
        )
        .order_by("-submitted_at")
    )
    ranks = teacher_attempt_ranks(all_attempts)
    selected = {
        "batch": request.GET.get("batch", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "subject": request.GET.get("subject", "").strip(),
        "exam": request.GET.get("exam", "").strip(),
        "student": request.GET.get("student", "").strip(),
        "date_from": request.GET.get("date_from", "").strip(),
        "date_to": request.GET.get("date_to", "").strip(),
        "performance": request.GET.get("performance", "").strip(),
        "min_percentage": request.GET.get("min_percentage", "").strip(),
        "max_percentage": request.GET.get("max_percentage", "").strip(),
    }
    date_from = parse_date(selected["date_from"])
    date_to = parse_date(selected["date_to"])
    try:
        minimum = Decimal(selected["min_percentage"]) if selected["min_percentage"] else None
    except (ValueError, ArithmeticError):
        minimum = None
    try:
        maximum = Decimal(selected["max_percentage"]) if selected["max_percentage"] else None
    except (ValueError, ArithmeticError):
        maximum = None

    rows = []
    for attempt in all_attempts:
        percentage = teacher_attempt_percentage(attempt)
        student_name = attempt.student.user.get_full_name() or attempt.student.user.username
        search_text = f"{student_name} {attempt.student.user.username} {attempt.academic_session.admission_number}"
        submitted_date = timezone.localtime(attempt.submitted_at).date()
        if selected["batch"] and str(attempt.exam.batch_id) != selected["batch"]:
            continue
        if selected["course"] and str(attempt.exam.course_id or "") != selected["course"]:
            continue
        if selected["subject"] and str(attempt.exam.subject_id or "") != selected["subject"]:
            continue
        if selected["exam"] and str(attempt.exam_id) != selected["exam"]:
            continue
        if selected["student"] and selected["student"].lower() not in search_text.lower():
            continue
        if date_from and submitted_date < date_from:
            continue
        if date_to and submitted_date > date_to:
            continue
        if minimum is not None and percentage < minimum:
            continue
        if maximum is not None and percentage > maximum:
            continue
        if selected["performance"] == "passed" and percentage < Decimal("40"):
            continue
        if selected["performance"] == "failed" and percentage >= Decimal("40"):
            continue
        rows.append(
            {
                "attempt": attempt,
                "student_name": student_name,
                "percentage": percentage,
                "rank": ranks[attempt.pk],
                "performance": "Passed" if percentage >= Decimal("40") else "Needs improvement",
            }
        )

    rows.sort(key=lambda row: (row["attempt"].exam.title.lower(), row["rank"], row["student_name"].lower()))
    percentages = [row["percentage"] for row in rows]
    passed_count = sum(1 for value in percentages if value >= Decimal("40"))
    courses = Course.objects.filter(
        institute=profile.institute,
        academic_year=academic_year,
        exams__in=exams,
    ).distinct().order_by("name")
    subjects = Subject.objects.filter(
        institute=profile.institute,
        academic_year=academic_year,
        exams__in=exams,
    ).distinct().order_by("name")
    filters = teacher_result_filter_labels(selected, batches, courses, subjects, exams)
    average = sum(percentages, Decimal("0.00")) / len(percentages) if percentages else Decimal("0.00")
    return {
        "institute": profile.institute,
        "academic_year": academic_year,
        "rows": rows,
        "batches": batches,
        "courses": courses,
        "subjects": subjects,
        "exams": exams.order_by("-exam_date", "title"),
        "filters": selected,
        "filter_labels": filters,
        "result_count": len(rows),
        "passed_count": passed_count,
        "needs_improvement_count": len(rows) - passed_count,
        "average_percentage": average.quantize(Decimal("0.01")),
        "highest_percentage": max(percentages, default=Decimal("0.00")),
    }


def teacher_result_filter_labels(selected, batches, courses, subjects, exams):
    def label(queryset, selected_id):
        item = queryset.filter(pk=selected_id).first() if selected_id else None
        return str(item) if item else "All"
    performance = {
        "passed": "Passed (40% and above)",
        "failed": "Needs improvement (below 40%)",
    }
    return [
        f"Batch: {label(batches, selected['batch'])}",
        f"Course: {label(courses, selected['course'])}",
        f"Subject: {label(subjects, selected['subject'])}",
        f"Exam: {label(exams, selected['exam'])}",
        f"Student: {selected['student'] or 'All'}",
        f"Submitted: {selected['date_from'] or 'Any'} to {selected['date_to'] or 'Any'}",
        f"Percentage: {selected['min_percentage'] or '0'} to {selected['max_percentage'] or '100'}",
        f"Performance: {performance.get(selected['performance'], 'All')}",
    ]


def build_teacher_results_excel(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exam Results"
    columns = [
        "Rank", "Exam", "Exam Date", "Student", "Admission No.", "Batch", "Course",
        "Subject", "Score", "Total Marks", "Percentage", "Correct", "Wrong",
        "Unattempted", "Performance", "Submitted At",
    ]
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title = sheet.cell(1, 1, "Teacher Exam Results Report")
    title.fill = PatternFill("solid", fgColor="7C3AED")
    title.font = Font(size=18, bold=True, color="FFFFFF")
    title.alignment = center
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sheet.cell(
        2, 1,
        f"{report['institute'].name} | Academic Year: {report['academic_year']} | Generated {timezone.localtime().strftime('%d-%m-%Y %I:%M %p')}",
    ).alignment = center
    summary = [
        ("Results", report["result_count"]), ("Passed", report["passed_count"]),
        ("Needs Improvement", report["needs_improvement_count"]),
        ("Average", f"{report['average_percentage']}%"), ("Highest", f"{report['highest_percentage']}%"),
    ]
    for index, (label, value) in enumerate(summary, start=1):
        sheet.cell(4, index, label).font = Font(bold=True)
        sheet.cell(4, index).fill = PatternFill("solid", fgColor="EDE9FE")
        sheet.cell(5, index, value).font = Font(size=13, bold=True)
        sheet.cell(4, index).alignment = sheet.cell(5, index).alignment = center
    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(columns))
    sheet.cell(7, 1, "Filters: " + " | ".join(report["filter_labels"])).font = Font(italic=True, color="64748B")
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(9, index, column)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border
    for row_index, row in enumerate(report["rows"], start=10):
        attempt = row["attempt"]
        values = [
            row["rank"], attempt.exam.title, attempt.exam.exam_date.strftime("%d-%m-%Y"),
            row["student_name"], attempt.academic_session.admission_number, attempt.exam.batch.name,
            attempt.exam.course.name if attempt.exam.course else "",
            attempt.exam.subject.name if attempt.exam.subject else "",
            float(attempt.score), float(attempt.total_marks), float(row["percentage"]),
            attempt.correct_count, attempt.wrong_count, attempt.unattempted_count,
            row["performance"], timezone.localtime(attempt.submitted_at).strftime("%d-%m-%Y %I:%M %p"),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, col_index, value)
            cell.border = border
            cell.alignment = center
    widths = [8, 28, 13, 25, 18, 20, 20, 20, 11, 12, 12, 10, 10, 13, 20, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:P{max(10, 9 + len(report['rows']))}"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_teacher_results_pdf(report):
    widths = [34, 112, 132, 86, 92, 60, 44, 28, 28, 28, 92]
    headers = ["Rank", "Exam", "Student", "Adm No", "Batch", "Score", "%", "C", "W", "U", "Status"]
    chunks = [report["rows"][i:i + 16] for i in range(0, len(report["rows"]), 16)] or [[]]
    pages = []
    for page_number, chunk in enumerate(chunks, start=1):
        stream = pdf_rect(0, 545, 842, 50, "0.486 0.227 0.929")
        stream += pdf_text(30, 570, "Teacher Exam Results Report", 18, "F2", "1 1 1")
        stream += pdf_text(30, 552, f"{report['institute'].name} | Academic Year: {report['academic_year']}", 8, "F1", "1 1 1")
        stream += pdf_text(730, 552, f"Page {page_number}/{len(chunks)}", 8, "F1", "1 1 1")
        summary = [
            ("Results", report["result_count"]), ("Passed", report["passed_count"]),
            ("Needs Improvement", report["needs_improvement_count"]),
            ("Average", f"{report['average_percentage']}%"), ("Highest", f"{report['highest_percentage']}%"),
        ]
        x = 30
        for label, value in summary:
            stream += pdf_rect(x, 502, 145, 31, "0.95 0.94 1")
            stream += pdf_text(x + 7, 520, label, 7, "F1", "0.39 0.45 0.55")
            stream += pdf_text(x + 7, 507, value, 11, "F2", "0.06 0.09 0.16")
            x += 155
        stream += pdf_text(30, 481, ("Filters: " + " | ".join(report["filter_labels"]))[:150], 7, "F1", "0.39 0.45 0.55")
        y, start_x = 452, 30
        stream += pdf_rect(start_x, y - 6, sum(widths), 22, "0.06 0.09 0.16")
        x = start_x
        for header, width in zip(headers, widths):
            stream += pdf_text(x + 3, y, header, 7, "F2", "1 1 1")
            x += width
        y -= 24
        for row in chunk:
            attempt = row["attempt"]
            values = [
                row["rank"], attempt.exam.title[:17], row["student_name"][:20],
                attempt.academic_session.admission_number[:13], attempt.exam.batch.name[:14],
                f"{attempt.score}/{attempt.total_marks}", row["percentage"], attempt.correct_count,
                attempt.wrong_count, attempt.unattempted_count, row["performance"][:15],
            ]
            stream += pdf_rect(start_x, y - 5, sum(widths), 20, "0.98 0.99 1")
            x = start_x
            for value, width in zip(values, widths):
                stream += pdf_text(x + 3, y, value, 6.5, "F1", "0.06 0.09 0.16")
                x += width
            y -= 21
        stream += pdf_text(30, 28, "C = Correct | W = Wrong | U = Unattempted | Rank is within each exam", 8, "F1", "0.39 0.45 0.55")
        pages.append(stream)
    return build_pdf_document(pages, width=842, height=595)


@teacher_required
def results_export(request):
    report = get_teacher_result_report(request)
    file_format = request.GET.get("format", "excel").lower()
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    if file_format == "pdf":
        response = HttpResponse(build_teacher_results_pdf(report), content_type="application/pdf")
        extension = "pdf"
    else:
        response = HttpResponse(
            build_teacher_results_excel(report),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        extension = "xlsx"
    response["Content-Disposition"] = f'attachment; filename="teacher_exam_results_{stamp}.{extension}"'
    return response
