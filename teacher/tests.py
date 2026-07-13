import json
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from institute_admin.models import AcademicYear, Batch, Course, Notice, NoticeRead, Subject
from student_parent.models import StudentAcademicSession, StudentEnrollment, StudentProfile
from super_admin.mobile_auth import create_access_token
from super_admin.models import Institute, UserProfile
from teacher.forms import TeacherExamForm, TeacherHomeworkForm
from teacher.models import Attendance, Exam, ExamAttempt, ExamAttemptUpload, ExamQuestion, ExamQuestionAttempt, ExamQuestionOption, ExamResult, Homework


class TeacherMobileReadApiTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(name="Mobile Institute", code="mobile")
        self.year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date="2026-04-01",
            end_date="2027-03-31",
        )
        self.teacher_user = User.objects.create_user(username="mobile-teacher", password="pass")
        UserProfile.objects.create(
            user=self.teacher_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        self.student_user = User.objects.create_user(
            username="mobile-student",
            password="pass",
            first_name="Mobile",
            last_name="Student",
        )
        UserProfile.objects.create(
            user=self.student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        self.student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=self.student_user,
            admission_number="M001",
        )
        self.session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year,
            admission_number="M001",
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Science",
        )
        self.subject = Subject.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Physics",
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Assigned Batch",
        )
        self.batch.courses.add(self.course)
        self.batch.teachers.add(self.teacher_user)
        enrollment = StudentEnrollment.objects.create(
            student=self.student,
            academic_session=self.session,
            batch=self.batch,
        )
        enrollment.courses.add(self.course)

        self.unassigned_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Private Batch",
        )
        self.unassigned_batch.courses.add(self.course)
        self.unassigned_homework = Homework.objects.create(
            batch=self.unassigned_batch,
            course=self.course,
            subject=self.subject,
            title="Private Homework",
            created_by=self.teacher_user,
        )
        self.unassigned_exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.unassigned_batch,
            course=self.course,
            subject=self.subject,
            title="Private Exam",
            exam_date="2026-07-02",
            is_published=True,
        )

        self.homework = Homework.objects.create(
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Assigned Homework",
            due_date="2026-06-20",
            created_by=self.teacher_user,
        )
        self.attendance = Attendance.objects.create(
            student=self.student,
            academic_session=self.session,
            batch=self.batch,
            date=timezone.localdate(),
            status=Attendance.Status.PRESENT,
            marked_by=self.teacher_user,
        )
        self.notice = Notice.objects.create(
            institute=self.institute,
            academic_year=self.year,
            title="Assigned Notice",
            message="Visible",
            audience=Notice.Audience.TEACHERS,
            push_to_app=True,
        )
        self.notice.target_batches.add(self.batch)
        self.hidden_notice = Notice.objects.create(
            institute=self.institute,
            academic_year=self.year,
            title="Hidden Notice",
            message="Hidden",
            audience=Notice.Audience.TEACHERS,
            push_to_app=True,
        )
        self.hidden_notice.target_batches.add(self.unassigned_batch)
        self.exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Assigned Exam",
            exam_date="2026-07-01",
            is_published=True,
        )
        self.question = ExamQuestion.objects.create(
            exam=self.exam,
            text="2 + 2?",
            marks=2,
            order=1,
        )
        ExamQuestionOption.objects.create(question=self.question, text="4", is_correct=True, order=1)
        self.attempt = ExamAttempt.objects.create(
            exam=self.exam,
            academic_session=self.session,
            student=self.student,
            submitted_at=timezone.now(),
            score=2,
            total_marks=2,
            correct_count=1,
        )

    def auth_headers(self, user=None):
        return {"HTTP_AUTHORIZATION": f"Bearer {create_access_token(user or self.teacher_user)}"}

    def get_json(self, path):
        response = self.client.get(path, **self.auth_headers())
        self.assertEqual(response.status_code, 200, response.content)
        return response.json()

    def post_json(self, path, payload=None):
        return self.client.post(
            path,
            data=json.dumps(payload or {}),
            content_type="application/json",
            **self.auth_headers(),
        )

    def put_json(self, path, payload=None):
        return self.client.put(
            path,
            data=json.dumps(payload or {}),
            content_type="application/json",
            **self.auth_headers(),
        )

    def patch_json(self, path, payload=None):
        return self.client.patch(
            path,
            data=json.dumps(payload or {}),
            content_type="application/json",
            **self.auth_headers(),
        )

    def result_titles(self, payload, key="title"):
        return [item[key] for item in payload["results"]]

    def test_phase_2_dashboard_returns_real_teacher_data(self):
        payload = self.get_json("/api/mobile/teacher/dashboard/")

        self.assertTrue(payload["success"])
        self.assertEqual(payload["batch_count"], 1)
        self.assertEqual(payload["student_count"], 1)
        self.assertEqual(payload["homework_count"], 1)
        self.assertEqual(payload["exam_count"], 1)
        self.assertEqual(payload["today_attendance_count"], 1)
        self.assertEqual(payload["submitted_attempt_count"], 1)
        self.assertEqual(payload["pending_submission_count"], 0)
        self.assertEqual(payload["assigned_batches"][0]["name"], "Assigned Batch")
        self.assertEqual(payload["recent_homework"][0]["title"], "Assigned Homework")
        self.assertEqual(payload["recent_exams"][0]["title"], "Assigned Exam")
        self.assertEqual(payload["recent_results"][0]["exam_title"], "Assigned Exam")

    def test_dashboard_reports_pending_submissions(self):
        self.attempt.submitted_at = None
        self.attempt.save(update_fields=["submitted_at"])

        payload = self.get_json("/api/mobile/teacher/dashboard/")

        self.assertEqual(payload["submitted_attempt_count"], 0)
        self.assertEqual(payload["pending_submission_count"], 1)
        self.assertEqual(payload["recent_results"], [])

    def test_phase_2_classes_and_students_are_teacher_scoped(self):
        classes = self.get_json("/api/mobile/teacher/classes/")
        students = self.get_json(f"/api/mobile/teacher/classes/{self.batch.pk}/students/")

        self.assertEqual(self.result_titles(classes, key="name"), ["Assigned Batch"])
        self.assertEqual(students["results"][0]["name"], "Mobile Student")
        self.assertEqual(
            self.client.get(
                f"/api/mobile/teacher/classes/{self.unassigned_batch.pk}/students/",
                **self.auth_headers(),
            ).status_code,
            404,
        )

    def test_phase_2_attendance_assignments_notices_are_teacher_scoped(self):
        attendance = self.get_json("/api/mobile/teacher/attendance/")
        assignments = self.get_json("/api/mobile/teacher/assignments/")
        notices = self.get_json("/api/mobile/teacher/notices/")

        self.assertEqual(attendance["total"], 1)
        self.assertEqual(attendance["rows"][0]["student_name"], "Mobile Student")
        self.assertEqual(self.result_titles(assignments), ["Assigned Homework"])
        self.assertNotIn("Private Homework", self.result_titles(assignments))
        self.assertEqual(self.result_titles(notices), ["Assigned Notice"])
        self.assertNotIn("Hidden Notice", self.result_titles(notices))

    def test_phase_2_exams_questions_submissions_results_are_teacher_scoped(self):
        exams = self.get_json("/api/mobile/teacher/exams/")
        questions = self.get_json("/api/mobile/teacher/questions/")
        submissions = self.get_json("/api/mobile/teacher/submissions/")
        results = self.get_json("/api/mobile/teacher/results/")

        self.assertEqual(self.result_titles(exams), ["Assigned Exam"])
        self.assertNotIn("Private Exam", self.result_titles(exams))
        self.assertEqual(questions["results"][0]["exam_title"], "Assigned Exam")
        self.assertEqual(submissions["results"][0]["exam_title"], "Assigned Exam")
        self.assertEqual(results["results"][0]["student_name"], "Mobile Student")

    def test_phase_2_read_apis_require_teacher_bearer_token(self):
        student_response = self.client.get(
            "/api/mobile/teacher/dashboard/",
            **self.auth_headers(self.student_user),
        )
        anonymous_response = self.client.get("/api/mobile/teacher/dashboard/")

        self.assertEqual(student_response.status_code, 403)
        self.assertFalse(student_response.json()["success"])
        self.assertEqual(anonymous_response.status_code, 401)
        self.assertFalse(anonymous_response.json()["success"])

    def test_phase_2_paginated_endpoints_expose_meta(self):
        payload = self.get_json("/api/mobile/teacher/classes/?page_size=1")

        self.assertEqual(payload["meta"]["page_size"], 1)
        self.assertEqual(payload["meta"]["count"], 1)

    def test_class_students_support_search_and_pagination(self):
        self.student.roll_number = "03"
        self.student.save(update_fields=["roll_number"])
        alpha_student, _ = self._create_extra_student("alpha-student", "Alpha", "Learner", "M101")
        beta_student, _ = self._create_extra_student("beta-student", "Beta", "Learner", "M102")
        alpha_student.roll_number = "01"
        beta_student.roll_number = "02"
        alpha_student.save(update_fields=["roll_number"])
        beta_student.save(update_fields=["roll_number"])

        page = self.get_json(f"/api/mobile/teacher/classes/{self.batch.pk}/students/?page_size=3")
        self.assertEqual(page["meta"]["page_size"], 3)
        self.assertEqual(page["meta"]["count"], 3)
        self.assertEqual([row["roll_number"] for row in page["results"]], ["01", "02", "03"])

        search = self.get_json(f"/api/mobile/teacher/classes/{self.batch.pk}/students/?search=Beta")
        self.assertEqual(search["meta"]["count"], 1)
        self.assertEqual(search["results"][0]["admission_number"], "M102")

    def test_classes_students_and_student_detail_support_filters_and_summary(self):
        classes = self.get_json(f"/api/mobile/teacher/classes/?search=Assigned&page_size=1&course_id={self.course.pk}")
        self.assertEqual(classes["meta"]["page_size"], 1)
        self.assertEqual(classes["meta"]["count"], 1)
        self.assertEqual(classes["results"][0]["id"], self.batch.pk)

        students = self.get_json(
            f"/api/mobile/teacher/classes/{self.batch.pk}/students/?course_id={self.course.pk}&search=M001"
        )
        self.assertEqual(students["meta"]["count"], 1)
        self.assertEqual(students["results"][0]["id"], self.session.pk)

        detail = self.get_json(f"/api/mobile/teacher/classes/{self.batch.pk}/students/{self.session.pk}/")
        self.assertEqual(detail["id"], self.session.pk)
        self.assertEqual(detail["summary"]["attendance_total"], 1)
        self.assertEqual(detail["summary"]["homework_count"], 1)
        self.assertEqual(detail["summary"]["exam_count"], 1)

        private = self.client.get(
            f"/api/mobile/teacher/classes/{self.unassigned_batch.pk}/students/{self.session.pk}/",
            **self.auth_headers(),
        )
        self.assertEqual(private.status_code, 404)

    def test_attendance_supports_search_status_and_pagination(self):
        self.student.roll_number = "03"
        self.student.save(update_fields=["roll_number"])
        alpha_student, alpha_session = self._create_extra_student("alpha-att", "Alpha", "Present", "M201")
        beta_student, beta_session = self._create_extra_student("beta-att", "Beta", "Absent", "M202")
        alpha_student.roll_number = "01"
        beta_student.roll_number = "02"
        alpha_student.save(update_fields=["roll_number"])
        beta_student.save(update_fields=["roll_number"])
        Attendance.objects.create(
            student=alpha_student,
            academic_session=alpha_session,
            batch=self.batch,
            date=timezone.localdate(),
            status=Attendance.Status.PRESENT,
            marked_by=self.teacher_user,
        )
        Attendance.objects.create(
            student=beta_student,
            academic_session=beta_session,
            batch=self.batch,
            date=timezone.localdate(),
            status=Attendance.Status.ABSENT,
            marked_by=self.teacher_user,
        )

        page = self.get_json("/api/mobile/teacher/attendance/?page_size=3")
        self.assertEqual(page["meta"]["page_size"], 3)
        self.assertEqual(page["meta"]["count"], 3)
        self.assertEqual([row["roll_number"] for row in page["rows"]], ["01", "02", "03"])

        absent = self.get_json("/api/mobile/teacher/attendance/?status=ABSENT")
        self.assertEqual(absent["total"], 1)
        self.assertEqual(absent["rows"][0]["student_name"], "Beta Absent")

        search = self.get_json("/api/mobile/teacher/attendance/?search=M201")
        self.assertEqual(search["meta"]["count"], 1)
        self.assertEqual(search["rows"][0]["student_name"], "Alpha Present")

    def test_reports_return_summaries_and_export_hooks(self):
        self.attendance.date = "2026-07-10"
        self.attendance.status = Attendance.Status.ABSENT
        self.attendance.save(update_fields=["date", "status"])
        self.student.roll_number = "02"
        self.student.save(update_fields=["roll_number"])
        no_attempt_student, _ = self._create_extra_student("no-attempt", "No", "Attempt", "M000")
        no_attempt_student.roll_number = "01"
        no_attempt_student.save(update_fields=["roll_number"])

        attendance = self.get_json(
            f"/api/mobile/teacher/reports/attendance/?class_id={self.batch.pk}&date_from=2026-07-01&date_to=2026-07-31"
        )
        self.assertEqual(attendance["report_type"], "attendance")
        self.assertEqual(attendance["summary"]["total"], 1)
        self.assertEqual(attendance["summary"]["absent"], 1)
        self.assertEqual(attendance["class_summaries"][0]["class_name"], "Assigned Batch")
        self.assertIn("format=pdf", attendance["export_hooks"]["pdf"]["url"])
        self.assertTrue(attendance["export_hooks"]["excel"]["requires_auth"])

        results = self.get_json(
            f"/api/mobile/teacher/reports/results/?class_id={self.batch.pk}&exam_id={self.exam.pk}&date_from=2026-07-01&date_to=2026-07-31"
        )
        self.assertEqual(results["report_type"], "results")
        self.assertEqual(results["summary"]["total"], 2)
        self.assertEqual(results["summary"]["submitted"], 1)
        self.assertEqual(results["summary"]["not_attempted"], 1)
        self.assertEqual(results["class_summaries"][0]["average_percentage"], 100)
        self.assertEqual(results["student_summaries"][0]["student_name"], "No Attempt")
        self.assertEqual(results["student_summaries"][0]["roll_number"], "01")
        self.assertEqual(results["student_summaries"][0]["status"], "not_attempted")
        self.assertEqual(results["student_summaries"][0]["attempted_questions"], 0)
        self.assertEqual(results["student_summaries"][0]["unattempted_questions"], 1)
        self.assertEqual(results["student_summaries"][0]["percentage"], 0)
        self.assertEqual(results["student_summaries"][1]["student_name"], "Mobile Student")
        self.assertEqual(results["student_summaries"][1]["attempted_questions"], 1)
        self.assertEqual(results["student_summaries"][1]["unattempted_questions"], 0)
        self.assertEqual(results["student_summaries"][1]["correct_count"], 1)
        self.assertEqual(results["student_summaries"][1]["wrong_count"], 0)
        self.assertEqual(results["student_summaries"][1]["percentage"], 100)
        self.assertIn("format=excel", results["export_hooks"]["excel"]["url"])

        attendance_pdf = self.client.get(
            "/api/mobile/teacher/reports/attendance/",
            {"class_id": self.batch.pk, "date_from": "2026-07-01", "date_to": "2026-07-31", "format": "pdf"},
            **self.auth_headers(),
        )
        self.assertEqual(attendance_pdf.status_code, 200)
        self.assertEqual(attendance_pdf["Content-Type"], "application/pdf")
        self.assertIn(b"Teacher Attendance Report", attendance_pdf.content)
        self.assertIn(b"Total Records", attendance_pdf.content)

        attendance_excel = self.client.get(
            "/api/mobile/teacher/reports/attendance/",
            {"class_id": self.batch.pk, "date_from": "2026-07-01", "date_to": "2026-07-31", "format": "excel"},
            **self.auth_headers(),
        )
        self.assertEqual(attendance_excel.status_code, 200)
        attendance_workbook = load_workbook(BytesIO(attendance_excel.content))
        self.assertIn("Summary", attendance_workbook.sheetnames)
        attendance_summary = attendance_workbook["Summary"]
        self.assertEqual(attendance_summary["A1"].value, "Teacher Attendance Report")
        self.assertEqual(attendance_summary["A4"].value, "Metric")
        self.assertEqual(attendance_summary["A5"].value, "Total")
        self.assertEqual(attendance_summary["B5"].value, 1)
        self.assertIn("Applied Filters", [cell.value for row in attendance_summary.iter_rows() for cell in row])

        results_excel = self.client.get(
            "/api/mobile/teacher/reports/results/",
            {"class_id": self.batch.pk, "exam_id": self.exam.pk, "date_from": "2026-07-01", "date_to": "2026-07-31", "format": "excel"},
            **self.auth_headers(),
        )
        self.assertEqual(results_excel.status_code, 200)
        self.assertIn("spreadsheetml.sheet", results_excel["Content-Type"])
        workbook = load_workbook(BytesIO(results_excel.content))
        self.assertIn("Student Summary", workbook.sheetnames)
        sheet = workbook["Student Summary"]
        self.assertEqual(sheet["A1"].value, "Roll Number")
        self.assertEqual(sheet["J1"].value, "Percentage")
        self.assertEqual(sheet["A2"].value, "01")
        self.assertEqual(sheet["C2"].value, "No Attempt")
        self.assertEqual(sheet["F2"].value, 0)
        self.assertEqual(sheet["G2"].value, 1)
        self.assertEqual(sheet["J2"].value, 0)
        self.assertEqual(sheet["A3"].value, "02")
        self.assertEqual(sheet["C3"].value, "Mobile Student")
        self.assertEqual(sheet["F3"].value, 1)
        self.assertEqual(sheet["G3"].value, 0)
        self.assertEqual(sheet["J3"].value, 100)

    def test_result_report_keeps_same_student_class_sessions_separate(self):
        next_year = AcademicYear.objects.create(
            institute=self.institute,
            name="2027-28",
            start_date="2027-04-01",
            end_date="2028-03-31",
        )
        next_course = Course.objects.create(
            institute=self.institute,
            academic_year=next_year,
            name="Science",
        )
        next_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=next_year,
            name="Assigned Batch",
        )
        next_batch.courses.add(next_course)
        next_batch.teachers.add(self.teacher_user)
        next_session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=next_year,
            admission_number="M001-27",
        )
        next_enrollment = StudentEnrollment.objects.create(
            student=self.student,
            academic_session=next_session,
            batch=next_batch,
        )
        next_exam = Exam.objects.create(
            academic_year=next_year,
            batch=next_batch,
            course=next_course,
            title="Next Session Exam",
            exam_date="2027-07-01",
            is_published=True,
        )
        ExamQuestion.objects.create(
            exam=next_exam,
            text="Next session question",
            marks=2,
            order=1,
        )
        ExamAttempt.objects.create(
            exam=next_exam,
            academic_session=next_session,
            student=self.student,
            submitted_at=timezone.now(),
            score=1,
            total_marks=2,
            correct_count=1,
        )

        # Simulate legacy/imported data where the same student appears in the same
        # class across sessions; report grouping must still keep sessions separate.
        StudentEnrollment.objects.filter(pk=next_enrollment.pk).update(batch=self.batch)
        Exam.objects.filter(pk=next_exam.pk).update(batch=self.batch)

        results = self.get_json(
            f"/api/mobile/teacher/reports/results/?class_id={self.batch.pk}&student_id={self.student.pk}"
        )

        self.assertEqual(results["summary"]["total"], 2)
        self.assertEqual(results["summary"]["submitted"], 2)
        rows_by_admission = {
            row["admission_number"]: row
            for row in results["student_summaries"]
        }
        self.assertEqual(set(rows_by_admission), {"M001", "M001-27"})
        self.assertEqual(rows_by_admission["M001"]["percentage"], 100)
        self.assertEqual(rows_by_admission["M001-27"]["percentage"], 50)

    def test_phase_3_attendance_post_creates_and_updates_rows(self):
        Attendance.objects.filter(pk=self.attendance.pk).delete()
        response = self.post_json(
            "/api/mobile/teacher/attendance/",
            {
                "class_id": self.batch.pk,
                "date": "2026-07-10",
                "rows": [
                    {
                        "academic_session_id": self.session.pk,
                        "status": Attendance.Status.ABSENT,
                        "note": "Sick",
                    }
                ],
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        record = Attendance.objects.get(batch=self.batch, academic_session=self.session, date="2026-07-10")
        self.assertEqual(record.status, Attendance.Status.ABSENT)
        self.assertEqual(record.note, "Sick")

        response = self.post_json(
            "/api/mobile/teacher/attendance/",
            {
                "class_id": self.batch.pk,
                "date": "2026-07-10",
                "rows": [
                    {
                        "academic_session_id": self.session.pk,
                        "status": Attendance.Status.PRESENT,
                        "note": "",
                    }
                ],
            },
        )
        self.assertEqual(response.status_code, 200, response.content)
        record.refresh_from_db()
        self.assertEqual(record.status, Attendance.Status.PRESENT)
        self.assertEqual(Attendance.objects.filter(batch=self.batch, academic_session=self.session, date="2026-07-10").count(), 1)

        missing_status = self.post_json(
            "/api/mobile/teacher/attendance/",
            {
                "class_id": self.batch.pk,
                "date": "2026-07-10",
                "rows": [{"academic_session_id": self.session.pk}],
            },
        )
        self.assertEqual(missing_status.status_code, 400)

    def _create_extra_student(self, username, first_name, last_name, admission_number):
        user = User.objects.create_user(
            username=username,
            password="pass",
            first_name=first_name,
            last_name=last_name,
        )
        UserProfile.objects.create(
            user=user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=user,
            admission_number=admission_number,
        )
        session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=student,
            academic_year=self.year,
            admission_number=admission_number,
        )
        enrollment = StudentEnrollment.objects.create(
            student=student,
            academic_session=session,
            batch=self.batch,
        )
        enrollment.courses.add(self.course)
        return student, session

    def test_phase_3_assignment_crud_is_teacher_scoped(self):
        create_response = self.post_json(
            "/api/mobile/teacher/assignments/",
            {
                "batch_id": self.batch.pk,
                "course_id": self.course.pk,
                "subject_id": self.subject.pk,
                "title": "API Homework",
                "instructions": "Solve page 10",
                "due_date": "2026-08-01",
            },
        )
        self.assertEqual(create_response.status_code, 201, create_response.content)
        homework = Homework.objects.get(title="API Homework")

        put_response = self.put_json(
            f"/api/mobile/teacher/assignments/{homework.pk}/",
            {"title": "API Homework Updated", "instructions": "Updated", "due_date": "2026-08-02"},
        )
        self.assertEqual(put_response.status_code, 200, put_response.content)
        homework.refresh_from_db()
        self.assertEqual(homework.title, "API Homework Updated")

        patch_response = self.patch_json(
            f"/api/mobile/teacher/assignments/{homework.pk}/",
            {"instructions": "Patch update"},
        )
        self.assertEqual(patch_response.status_code, 200, patch_response.content)
        homework.refresh_from_db()
        self.assertEqual(homework.instructions, "Patch update")

        self.assertEqual(
            self.put_json(f"/api/mobile/teacher/assignments/{self.unassigned_homework.pk}/", {"title": "Nope"}).status_code,
            404,
        )
        delete_response = self.client.delete(f"/api/mobile/teacher/assignments/{homework.pk}/", **self.auth_headers())
        self.assertEqual(delete_response.status_code, 200, delete_response.content)
        self.assertFalse(Homework.objects.filter(pk=homework.pk).exists())

        upload_response = self.client.post(
            "/api/mobile/teacher/assignments/",
            {
                "class_id": self.batch.pk,
                "course_id": self.course.pk,
                "subject_id": self.subject.pk,
                "title": "Photo Homework",
                "instructions": "See attached photo",
                "due_date": "2026-08-03",
                "files": [
                    SimpleUploadedFile("homework.png", b"image-bytes", content_type="image/png"),
                ],
            },
            **self.auth_headers(),
        )
        self.assertEqual(upload_response.status_code, 201, upload_response.content)
        uploaded_homework = Homework.objects.get(title="Photo Homework")
        self.assertEqual(uploaded_homework.attachments.count(), 1)
        self.assertEqual(upload_response.json()["assignment"]["attachment_count"], 1)
        self.assertEqual(len(upload_response.json()["assignment"]["attachments"]), 1)
        attachment_id = uploaded_homework.attachments.first().pk
        remove_response = self.patch_json(
            f"/api/mobile/teacher/assignments/{uploaded_homework.pk}/",
            {"remove_attachment_ids": [attachment_id]},
        )
        self.assertEqual(remove_response.status_code, 200, remove_response.content)
        self.assertEqual(uploaded_homework.attachments.count(), 0)
        self.assertEqual(remove_response.json()["assignment"]["attachment_count"], 0)

    def test_assignments_support_pagination_search_and_filters(self):
        overdue = Homework.objects.create(
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Old Algebra Sheet",
            due_date=timezone.localdate() - timedelta(days=2),
            created_by=self.teacher_user,
        )
        Homework.objects.create(
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Upcoming Physics Sheet",
            due_date=timezone.localdate() + timedelta(days=3),
            created_by=self.teacher_user,
        )
        today_homework = Homework.objects.create(
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Today Worksheet",
            due_date=timezone.localdate(),
            created_by=self.teacher_user,
        )

        page = self.get_json("/api/mobile/teacher/assignments/?page_size=1")
        self.assertEqual(page["meta"]["page_size"], 1)
        self.assertEqual(page["meta"]["count"], 1)
        self.assertEqual(len(page["results"]), 1)
        self.assertEqual(page["results"][0]["id"], today_homework.pk)

        ranged = self.get_json(
            "/api/mobile/teacher/assignments/?page_size=1&due_from=2026-01-01&due_to=2026-12-31"
        )
        self.assertGreaterEqual(ranged["meta"]["count"], 3)

        search = self.get_json("/api/mobile/teacher/assignments/?search=Algebra")
        self.assertEqual(search["meta"]["count"], 1)
        self.assertEqual(search["results"][0]["id"], overdue.pk)

        filtered = self.get_json(
            f"/api/mobile/teacher/assignments/?class_id={self.batch.pk}&subject_id={self.subject.pk}&status=overdue&search=Algebra"
        )
        self.assertEqual(filtered["meta"]["count"], 1)
        self.assertEqual(filtered["results"][0]["title"], "Old Algebra Sheet")

    def test_assignment_update_restricts_unassigned_class(self):
        response = self.patch_json(
            f"/api/mobile/teacher/assignments/{self.homework.pk}/",
            {"class_id": self.unassigned_batch.pk, "title": "Moved"},
        )

        self.assertEqual(response.status_code, 404)
        self.homework.refresh_from_db()
        self.assertEqual(self.homework.batch_id, self.batch.pk)

    def test_phase_3_exam_crud_and_publish_are_teacher_scoped(self):
        create_response = self.post_json(
            "/api/mobile/teacher/exams/",
            {
                "batch_id": self.batch.pk,
                "course_id": self.course.pk,
                "subject_id": self.subject.pk,
                "title": "API Exam",
                "exam_date": "2026-08-10",
                "duration_minutes": 45,
                "instructions": "No books",
            },
        )
        self.assertEqual(create_response.status_code, 201, create_response.content)
        exam = Exam.objects.get(title="API Exam")
        self.assertFalse(exam.is_published)

        put_response = self.put_json(
            f"/api/mobile/teacher/exams/{exam.pk}/",
            {
                "title": "API Exam Updated",
                "exam_date": "2026-08-11",
                "duration_minutes": 50,
                "is_published": False,
            },
        )
        self.assertEqual(put_response.status_code, 200, put_response.content)
        exam.refresh_from_db()
        self.assertEqual(exam.title, "API Exam Updated")

        publish_response = self.post_json(f"/api/mobile/teacher/exams/{exam.pk}/publish/")
        self.assertEqual(publish_response.status_code, 200, publish_response.content)
        exam.refresh_from_db()
        self.assertTrue(exam.is_published)
        self.assertEqual(
            self.post_json(f"/api/mobile/teacher/exams/{self.unassigned_exam.pk}/publish/").status_code,
            404,
        )

        delete_response = self.client.delete(f"/api/mobile/teacher/exams/{exam.pk}/", **self.auth_headers())
        self.assertEqual(delete_response.status_code, 200, delete_response.content)
        self.assertFalse(Exam.objects.filter(pk=exam.pk).exists())

    def test_exams_support_pagination_search_filters_and_summary(self):
        draft = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Draft Algebra Test",
            exam_date=timezone.localdate() + timedelta(days=5),
            is_published=False,
        )
        ExamQuestion.objects.create(exam=draft, text="Q1", marks=5, order=1)
        ExamAttempt.objects.create(
            exam=draft,
            academic_session=self.session,
            student=self.student,
            submitted_at=timezone.now(),
            score=5,
            total_marks=5,
        )
        Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Published Biology Test",
            exam_date=timezone.localdate() + timedelta(days=2),
            is_published=True,
        )

        page = self.get_json("/api/mobile/teacher/exams/?page_size=1")
        self.assertEqual(page["meta"]["page_size"], 1)
        self.assertGreaterEqual(page["meta"]["count"], 3)
        self.assertEqual(len(page["results"]), 1)

        filtered = self.get_json(
            f"/api/mobile/teacher/exams/?status=draft&search=Algebra&class_id={self.batch.pk}"
        )
        self.assertEqual(filtered["meta"]["count"], 1)
        self.assertEqual(filtered["results"][0]["id"], draft.pk)
        self.assertEqual(filtered["results"][0]["question_count"], 1)
        self.assertEqual(filtered["results"][0]["submission_count"], 1)
        self.assertEqual(filtered["results"][0]["result_count"], 1)

    def test_exam_update_restricts_unassigned_class(self):
        response = self.patch_json(
            f"/api/mobile/teacher/exams/{self.exam.pk}/",
            {"batch_id": self.unassigned_batch.pk, "title": "Moved Exam"},
        )

        self.assertEqual(response.status_code, 404)
        self.exam.refresh_from_db()
        self.assertEqual(self.exam.batch_id, self.batch.pk)

    def test_phase_3_question_crud_resyncs_exam_marks(self):
        create_response = self.post_json(
            "/api/mobile/teacher/questions/",
            {
                "exam_id": self.exam.pk,
                "text": "Capital of India?",
                "marks": 3,
                "options": [
                    {"text": "Delhi", "is_correct": True},
                    {"text": "Mumbai", "is_correct": False},
                    {"text": "Pune", "is_correct": False},
                    {"text": "Chennai", "is_correct": False},
                ],
            },
        )
        self.assertEqual(create_response.status_code, 201, create_response.content)
        question = ExamQuestion.objects.get(text="Capital of India?")
        self.assertEqual(question.options.count(), 4)
        self.exam.refresh_from_db()
        self.assertEqual(self.exam.total_marks, 5)

        patch_response = self.patch_json(
            f"/api/mobile/teacher/questions/{question.pk}/",
            {
                "marks": 4,
                "options": [
                    {"text": "New Delhi", "is_correct": True},
                    {"text": "Mumbai", "is_correct": False},
                    {"text": "Pune", "is_correct": False},
                    {"text": "Chennai", "is_correct": False},
                ],
            },
        )
        self.assertEqual(patch_response.status_code, 200, patch_response.content)
        question.refresh_from_db()
        self.assertEqual(question.marks, 4)
        self.exam.refresh_from_db()
        self.assertEqual(self.exam.total_marks, 6)

        self.assertEqual(
            self.post_json(
                "/api/mobile/teacher/questions/",
                {
                    "exam_id": self.unassigned_exam.pk,
                    "text": "Hidden?",
                    "marks": 1,
                    "options": [
                        {"text": "A", "is_correct": True},
                        {"text": "B"},
                        {"text": "C"},
                        {"text": "D"},
                    ],
                },
            ).status_code,
            404,
        )
        delete_response = self.client.delete(f"/api/mobile/teacher/questions/{question.pk}/", **self.auth_headers())
        self.assertEqual(delete_response.status_code, 200, delete_response.content)
        self.assertFalse(ExamQuestion.objects.filter(pk=question.pk).exists())

    def test_questions_support_pagination_search_filter_and_validation(self):
        algebra = ExamQuestion.objects.create(exam=self.exam, text="Algebra MCQ", marks=2, order=2)
        for index, text in enumerate(["A", "B", "C", "D"], start=1):
            ExamQuestionOption.objects.create(
                question=algebra,
                text=text,
                is_correct=index == 1,
                order=index,
            )
        other_exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.subject,
            title="Other Exam",
            exam_date="2026-09-01",
        )
        ExamQuestion.objects.create(exam=other_exam, text="Other Question", marks=1, order=1)

        page = self.get_json(f"/api/mobile/teacher/questions/?exam_id={self.exam.pk}&page_size=1")
        self.assertEqual(page["meta"]["page_size"], 1)
        self.assertGreaterEqual(page["meta"]["count"], 2)
        self.assertEqual(len(page["results"]), 1)
        self.assertEqual(page["results"][0]["question_type"], "MCQ")

        search = self.get_json(f"/api/mobile/teacher/questions/?exam_id={self.exam.pk}&search=Algebra")
        self.assertEqual(search["meta"]["count"], 1)
        self.assertEqual(search["results"][0]["id"], algebra.pk)

        invalid = self.post_json(
            "/api/mobile/teacher/questions/",
            {
                "exam_id": self.exam.pk,
                "text": "",
                "marks": 1,
                "options": [{"text": "Only one", "is_correct": True}],
            },
        )
        self.assertEqual(invalid.status_code, 400)

    def test_questions_support_mobile_image_upload_and_remove(self):
        response = self.client.post(
            "/api/mobile/teacher/questions/",
            data={
                "exam_id": self.exam.pk,
                "text": "",
                "marks": 3,
                "order": 4,
                "options": json.dumps(
                    [
                        {"text": "A", "is_correct": True, "order": 1},
                        {"text": "B", "is_correct": False, "order": 2},
                        {"text": "C", "is_correct": False, "order": 3},
                        {"text": "D", "is_correct": False, "order": 4},
                    ]
                ),
                "image": SimpleUploadedFile("question.png", b"fake-image", content_type="image/png"),
            },
            **self.auth_headers(),
        )
        self.assertEqual(response.status_code, 201, response.content)
        payload = response.json()["question"]
        self.assertEqual(payload["text"], "")
        self.assertTrue(payload["image_url"])
        question = ExamQuestion.objects.get(pk=payload["id"])
        self.assertTrue(question.image)

        remove_response = self.patch_json(
            f"/api/mobile/teacher/questions/{question.pk}/",
            {"text": "Text only now", "remove_image": True},
        )
        self.assertEqual(remove_response.status_code, 200, remove_response.content)
        question.refresh_from_db()
        self.assertEqual(question.text, "Text only now")
        self.assertFalse(question.image)

    def test_phase_3_submission_and_result_actions(self):
        pending_attempt = self.attempt
        pending_attempt.submitted_at = None
        pending_attempt.save(update_fields=["submitted_at"])
        force_response = self.post_json(f"/api/mobile/teacher/submissions/{pending_attempt.pk}/force-submit/")
        self.assertEqual(force_response.status_code, 200, force_response.content)
        pending_attempt.refresh_from_db()
        self.assertIsNotNone(pending_attempt.submitted_at)
        self.assertTrue(ExamResult.objects.filter(exam=self.exam, student=self.student).exists())

        hide_response = self.post_json(f"/api/mobile/teacher/results/{self.exam.pk}/hide/")
        self.assertEqual(hide_response.status_code, 200, hide_response.content)
        self.exam.refresh_from_db()
        self.assertFalse(self.exam.show_result_after_submit)

        with (
            patch("teacher.api.views.notify_exam_results_declared") as notify_results,
            self.captureOnCommitCallbacks(execute=True),
        ):
            publish_response = self.post_json(f"/api/mobile/teacher/results/{self.exam.pk}/publish/")
        self.assertEqual(publish_response.status_code, 200, publish_response.content)
        self.exam.refresh_from_db()
        self.assertTrue(self.exam.show_result_after_submit)
        notify_results.assert_called_once_with(self.exam.pk)

        reset_response = self.post_json(f"/api/mobile/teacher/submissions/{pending_attempt.pk}/reset/")
        self.assertEqual(reset_response.status_code, 200, reset_response.content)
        self.assertFalse(ExamAttempt.objects.filter(pk=pending_attempt.pk).exists())

    def test_submissions_and_results_support_filters_pagination_and_summary(self):
        extra_student, extra_session = self._create_extra_student(
            "attempt-student",
            "Attempt",
            "Student",
            "M301",
        )
        self.student.roll_number = "02"
        extra_student.roll_number = "03"
        self.student.save(update_fields=["roll_number"])
        extra_student.save(update_fields=["roll_number"])
        in_progress = ExamAttempt.objects.create(
            exam=self.exam,
            academic_session=extra_session,
            student=extra_student,
            score=0,
            total_marks=2,
        )
        first_student, first_session = self._create_extra_student(
            "first-result-student",
            "First",
            "Result",
            "M300",
        )
        first_student.roll_number = "01"
        first_student.save(update_fields=["roll_number"])
        first_attempt = ExamAttempt.objects.create(
            exam=self.exam,
            academic_session=first_session,
            student=first_student,
            submitted_at=timezone.now(),
            score=1,
            total_marks=2,
            correct_count=1,
        )

        submissions = self.get_json(
            f"/api/mobile/teacher/submissions/?exam_id={self.exam.pk}&class_id={self.batch.pk}&status=in_progress&page_size=1"
        )
        self.assertEqual(submissions["meta"]["page_size"], 1)
        self.assertEqual(submissions["summary"]["in_progress"], 1)
        self.assertEqual(submissions["results"][0]["id"], in_progress.pk)

        results = self.get_json(
            f"/api/mobile/teacher/results/?exam_id={self.exam.pk}&status=submitted&page_size=5"
        )
        self.assertEqual(results["summary"]["submitted"], 2)
        self.assertEqual(results["summary"]["not_attempted"], 1)
        self.assertEqual([row["roll_number"] for row in results["results"]], ["01", "02"])
        self.assertEqual(results["results"][0]["id"], first_attempt.pk)
        self.assertGreaterEqual(results["summary"]["average_percentage"], 0)

        search = self.get_json(
            "/api/mobile/teacher/results/?search=02"
        )
        self.assertEqual(search["summary"]["submitted"], 1)
        self.assertEqual(search["results"][0]["id"], self.attempt.pk)

    def test_phase_3_notice_read_and_message_post(self):
        read_response = self.post_json(f"/api/mobile/teacher/notices/{self.notice.pk}/read/")
        self.assertEqual(read_response.status_code, 200, read_response.content)
        self.assertTrue(NoticeRead.objects.filter(notice=self.notice, user=self.teacher_user).exists())
        self.assertEqual(
            self.post_json(f"/api/mobile/teacher/notices/{self.hidden_notice.pk}/read/").status_code,
            404,
        )

        message_response = self.post_json(
            "/api/mobile/teacher/messages/",
            {"batch_id": self.batch.pk, "subject": "Hello", "message": "Class update"},
        )
        self.assertEqual(message_response.status_code, 202, message_response.content)
        self.assertEqual(message_response.json()["message"]["message"], "Class update")

    def test_notices_support_pagination_filters_and_search(self):
        NoticeRead.objects.get_or_create(notice=self.notice, user=self.teacher_user)
        urgent = Notice.objects.create(
            institute=self.institute,
            academic_year=self.year,
            title="Urgent Exam Notice",
            message="Exam tomorrow",
            audience=Notice.Audience.TEACHERS,
            category=Notice.Category.EXAM,
            priority=Notice.Priority.URGENT,
            push_to_app=True,
        )
        urgent.target_batches.add(self.batch)

        page = self.get_json("/api/mobile/teacher/notices/?page_size=1")
        self.assertEqual(page["meta"]["page_size"], 1)
        self.assertGreaterEqual(page["meta"]["count"], 2)

        unread = self.get_json("/api/mobile/teacher/notices/?read=unread&category=EXAM&priority=URGENT&search=Exam")
        self.assertEqual(unread["meta"]["count"], 1)
        self.assertEqual(unread["results"][0]["id"], urgent.pk)
        self.assertFalse(unread["results"][0]["is_read"])

    def test_message_restricts_unassigned_student_recipient(self):
        unassigned_student, session = self._create_extra_student(
            "message-private-student",
            "Message",
            "Private",
            "M401",
        )
        StudentEnrollment.objects.filter(student=unassigned_student, batch=self.batch).delete()
        StudentEnrollment.objects.create(
            student=unassigned_student,
            academic_session=session,
            batch=self.unassigned_batch,
        )

        response = self.post_json(
            "/api/mobile/teacher/messages/",
            {"recipient_id": unassigned_student.pk, "subject": "Hi", "message": "Private"},
        )

        self.assertEqual(response.status_code, 404)

    def test_phase_4_attendance_cannot_mark_unassigned_students(self):
        unassigned_user = User.objects.create_user(username="phase4-private-student", password="pass")
        UserProfile.objects.create(
            user=unassigned_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        unassigned_student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=unassigned_user,
            admission_number="P401",
        )
        unassigned_session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=unassigned_student,
            academic_year=self.year,
            admission_number="P401",
        )
        StudentEnrollment.objects.create(
            student=unassigned_student,
            academic_session=unassigned_session,
            batch=self.unassigned_batch,
        )

        response = self.post_json(
            "/api/mobile/teacher/attendance/",
            {
                "class_id": self.batch.pk,
                "date": "2026-09-01",
                "rows": [
                    {
                        "academic_session_id": unassigned_session.pk,
                        "status": Attendance.Status.ABSENT,
                    }
                ],
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["saved_count"], 0)
        self.assertEqual(response.json()["skipped_count"], 1)
        self.assertFalse(
            Attendance.objects.filter(
                academic_session=unassigned_session,
                date="2026-09-01",
            ).exists()
        )

    def test_phase_4_cannot_use_other_institute_course_or_subject(self):
        other_institute = Institute.objects.create(name="Other API Institute", code="other-api")
        other_year = AcademicYear.objects.create(
            institute=other_institute,
            name="2026-27",
            start_date="2026-04-01",
            end_date="2027-03-31",
        )
        other_course = Course.objects.create(
            institute=other_institute,
            academic_year=other_year,
            name="Private Course",
        )
        other_subject = Subject.objects.create(
            institute=other_institute,
            academic_year=other_year,
            name="Private Subject",
        )

        homework_response = self.post_json(
            "/api/mobile/teacher/assignments/",
            {
                "batch_id": self.batch.pk,
                "course_id": other_course.pk,
                "subject_id": self.subject.pk,
                "title": "Cross Institute Homework",
            },
        )
        exam_response = self.post_json(
            "/api/mobile/teacher/exams/",
            {
                "batch_id": self.batch.pk,
                "course_id": self.course.pk,
                "subject_id": other_subject.pk,
                "title": "Cross Institute Exam",
                "exam_date": "2026-09-10",
                "duration_minutes": 30,
            },
        )

        self.assertEqual(homework_response.status_code, 400)
        self.assertEqual(exam_response.status_code, 400)
        self.assertFalse(Homework.objects.filter(title="Cross Institute Homework").exists())
        self.assertFalse(Exam.objects.filter(title="Cross Institute Exam").exists())

    def test_phase_4_unassigned_mutations_are_blocked(self):
        unassigned_attempt = ExamAttempt.objects.create(
            exam=self.unassigned_exam,
            academic_session=self.session,
            student=self.student,
            submitted_at=timezone.now(),
            score=1,
            total_marks=1,
        )

        blocked_requests = [
            self.patch_json(f"/api/mobile/teacher/exams/{self.unassigned_exam.pk}/", {"title": "No"}),
            self.client.delete(f"/api/mobile/teacher/exams/{self.unassigned_exam.pk}/", **self.auth_headers()),
            self.post_json(f"/api/mobile/teacher/results/{self.unassigned_exam.pk}/publish/"),
            self.post_json(f"/api/mobile/teacher/results/{self.unassigned_exam.pk}/hide/"),
            self.post_json(f"/api/mobile/teacher/submissions/{unassigned_attempt.pk}/force-submit/"),
            self.post_json(f"/api/mobile/teacher/submissions/{unassigned_attempt.pk}/reset/"),
            self.post_json(
                "/api/mobile/teacher/messages/",
                {"batch_id": self.unassigned_batch.pk, "message": "No access"},
            ),
        ]

        for response in blocked_requests:
            with self.subTest(response=response.content):
                self.assertEqual(response.status_code, 404)

    def test_phase_6_invalid_payloads_return_400(self):
        invalid_attendance = self.post_json(
            "/api/mobile/teacher/attendance/",
            {"class_id": self.batch.pk, "date": "2026-09-01", "rows": []},
        )
        invalid_assignment = self.post_json(
            "/api/mobile/teacher/assignments/",
            {"title": "Missing class"},
        )
        invalid_exam = self.post_json(
            "/api/mobile/teacher/exams/",
            {"batch_id": self.batch.pk, "title": "Missing date"},
        )
        invalid_question = self.post_json(
            "/api/mobile/teacher/questions/",
            {
                "exam_id": self.exam.pk,
                "text": "Invalid options",
                "marks": 1,
                "options": [{"text": "Only one", "is_correct": True}],
            },
        )
        invalid_message = self.post_json("/api/mobile/teacher/messages/", {"subject": "No body"})

        for response in [
            invalid_attendance,
            invalid_assignment,
            invalid_exam,
            invalid_question,
            invalid_message,
        ]:
            with self.subTest(response=response.content):
                self.assertEqual(response.status_code, 400)
                self.assertFalse(response.json()["success"])

    def test_teacher_mobile_password_change_updates_password(self):
        response = self.post_json(
            "/api/mobile/auth/password/",
            {
                "current_password": "pass",
                "new_password": "NewTeacherPass123!",
                "confirm_password": "NewTeacherPass123!",
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.teacher_user.refresh_from_db()
        self.assertTrue(self.teacher_user.check_password("NewTeacherPass123!"))

        invalid = self.post_json(
            "/api/mobile/auth/password/",
            {
                "current_password": "pass",
                "new_password": "AnotherTeacherPass123!",
                "confirm_password": "AnotherTeacherPass123!",
            },
        )
        self.assertEqual(invalid.status_code, 400)


class ExamYearScopeTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(name="Scope Institute", code="scope")
        self.year_a = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date="2026-04-01",
            end_date="2027-03-31",
        )
        self.year_b = AcademicYear.objects.create(
            institute=self.institute,
            name="2027-28",
            start_date="2027-04-01",
            end_date="2028-03-31",
        )
        self.teacher_user = User.objects.create_user(username="teacher", password="pass")
        UserProfile.objects.create(
            user=self.teacher_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        self.student_user = User.objects.create_user(username="student", password="pass")
        UserProfile.objects.create(
            user=self.student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        self.student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year_a,
            user=self.student_user,
            admission_number="A001",
        )
        self.course_a = Course.objects.create(institute=self.institute, academic_year=self.year_a, name="Physics")
        self.course_b = Course.objects.create(institute=self.institute, academic_year=self.year_b, name="Physics")
        self.subject_a = Subject.objects.create(institute=self.institute, academic_year=self.year_a, name="Physics")
        self.subject_b = Subject.objects.create(institute=self.institute, academic_year=self.year_b, name="Maths")
        self.batch_a = Batch.objects.create(institute=self.institute, academic_year=self.year_a, name="Batch A")
        self.batch_a.courses.add(self.course_a)
        self.batch_a.teachers.add(self.teacher_user)
        self.unassigned_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_a,
            name="Unassigned Batch",
        )
        self.unassigned_batch.courses.add(self.course_a)
        self.batch_b = Batch.objects.create(institute=self.institute, academic_year=self.year_b, name="Batch B")
        self.batch_b.courses.add(self.course_b)
        self.batch_b.teachers.add(self.teacher_user)
        self.session_a = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year_a,
            admission_number="A001",
        )
        self.session_b = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year_b,
            admission_number="B001",
        )
        StudentEnrollment.objects.create(student=self.student, academic_session=self.session_a, batch=self.batch_a)
        StudentEnrollment.objects.create(student=self.student, academic_session=self.session_b, batch=self.batch_b)
        self.second_student_user = User.objects.create_user(username="student-two", password="pass")
        UserProfile.objects.create(
            user=self.second_student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        self.second_student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year_a,
            user=self.second_student_user,
            admission_number="A002",
        )
        self.second_session_a = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.second_student,
            academic_year=self.year_a,
            admission_number="A002",
        )
        StudentEnrollment.objects.create(student=self.second_student, academic_session=self.second_session_a, batch=self.batch_a)
        self.unassigned_student_user = User.objects.create_user(username="unassigned-student", password="pass")
        UserProfile.objects.create(
            user=self.unassigned_student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        self.unassigned_student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year_a,
            user=self.unassigned_student_user,
            admission_number="U001",
        )
        self.unassigned_session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.unassigned_student,
            academic_year=self.year_a,
            admission_number="U001",
        )
        StudentEnrollment.objects.create(
            student=self.unassigned_student,
            academic_session=self.unassigned_session,
            batch=self.unassigned_batch,
        )
        self.exam_a = Exam.objects.create(
            academic_year=self.year_a,
            batch=self.batch_a,
            course=self.course_a,
            title="Year A Exam",
            exam_date="2026-06-10",
            is_published=True,
        )
        self.exam_b = Exam.objects.create(
            academic_year=self.year_b,
            batch=self.batch_b,
            course=self.course_b,
            title="Year B Exam",
            exam_date="2027-06-10",
            is_published=True,
        )
        self.unassigned_exam = Exam.objects.create(
            academic_year=self.year_a,
            batch=self.unassigned_batch,
            course=self.course_a,
            title="Unassigned Exam",
            exam_date="2026-06-11",
            is_published=True,
        )

    def test_teacher_exams_use_selected_academic_year(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:exams"))

        self.assertContains(response, "Year A Exam")
        self.assertNotContains(response, "Year B Exam")
        self.assertNotContains(response, "Unassigned Exam")

    def test_teacher_dashboard_counts_only_assigned_batches(self):
        Homework.objects.create(
            batch=self.batch_a,
            course=self.course_a,
            title="Assigned Homework",
            due_date="2026-06-20",
            created_by=self.teacher_user,
        )
        Homework.objects.create(
            batch=self.unassigned_batch,
            course=self.course_a,
            title="Unassigned Homework",
            due_date="2026-06-20",
            created_by=self.teacher_user,
        )
        Attendance.objects.create(
            student=self.student,
            academic_session=self.session_a,
            batch=self.batch_a,
            date=timezone.localdate(),
            status=Attendance.Status.PRESENT,
            marked_by=self.teacher_user,
        )
        Attendance.objects.create(
            student=self.unassigned_student,
            academic_session=self.unassigned_session,
            batch=self.unassigned_batch,
            date=timezone.localdate(),
            status=Attendance.Status.PRESENT,
            marked_by=self.teacher_user,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:dashboard"))

        self.assertEqual(response.context["batch_count"], 1)
        self.assertEqual(response.context["student_count"], 2)
        self.assertEqual(response.context["homework_count"], 1)
        self.assertEqual(response.context["exam_count"], 1)
        self.assertEqual(response.context["today_attendance_count"], 1)

    def test_teacher_homework_is_limited_to_assigned_batches(self):
        assigned_homework = Homework.objects.create(
            batch=self.batch_a,
            course=self.course_a,
            title="Assigned Homework",
            due_date="2026-06-20",
            created_by=self.teacher_user,
        )
        unassigned_homework = Homework.objects.create(
            batch=self.unassigned_batch,
            course=self.course_a,
            title="Unassigned Homework",
            due_date="2026-06-20",
            created_by=self.teacher_user,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:homework"))

        self.assertContains(response, assigned_homework.title)
        self.assertNotContains(response, unassigned_homework.title)
        self.assertEqual(self.client.get(reverse("teacher:homework_update", args=[unassigned_homework.pk])).status_code, 404)
        self.assertEqual(
            self.client.post(reverse("teacher:homework_delete", args=[unassigned_homework.pk])).status_code,
            404,
        )

    def test_teacher_homework_form_rejects_unassigned_batch(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(
            reverse("teacher:homework_create"),
            {
                "batch": str(self.unassigned_batch.pk),
                "course": str(self.course_a.pk),
                "title": "Should Not Save",
                "instructions": "Private",
                "due_date": "2026-06-20",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Homework.objects.filter(title="Should Not Save").exists())

    def test_teacher_notices_follow_institute_targeting_and_live_updates(self):
        visible = Notice.objects.create(
            institute=self.institute,
            title="Assigned batch notice",
            message="Original message",
            audience=Notice.Audience.TEACHERS,
            push_to_app=False,
        )
        visible.target_batches.add(self.batch_a)
        unassigned = Notice.objects.create(
            institute=self.institute,
            title="Unassigned batch notice",
            message="Private",
            audience=Notice.Audience.TEACHERS,
        )
        unassigned.target_batches.add(self.unassigned_batch)
        student_only = Notice.objects.create(
            institute=self.institute,
            title="Student only notice",
            message="Students",
            audience=Notice.Audience.STUDENTS_PARENTS,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:notices"))

        self.assertContains(response, visible.title)
        self.assertContains(response, "Original message")
        self.assertNotContains(response, unassigned.title)
        self.assertNotContains(response, student_only.title)

        visible.message = "Updated by institute"
        visible.save(update_fields=["message"])
        response = self.client.get(reverse("teacher:notices"))
        self.assertContains(response, "Updated by institute")
        self.assertNotContains(response, "Original message")

    def test_teacher_notices_exclude_future_expired_and_other_institute(self):
        future = Notice.objects.create(
            institute=self.institute,
            title="Future notice",
            message="Later",
            audience=Notice.Audience.EVERYONE,
            publish_at=timezone.now() + timedelta(days=1),
        )
        expired = Notice.objects.create(
            institute=self.institute,
            title="Expired notice",
            message="Past",
            audience=Notice.Audience.EVERYONE,
            expires_at=timezone.now() - timedelta(days=1),
        )
        other_institute = Institute.objects.create(name="Other Institute", code="other-notice")
        other = Notice.objects.create(
            institute=other_institute,
            title="Other institute notice",
            message="Private",
            audience=Notice.Audience.EVERYONE,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:notices"))

        self.assertNotContains(response, future.title)
        self.assertNotContains(response, expired.title)
        self.assertNotContains(response, other.title)

    def test_teacher_can_mark_only_accessible_notice_as_read(self):
        visible = Notice.objects.create(
            institute=self.institute,
            title="Readable notice",
            message="Read this",
            audience=Notice.Audience.TEACHERS,
        )
        hidden = Notice.objects.create(
            institute=self.institute,
            title="Hidden notice",
            message="No access",
            audience=Notice.Audience.TEACHERS,
        )
        hidden.target_batches.add(self.unassigned_batch)
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(
            reverse("teacher:notice_mark_read", args=[visible.pk]),
            {"next": reverse("teacher:notices")},
        )

        self.assertRedirects(response, reverse("teacher:notices"))
        self.assertTrue(NoticeRead.objects.filter(notice=visible, user=self.teacher_user).exists())
        self.assertEqual(
            self.client.post(reverse("teacher:notice_mark_read", args=[hidden.pk])).status_code,
            404,
        )

    def test_teacher_attendance_ignores_unassigned_batch_and_students(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:attendance"), {"batch": self.unassigned_batch.pk})

        self.assertContains(response, "Batch A")
        self.assertNotContains(response, "Unassigned Batch")
        self.assertNotContains(response, "unassigned-student")

        response = self.client.post(
            f"{reverse('teacher:attendance')}?batch={self.unassigned_batch.pk}&date=2026-06-15",
            {
                "student_ids": [str(self.unassigned_student.pk)],
                f"status_{self.unassigned_student.pk}": Attendance.Status.ABSENT,
                f"note_{self.unassigned_student.pk}": "Forged",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            Attendance.objects.filter(
                batch=self.unassigned_batch,
                academic_session=self.unassigned_session,
                date="2026-06-15",
            ).exists()
        )

    def test_teacher_attendance_export_excludes_unassigned_batch(self):
        Attendance.objects.create(
            student=self.student,
            academic_session=self.session_a,
            batch=self.batch_a,
            date="2026-06-15",
            status=Attendance.Status.PRESENT,
            marked_by=self.teacher_user,
        )
        Attendance.objects.create(
            student=self.unassigned_student,
            academic_session=self.unassigned_session,
            batch=self.unassigned_batch,
            date="2026-06-15",
            status=Attendance.Status.ABSENT,
            marked_by=self.teacher_user,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:attendance_export"), {"format": "excel"})

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        worksheet = workbook.active
        values = "\n".join(str(cell.value or "") for row in worksheet.iter_rows() for cell in row)
        self.assertIn("A001", values)
        self.assertNotIn("U001", values)

    def test_teacher_exam_direct_urls_reject_unassigned_batch_exam(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        protected_urls = [
            reverse("teacher:exam_update", args=[self.unassigned_exam.pk]),
            reverse("teacher:exam_questions", args=[self.unassigned_exam.pk]),
            reverse("teacher:exam_question_create", args=[self.unassigned_exam.pk]),
            reverse("teacher:exam_question_import_template", args=[self.unassigned_exam.pk]),
            reverse("teacher:exam_submissions", args=[self.unassigned_exam.pk]),
        ]

        for url in protected_urls:
            with self.subTest(url=url):
                self.assertEqual(self.client.get(url).status_code, 404)

        self.assertEqual(
            self.client.post(reverse("teacher:exam_toggle_result_publish", args=[self.unassigned_exam.pk]), {"action": "publish"}).status_code,
            404,
        )
        self.assertEqual(
            self.client.post(reverse("teacher:exam_publish", args=[self.unassigned_exam.pk])).status_code,
            404,
        )

    def test_teacher_exam_form_rejects_unassigned_batch(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(
            reverse("teacher:exam_create"),
            {
                "batch": str(self.unassigned_batch.pk),
                "subject": str(self.subject_a.pk),
                "title": "Forbidden Exam",
                "exam_date": "2026-07-01",
                "duration_minutes": "60",
                "instructions": "Nope",
                "allow_rough_work_uploads": "on",
                "is_published": "on",
                "show_result_after_submit": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Exam.objects.filter(title="Forbidden Exam").exists())

    def test_teacher_results_are_limited_to_assigned_batch_exams(self):
        ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.session_a,
            student=self.student,
            submitted_at=timezone.now(),
            score=8,
            total_marks=10,
        )
        ExamAttempt.objects.create(
            exam=self.unassigned_exam,
            academic_session=self.unassigned_session,
            student=self.unassigned_student,
            submitted_at=timezone.now(),
            score=9,
            total_marks=10,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:results"))

        self.assertContains(response, "Year A Exam")
        self.assertNotContains(response, "Unassigned Exam")
        self.assertNotContains(response, "Add Result")

    def test_teacher_result_filters_and_exports_use_assigned_attempts(self):
        ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.session_a,
            student=self.student,
            submitted_at=timezone.now(),
            score=8,
            total_marks=10,
            correct_count=8,
            wrong_count=2,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        query = {"exam": str(self.exam_a.pk), "performance": "passed", "student": "student"}
        response = self.client.get(reverse("teacher:results"), query)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["result_count"], 1)
        self.assertContains(response, "student")

        excel_response = self.client.get(
            reverse("teacher:results_export"),
            {**query, "format": "excel"},
        )
        self.assertEqual(excel_response.status_code, 200)
        workbook = load_workbook(BytesIO(excel_response.content))
        values = [cell.value for row in workbook["Exam Results"].iter_rows() for cell in row]
        self.assertIn("student", values)
        self.assertNotIn("Unassigned Exam", values)

        pdf_response = self.client.get(
            reverse("teacher:results_export"),
            {**query, "format": "pdf"},
        )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertIn(b"student", pdf_response.content)
        self.assertNotIn(b"Unassigned Exam", pdf_response.content)

    def test_student_exams_use_selected_academic_session_year(self):
        self.client.login(username="student", password="pass")

        response = self.client.get(
            reverse("student_parent:exams"),
            {"academic_session_id": self.session_b.pk},
        )

        self.assertContains(response, "Year B Exam")
        self.assertNotContains(response, "Year A Exam")

    def test_exam_submissions_show_attempted_and_not_attempted_students(self):
        attempt = ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.session_a,
            student=self.student,
            submitted_at=timezone.now(),
            score=1,
            total_marks=1,
            correct_count=1,
        )
        ExamAttemptUpload.objects.create(
            attempt=attempt,
            image="exams/rough-work/teacher-private-upload.png",
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:exam_submissions", args=[self.exam_a.pk]))

        self.assertContains(response, "student")
        self.assertContains(response, "student-two")
        self.assertContains(response, "Attempted")
        self.assertContains(response, "Not Attempted")
        attempt_row = next(row for row in response.context["rows"] if row["attempt"] == attempt)
        self.assertEqual(attempt_row["upload_count"], 1)
        self.assertNotContains(response, "teacher-private-upload.png")
        self.assertContains(response, 'class="submission-action manage"')
        self.assertContains(response, 'class="submission-action reset"')

    def test_bulk_question_template_download_contains_samples(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:exam_question_import_template", args=[self.exam_a.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        worksheet = workbook.active
        self.assertEqual(worksheet["A1"].value, "Question Text")
        self.assertEqual(worksheet["F1"].value, "Correct Answer")
        self.assertIn(worksheet["F2"].value, {"A", "B", "C", "D"})

    def test_bulk_question_import_creates_questions_options_and_correct_answer(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Marks"])
        worksheet.append(["2 + 2 equals?", "1", "2", "4", "8", "C", 2])
        worksheet.append(["Unit of force?", "Newton", "Joule", "Watt", "Pascal", "A", 1])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            "questions.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("teacher:exam_question_bulk_import", args=[self.exam_a.pk]),
            {"question_file": upload},
        )

        self.assertRedirects(response, reverse("teacher:exam_questions", args=[self.exam_a.pk]))
        self.assertEqual(self.exam_a.questions.count(), 2)
        first_question = ExamQuestion.objects.get(exam=self.exam_a, text="2 + 2 equals?")
        self.assertEqual(first_question.options.count(), 4)
        self.assertEqual(first_question.correct_option.text, "4")
        self.exam_a.refresh_from_db()
        self.assertEqual(self.exam_a.total_marks, 3)

    def test_bulk_question_import_rejects_invalid_correct_answer_without_partial_save(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Marks"])
        worksheet.append(["2 + 2 equals?", "1", "2", "4", "8", "E", 1])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            "questions.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("teacher:exam_question_bulk_import", args=[self.exam_a.pk]),
            {"question_file": upload},
        )

        self.assertRedirects(response, reverse("teacher:exam_questions", args=[self.exam_a.pk]))
        self.assertEqual(self.exam_a.questions.count(), 0)

    def test_question_create_success_closes_popup_window(self):
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(
            reverse("teacher:exam_question_create", args=[self.exam_a.pk]),
            {
                "text": "What is 3 + 3?",
                "marks": "1",
                "options-TOTAL_FORMS": "4",
                "options-INITIAL_FORMS": "0",
                "options-MIN_NUM_FORMS": "0",
                "options-MAX_NUM_FORMS": "4",
                "options-0-text": "4",
                "options-0-is_correct": "",
                "options-1-text": "5",
                "options-1-is_correct": "",
                "options-2-text": "6",
                "options-2-is_correct": "on",
                "options-3-text": "7",
                "options-3-is_correct": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "window.close()")
        self.assertContains(response, reverse("teacher:exam_questions", args=[self.exam_a.pk]))
        self.assertEqual(self.exam_a.questions.count(), 1)

    def test_exam_form_subjects_are_scoped_to_teacher_batch_institute_and_year(self):
        other_institute = Institute.objects.create(name="Other Institute", code="other")
        other_year = AcademicYear.objects.create(
            institute=other_institute,
            name="2026-27",
            start_date="2026-04-01",
            end_date="2027-03-31",
        )
        Subject.objects.create(institute=other_institute, academic_year=other_year, name="Other Physics")

        form = TeacherExamForm(batches=Batch.objects.filter(pk=self.batch_a.pk))

        subject_names = list(form.fields["subject"].queryset.values_list("name", flat=True))
        self.assertEqual(subject_names, ["Physics"])
        rendered = str(form["subject"])
        self.assertIn(f'data-year-id="{self.year_a.pk}"', rendered)

    def test_exam_update_form_keeps_selected_inactive_subject_visible(self):
        self.subject_a.is_active = False
        self.subject_a.save()
        self.exam_a.subject = self.subject_a
        self.exam_a.save()

        form = TeacherExamForm(instance=self.exam_a, batches=Batch.objects.filter(pk=self.batch_a.pk))

        self.assertIn(self.subject_a, form.fields["subject"].queryset)

    def test_teacher_can_publish_and_hide_exam_results_from_submissions(self):
        self.exam_a.show_result_after_submit = False
        self.exam_a.save()
        result = ExamResult.objects.create(
            exam=self.exam_a,
            student=self.student,
            marks_obtained="82.00",
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:exam_submissions", args=[self.exam_a.pk]))

        self.assertContains(response, "Results are hidden")
        self.assertContains(response, "Publish Results")

        with (
            patch(
                "student_parent.notifications.notify_result_declared"
            ) as notify_result,
            self.captureOnCommitCallbacks(execute=True),
        ):
            response = self.client.post(
                reverse("teacher:exam_toggle_result_publish", args=[self.exam_a.pk]),
                {"action": "publish"},
            )

        self.assertRedirects(response, reverse("teacher:exam_submissions", args=[self.exam_a.pk]))
        self.exam_a.refresh_from_db()
        self.assertTrue(self.exam_a.show_result_after_submit)
        notify_result.assert_called_once()
        self.assertEqual(notify_result.call_args.args[0].pk, result.pk)

        with (
            patch(
                "student_parent.notifications.notify_result_declared"
            ) as notify_result,
            self.captureOnCommitCallbacks(execute=True),
        ):
            self.client.post(
                reverse("teacher:exam_toggle_result_publish", args=[self.exam_a.pk]),
                {"action": "publish"},
            )
        notify_result.assert_not_called()

        self.client.post(
            reverse("teacher:exam_toggle_result_publish", args=[self.exam_a.pk]),
            {"action": "hide"},
        )
        self.exam_a.refresh_from_db()
        self.assertFalse(self.exam_a.show_result_after_submit)

    def test_teacher_can_publish_exam_from_submissions(self):
        self.exam_a.is_published = False
        self.exam_a.save()
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.get(reverse("teacher:exam_submissions", args=[self.exam_a.pk]))

        self.assertContains(response, "Exam Not Published")
        self.assertContains(response, "Publish Exam")

        response = self.client.post(reverse("teacher:exam_publish", args=[self.exam_a.pk]))

        self.assertRedirects(response, reverse("teacher:exam_submissions", args=[self.exam_a.pk]))
        self.exam_a.refresh_from_db()
        self.assertTrue(self.exam_a.is_published)

    def test_teacher_can_manage_attempt_answer_and_recalculate_score(self):
        question = ExamQuestion.objects.create(exam=self.exam_a, text="2 + 2?", marks=2, order=1)
        wrong_option = ExamQuestionOption.objects.create(question=question, text="3", order=1)
        correct_option = ExamQuestionOption.objects.create(question=question, text="4", order=2, is_correct=True)
        attempt = ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.session_a,
            student=self.student,
            submitted_at=timezone.now(),
            score=0,
            total_marks=2,
            wrong_count=1,
        )
        ExamQuestionAttempt.objects.create(
            attempt=attempt,
            question=question,
            selected_option=wrong_option,
            is_correct=False,
        )
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(
            reverse("teacher:exam_attempt_manage", args=[self.exam_a.pk, attempt.pk]),
            {f"answer_{question.pk}": str(correct_option.pk), "action": "save_answers"},
        )

        self.assertRedirects(response, reverse("teacher:exam_attempt_manage", args=[self.exam_a.pk, attempt.pk]))
        attempt.refresh_from_db()
        self.assertEqual(attempt.score, 2)
        self.assertEqual(attempt.correct_count, 1)
        self.assertEqual(attempt.wrong_count, 0)
        self.assertEqual(attempt.unattempted_count, 0)
        answer = ExamQuestionAttempt.objects.get(attempt=attempt, question=question)
        self.assertEqual(answer.selected_option, correct_option)
        self.assertTrue(answer.is_correct)
        self.assertTrue(ExamResult.objects.filter(exam=self.exam_a, student=self.student, marks_obtained=2).exists())

    def test_teacher_can_reset_attempt_so_student_can_reattend(self):
        attempt = ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.session_a,
            student=self.student,
            submitted_at=timezone.now(),
            score=1,
            total_marks=1,
            correct_count=1,
        )
        ExamResult.objects.create(exam=self.exam_a, student=self.student, marks_obtained=1)
        self.client.login(username="teacher", password="pass")
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

        response = self.client.post(reverse("teacher:exam_attempt_reset", args=[self.exam_a.pk, attempt.pk]))

        self.assertRedirects(response, reverse("teacher:exam_submissions", args=[self.exam_a.pk]))
        self.assertFalse(ExamAttempt.objects.filter(pk=attempt.pk).exists())
        self.assertFalse(ExamResult.objects.filter(exam=self.exam_a, student=self.student).exists())
