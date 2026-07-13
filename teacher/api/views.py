import json
from datetime import timedelta
from io import BytesIO

from django.http import HttpResponse
from django.db import transaction
from django.db.models import Case, Count, IntegerField, Q, Value, When
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework import status
from rest_framework.views import APIView

from institute_admin.models import Course, NoticeRead, Subject
from student_parent.notifications import notify_exam_results_declared
from student_parent.models import StudentAcademicSession, StudentEnrollment
from ..models import Attendance, Exam, ExamAttempt, ExamQuestion, ExamQuestionOption, ExamResult, Homework, HomeworkAttachment
from ..views import build_pdf_document, pdf_rect, pdf_text, recalculate_exam_attempt, sync_exam_total_marks, teacher_attempt_percentage, teacher_students_for_batches
from .authentication import MobileBearerAuthentication
from .pagination import TeacherMobilePagination
from .permissions import IsTeacherMobileUser
from .responses import api_response, list_response
from .serializers import (
    AcademicYearSerializer,
    AttemptSerializer,
    AttendanceRowSerializer,
    AttendanceSaveSerializer,
    BatchSerializer,
    ExamUpdateSerializer,
    ExamSerializer,
    ExamWriteSerializer,
    HomeworkSerializer,
    HomeworkUpdateSerializer,
    HomeworkWriteSerializer,
    MessageSerializer,
    NoticeSerializer,
    QuestionUpdateSerializer,
    QuestionSerializer,
    QuestionWriteSerializer,
    StudentDetailSerializer,
    StudentSerializer,
)
from .services import (
    assigned_batches_with_counts,
    attempts_queryset,
    exam_queryset,
    homework_queryset,
    parse_api_date,
    question_queryset,
    submitted_attempts_queryset,
    teacher_attempt_or_none,
    teacher_assigned_batches,
    teacher_exam_or_none,
    teacher_homework_or_none,
    teacher_notice_queryset,
    teacher_question_or_none,
    teacher_students_for_batch,
    student_display_name,
    today_date,
)


class TeacherMobileAPIView(APIView):
    authentication_classes = [MobileBearerAuthentication]
    permission_classes = [IsTeacherMobileUser]
    pagination_class = TeacherMobilePagination

    def paginate_queryset(self, queryset):
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, self.request, view=self)
        return paginator, page

    def paginated_response(self, queryset, serializer_class, **serializer_kwargs):
        paginator, page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = serializer_class(page, many=True, **serializer_kwargs)
            return paginator.get_paginated_response(serializer.data)
        serializer = serializer_class(queryset, many=True, **serializer_kwargs)
        return list_response(serializer.data)


class TeacherPdfRenderer(BaseRenderer):
    media_type = "application/pdf"
    format = "pdf"
    charset = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class TeacherExcelRenderer(BaseRenderer):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "excel"
    charset = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class TeacherXlsxRenderer(TeacherExcelRenderer):
    format = "xlsx"


class TeacherDashboardAPI(TeacherMobileAPIView):
    def get(self, request):
        batches = teacher_assigned_batches(request)
        today = today_date()
        week_end = today + timedelta(days=7)
        students_qs = teacher_students_for_batches(batches)
        homework_qs = Homework.objects.filter(batch__in=batches)
        exam_qs = exam_queryset(request)
        today_attendance = Attendance.objects.filter(batch__in=batches, date=today)
        present_today = today_attendance.filter(status=Attendance.Status.PRESENT).count()
        marked_count = today_attendance.count()
        attempts_qs = ExamAttempt.objects.filter(exam__in=exam_qs)
        submitted_attempts = attempts_qs.filter(submitted_at__isnull=False)

        payload = {
            "batch_count": batches.count(),
            "student_count": students_qs.count(),
            "homework_count": homework_qs.count(),
            "exam_count": exam_qs.count(),
            "today_attendance_count": marked_count,
            "present_today": present_today,
            "absent_today": today_attendance.filter(status=Attendance.Status.ABSENT).count(),
            "late_today": today_attendance.filter(status=Attendance.Status.LATE).count(),
            "attendance_rate": round((present_today / marked_count) * 100, 1) if marked_count else 0,
            "open_homework_count": homework_qs.filter(Q(due_date__isnull=True) | Q(due_date__gte=today)).count(),
            "due_soon_homework_count": homework_qs.filter(due_date__range=(today, week_end)).count(),
            "overdue_homework_count": homework_qs.filter(due_date__lt=today).count(),
            "published_exam_count": exam_qs.filter(is_published=True).count(),
            "draft_exam_count": exam_qs.filter(is_published=False).count(),
            "upcoming_exam_count": exam_qs.filter(exam_date__gte=today).count(),
            "submitted_attempt_count": submitted_attempts.count(),
            "pending_submission_count": attempts_qs.filter(submitted_at__isnull=True).count(),
            "result_count": submitted_attempts.count(),
            "assigned_batches": BatchSerializer(assigned_batches_with_counts(request), many=True).data,
            "recent_homework": HomeworkSerializer(homework_queryset(request)[:6], many=True).data,
            "recent_exams": ExamSerializer(exam_qs.order_by("exam_date")[:6], many=True).data,
            "recent_results": AttemptSerializer(submitted_attempts_queryset(request)[:6], many=True).data,
        }
        return api_response(payload)


class TeacherAcademicYearsAPI(TeacherMobileAPIView):
    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.institute_id:
            return list_response([])
        years = (
            profile.institute.academic_years.filter(
                batches__teachers=request.user,
                batches__is_active=True,
            )
            .distinct()
            .order_by("-start_date", "-id")
        )
        return list_response(AcademicYearSerializer(years, many=True).data)


class TeacherClassesAPI(TeacherMobileAPIView):
    def get(self, request):
        batches = assigned_batches_with_counts(request)
        search = (request.query_params.get("search") or "").strip()
        if search:
            batches = batches.filter(
                Q(name__icontains=search)
                | Q(courses__name__icontains=search)
            ).distinct()
        course_id = request.query_params.get("course_id")
        if course_id:
            batches = batches.filter(courses__pk=course_id)
        subjects = Subject.objects.filter(
            institute_id__in=batches.values_list("institute_id", flat=True),
            academic_year_id__in=batches.values_list("academic_year_id", flat=True),
            is_active=True,
        ).order_by("name")
        subjects_by_year = {}
        for subject in subjects:
            subjects_by_year.setdefault(subject.academic_year_id, []).append(subject)
        return self.paginated_response(
            batches,
            BatchSerializer,
            context={"subjects_by_year": subjects_by_year},
        )


class TeacherClassStudentsAPI(TeacherMobileAPIView):
    def get(self, request, batch_id):
        batch = teacher_assigned_batches(request).filter(pk=batch_id).first()
        if not batch:
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        sessions = teacher_students_for_batch(request, batch)
        search = (request.query_params.get("search") or "").strip()
        if search:
            sessions = sessions.filter(
                Q(admission_number__icontains=search)
                | Q(student__admission_number__icontains=search)
                | Q(student__user__first_name__icontains=search)
                | Q(student__user__last_name__icontains=search)
                | Q(student__user__username__icontains=search)
            )
        course_id = request.query_params.get("course_id")
        if course_id:
            sessions = sessions.filter(enrollments__courses__pk=course_id)
        return self.paginated_response(sessions, StudentSerializer, context={"batch": batch})


class TeacherClassStudentDetailAPI(TeacherMobileAPIView):
    def get(self, request, batch_id, session_id):
        batch = teacher_assigned_batches(request).filter(pk=batch_id).first()
        if not batch:
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        session = teacher_students_for_batch(request, batch).filter(pk=session_id).first()
        if not session:
            return api_response(message="Student not found.", status_code=status.HTTP_404_NOT_FOUND)
        attendance = Attendance.objects.filter(batch=batch, academic_session=session)
        attendance_total = attendance.count()
        present = attendance.filter(status=Attendance.Status.PRESENT).count()
        exams = Exam.objects.filter(batch=batch)
        attempts = ExamAttempt.objects.filter(exam__in=exams, academic_session=session)
        submitted = attempts.filter(submitted_at__isnull=False)
        homework = Homework.objects.filter(batch=batch)
        summary = {
            "attendance_total": attendance_total,
            "present": present,
            "absent": attendance.filter(status=Attendance.Status.ABSENT).count(),
            "late": attendance.filter(status=Attendance.Status.LATE).count(),
            "attendance_rate": round((present / attendance_total) * 100, 1) if attendance_total else 0,
            "homework_count": homework.count(),
            "exam_count": exams.count(),
            "attempt_count": attempts.count(),
            "submitted_attempt_count": submitted.count(),
            "average_percentage": _attempts_summary(submitted)["average_percentage"],
        }
        return api_response(
            StudentDetailSerializer(session, context={"batch": batch, "summary": summary}).data
        )


class TeacherAttendanceAPI(TeacherMobileAPIView):
    def get(self, request):
        selected_date = parse_api_date(request.query_params.get("date"), today_date())
        batch_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
        records = Attendance.objects.filter(
            batch__in=teacher_assigned_batches(request),
            date=selected_date,
        ).select_related("academic_session", "student", "student__user", "batch")
        if batch_id:
            records = records.filter(batch_id=batch_id)
        attendance_status = request.query_params.get("status", "")
        if attendance_status in Attendance.Status.values:
            records = records.filter(status=attendance_status)
        search = (request.query_params.get("search") or "").strip()
        if search:
            records = records.filter(
                Q(academic_session__admission_number__icontains=search)
                | Q(student__admission_number__icontains=search)
                | Q(student__user__first_name__icontains=search)
                | Q(student__user__last_name__icontains=search)
                | Q(student__user__username__icontains=search)
            )
        counts = records.aggregate(
            total=Count("pk"),
            present=Count("pk", filter=Q(status=Attendance.Status.PRESENT)),
            absent=Count("pk", filter=Q(status=Attendance.Status.ABSENT)),
            late=Count("pk", filter=Q(status=Attendance.Status.LATE)),
        )
        total = counts["total"]
        records = records.annotate(
            roll_missing=Case(
                When(student__roll_number="", then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
        paginator, page = self.paginate_queryset(
            records.order_by("batch__name", "roll_missing", "student__roll_number", "academic_session__admission_number")
        )
        serializer = AttendanceRowSerializer(page if page is not None else records, many=True)
        meta = {}
        if page is not None:
            meta = {
                "count": paginator.page.paginator.count,
                "page": paginator.page.number,
                "page_size": paginator.get_page_size(request),
                "total_pages": paginator.page.paginator.num_pages,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
            }
        return api_response(
            {
                "total": total,
                "present": counts["present"],
                "absent": counts["absent"],
                "late": counts["late"],
                "rate": round((counts["present"] / total) * 100, 1) if total else 0,
                "rows": serializer.data,
                "meta": meta,
            }
        )

    def post(self, request):
        serializer = AttendanceSaveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        batch = teacher_assigned_batches(request).filter(pk=serializer.validated_data["batch_id"]).first()
        if not batch:
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        sessions = {
            session.pk: session
            for session in teacher_students_for_batch(request, batch)
        }
        sessions_by_student = {
            session.student_id: session
            for session in teacher_students_for_batch(request, batch)
        }
        saved = 0
        skipped = 0
        for row in serializer.validated_data["rows"]:
            session = sessions.get(row.get("academic_session_id")) or sessions_by_student.get(row.get("student_id"))
            if not session:
                skipped += 1
                continue
            Attendance.objects.update_or_create(
                academic_session=session,
                batch=batch,
                date=serializer.validated_data["date"],
                defaults={
                    "student": session.student,
                    "status": row["status"],
                    "note": row.get("note", ""),
                    "marked_by": request.user,
                },
            )
            saved += 1
        return api_response({"saved_count": saved, "skipped_count": skipped}, message="Attendance saved.")


class TeacherAssignmentsAPI(TeacherMobileAPIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request):
        homework = homework_queryset(request)
        class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
        if class_id:
            homework = homework.filter(batch_id=class_id)
        course_id = request.query_params.get("course_id")
        if course_id:
            homework = homework.filter(course_id=course_id)
        subject_id = request.query_params.get("subject_id")
        if subject_id:
            homework = homework.filter(subject_id=subject_id)
        search = (request.query_params.get("search") or "").strip()
        if search:
            homework = homework.filter(
                Q(title__icontains=search)
                | Q(instructions__icontains=search)
                | Q(batch__name__icontains=search)
                | Q(course__name__icontains=search)
                | Q(subject__name__icontains=search)
            )
        due_from = parse_api_date(request.query_params.get("due_from"))
        if due_from:
            homework = homework.filter(due_date__gte=due_from)
        due_to = parse_api_date(request.query_params.get("due_to"))
        if due_to:
            homework = homework.filter(due_date__lte=due_to)
        homework_status = (request.query_params.get("status") or "").lower()
        today = today_date()
        if homework_status == "overdue":
            homework = homework.filter(due_date__lt=today)
        elif homework_status == "due_soon":
            homework = homework.filter(due_date__range=(today, today + timedelta(days=7)))
        elif homework_status == "open":
            homework = homework.filter(Q(due_date__isnull=True) | Q(due_date__gte=today))
        elif homework_status == "no_due":
            homework = homework.filter(due_date__isnull=True)
        if not (due_from or due_to or homework_status or search or class_id or course_id or subject_id):
            homework = homework.filter(due_date=today)
        return self.paginated_response(
            homework.prefetch_related("attachments").order_by("-created_at"),
            HomeworkSerializer,
            context={"request": request},
        )

    def post(self, request):
        serializer = HomeworkWriteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        batch = teacher_assigned_batches(request).filter(pk=data["batch_id"]).first()
        if not batch:
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        course, subject, error = _resolve_batch_course_subject(batch, data)
        if error:
            return error
        homework = Homework.objects.create(
            batch=batch,
            course=course,
            subject=subject,
            title=(data.get("title") or "Homework").strip() or "Homework",
            instructions=data.get("instructions", ""),
            due_date=data.get("due_date"),
            created_by=request.user,
        )
        _save_homework_attachments(homework, request)
        return api_response(
            {"assignment": HomeworkSerializer(homework, context={"request": request}).data},
            message="Homework created.",
            status_code=status.HTTP_201_CREATED,
        )


class TeacherAssignmentDetailAPI(TeacherMobileAPIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def put(self, request, assignment_id):
        return self._update(request, assignment_id)

    def patch(self, request, assignment_id):
        return self._update(request, assignment_id)

    def delete(self, request, assignment_id):
        homework = teacher_homework_or_none(request, assignment_id)
        if not homework:
            return api_response(message="Homework not found.", status_code=status.HTTP_404_NOT_FOUND)
        homework.delete()
        return api_response(message="Homework deleted.")

    def _update(self, request, assignment_id):
        homework = teacher_homework_or_none(request, assignment_id)
        if not homework:
            return api_response(message="Homework not found.", status_code=status.HTTP_404_NOT_FOUND)
        serializer = HomeworkUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if "batch_id" in data:
            batch = teacher_assigned_batches(request).filter(pk=data["batch_id"]).first()
            if not batch:
                return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
            course, subject, error = _resolve_batch_course_subject(
                batch,
                data,
                default_course=homework.course,
                default_subject=homework.subject,
            )
            if error:
                return error
            homework.batch = batch
            homework.course = course
            homework.subject = subject
        elif "course_id" in data or "subject_id" in data:
            course, subject, error = _resolve_batch_course_subject(
                homework.batch,
                data,
                default_course=homework.course,
                default_subject=homework.subject,
            )
            if error:
                return error
            homework.course = course
            homework.subject = subject
        if "title" in data:
            homework.title = data["title"].strip() or homework.title
        if "instructions" in data:
            homework.instructions = data["instructions"]
        if "due_date" in data:
            homework.due_date = data["due_date"]
        homework.save()
        _delete_homework_attachments(homework, request, data.get("remove_attachment_ids"))
        _save_homework_attachments(homework, request)
        return api_response(
            {"assignment": HomeworkSerializer(homework, context={"request": request}).data},
            message="Homework updated.",
        )


def _save_homework_attachments(homework, request):
    files = request.FILES.getlist("files")
    for uploaded_file in files:
        HomeworkAttachment.objects.create(homework=homework, file=uploaded_file)
    _clear_homework_attachment_cache(homework)


def _delete_homework_attachments(homework, request, validated_ids=None):
    raw_ids = []
    if validated_ids is not None:
        raw_ids = validated_ids
    if hasattr(request.data, "getlist"):
        raw_ids = raw_ids or request.data.getlist("remove_attachment_ids")
        if not raw_ids:
            raw_ids = request.data.getlist("remove_attachment_ids[]")
    if not raw_ids:
        value = request.data.get("remove_attachment_ids") if hasattr(request, "data") else None
        if isinstance(value, list):
            raw_ids = value
        elif value:
            raw_ids = str(value).split(",")
    attachment_ids = []
    for raw_id in raw_ids:
        if isinstance(raw_id, (list, tuple)):
            raw_values = raw_id
        else:
            raw_values = [raw_id]
        for raw_value in raw_values:
            try:
                attachment_ids.append(int(raw_value))
            except (TypeError, ValueError):
                continue
    if attachment_ids:
        homework.attachments.filter(pk__in=attachment_ids).delete()
        _clear_homework_attachment_cache(homework)


def _clear_homework_attachment_cache(homework):
    if hasattr(homework, "_prefetched_objects_cache"):
        homework._prefetched_objects_cache.pop("attachments", None)


class TeacherNoticesAPI(TeacherMobileAPIView):
    def get(self, request):
        notices = teacher_notice_queryset(request)
        read_ids = set(
            NoticeRead.objects.filter(
                user=request.user,
                notice__in=notices,
            ).values_list("notice_id", flat=True)
        )
        category = (request.query_params.get("category") or "").strip().upper()
        if category:
            notices = notices.filter(category=category)
        priority = (request.query_params.get("priority") or "").strip().upper()
        if priority:
            notices = notices.filter(priority=priority)
        read_filter = (request.query_params.get("read") or "").strip().lower()
        if read_filter in {"false", "0", "unread"}:
            notices = notices.exclude(pk__in=read_ids)
        elif read_filter in {"true", "1", "read"}:
            notices = notices.filter(pk__in=read_ids)
        search = (request.query_params.get("search") or "").strip()
        if search:
            notices = notices.filter(
                Q(title__icontains=search)
                | Q(message__icontains=search)
                | Q(category__icontains=search)
                | Q(priority__icontains=search)
            )
        return self.paginated_response(notices, NoticeSerializer, context={"read_ids": read_ids})


class TeacherNoticeReadAPI(TeacherMobileAPIView):
    def post(self, request, notice_id):
        notice = teacher_notice_queryset(request).filter(pk=notice_id).first()
        if not notice:
            return api_response(message="Notice not found.", status_code=status.HTTP_404_NOT_FOUND)
        receipt, _created = NoticeRead.objects.get_or_create(notice=notice, user=request.user)
        return api_response(
            {"notice_id": notice.pk, "read_at": receipt.read_at},
            message="Notice marked as read.",
        )


class TeacherMessagesAPI(TeacherMobileAPIView):
    def post(self, request):
        serializer = MessageSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        batch_id = serializer.validated_data.get("batch_id")
        if batch_id and not teacher_assigned_batches(request).filter(pk=batch_id).exists():
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        recipient_id = serializer.validated_data.get("recipient_id")
        if recipient_id and not teacher_students_for_batches(teacher_assigned_batches(request)).filter(student_id=recipient_id).exists():
            return api_response(message="Student not found.", status_code=status.HTTP_404_NOT_FOUND)
        return api_response(
            {"message": serializer.validated_data},
            message="Message accepted.",
            status_code=status.HTTP_202_ACCEPTED,
        )


class TeacherExamsAPI(TeacherMobileAPIView):
    def get(self, request):
        exams = exam_queryset(request)
        class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
        if class_id:
            exams = exams.filter(batch_id=class_id)
        course_id = request.query_params.get("course_id")
        if course_id:
            exams = exams.filter(course_id=course_id)
        subject_id = request.query_params.get("subject_id")
        if subject_id:
            exams = exams.filter(subject_id=subject_id)
        search = (request.query_params.get("search") or "").strip()
        if search:
            exams = exams.filter(
                Q(title__icontains=search)
                | Q(instructions__icontains=search)
                | Q(batch__name__icontains=search)
                | Q(course__name__icontains=search)
                | Q(subject__name__icontains=search)
            )
        exam_status = (request.query_params.get("status") or "").lower()
        today = today_date()
        if exam_status == "draft":
            exams = exams.filter(is_published=False)
        elif exam_status == "published":
            exams = exams.filter(is_published=True)
        elif exam_status == "upcoming":
            exams = exams.filter(exam_date__gte=today)
        elif exam_status == "past":
            exams = exams.filter(exam_date__lt=today)
        date_from = parse_api_date(request.query_params.get("date_from"))
        if date_from:
            exams = exams.filter(exam_date__gte=date_from)
        date_to = parse_api_date(request.query_params.get("date_to"))
        if date_to:
            exams = exams.filter(exam_date__lte=date_to)
        exams = exams.annotate(
            question_count=Count("questions", distinct=True),
            submission_count=Count("attempts", filter=Q(attempts__submitted_at__isnull=False), distinct=True),
            result_count=Count("attempts", filter=Q(attempts__submitted_at__isnull=False), distinct=True),
        )
        return self.paginated_response(exams.order_by("-exam_date"), ExamSerializer)

    def post(self, request):
        serializer = ExamWriteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        batch = teacher_assigned_batches(request).filter(pk=data["batch_id"]).first()
        if not batch:
            return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
        course, subject, error = _resolve_exam_relations(batch, data)
        if error:
            return error
        exam = Exam.objects.create(
            academic_year=batch.academic_year,
            batch=batch,
            course=course,
            subject=subject,
            title=data["title"],
            exam_date=data["exam_date"],
            duration_minutes=data["duration_minutes"],
            instructions=data.get("instructions", ""),
            allow_rough_work_uploads=data.get("allow_rough_work_uploads", True),
            is_published=data.get("is_published", False),
            show_result_after_submit=data.get("show_result_after_submit", True),
            created_by=request.user,
        )
        return api_response(
            {"exam": ExamSerializer(exam).data},
            message="Exam created.",
            status_code=status.HTTP_201_CREATED,
        )


class TeacherExamDetailAPI(TeacherMobileAPIView):
    def put(self, request, exam_id):
        return self._update(request, exam_id)

    def patch(self, request, exam_id):
        return self._update(request, exam_id)

    def delete(self, request, exam_id):
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        exam.delete()
        return api_response(message="Exam deleted.")

    def _update(self, request, exam_id):
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        serializer = ExamUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        batch = exam.batch
        batch_changed = False
        if "batch_id" in data:
            batch = teacher_assigned_batches(request).filter(pk=data["batch_id"]).first()
            if not batch:
                return api_response(message="Class not found.", status_code=status.HTTP_404_NOT_FOUND)
            batch_changed = batch.pk != exam.batch_id
        course, subject, error = _resolve_exam_relations(batch, data, exam=exam, batch_changed=batch_changed)
        if error:
            return error
        for field in (
            "title",
            "exam_date",
            "duration_minutes",
            "instructions",
            "allow_rough_work_uploads",
            "is_published",
            "show_result_after_submit",
        ):
            if field in data:
                setattr(exam, field, data[field])
        exam.batch = batch
        exam.academic_year = batch.academic_year
        exam.course = course
        exam.subject = subject
        exam.save()
        return api_response({"exam": ExamSerializer(exam).data}, message="Exam updated.")


class TeacherExamPublishAPI(TeacherMobileAPIView):
    def post(self, request, exam_id):
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        if not exam.is_published:
            exam.is_published = True
            exam.save(update_fields=["is_published"])
        return api_response({"exam": ExamSerializer(exam).data}, message="Exam published.")


class TeacherQuestionsAPI(TeacherMobileAPIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request):
        questions = question_queryset(request)
        exam_id = request.query_params.get("exam_id")
        if exam_id:
            questions = questions.filter(exam_id=exam_id)
        search = (request.query_params.get("search") or "").strip()
        if search:
            questions = questions.filter(
                Q(text__icontains=search)
                | Q(exam__title__icontains=search)
            )
        questions = questions.order_by("exam__title", "order")
        return self.paginated_response(questions, QuestionSerializer, context={"request": request})

    def post(self, request):
        serializer = QuestionWriteSerializer(data=_question_payload_data(request))
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        image = request.FILES.get("image")
        if not data.get("text") and not image:
            return api_response(message="Question text or image is required.", status_code=status.HTTP_400_BAD_REQUEST)
        exam = teacher_exam_or_none(request, data.get("exam_id"))
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        with transaction.atomic():
            question = ExamQuestion.objects.create(
                exam=exam,
                text=data.get("text", ""),
                image=image,
                marks=data["marks"],
                order=data.get("order") or (exam.questions.count() + 1),
            )
            _replace_question_options(question, data["options"])
            sync_exam_total_marks(exam)
        return api_response(
            {"question": QuestionSerializer(question, context={"request": request}).data},
            message="Question created.",
            status_code=status.HTTP_201_CREATED,
        )


class TeacherQuestionDetailAPI(TeacherMobileAPIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def put(self, request, question_id):
        return self._update(request, question_id)

    def patch(self, request, question_id):
        return self._update(request, question_id)

    def delete(self, request, question_id):
        question = teacher_question_or_none(request, question_id)
        if not question:
            return api_response(message="Question not found.", status_code=status.HTTP_404_NOT_FOUND)
        exam = question.exam
        question.delete()
        sync_exam_total_marks(exam)
        return api_response(message="Question deleted.")

    def _update(self, request, question_id):
        question = teacher_question_or_none(request, question_id)
        if not question:
            return api_response(message="Question not found.", status_code=status.HTTP_404_NOT_FOUND)
        serializer = QuestionUpdateSerializer(data=_question_payload_data(request), partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        image = request.FILES.get("image")
        if "text" in data and not data.get("text") and not (image or question.image):
            return api_response(message="Question text or image is required.", status_code=status.HTTP_400_BAD_REQUEST)
        exam = question.exam
        if "exam_id" in data:
            exam = teacher_exam_or_none(request, data["exam_id"])
            if not exam:
                return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        with transaction.atomic():
            old_exam = question.exam
            if "text" in data:
                question.text = data["text"]
            if data.get("remove_image") and question.image:
                question.image.delete(save=False)
                question.image = ""
            if image:
                if question.image:
                    question.image.delete(save=False)
                question.image = image
            if "marks" in data:
                question.marks = data["marks"]
            if "order" in data:
                question.order = data["order"]
            question.exam = exam
            question.save()
            if "options" in data:
                _replace_question_options(question, data["options"])
            sync_exam_total_marks(exam)
            if old_exam.pk != exam.pk:
                sync_exam_total_marks(old_exam)
        return api_response({"question": QuestionSerializer(question, context={"request": request}).data}, message="Question updated.")


class TeacherSubmissionsAPI(TeacherMobileAPIView):
    def get(self, request):
        attempts = _filter_attempts(request, attempts_queryset(request))
        return _attempts_paginated_response(self, attempts.order_by("-started_at"))


class TeacherSubmissionForceSubmitAPI(TeacherMobileAPIView):
    def post(self, request, attempt_id):
        attempt = teacher_attempt_or_none(request, attempt_id)
        if not attempt:
            return api_response(message="Submission not found.", status_code=status.HTTP_404_NOT_FOUND)
        if not attempt.is_submitted:
            attempt.submitted_at = timezone.now()
            attempt.save(update_fields=["submitted_at"])
        recalculate_exam_attempt(attempt)
        return api_response({"submission": AttemptSerializer(attempt).data}, message="Submission force submitted.")


class TeacherSubmissionResetAPI(TeacherMobileAPIView):
    def post(self, request, attempt_id):
        attempt = teacher_attempt_or_none(request, attempt_id)
        if not attempt:
            return api_response(message="Submission not found.", status_code=status.HTTP_404_NOT_FOUND)
        exam_id = attempt.exam_id
        student_id = attempt.student_id
        ExamResult.objects.filter(exam_id=exam_id, student_id=student_id).delete()
        attempt.delete()
        return api_response(message="Submission reset.")


class TeacherResultsAPI(TeacherMobileAPIView):
    def get(self, request):
        attempts = _filter_attempts(request, submitted_attempts_queryset(request))
        return _attempts_paginated_response(
            self,
            attempts.order_by(
                "exam__batch__name",
                "student__roll_number",
                "academic_session__admission_number",
                "student__user__first_name",
                "student__user__last_name",
                "-submitted_at",
            ),
            extra_summary={"not_attempted": _result_not_attempted_count(request)},
        )


class TeacherResultPublishAPI(TeacherMobileAPIView):
    def post(self, request, exam_id):
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        if not exam.show_result_after_submit:
            exam.show_result_after_submit = True
            exam.save(update_fields=["show_result_after_submit"])
            transaction.on_commit(lambda exam_id=exam.pk: notify_exam_results_declared(exam_id))
        return api_response({"exam": ExamSerializer(exam).data}, message="Results published.")


class TeacherResultHideAPI(TeacherMobileAPIView):
    def post(self, request, exam_id):
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return api_response(message="Exam not found.", status_code=status.HTTP_404_NOT_FOUND)
        if exam.show_result_after_submit:
            exam.show_result_after_submit = False
            exam.save(update_fields=["show_result_after_submit"])
        return api_response({"exam": ExamSerializer(exam).data}, message="Results hidden.")


class TeacherAttendanceReportAPI(TeacherMobileAPIView):
    renderer_classes = [JSONRenderer, TeacherPdfRenderer, TeacherExcelRenderer, TeacherXlsxRenderer]

    def get(self, request):
        records = _filter_attendance_report(request)
        counts = records.aggregate(
            total=Count("pk"),
            present=Count("pk", filter=Q(status=Attendance.Status.PRESENT)),
            absent=Count("pk", filter=Q(status=Attendance.Status.ABSENT)),
            late=Count("pk", filter=Q(status=Attendance.Status.LATE)),
        )
        total = counts["total"]
        payload = {
            "report_type": "attendance",
            "filters": _report_filters_payload(request),
            "summary": {
                "total": total,
                "present": counts["present"],
                "absent": counts["absent"],
                "late": counts["late"],
                "rate": round((counts["present"] / total) * 100, 1) if total else 0,
            },
            "class_summaries": _attendance_class_summaries(records),
            "student_summaries": _attendance_student_summaries(records),
            "export_hooks": _export_hooks(request, "attendance"),
        }
        export_response = _report_export_response(request, payload)
        if export_response:
            return export_response
        return api_response(payload)


class TeacherResultReportAPI(TeacherMobileAPIView):
    renderer_classes = [JSONRenderer, TeacherPdfRenderer, TeacherExcelRenderer, TeacherXlsxRenderer]

    def get(self, request):
        attempts = _filter_result_report_attempts(request)
        student_summaries = _result_student_summaries(request, attempts)
        payload = {
            "report_type": "results",
            "filters": _report_filters_payload(request),
            "summary": _result_report_summary(student_summaries),
            "class_summaries": _result_class_summaries(student_summaries),
            "student_summaries": _public_result_student_rows(student_summaries),
            "export_hooks": _export_hooks(request, "results"),
        }
        export_response = _report_export_response(request, payload)
        if export_response:
            return export_response
        return api_response(payload)


def _resolve_exam_relations(batch, data, *, exam=None, batch_changed=False):
    course = exam.course if exam and not batch_changed and "course_id" not in data else None
    subject = exam.subject if exam and not batch_changed and "subject_id" not in data else None
    course, subject, error = _resolve_batch_course_subject(
        batch,
        data,
        default_course=course,
        default_subject=subject,
    )
    if error:
        return None, None, error
    return course, subject, None


def _filter_attempts(request, attempts):
    exam_id = request.query_params.get("exam_id")
    if exam_id:
        attempts = attempts.filter(exam_id=exam_id)
    class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
    if class_id:
        attempts = attempts.filter(exam__batch_id=class_id)
    student_id = request.query_params.get("student_id")
    if student_id:
        attempts = attempts.filter(student_id=student_id)
    status_filter = (request.query_params.get("status") or "").lower()
    if status_filter in {"submitted", "completed"}:
        attempts = attempts.filter(submitted_at__isnull=False)
    elif status_filter in {"in_progress", "pending"}:
        attempts = attempts.filter(submitted_at__isnull=True)
    search = (request.query_params.get("search") or "").strip()
    if search:
        attempts = attempts.filter(
            Q(exam__title__icontains=search)
            | Q(academic_session__admission_number__icontains=search)
            | Q(student__admission_number__icontains=search)
            | Q(student__roll_number__icontains=search)
            | Q(student__user__first_name__icontains=search)
            | Q(student__user__last_name__icontains=search)
            | Q(student__user__username__icontains=search)
        )
    return attempts


def _filter_result_report_attempts(request):
    attempts = attempts_queryset(request)
    exam_id = request.query_params.get("exam_id")
    if exam_id:
        attempts = attempts.filter(exam_id=exam_id)
    class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
    if class_id:
        attempts = attempts.filter(exam__batch_id=class_id)
    student_id = request.query_params.get("student_id")
    if student_id:
        attempts = attempts.filter(student_id=student_id)
    date_from = parse_api_date(request.query_params.get("date_from"))
    if date_from:
        attempts = attempts.filter(exam__exam_date__gte=date_from)
    date_to = parse_api_date(request.query_params.get("date_to"))
    if date_to:
        attempts = attempts.filter(exam__exam_date__lte=date_to)
    search = (request.query_params.get("search") or "").strip()
    if search:
        attempts = attempts.filter(
            Q(exam__title__icontains=search)
            | Q(academic_session__admission_number__icontains=search)
            | Q(student__admission_number__icontains=search)
            | Q(student__roll_number__icontains=search)
            | Q(student__user__first_name__icontains=search)
            | Q(student__user__last_name__icontains=search)
            | Q(student__user__username__icontains=search)
        )
    return attempts


def _filter_attendance_report(request):
    date_from = parse_api_date(request.query_params.get("date_from"))
    date_to = parse_api_date(request.query_params.get("date_to"))
    selected_date = parse_api_date(request.query_params.get("date"))
    records = Attendance.objects.filter(
        batch__in=teacher_assigned_batches(request),
    ).select_related("academic_session", "student", "student__user", "batch")
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)
    if selected_date and not date_from and not date_to:
        records = records.filter(date=selected_date)
    if not (date_from or date_to or selected_date):
        records = records.filter(date=today_date())
    class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
    if class_id:
        records = records.filter(batch_id=class_id)
    status_filter = request.query_params.get("status", "")
    if status_filter in Attendance.Status.values:
        records = records.filter(status=status_filter)
    search = (request.query_params.get("search") or "").strip()
    if search:
        records = records.filter(
            Q(academic_session__admission_number__icontains=search)
            | Q(student__admission_number__icontains=search)
            | Q(student__roll_number__icontains=search)
            | Q(student__user__first_name__icontains=search)
            | Q(student__user__last_name__icontains=search)
            | Q(student__user__username__icontains=search)
        )
    return records


def _attendance_class_summaries(records):
    return [
        {
            "class_id": row["batch_id"],
            "class_name": row["batch__name"],
            "total": row["total"],
            "present": row["present"],
            "absent": row["absent"],
            "late": row["late"],
            "rate": round((row["present"] / row["total"]) * 100, 1) if row["total"] else 0,
        }
        for row in records.values("batch_id", "batch__name")
        .annotate(
            total=Count("pk"),
            present=Count("pk", filter=Q(status=Attendance.Status.PRESENT)),
            absent=Count("pk", filter=Q(status=Attendance.Status.ABSENT)),
            late=Count("pk", filter=Q(status=Attendance.Status.LATE)),
        )
        .order_by("batch__name")
    ]


def _attendance_student_summaries(records):
    summaries = {}
    for record in records.order_by("batch__name", "student__roll_number", "academic_session__admission_number"):
        key = (record.student_id, record.batch_id)
        item = summaries.setdefault(
            key,
            {
                "student_id": record.student_id,
                "student_name": student_display_name(record.student),
                "roll_number": record.student.roll_number or "",
                "admission_number": record.academic_session.admission_number or record.student.admission_number,
                "class_name": record.batch.name,
                "total": 0,
                "present": 0,
                "absent": 0,
                "late": 0,
                "rate": 0,
            },
        )
        item["total"] += 1
        if record.status == Attendance.Status.PRESENT:
            item["present"] += 1
        elif record.status == Attendance.Status.ABSENT:
            item["absent"] += 1
        elif record.status == Attendance.Status.LATE:
            item["late"] += 1
    for item in summaries.values():
        item["rate"] = round((item["present"] / item["total"]) * 100, 1) if item["total"] else 0
    return sorted(summaries.values(), key=_attendance_row_sort_key)


def _attendance_row_sort_key(row):
    roll_number = str(row.get("roll_number") or "")
    roll_digits = "".join(ch for ch in roll_number if ch.isdigit())
    roll_value = int(roll_digits) if roll_digits else 10**9
    return (
        row.get("class_name") or "",
        roll_value,
        roll_number,
        row.get("admission_number") or "",
        row.get("student_name") or "",
    )


def _result_class_summaries(student_rows):
    summaries = {}
    for row in student_rows:
        item = summaries.setdefault(
            row["class_id"],
            {
                "class_id": row["class_id"],
                "class_name": row["class_name"],
                "total": 0,
                "submitted": 0,
                "_score": 0,
                "_marks": 0,
                "average_percentage": 0,
            },
        )
        item["total"] += 1
        item["submitted"] += 1 if row["status"] == "submitted" else 0
        item["_score"] += row.get("_score", 0)
        item["_marks"] += row.get("_marks", 0)
    rows = []
    for item in summaries.values():
        item["average_percentage"] = _percentage_from_totals(item.pop("_score"), item.pop("_marks"))
        rows.append(item)
    return sorted(rows, key=lambda row: row["class_name"])


def _result_student_summaries(request, attempts):
    batches = _result_report_batches(request)
    enrollments = StudentEnrollment.objects.filter(
        batch__in=batches,
        status=StudentEnrollment.Status.ACTIVE,
        academic_session__status=StudentAcademicSession.Status.ACTIVE,
        student__is_active=True,
    ).select_related(
        "batch",
        "academic_session",
        "student",
        "student__user",
    ).distinct()
    student_id = request.query_params.get("student_id")
    if student_id:
        enrollments = enrollments.filter(student_id=student_id)
    search = (request.query_params.get("search") or "").strip()
    if search:
        enrollments = enrollments.filter(
            Q(academic_session__admission_number__icontains=search)
            | Q(student__admission_number__icontains=search)
            | Q(student__roll_number__icontains=search)
            | Q(student__user__first_name__icontains=search)
            | Q(student__user__last_name__icontains=search)
            | Q(student__user__username__icontains=search)
        )
    attempts_by_scope = {}
    for attempt in attempts.select_related("exam", "exam__batch", "academic_session", "student", "student__user"):
        key = (attempt.student_id, attempt.exam.batch_id, attempt.academic_session_id)
        attempts_by_scope.setdefault(key, []).append(attempt)
    question_counts, exam_question_counts = _result_report_question_counts(request, batches)
    rows = []
    seen = set()
    for enrollment in enrollments:
        key = (enrollment.student_id, enrollment.batch_id, enrollment.academic_session_id)
        if key in seen:
            continue
        seen.add(key)
        student = enrollment.student
        session = enrollment.academic_session
        scoped_attempts = attempts_by_scope.get(key, [])
        row = _empty_result_student_row(enrollment, question_counts.get(enrollment.batch_id, 0))
        for attempt in scoped_attempts:
            correct = int(attempt.correct_count or 0)
            wrong = int(attempt.wrong_count or 0)
            row["unattempted_questions"] -= exam_question_counts.get(attempt.exam_id, 0)
            row["total"] += 1
            row["submitted"] += 1 if attempt.is_submitted else 0
            row["attempted_questions"] += correct + wrong
            row["unattempted_questions"] += int(attempt.unattempted_count or 0)
            row["correct_count"] += correct
            row["wrong_count"] += wrong
            row["_score"] += float(attempt.score or 0)
            row["_marks"] += float(attempt.total_marks or 0)
        if scoped_attempts:
            row["status"] = "submitted" if row["submitted"] else "in_progress"
            row["unattempted_questions"] = max(row["unattempted_questions"], 0)
        row["student_name"] = student_display_name(student)
        row["admission_number"] = session.admission_number or student.admission_number
        row["percentage"] = _percentage_from_totals(row["_score"], row["_marks"])
        row["average_percentage"] = row["percentage"]
        rows.append(row)
    return sorted(rows, key=_result_row_sort_key)


def _result_report_batches(request):
    batches = teacher_assigned_batches(request)
    exam_id = request.query_params.get("exam_id")
    if exam_id:
        exam = teacher_exam_or_none(request, exam_id)
        if not exam:
            return batches.none()
        return batches.filter(pk=exam.batch_id)
    class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
    if class_id:
        batches = batches.filter(pk=class_id)
    return batches


def _result_report_question_counts(request, batches):
    exams = Exam.objects.filter(batch__in=batches)
    exam_id = request.query_params.get("exam_id")
    if exam_id:
        exams = exams.filter(pk=exam_id)
    date_from = parse_api_date(request.query_params.get("date_from"))
    if date_from:
        exams = exams.filter(exam_date__gte=date_from)
    date_to = parse_api_date(request.query_params.get("date_to"))
    if date_to:
        exams = exams.filter(exam_date__lte=date_to)
    batch_counts = {}
    exam_counts = {}
    for row in exams.values("pk").annotate(question_count=Count("questions")):
        exam_counts[row["pk"]] = row["question_count"]
    for row in exams.values("batch_id").annotate(question_count=Count("questions")):
        batch_counts[row["batch_id"]] = row["question_count"]
    return batch_counts, exam_counts


def _empty_result_student_row(enrollment, question_count):
    student = enrollment.student
    return {
        "student_id": student.pk,
        "student_name": student_display_name(student),
        "roll_number": student.roll_number or "",
        "admission_number": enrollment.academic_session.admission_number or student.admission_number,
        "class_id": enrollment.batch_id,
        "class_name": enrollment.batch.name,
        "status": "not_attempted",
        "total": 0,
        "submitted": 0,
        "attempted_questions": 0,
        "unattempted_questions": question_count,
        "correct_count": 0,
        "wrong_count": 0,
        "_score": 0,
        "_marks": 0,
        "percentage": 0,
        "average_percentage": 0,
    }


def _result_row_sort_key(row):
    roll_number = str(row.get("roll_number") or "")
    roll_digits = "".join(ch for ch in roll_number if ch.isdigit())
    roll_value = int(roll_digits) if roll_digits else 10**9
    return (
        row.get("class_name") or "",
        roll_value,
        roll_number,
        row.get("admission_number") or "",
        row.get("student_name") or "",
    )


def _result_report_summary(student_rows):
    total = len(student_rows)
    submitted = sum(1 for row in student_rows if row["status"] == "submitted")
    in_progress = sum(1 for row in student_rows if row["status"] == "in_progress")
    not_attempted = sum(1 for row in student_rows if row["status"] == "not_attempted")
    score = sum(row.get("_score", 0) for row in student_rows)
    marks = sum(row.get("_marks", 0) for row in student_rows)
    return {
        "total": total,
        "submitted": submitted,
        "in_progress": in_progress,
        "not_attempted": not_attempted,
        "average_percentage": _percentage_from_totals(score, marks),
    }


def _public_result_student_rows(student_rows):
    return [
        {key: value for key, value in row.items() if key not in {"_score", "_marks"}}
        for row in student_rows
    ]


def _percentage_from_totals(score, marks):
    return round((float(score or 0) / float(marks or 0)) * 100, 1) if marks else 0


def _report_filters_payload(request):
    allowed = ("class_id", "batch_id", "student_id", "exam_id", "status", "search", "date", "date_from", "date_to")
    return {key: request.query_params.get(key) for key in allowed if request.query_params.get(key)}


def _export_hooks(request, report_type):
    hooks = {}
    for export_format, label in (("pdf", "PDF"), ("excel", "Excel")):
        params = request.query_params.copy()
        params["format"] = export_format
        hooks[export_format] = {
            "label": label,
            "method": "GET",
            "url": request.build_absolute_uri(f"{request.path}?{params.urlencode()}"),
            "filename": f"teacher-{report_type}-report.{ 'xlsx' if export_format == 'excel' else 'pdf' }",
            "requires_auth": True,
        }
    return hooks


def _report_export_response(request, payload):
    export_format = (request.query_params.get("format") or "").strip().lower()
    if export_format not in {"pdf", "excel", "xlsx"}:
        return None
    report_type = payload["report_type"]
    extension = "xlsx" if export_format in {"excel", "xlsx"} else "pdf"
    filename = f"teacher-{report_type}-report.{extension}"
    if extension == "xlsx":
        content = _build_report_workbook(payload)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _build_report_pdf(payload)
        content_type = "application/pdf"
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["X-Export-Filename"] = filename
    return response


def _build_report_workbook(payload):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_report_summary_sheet(summary, payload)
    class_sheet = workbook.create_sheet("Class Summary")
    _write_sheet(class_sheet, _summary_headers(payload["report_type"], class_rows=True), payload["class_summaries"])
    student_sheet = workbook.create_sheet("Student Summary")
    _write_sheet(student_sheet, _summary_headers(payload["report_type"], class_rows=False), payload["student_summaries"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_sheet(sheet, headers, rows):
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for row in rows:
        if isinstance(row, dict):
            sheet.append([row.get(_header_key(header), "") for header in headers])
        else:
            sheet.append(list(row))
    for data_row in sheet.iter_rows(min_row=2):
        for cell in data_row:
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
            cell.alignment = Alignment(vertical="center")
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(14, len(header) + 2)


def _write_report_summary_sheet(sheet, payload):
    title = _report_title(payload)
    generated_at = timezone.now().strftime("%d-%m-%Y %I:%M %p")
    sheet.merge_cells("A1:D1")
    sheet["A1"] = f"Teacher {title}"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="111827")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet["A2"] = "Generated"
    sheet["B2"] = generated_at
    sheet["A2"].font = Font(bold=True)

    sheet.append([])
    _write_summary_section(sheet, "Overall Totals", _summary_rows(payload["summary"]))
    sheet.append([])
    _write_summary_section(sheet, "Applied Filters", _filter_rows(payload["filters"]))
    sheet.append([])
    sheet.append(["Included Sheets", "Class Summary", "Student Summary"])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
    for column, width in {"A": 24, "B": 24, "C": 24, "D": 24}.items():
        sheet.column_dimensions[column].width = width


def _write_summary_section(sheet, title, rows):
    start_row = sheet.max_row + 1
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0F766E")
    sheet.append(["Metric", "Value"])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
    for label, value in rows:
        sheet.append([label, value])


def _summary_rows(summary):
    return [(key.replace("_", " ").title(), value) for key, value in summary.items()]


def _filter_rows(filters):
    if not filters:
        return [("Filters", "None")]
    return [(key.replace("_", " ").title(), value) for key, value in filters.items()]


def _report_title(payload):
    return "Attendance Report" if payload["report_type"] == "attendance" else "Result Report"


def _summary_headers(report_type, *, class_rows):
    if report_type == "attendance":
        if class_rows:
            return ["Class Name", "Total", "Present", "Absent", "Late", "Rate"]
        return ["Roll Number", "Student Name", "Admission Number", "Class Name", "Total", "Present", "Absent", "Late", "Rate"]
    if class_rows:
        return ["Class Name", "Total", "Submitted", "Average Percentage"]
    return [
        "Roll Number",
        "Admission Number",
        "Student Name",
        "Class Name",
        "Status",
        "Attempted Questions",
        "Unattempted Questions",
        "Correct Count",
        "Wrong Count",
        "Percentage",
    ]


def _header_key(header):
    return header.lower().replace(" ", "_")


def _build_report_pdf(payload):
    title = _report_title(payload)
    width = 842
    height = 595
    rows = payload["student_summaries"] or payload["class_summaries"]
    chunks = [rows[index:index + 16] for index in range(0, len(rows), 16)] or [[]]
    pages = []
    if payload["report_type"] == "attendance":
        pages.append(_build_attendance_pdf_cover(payload, title, width, height))
    for page_number, chunk in enumerate(chunks, start=1):
        display_page = page_number + (1 if payload["report_type"] == "attendance" else 0)
        total_pages = len(chunks) + (1 if payload["report_type"] == "attendance" else 0)
        stream = pdf_rect(0, 545, width, 50, "0.278 0.333 0.412")
        stream += pdf_text(30, 570, f"Teacher {title}", 18, "F2", "1 1 1")
        stream += pdf_text(30, 552, f"Generated {timezone.now().strftime('%d-%m-%Y %I:%M %p')}", 8, "F1", "1 1 1")
        stream += pdf_text(730, 552, f"Page {display_page}/{total_pages}", 8, "F1", "1 1 1")
        summary_items = list(payload["summary"].items())[:4]
        for index, (key, value) in enumerate(summary_items):
            x = 30 + (index * 150)
            stream += pdf_rect(x, 502, 136, 31, "0.95 0.97 1")
            stream += pdf_text(x + 7, 520, key.replace("_", " ").title(), 7, "F1", "0.39 0.45 0.55")
            stream += pdf_text(x + 7, 507, str(value), 11, "F2", "0.06 0.09 0.16")
        headers = _pdf_headers(payload["report_type"])
        widths = _pdf_widths(payload["report_type"])
        y = 470
        stream += pdf_rect(30, y - 6, sum(widths), 22, "0.06 0.09 0.16")
        x = 30
        for header, col_width in zip(headers, widths):
            stream += pdf_text(x + 4, y, header, 7, "F2", "1 1 1")
            x += col_width
        y -= 25
        for row in chunk:
            stream += pdf_rect(30, y - 5, sum(widths), 20, "0.98 0.99 1")
            values = _pdf_row_values(payload["report_type"], row)
            x = 30
            for value, col_width in zip(values, widths):
                stream += pdf_text(x + 4, y, str(value)[:28], 6.5, "F1", "0.06 0.09 0.16")
                x += col_width
            y -= 22
        stream += pdf_text(30, 28, "UltraCoachMatrix teacher mobile export", 8, "F1", "0.39 0.45 0.55")
        pages.append(stream)
    return build_pdf_document(pages, width=width, height=height)


def _build_attendance_pdf_cover(payload, title, width, height):
    generated_at = timezone.now().strftime("%d-%m-%Y %I:%M %p")
    summary = payload["summary"]
    stream = pdf_rect(0, 520, width, 75, "0.06 0.09 0.16")
    stream += pdf_text(34, 565, f"Teacher {title}", 22, "F2", "1 1 1")
    stream += pdf_text(34, 545, f"Generated {generated_at}", 9, "F1", "0.86 0.91 0.98")
    stream += pdf_text(730, 545, "Page 1", 9, "F1", "0.86 0.91 0.98")

    cards = [
        ("Total Records", summary.get("total", 0), "0.91 0.95 1"),
        ("Present", summary.get("present", 0), "0.88 0.98 0.92"),
        ("Absent", summary.get("absent", 0), "1 0.91 0.92"),
        ("Late", summary.get("late", 0), "1 0.96 0.86"),
        ("Attendance Rate", f"{summary.get('rate', 0)}%", "0.9 0.98 1"),
    ]
    for index, (label, value, color) in enumerate(cards):
        x = 34 + (index * 154)
        stream += pdf_rect(x, 438, 138, 58, color)
        stream += pdf_text(x + 10, 475, label, 8, "F1", "0.39 0.45 0.55")
        stream += pdf_text(x + 10, 452, str(value), 18, "F2", "0.06 0.09 0.16")

    stream += pdf_text(34, 400, "Applied Filters", 14, "F2", "0.06 0.09 0.16")
    filter_rows = _filter_rows(payload["filters"])
    y = 372
    for index, (label, value) in enumerate(filter_rows[:10]):
        x = 34 if index < 5 else 430
        row_y = y - ((index % 5) * 36)
        stream += pdf_rect(x, row_y - 8, 350, 26, "0.98 0.99 1")
        stream += pdf_text(x + 10, row_y + 6, label, 8, "F2", "0.39 0.45 0.55")
        stream += pdf_text(x + 135, row_y + 6, str(value)[:32], 8, "F1", "0.06 0.09 0.16")

    stream += pdf_text(34, 150, "Report Sections", 14, "F2", "0.06 0.09 0.16")
    stream += pdf_text(34, 127, "1. Overall attendance totals for the selected date range and filters.", 9, "F1", "0.22 0.27 0.35")
    stream += pdf_text(34, 108, "2. Class-wise summary with present, absent, late and attendance rate.", 9, "F1", "0.22 0.27 0.35")
    stream += pdf_text(34, 89, "3. Student-wise summary for the exported records.", 9, "F1", "0.22 0.27 0.35")
    stream += pdf_text(34, 36, "UltraCoachMatrix teacher mobile export", 8, "F1", "0.39 0.45 0.55")
    return stream


def _pdf_headers(report_type):
    if report_type == "attendance":
        return ["Roll", "Student", "Admission", "Class", "Present", "Absent", "Rate"]
    return ["Roll", "Student", "Class", "Attempted", "Unattempted", "Correct", "Wrong", "%"]


def _pdf_widths(report_type):
    if report_type == "attendance":
        return [55, 135, 90, 130, 65, 65, 60]
    return [60, 150, 130, 80, 90, 70, 70, 55]


def _pdf_row_values(report_type, row):
    if report_type == "attendance":
        return [
            row.get("roll_number", ""),
            row.get("student_name") or row.get("class_name", ""),
            row.get("admission_number", ""),
            row.get("class_name", ""),
            row.get("present", ""),
            row.get("absent", ""),
            f"{row.get('rate', 0)}%",
        ]
    return [
        row.get("roll_number", ""),
        row.get("student_name") or row.get("class_name", ""),
        row.get("class_name", ""),
        row.get("attempted_questions", ""),
        row.get("unattempted_questions", ""),
        row.get("correct_count", ""),
        row.get("wrong_count", ""),
        f"{row.get('percentage', 0)}%",
    ]


def _attempts_paginated_response(view, attempts, *, extra_summary=None):
    summary = _attempts_summary(attempts)
    if extra_summary:
        summary.update(extra_summary)
    paginator, page = view.paginate_queryset(attempts)
    serializer = AttemptSerializer(page if page is not None else attempts, many=True)
    meta = {}
    if page is not None:
        meta = {
            "count": paginator.page.paginator.count,
            "page": paginator.page.number,
            "page_size": paginator.get_page_size(view.request),
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
        }
    return api_response({"summary": summary, "results": serializer.data, "meta": meta})


def _attempts_summary(attempts):
    total = attempts.count()
    submitted = attempts.filter(submitted_at__isnull=False).count()
    in_progress = total - submitted
    scores = [
        teacher_attempt_percentage(attempt)
        for attempt in attempts.filter(submitted_at__isnull=False)
    ]
    average = round(sum(scores) / len(scores), 1) if scores else 0
    return {
        "total": total,
        "submitted": submitted,
        "in_progress": in_progress,
        "average_percentage": average,
    }


def _result_not_attempted_count(request):
    batches = teacher_assigned_batches(request)
    exams = exam_queryset(request)
    exam_id = request.query_params.get("exam_id")
    if exam_id:
        exams = exams.filter(pk=exam_id)
        batches = batches.filter(pk__in=exams.values("batch_id"))
    class_id = request.query_params.get("class_id") or request.query_params.get("batch_id")
    if class_id:
        batches = batches.filter(pk=class_id)
        exams = exams.filter(batch_id=class_id)
    student_id = request.query_params.get("student_id")
    enrollments = StudentEnrollment.objects.filter(
        batch__in=batches,
        status=StudentEnrollment.Status.ACTIVE,
        academic_session__status=StudentAcademicSession.Status.ACTIVE,
        student__is_active=True,
    ).select_related("student", "student__user", "academic_session", "batch")
    if student_id:
        enrollments = enrollments.filter(student_id=student_id)
    search = (request.query_params.get("search") or "").strip()
    if search:
        enrollments = enrollments.filter(
            Q(academic_session__admission_number__icontains=search)
            | Q(student__admission_number__icontains=search)
            | Q(student__roll_number__icontains=search)
            | Q(student__user__first_name__icontains=search)
            | Q(student__user__last_name__icontains=search)
            | Q(student__user__username__icontains=search)
        )
    exam_ids_by_batch = {}
    for exam in exams.values("pk", "batch_id"):
        exam_ids_by_batch.setdefault(exam["batch_id"], []).append(exam["pk"])
    submitted_pairs = set(
        ExamAttempt.objects.filter(
            exam__in=exams,
            submitted_at__isnull=False,
        ).values_list("student_id", "exam_id")
    )
    not_attempted = 0
    seen = set()
    for enrollment in enrollments:
        key = (enrollment.student_id, enrollment.batch_id)
        if key in seen:
            continue
        seen.add(key)
        for scoped_exam_id in exam_ids_by_batch.get(enrollment.batch_id, []):
            if (enrollment.student_id, scoped_exam_id) not in submitted_pairs:
                not_attempted += 1
    return not_attempted


def _resolve_batch_course_subject(batch, data, *, default_course=None, default_subject=None):
    course = default_course
    subject = default_subject
    if data.get("course_id"):
        course = Course.objects.filter(
            pk=data["course_id"],
            institute=batch.institute,
            academic_year=batch.academic_year,
        ).first()
        if not course:
            return None, None, api_response(message="Course not found for this class.", status_code=status.HTTP_400_BAD_REQUEST)
    elif "course_id" in data:
        course = None
    if data.get("subject_id"):
        subject = Subject.objects.filter(
            pk=data["subject_id"],
            institute=batch.institute,
            academic_year=batch.academic_year,
        ).first()
        if not subject:
            return None, None, api_response(message="Subject not found for this class.", status_code=status.HTTP_400_BAD_REQUEST)
    elif "subject_id" in data:
        subject = None
    return course, subject, None


def _replace_question_options(question, options):
    question.options.all().delete()
    ExamQuestionOption.objects.bulk_create(
        [
            ExamQuestionOption(
                question=question,
                text=option["text"],
                is_correct=option.get("is_correct", False),
                order=option.get("order") or index,
            )
            for index, option in enumerate(options, start=1)
        ]
    )


def _question_payload_data(request):
    data = {key: request.data.get(key) for key in request.data.keys()}
    options = data.get("options")
    if isinstance(options, str):
        try:
            data["options"] = json.loads(options)
        except json.JSONDecodeError:
            data["options"] = []
    remove_image = data.get("remove_image")
    if isinstance(remove_image, str):
        data["remove_image"] = remove_image.lower() in {"1", "true", "yes", "on"}
    return data
