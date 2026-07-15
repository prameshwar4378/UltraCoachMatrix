from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import CareerApplication, ContactEnquiry, WebsiteFeedback


@admin.register(ContactEnquiry)
class ContactEnquiryAdmin(admin.ModelAdmin):
    list_display = ("name", "school", "phone", "email", "enquiry_type", "status", "created_at")
    list_filter = ("enquiry_type", "institution_size", "status", "created_at")
    search_fields = ("name", "school", "phone", "email", "message")
    readonly_fields = ("created_at", "updated_at")


@admin.register(WebsiteFeedback)
class WebsiteFeedbackAdmin(admin.ModelAdmin):
    list_display = ("name", "institute", "rating", "is_visible", "created_at")
    list_filter = ("rating", "is_visible", "created_at")
    search_fields = ("name", "institute", "feedback")
    readonly_fields = ("created_at",)


@admin.register(CareerApplication)
class CareerApplicationAdmin(admin.ModelAdmin):
    list_display = ("full_name", "role", "phone", "email", "experience", "city", "status", "created_at")
    list_filter = ("role", "experience", "status", "created_at")
    search_fields = (
        "full_name",
        "email",
        "phone",
        "city",
        "qualification",
        "cover_letter",
    )
    readonly_fields = ("created_at", "updated_at")
    actions = ("export_selected_to_excel",)

    @admin.action(description="Export selected applications to Excel")
    def export_selected_to_excel(self, request, queryset):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Career Applications"

        headers = [
            "Full name",
            "Role",
            "Phone",
            "Email",
            "Experience",
            "Qualification",
            "City",
            "Notice period",
            "Portfolio link",
            "Resume file",
            "Status",
            "Message",
            "Created at",
            "Updated at",
        ]
        sheet.append(headers)

        header_fill = PatternFill("solid", fgColor="123C7A")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for application in queryset.order_by("-created_at"):
            sheet.append(
                [
                    application.full_name,
                    application.get_role_display(),
                    application.phone,
                    application.email,
                    application.get_experience_display(),
                    application.qualification,
                    application.city,
                    application.notice_period,
                    application.portfolio_link,
                    application.resume.name if application.resume else "",
                    application.get_status_display(),
                    application.cover_letter,
                    timezone.localtime(application.created_at).strftime("%Y-%m-%d %H:%M"),
                    timezone.localtime(application.updated_at).strftime("%Y-%m-%d %H:%M"),
                ]
            )

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 42)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        response["Content-Disposition"] = (
            f'attachment; filename="career_applications_{stamp}.xlsx"'
        )
        workbook.save(response)
        return response
