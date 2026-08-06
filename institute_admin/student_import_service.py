from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook

from student_parent.models import GuardianProfile, StudentAcademicSession, StudentProfile
from super_admin.models import UserProfile

from .models import AcademicYear
from UltraCoachMatrix.email_notifications import on_commit_email, send_bulk_student_welcomes


def process_student_import_job(job):
    from .forms import build_student_username, get_last_student_admission_sequence, get_student_admission_prefix
    from .views import (
        match_student_import_headers,
        parse_student_import_row,
        student_import_columns,
    )

    institute = job.institute
    academic_year = AcademicYear.objects.get(pk=job.academic_year_id, institute=institute)
    workbook = load_workbook(job.input_file.path, data_only=True)
    sheet = workbook["Students"] if "Students" in workbook.sheetnames else workbook.active
    expected = student_import_columns()
    headers = [str(cell.value or "").strip() for cell in sheet[3]]
    headers_match, _missing_headers = match_student_import_headers(headers)
    if not headers_match:
        raise ValueError("Invalid student import template.")

    rows = []
    for row_number in range(4, sheet.max_row + 1):
        values = [sheet.cell(row=row_number, column=column).value for column in range(1, len(expected) + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        data = dict(zip(expected, values))
        try:
            parsed = parse_student_import_row(data)
            parsed["row_number"] = row_number
            rows.append(parsed)
        except Exception as exc:
            raise ValueError(f"Row {row_number}: {exc}")

    with transaction.atomic():
        academic_year = AcademicYear.objects.select_for_update().get(pk=academic_year.pk)
        prefix = get_student_admission_prefix(institute, academic_year)
        sequence = get_last_student_admission_sequence(institute, academic_year) + 1
        username_prefix = build_student_username(institute, prefix)
        reserved_usernames = set(
            User.objects.filter(username__startswith=username_prefix).values_list("username", flat=True)
        )
        password_hashes = {}
        users = []
        for row in rows:
            while True:
                admission_number = f"{prefix}{sequence:04d}"
                username = build_student_username(institute, admission_number)
                sequence += 1
                if username not in reserved_usernames:
                    reserved_usernames.add(username)
                    break
            row["admission_number"] = admission_number
            row["username"] = username
            password_hashes.setdefault(row["password"], make_password(row["password"]))
            users.append(
                User(
                    username=username,
                    first_name=row["first_name"],
                    last_name=row["last_name"],
                    email=row["email"],
                    password=password_hashes[row["password"]],
                    is_active=row["is_active"],
                )
            )

        User.objects.bulk_create(users, batch_size=500)
        users_by_username = User.objects.in_bulk(
            [row["username"] for row in rows],
            field_name="username",
        )
        UserProfile.objects.bulk_create(
            [
                UserProfile(
                    user=users_by_username[row["username"]],
                    institute=institute,
                    role=UserProfile.Role.STUDENT_PARENT,
                    phone=row["phone"],
                )
                for row in rows
            ],
            batch_size=500,
        )
        StudentProfile.objects.bulk_create(
            [
                StudentProfile(
                    institute=institute,
                    academic_year=academic_year,
                    user=users_by_username[row["username"]],
                    admission_number=row["admission_number"],
                    pen_no=row["pen_no"],
                    appar_id=row["appar_id"],
                    gr_number_udise=row["gr_number_udise"],
                    udise_number=row["udise_number"],
                    roll_number=row["roll_number"],
                    middle_name=row["middle_name"],
                    gender=row["gender"],
                    date_of_birth=row["date_of_birth"],
                    blood_group=row["blood_group"],
                    religion=row["religion"],
                    cast=row["cast"],
                    caste_category=row["caste_category"],
                    nationality=row["nationality"],
                    aadhaar_number=row["aadhaar_number"],
                    birth_certificate_number=row["birth_certificate_number"],
                    place_of_birth=row["place_of_birth"],
                    mother_tongue=row["mother_tongue"],
                    father_name=row["father_name"],
                    father_occupation=row["father_occupation"],
                    father_qualification=row["father_qualification"],
                    father_mobile_number=row["father_mobile_number"],
                    father_email=row["father_email"],
                    father_aadhaar_number=row["father_aadhaar_number"],
                    father_annual_income=row["father_annual_income"],
                    mother_name=row["mother_name"],
                    mother_occupation=row["mother_occupation"],
                    mother_qualification=row["mother_qualification"],
                    mother_mobile_number=row["mother_mobile_number"],
                    mother_aadhaar_number=row["mother_aadhaar_number"],
                    mother_annual_income=row["mother_annual_income"],
                    guardian_address=row["guardian_address"],
                    current_house_number=row["current_house_number"],
                    current_street_area=row["current_street_area"],
                    current_village_city=row["current_village_city"],
                    current_taluka=row["current_taluka"],
                    current_district=row["current_district"],
                    current_state=row["current_state"],
                    current_pin_code=row["current_pin_code"],
                    permanent_house_number=row["permanent_house_number"],
                    permanent_street_area=row["permanent_street_area"],
                    permanent_village_city=row["permanent_village_city"],
                    permanent_taluka=row["permanent_taluka"],
                    permanent_district=row["permanent_district"],
                    permanent_state=row["permanent_state"],
                    permanent_pin_code=row["permanent_pin_code"],
                    joined_on=row["joined_on"],
                    address=row["address"],
                    admission_class=row["admission_class"],
                    current_class=row["current_class"],
                    division=row["division"],
                    medium=row["medium"],
                    current_school_name=row["current_school_name"],
                    current_school_address=row["current_school_address"],
                    previous_school_name=row["previous_school_name"],
                    previous_school_address=row["previous_school_address"],
                    previous_school_udise_code=row["previous_school_udise_code"],
                    previous_class=row["previous_class"],
                    previous_class_passed=row["previous_class_passed"],
                    last_exam_result=row["last_exam_result"],
                    result=row["result"],
                    conduct=row["conduct"],
                    reason_for_leaving=row["reason_for_leaving"],
                    date_of_leaving_school=row["date_of_leaving_school"],
                    tc_issue_date=row["tc_issue_date"],
                    bonafide_purpose=row["bonafide_purpose"],
                    emergency_contact_number=row["emergency_contact_number"],
                    is_active=row["is_active"],
                )
                for row in rows
            ],
            batch_size=500,
        )
        students_by_user_id = StudentProfile.objects.in_bulk(
            [user.pk for user in users_by_username.values()],
            field_name="user_id",
        )
        StudentAcademicSession.objects.bulk_create(
            [
                StudentAcademicSession(
                    institute=institute,
                    student=students_by_user_id[users_by_username[row["username"]].pk],
                    academic_year=academic_year,
                    admission_number=row["admission_number"],
                    joined_on=row["joined_on"],
                    status=(
                        StudentAcademicSession.Status.ACTIVE
                        if row["is_active"]
                        else StudentAcademicSession.Status.LEFT
                    ),
                    current_school_name=row["current_school_name"],
                    current_school_address=row["current_school_address"],
                    previous_school_name=row["previous_school_name"],
                    previous_class=row["previous_class"],
                )
                for row in rows
            ],
            batch_size=500,
        )
        guardians = [
            GuardianProfile(
                student=students_by_user_id[users_by_username[row["username"]].pk],
                name=row["guardian_name"] or "Primary Guardian",
                relation=row["guardian_relation"],
                phone=row["guardian_phone"],
                email=row["guardian_email"],
                is_primary=True,
            )
            for row in rows
            if row["guardian_name"] or row["guardian_phone"]
        ]
        if guardians:
            GuardianProfile.objects.bulk_create(guardians, batch_size=500)
        student_credentials = [
            (
                students_by_user_id[users_by_username[row["username"]].pk].pk,
                row["password"],
            )
            for row in rows
        ]
        on_commit_email(send_bulk_student_welcomes, student_credentials)

    return {"created_count": len(rows)}
