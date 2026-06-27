from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.sessions.models import Session
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError, connection, transaction
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from accountant.models import Expense, ExpenseDocument, FeeCategory, FeeInvoice, Payment
from student_parent.models import GuardianProfile, StudentAcademicSession, StudentEnrollment, StudentProfile
from super_admin.models import (
    Institute,
    InstituteSubscription,
    SubscriptionPayment,
    UserProfile,
)
from teacher.models import Attendance, Exam, ExamAttempt, ExamAttemptUpload, ExamResult, Homework, TeacherProfile

from . import views
from .forms import (
    AcademicYearForm,
    BatchForm,
    PaymentUpdateForm,
    ReceiveFeeForm,
    StudentForm,
    build_student_username,
    get_student_admission_prefix,
)
from .lookup_cache import (
    get_cached_academic_years,
    get_cached_batch_course_data,
)
from .models import AcademicYear, BackgroundJob, Batch, Course, Lead, Notice, SupportTicket, Visitor


class TenantValidationTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(name="Tenant A", code="tenant-a")
        self.other_institute = Institute.objects.create(name="Tenant B", code="tenant-b")
        self.year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.other_year = AcademicYear.objects.create(
            institute=self.other_institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Math",
            fee_amount=Decimal("1000.00"),
        )
        self.other_course = Course.objects.create(
            institute=self.other_institute,
            academic_year=self.other_year,
            name="Science",
            fee_amount=Decimal("1000.00"),
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Morning",
        )

    def test_course_cannot_use_another_institute_academic_year(self):
        course = Course(
            institute=self.institute,
            academic_year=self.other_year,
            name="Invalid",
            fee_amount=Decimal("10.00"),
        )

        with self.assertRaises(ValidationError):
            course.save()

    def test_batch_courses_reject_cross_institute_course(self):
        with self.assertRaises(ValidationError):
            self.batch.courses.add(self.other_course)

    def test_notice_targets_reject_cross_institute_course(self):
        notice = Notice.objects.create(
            institute=self.institute,
            title="Notice",
            message="Tenant scoped",
        )

        with self.assertRaises(ValidationError):
            notice.target_courses.add(self.other_course)


class AcademicSessionSettingsTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Session Institute",
            code="session-institute",
            status=Institute.Status.ACTIVE,
        )
        self.other_institute = Institute.objects.create(
            name="Other Institute",
            code="other-session-institute",
            status=Institute.Status.ACTIVE,
        )
        self.admin_user = User.objects.create_user(
            username="session-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.session = AcademicYear.objects.create(
            institute=self.institute,
            name="2025-26",
            start_date=date(2025, 4, 1),
            end_date=date(2026, 3, 31),
        )
        self.other_session = AcademicYear.objects.create(
            institute=self.other_institute,
            name="2025-26",
            start_date=date(2025, 4, 1),
            end_date=date(2026, 3, 31),
        )
        self.client.force_login(self.admin_user)

    def test_page_lists_only_current_institute_sessions(self):
        response = self.client.get(reverse("institute_admin:academic_session_settings"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.session.name)
        self.assertNotContains(response, self.other_institute.name)

    def test_create_session_assigns_current_institute_without_syncing_students(self):
        student_user = User.objects.create_user(
            username="mobile-session-student",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        student = StudentProfile.objects.create(
            institute=self.institute,
            user=student_user,
            admission_number="OLD-001",
            is_active=True,
        )
        response = self.client.post(
            reverse("institute_admin:academic_session_settings"),
            {
                "action": "save",
                "name": "2026-27",
                "start_date": "2026-04-01",
                "end_date": "2027-03-31",
                "is_active": "on",
            },
        )

        self.assertRedirects(response, reverse("institute_admin:academic_session_settings"))
        self.assertTrue(
            AcademicYear.objects.filter(
                institute=self.institute,
                name="2026-27",
            ).exists()
        )
        created_year = AcademicYear.objects.get(institute=self.institute, name="2026-27")
        self.assertFalse(
            StudentAcademicSession.objects.filter(
                student=student,
                academic_year=created_year,
            ).exists()
        )

    def test_sync_students_action_is_not_available(self):
        student_user = User.objects.create_user(
            username="existing-session-student",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        student = StudentProfile.objects.create(
            institute=self.institute,
            user=student_user,
            admission_number="OLD-002",
            is_active=True,
        )

        response = self.client.post(
            reverse("institute_admin:academic_session_settings"),
            {
                "action": "sync_students",
                "session_id": self.session.pk,
            },
        )

        self.assertRedirects(response, reverse("institute_admin:academic_session_settings"))
        self.assertEqual(
            StudentAcademicSession.objects.filter(
                student=student,
                academic_year=self.session,
            ).count(),
            0,
        )

    def test_form_rejects_overlapping_dates_and_duplicate_name(self):
        overlap_form = AcademicYearForm(
            {
                "name": "Overlap",
                "start_date": "2026-03-01",
                "end_date": "2027-02-28",
                "is_active": "on",
            },
            institute=self.institute,
        )
        duplicate_form = AcademicYearForm(
            {
                "name": "2025-26",
                "start_date": "2027-04-01",
                "end_date": "2028-03-31",
                "is_active": "on",
            },
            institute=self.institute,
        )

        self.assertFalse(overlap_form.is_valid())
        self.assertIn("start_date", overlap_form.errors)
        self.assertFalse(duplicate_form.is_valid())
        self.assertIn("name", duplicate_form.errors)

    def test_cannot_modify_another_institute_session(self):
        response = self.client.post(
            reverse("institute_admin:academic_session_settings"),
            {
                "action": "toggle_active",
                "session_id": self.other_session.pk,
            },
        )

        self.assertEqual(response.status_code, 404)
        self.other_session.refresh_from_db()
        self.assertTrue(self.other_session.is_active)

    def test_set_current_updates_selected_session(self):
        response = self.client.post(
            reverse("institute_admin:academic_session_settings"),
            {
                "action": "set_current",
                "session_id": self.session.pk,
            },
        )

        self.assertRedirects(response, reverse("institute_admin:academic_session_settings"))
        self.assertEqual(self.client.session["academic_year_id"], self.session.pk)

    def test_linked_session_is_not_deleted(self):
        Course.objects.create(
            institute=self.institute,
            academic_year=self.session,
            name="Protected Course",
        )

        response = self.client.post(
            reverse("institute_admin:academic_session_settings"),
            {
                "action": "delete",
                "session_id": self.session.pk,
            },
            follow=True,
        )

        self.assertTrue(AcademicYear.objects.filter(pk=self.session.pk).exists())
        self.assertContains(response, "cannot be deleted")


class InstituteProfileTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Original Institute",
            code="original-institute",
            owner_name="Original Owner",
            phone="9000000000",
            email="original@example.com",
            address="Original address",
            status=Institute.Status.ACTIVE,
        )
        self.admin_user = User.objects.create_user(
            username="profile-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)

    def profile_data(self, **overrides):
        data = {
            "name": "Updated Institute",
            "code": "updated-institute",
            "institute_type": Institute.InstituteType.SCHOOL,
            "owner_name": "Updated Owner",
            "phone": "9111111111",
            "email": "updated@example.com",
            "address": "Updated address",
        }
        data.update(overrides)
        return data

    def test_profile_page_displays_current_institute_data(self):
        response = self.client.get(reverse("institute_profile"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Original Institute")
        self.assertContains(response, "Original Owner")
        self.assertContains(response, "Active")

    def test_profile_update_saves_only_current_institute(self):
        other_institute = Institute.objects.create(
            name="Other Institute",
            code="other-institute",
        )

        response = self.client.post(
            reverse("institute_profile"),
            self.profile_data(),
        )

        self.assertRedirects(response, reverse("institute_profile"))
        self.institute.refresh_from_db()
        other_institute.refresh_from_db()
        self.assertEqual(self.institute.name, "Updated Institute")
        self.assertEqual(self.institute.code, "updated-institute")
        self.assertEqual(self.institute.institute_type, Institute.InstituteType.SCHOOL)
        self.assertEqual(self.institute.address, "Updated address")
        self.assertEqual(other_institute.name, "Other Institute")

    def test_profile_rejects_duplicate_institute_code(self):
        Institute.objects.create(name="Existing Institute", code="existing-code")

        response = self.client.post(
            reverse("institute_profile"),
            self.profile_data(code="existing-code"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This institute code is already in use.")
        self.institute.refresh_from_db()
        self.assertEqual(self.institute.code, "original-institute")


class SubscriptionBillingPageTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Billing School",
            code="billing-school",
            status=Institute.Status.ACTIVE,
        )
        self.subscription = InstituteSubscription.objects.create(
            institute=self.institute,
            plan=InstituteSubscription.Plan.PREMIUM,
            starts_on=date(2026, 6, 1),
            ends_on=date(2026, 12, 31),
        )
        self.admin_user = User.objects.create_user(
            username="billing-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)

    def test_billing_page_shows_subscription_and_payment_history(self):
        SubscriptionPayment.objects.create(
            institute=self.institute,
            amount=Decimal("12000.00"),
            paid_on=date(2026, 6, 1),
            method=SubscriptionPayment.Method.BANK_TRANSFER,
            transaction_id="TXN-PREMIUM-001",
            notes="Annual renewal",
        )

        response = self.client.get(reverse("subscription_billing"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Premium")
        self.assertContains(response, "01 Jun 2026")
        self.assertContains(response, "31 Dec 2026")
        self.assertContains(response, "TXN-PREMIUM-001")
        self.assertContains(response, "12000.00")

    def test_billing_page_does_not_show_another_institutes_payments(self):
        other_institute = Institute.objects.create(name="Other School", code="other-school")
        SubscriptionPayment.objects.create(
            institute=other_institute,
            amount=Decimal("5000.00"),
            method=SubscriptionPayment.Method.UPI,
            transaction_id="PRIVATE-TXN",
        )

        response = self.client.get(reverse("subscription_billing"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "PRIVATE-TXN")
        self.assertContains(response, "No payments recorded yet")

    def test_expired_institute_can_still_open_billing_page(self):
        self.subscription.ends_on = date(2026, 5, 31)
        self.subscription.save(update_fields=["ends_on"])

        response = self.client.get(reverse("subscription_billing"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Expired")

    def test_payment_row_links_to_printable_bill(self):
        payment = SubscriptionPayment.objects.create(
            institute=self.institute,
            amount=Decimal("5000.00"),
            paid_on=date(2026, 6, 7),
            method=SubscriptionPayment.Method.UPI,
            transaction_id="UPI-5000",
        )

        response = self.client.get(reverse("subscription_billing"))
        bill_url = reverse(
            "institute_admin:subscription_payment_bill",
            args=[payment.pk],
        )

        self.assertContains(response, "View Bill")
        self.assertContains(response, bill_url)

        bill_response = self.client.get(bill_url)
        self.assertEqual(bill_response.status_code, 200)
        self.assertContains(bill_response, "Payment Receipt")
        self.assertContains(bill_response, "Billing School")
        self.assertContains(bill_response, "UPI-5000")
        self.assertContains(bill_response, "Amount Paid: INR 5000.00")

    def test_cannot_open_another_institutes_payment_bill(self):
        other_institute = Institute.objects.create(name="Other School", code="private-school")
        other_payment = SubscriptionPayment.objects.create(
            institute=other_institute,
            amount=Decimal("9000.00"),
            method=SubscriptionPayment.Method.CASH,
        )

        response = self.client.get(
            reverse(
                "institute_admin:subscription_payment_bill",
                args=[other_payment.pk],
            )
        )

        self.assertEqual(response.status_code, 404)


class SecuritySettingsTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Secure Institute",
            code="secure-institute",
            status=Institute.Status.ACTIVE,
        )
        self.admin_user = User.objects.create_user(
            username="security-admin",
            password="CurrentPass123!",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)

    def test_security_page_shows_only_current_institute_accounts(self):
        teacher = User.objects.create_user(username="secure-teacher", password="TeacherPass123!")
        UserProfile.objects.create(
            user=teacher,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        other_institute = Institute.objects.create(name="Other Institute", code="other-secure")
        other_user = User.objects.create_user(username="private-user", password="PrivatePass123!")
        UserProfile.objects.create(
            user=other_user,
            institute=other_institute,
            role=UserProfile.Role.TEACHER,
        )

        response = self.client.get(reverse("security_settings"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Security Settings")
        self.assertContains(response, "secure-teacher")
        self.assertNotContains(response, "private-user")

    def test_password_change_requires_current_password(self):
        response = self.client.post(
            reverse("security_settings"),
            {
                "action": "change_password",
                "current_password": "WrongPass123!",
                "new_password": "UpdatedPass123!",
                "confirm_password": "UpdatedPass123!",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Current password is incorrect.")
        self.admin_user.refresh_from_db()
        self.assertTrue(self.admin_user.check_password("CurrentPass123!"))

    def test_password_change_keeps_current_session_signed_in(self):
        response = self.client.post(
            reverse("security_settings"),
            {
                "action": "change_password",
                "current_password": "CurrentPass123!",
                "new_password": "UpdatedPass123!",
                "confirm_password": "UpdatedPass123!",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.admin_user.refresh_from_db()
        self.assertTrue(self.admin_user.check_password("UpdatedPass123!"))
        self.assertEqual(int(self.client.session["_auth_user_id"]), self.admin_user.pk)

    def test_sign_out_other_sessions_keeps_current_session(self):
        other_client = self.client_class()
        other_client.force_login(self.admin_user)
        other_session_key = other_client.session.session_key

        response = self.client.post(
            reverse("security_settings"),
            {"action": "logout_other_sessions"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Session.objects.filter(session_key=other_session_key).exists())
        self.assertEqual(int(self.client.session["_auth_user_id"]), self.admin_user.pk)

    def test_security_page_shows_browser_device_ip_and_login_activity(self):
        response = self.client.get(
            reverse("security_settings"),
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36"
            ),
            REMOTE_ADDR="192.0.2.25",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Logged In Devices")
        self.assertContains(response, "Google Chrome")
        self.assertContains(response, "Windows")
        self.assertContains(response, "Computer")
        self.assertContains(response, "192.0.2.25")
        self.assertContains(response, "Current")

    def test_can_sign_out_one_selected_device(self):
        other_client = self.client_class()
        other_client.force_login(self.admin_user)
        other_session_key = other_client.session.session_key
        other_client.get(
            reverse("security_settings"),
            HTTP_USER_AGENT="Mozilla/5.0 (Android 14; Mobile; rv:125.0) Firefox/125.0",
            REMOTE_ADDR="198.51.100.8",
        )

        response = self.client.get(reverse("security_settings"))
        other_session = next(
            session
            for session in response.context["web_sessions"]
            if session["session_key"] == other_session_key
        )
        sign_out_response = self.client.post(
            reverse("security_settings"),
            {
                "action": "logout_session",
                "session_identifier": other_session["identifier"],
            },
            follow=True,
        )

        self.assertEqual(sign_out_response.status_code, 200)
        self.assertFalse(Session.objects.filter(session_key=other_session_key).exists())
        self.assertEqual(int(self.client.session["_auth_user_id"]), self.admin_user.pk)
        self.assertContains(sign_out_response, "The selected device has been signed out.")

    def test_non_institute_admin_cannot_open_security_settings(self):
        teacher = User.objects.create_user(username="blocked-teacher", password="TeacherPass123!")
        UserProfile.objects.create(
            user=teacher,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        self.client.force_login(teacher)

        response = self.client.get(reverse("security_settings"))

        self.assertEqual(response.status_code, 302)

    def test_expired_institute_can_still_open_security_settings(self):
        InstituteSubscription.objects.create(
            institute=self.institute,
            plan=InstituteSubscription.Plan.PREMIUM,
            starts_on=date(2026, 1, 1),
            ends_on=date(2026, 1, 31),
        )

        response = self.client.get(reverse("security_settings"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Security Settings")


class HelpSupportTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Supported Institute",
            code="supported-institute",
            status=Institute.Status.ACTIVE,
        )
        self.admin_user = User.objects.create_user(
            username="support-admin",
            password="SupportPass123!",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)

    def test_help_page_contains_guides_contacts_and_ticket_form(self):
        response = self.client.get(reverse("help_support"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Quick Help Topics")
        self.assertContains(response, "Submit a Support Request")
        self.assertContains(response, "+91 7776824564")
        self.assertContains(response, "ultoxy.tech@gmail.com")

    def test_institute_can_submit_support_ticket(self):
        response = self.client.post(
            reverse("help_support"),
            {
                "category": SupportTicket.Category.FEES,
                "priority": SupportTicket.Priority.NORMAL,
                "subject": "Payment receipt issue",
                "message": "The payment receipt does not open for one student.",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        ticket = SupportTicket.objects.get()
        self.assertEqual(ticket.institute, self.institute)
        self.assertEqual(ticket.created_by, self.admin_user)
        self.assertContains(response, "Payment receipt issue")

    def test_help_page_does_not_show_another_institutes_tickets(self):
        other_institute = Institute.objects.create(
            name="Private Institute",
            code="private-support",
        )
        SupportTicket.objects.create(
            institute=other_institute,
            category=SupportTicket.Category.TECHNICAL,
            subject="PRIVATE SUPPORT ISSUE",
            message="This request must only be visible to the other institute.",
        )

        response = self.client.get(reverse("help_support"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "PRIVATE SUPPORT ISSUE")

    def test_admin_response_is_visible_in_ticket_history(self):
        SupportTicket.objects.create(
            institute=self.institute,
            created_by=self.admin_user,
            category=SupportTicket.Category.ACCOUNT,
            subject="Login assistance",
            message="A teacher account cannot sign in after password reset.",
            status=SupportTicket.Status.IN_PROGRESS,
            admin_response="The account has been checked. Please try the new password.",
        )

        response = self.client.get(reverse("help_support"))

        self.assertContains(response, "Support response:")
        self.assertContains(response, "Please try the new password.")

    def test_expired_institute_can_open_help_support(self):
        InstituteSubscription.objects.create(
            institute=self.institute,
            plan=InstituteSubscription.Plan.PREMIUM,
            starts_on=date(2026, 1, 1),
            ends_on=date(2026, 1, 31),
        )

        response = self.client.get(reverse("help_support"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Quick Help Topics")


class BatchTimetableFormTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Timetable Institute",
            code="timetable-institute",
            status=Institute.Status.ACTIVE,
        )
        self.academic_year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Science",
        )
        self.admin_user = User.objects.create_user(
            username="timetable-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)
        session = self.client.session
        session["academic_year_id"] = self.academic_year.pk
        session.save()

    def form_data(self, timetable="{}"):
        return {
            "courses": [self.course.pk],
            "name": "Morning Batch",
            "teachers": [],
            "start_date": "",
            "end_date": "",
            "timing": "",
            "weekly_timetable": timetable,
            "is_active": "on",
        }

    def test_timetable_is_optional(self):
        form = BatchForm(
            data=self.form_data(),
            institute=self.institute,
            academic_year=self.academic_year,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["weekly_timetable"], {})

    def test_valid_timetable_is_normalized(self):
        form = BatchForm(
            data=self.form_data(
                '{"monday":{"start":"09:00","end":"11:00"},'
                '"friday":{"start":"14:30","end":"16:00"}}'
            ),
            institute=self.institute,
            academic_year=self.academic_year,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(
            form.cleaned_data["weekly_timetable"]["monday"],
            {"start": "09:00", "end": "11:00"},
        )

    def test_timetable_rejects_end_before_start(self):
        form = BatchForm(
            data=self.form_data(
                '{"monday":{"start":"11:00","end":"09:00"}}'
            ),
            institute=self.institute,
            academic_year=self.academic_year,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("weekly_timetable", form.errors)

    def test_create_page_renders_optional_collapsed_timetable(self):
        response = self.client.get(reverse("institute_admin:batch_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Weekly Timetable")
        self.assertContains(response, "Optional")
        self.assertContains(response, 'aria-expanded="false"')
        self.assertContains(response, 'id="timetableBuilder"')
        self.assertContains(response, "hidden")

    def test_create_view_saves_client_generated_timetable(self):
        response = self.client.post(
            reverse("institute_admin:batch_create"),
            self.form_data(
                '{"monday":{"start":"09:00","end":"11:00"},'
                '"wednesday":{"start":"13:00","end":"15:00"}}'
            ),
        )

        self.assertEqual(response.status_code, 200)
        batch = Batch.objects.get(name="Morning Batch")
        self.assertEqual(
            batch.weekly_timetable,
            {
                "monday": {"start": "09:00", "end": "11:00"},
                "wednesday": {"start": "13:00", "end": "15:00"},
            },
        )

    def create_batch(self, timetable=None):
        batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Existing Batch",
            weekly_timetable=timetable or {},
        )
        batch.courses.add(self.course)
        return batch

    def test_update_page_shows_collapsed_builder_when_timetable_is_empty(self):
        batch = self.create_batch()

        response = self.client.get(
            reverse("institute_admin:batch_update", args=[batch.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Weekly Timetable")
        self.assertContains(response, 'aria-expanded="false"')
        self.assertContains(response, 'id="timetableBuilder"')

    def test_update_page_loads_existing_timetable_for_client_expansion(self):
        batch = self.create_batch(
            {"tuesday": {"start": "10:00", "end": "12:00"}}
        )

        response = self.client.get(
            reverse("institute_admin:batch_update", args=[batch.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '{&quot;tuesday&quot;: {&quot;start&quot;: &quot;10:00&quot;, '
            '&quot;end&quot;: &quot;12:00&quot;}}',
        )
        self.assertContains(
            response,
            "if(Object.keys(initialSchedule).length",
        )

    def test_update_view_can_add_or_replace_timetable(self):
        batch = self.create_batch()
        data = self.form_data(
            '{"tuesday":{"start":"10:00","end":"12:00"},'
            '"thursday":{"start":"14:00","end":"16:30"}}'
        )
        data["name"] = batch.name

        response = self.client.post(
            reverse("institute_admin:batch_update", args=[batch.pk]),
            data,
        )

        self.assertEqual(response.status_code, 200)
        batch.refresh_from_db()
        self.assertEqual(
            batch.weekly_timetable,
            {
                "tuesday": {"start": "10:00", "end": "12:00"},
                "thursday": {"start": "14:00", "end": "16:30"},
            },
        )


class LeadCrudTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Lead Institute",
            code="lead-institute",
            status=Institute.Status.ACTIVE,
        )
        self.other_institute = Institute.objects.create(
            name="Other Institute",
            code="other-lead-institute",
            status=Institute.Status.ACTIVE,
        )
        self.academic_year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 10",
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Morning Batch",
        )
        self.batch.courses.add(self.course)
        self.admin_user = User.objects.create_user(
            username="lead-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)
        session = self.client.session
        session["academic_year_id"] = self.academic_year.pk
        session.save()

    def lead_data(self, **overrides):
        data = {
            "first_name": "Rahul",
            "last_name": "Sharma",
            "mobile_number": "9876543210",
            "email": "rahul@example.com",
            "interested_class": self.course.pk,
            "interested_batch": self.batch.pk,
            "source": Lead.Source.PHONE,
            "status": Lead.Status.NEW,
            "follow_up_on": "2026-06-15",
            "message": "Interested in weekday classes.",
        }
        data.update(overrides)
        return data

    def test_front_office_lead_list_is_available(self):
        response = self.client.get(reverse("institute_admin:lead_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Lead Management")
        self.assertContains(response, reverse("institute_admin:lead_create"))
        self.assertContains(response, "openLeadWindow(this)")
        self.assertContains(response, 'window.open(url, "leadWindow", features)')
        self.assertContains(response, 'id="sidebarMenuSearch"')
        self.assertContains(response, 'data-bs-target="#front-office-nav"')
        self.assertContains(response, 'data-bs-target="#accounts-nav"')
        self.assertContains(response, "bi bi-person-plus")
        self.assertContains(response, "No matching menu found.")
        self.assertContains(response, "sidebar-logout-item")
        self.assertContains(response, 'class="menu-green" style="text-decoration:none;"')

    def test_lead_list_shows_conversion_modal_details(self):
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Modal",
            last_name="Student",
            mobile_number="9000000099",
            interested_class=self.course,
            interested_batch=self.batch,
        )

        response = self.client.get(reverse("institute_admin:lead_list"))

        self.assertContains(response, "Convert Lead to Student")
        self.assertContains(
            response,
            reverse("institute_admin:lead_convert", args=[lead.pk]),
        )
        self.assertContains(response, "Student@123")
        self.assertContains(response, 'id="convertLeadModal"')

    def test_lead_list_shows_delete_confirmation_modal(self):
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Delete",
            last_name="Candidate",
            mobile_number="9000000088",
        )

        response = self.client.get(reverse("institute_admin:lead_list"))

        self.assertContains(response, 'id="deleteLeadModal"')
        self.assertContains(response, "openDeleteLeadModal(this)")
        self.assertContains(response, 'window.jQuery(modalElement).modal("show")')
        self.assertContains(
            response,
            reverse("institute_admin:lead_delete", args=[lead.pk]),
        )
        self.assertContains(response, "Delete Candidate")
        self.assertNotContains(response, "confirm('Delete this lead?')")

    def test_create_form_includes_class_specific_batch_mapping(self):
        second_matching_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Evening Batch",
        )
        second_matching_batch.courses.add(self.course)
        other_course = Course.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 11",
        )
        other_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 11 Batch",
        )
        other_batch.courses.add(other_course)

        response = self.client.get(reverse("institute_admin:lead_create"))

        self.assertEqual(response.status_code, 200)
        batch_data = response.context["course_batch_data"][str(self.course.pk)]
        self.assertEqual(
            {batch["name"] for batch in batch_data},
            {"Morning Batch", "Evening Batch"},
        )
        self.assertNotIn("Class 11 Batch", {batch["name"] for batch in batch_data})
        self.assertContains(response, 'id="course-batch-data"')
        self.assertEqual(
            response.context["form"].fields["interested_class"].widget.attrs[
                "data-searchable"
            ],
            "false",
        )
        self.assertEqual(
            response.context["form"].fields["interested_batch"].widget.attrs[
                "data-searchable"
            ],
            "false",
        )
        self.assertContains(response, 'batchSelect.disabled = !classId')
        self.assertContains(response, 'setBatchOptions(classSelect.value, "")')

    def test_edit_form_preserves_existing_valid_batch_on_initial_load(self):
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Existing",
            last_name="Lead",
            mobile_number="9000000010",
            interested_class=self.course,
            interested_batch=self.batch,
        )

        response = self.client.get(
            reverse("institute_admin:lead_update", args=[lead.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "const initialBatchId = batchSelect.value")
        self.assertContains(
            response,
            "setBatchOptions(classSelect.value, initialBatchId)",
        )

    def test_create_update_and_delete_lead(self):
        create_response = self.client.post(
            reverse("institute_admin:lead_create"),
            self.lead_data(),
        )
        self.assertEqual(create_response.status_code, 200)
        self.assertContains(create_response, "window.opener.location.reload()")
        self.assertContains(create_response, "window.close()")

        lead = Lead.objects.get(first_name="Rahul")
        self.assertEqual(lead.last_name, "Sharma")
        self.assertEqual(lead.institute, self.institute)
        self.assertEqual(lead.created_by, self.admin_user)
        self.assertEqual(lead.interested_batch, self.batch)

        update_response = self.client.post(
            reverse("institute_admin:lead_update", args=[lead.pk]),
            self.lead_data(status=Lead.Status.FOLLOW_UP, message="Call tomorrow."),
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertContains(update_response, "window.opener.location.reload()")
        self.assertContains(update_response, "window.close()")
        lead.refresh_from_db()
        self.assertEqual(lead.status, Lead.Status.FOLLOW_UP)
        self.assertEqual(lead.message, "Call tomorrow.")

        delete_response = self.client.post(
            reverse("institute_admin:lead_delete", args=[lead.pk])
        )
        self.assertRedirects(delete_response, reverse("institute_admin:lead_list"))
        self.assertFalse(Lead.objects.filter(pk=lead.pk).exists())

    def test_list_filters_search_and_status(self):
        Lead.objects.create(
            institute=self.institute,
            first_name="Matching",
            last_name="Student",
            mobile_number="9000000001",
            status=Lead.Status.FOLLOW_UP,
        )
        Lead.objects.create(
            institute=self.institute,
            first_name="Other",
            last_name="Student",
            mobile_number="9000000002",
            status=Lead.Status.NEW,
        )

        response = self.client.get(
            reverse("institute_admin:lead_list"),
            {"search": "Matching", "status": Lead.Status.FOLLOW_UP},
        )

        self.assertContains(response, "Matching Student")
        self.assertNotContains(response, "Other Student")

    def test_cannot_access_another_institutes_lead(self):
        lead = Lead.objects.create(
            institute=self.other_institute,
            first_name="Private",
            last_name="Lead",
            mobile_number="9000000003",
        )

        edit_response = self.client.get(
            reverse("institute_admin:lead_update", args=[lead.pk])
        )
        delete_response = self.client.post(
            reverse("institute_admin:lead_delete", args=[lead.pk])
        )
        convert_response = self.client.post(
            reverse("institute_admin:lead_convert", args=[lead.pk])
        )

        self.assertEqual(edit_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)
        self.assertEqual(convert_response.status_code, 404)
        self.assertTrue(Lead.objects.filter(pk=lead.pk).exists())

    def test_batch_must_belong_to_interested_class(self):
        other_course = Course.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 11",
        )
        other_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 11 Batch",
        )
        other_batch.courses.add(other_course)

        response = self.client.post(
            reverse("institute_admin:lead_create"),
            self.lead_data(interested_batch=other_batch.pk),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selected batch must include the interested class.")
        self.assertFalse(Lead.objects.exists())

    def test_convert_lead_creates_student_session_and_enrollment(self):
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Converted",
            last_name="Student",
            mobile_number="9123456789",
            email="converted@example.com",
            interested_class=self.course,
            interested_batch=self.batch,
        )

        response = self.client.post(
            reverse("institute_admin:lead_convert", args=[lead.pk])
        )

        self.assertRedirects(response, reverse("institute_admin:lead_list"))
        user = User.objects.get(username="LEADINSTITUTE26270001")
        self.assertEqual(user.first_name, "Converted")
        self.assertEqual(user.last_name, "Student")
        self.assertEqual(user.email, "converted@example.com")
        self.assertTrue(user.check_password("Student@123"))
        self.assertEqual(user.profile.role, UserProfile.Role.STUDENT_PARENT)
        self.assertEqual(user.profile.phone, "9123456789")

        student = StudentProfile.objects.get(user=user)
        academic_session = StudentAcademicSession.objects.get(
            student=student,
            academic_year=self.academic_year,
        )
        enrollment = StudentEnrollment.objects.get(
            student=student,
            academic_session=academic_session,
            batch=self.batch,
        )
        self.assertQuerySetEqual(enrollment.courses.all(), [self.course])

        lead.refresh_from_db()
        self.assertEqual(lead.status, Lead.Status.CONVERTED)
        self.assertEqual(lead.converted_student, student)
        self.assertIsNotNone(lead.converted_at)

        repeat_response = self.client.post(
            reverse("institute_admin:lead_convert", args=[lead.pk])
        )
        self.assertRedirects(repeat_response, reverse("institute_admin:lead_list"))
        self.assertEqual(User.objects.filter(username="LEADINSTITUTE26270001").count(), 1)
        self.assertEqual(StudentEnrollment.objects.filter(student=student).count(), 1)

    def test_convert_lead_requires_class_and_batch(self):
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Incomplete",
            mobile_number="9234567890",
            interested_class=self.course,
        )

        response = self.client.post(
            reverse("institute_admin:lead_convert", args=[lead.pk])
        )

        self.assertRedirects(response, reverse("institute_admin:lead_list"))
        lead.refresh_from_db()
        self.assertEqual(lead.status, Lead.Status.NEW)
        self.assertFalse(User.objects.filter(username="9234567890").exists())

    def test_existing_mobile_username_does_not_block_lead_conversion(self):
        User.objects.create_user(username="9345678901", password="existing-pass")
        lead = Lead.objects.create(
            institute=self.institute,
            first_name="Duplicate",
            mobile_number="9345678901",
            interested_class=self.course,
            interested_batch=self.batch,
        )

        response = self.client.post(
            reverse("institute_admin:lead_convert", args=[lead.pk])
        )

        self.assertRedirects(response, reverse("institute_admin:lead_list"))
        lead.refresh_from_db()
        self.assertEqual(lead.status, Lead.Status.CONVERTED)
        self.assertIsNotNone(lead.converted_student)
        self.assertTrue(
            StudentProfile.objects.filter(user__username="LEADINSTITUTE26270001").exists()
        )


class VisitorCrudTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Visitor Institute",
            code="visitor-institute",
            status=Institute.Status.ACTIVE,
        )
        self.other_institute = Institute.objects.create(
            name="Other Visitor Institute",
            code="other-visitor-institute",
            status=Institute.Status.ACTIVE,
        )
        self.academic_year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Class 8",
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.academic_year,
            name="Evening Batch",
        )
        self.batch.courses.add(self.course)
        self.admin_user = User.objects.create_user(
            username="visitor-admin",
            password="pass12345",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin_user)
        session = self.client.session
        session["academic_year_id"] = self.academic_year.pk
        session.save()

    def visitor_data(self, **overrides):
        data = {
            "visitor_name": "Asha Patil",
            "phone_number": "9876501234",
            "id_card_number": "PASS-101",
            "meeting_with": "Admission Counsellor",
            "total_person": 2,
            "visit_date": "2026-06-10",
            "entry_time": "11:30",
            "exit_time": "12:15",
            "purpose": "Course admission enquiry",
        }
        data.update(overrides)
        return data

    def test_create_and_list_visitor(self):
        response = self.client.post(
            reverse("institute_admin:visitor_create"),
            self.visitor_data(),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "window.opener.location.reload()")
        self.assertContains(response, "window.close()")
        visitor = Visitor.objects.get(phone_number="9876501234")
        self.assertEqual(visitor.institute, self.institute)
        self.assertEqual(visitor.created_by, self.admin_user)
        self.assertEqual(visitor.meeting_with, "Admission Counsellor")
        self.assertEqual(visitor.total_person, 2)

        list_response = self.client.get(reverse("institute_admin:visitor_list"))
        self.assertContains(list_response, "Asha Patil")
        self.assertContains(list_response, "Course admission enquiry")
        self.assertContains(list_response, "PASS-101")
        self.assertContains(list_response, reverse("institute_admin:visitor_create"))
        self.assertContains(list_response, "openVisitorWindow(this)")
        self.assertContains(list_response, 'id="deleteVisitorModal"')
        self.assertContains(list_response, "openDeleteVisitorModal(this)")

    def test_update_and_delete_visitor(self):
        visitor = Visitor.objects.create(
            institute=self.institute,
            visitor_name="Old Visitor",
            phone_number="9000000011",
            meeting_with="Reception",
            visit_date="2026-06-10",
            entry_time="10:00",
            purpose="Initial visit",
        )

        update_response = self.client.post(
            reverse("institute_admin:visitor_update", args=[visitor.pk]),
            self.visitor_data(
                visitor_name="Updated Visitor",
                phone_number="9000000011",
                meeting_with="Principal",
            ),
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertContains(update_response, "window.opener.location.reload()")
        self.assertContains(update_response, "window.close()")
        visitor.refresh_from_db()
        self.assertEqual(visitor.visitor_name, "Updated Visitor")
        self.assertEqual(visitor.meeting_with, "Principal")

        delete_response = self.client.post(
            reverse("institute_admin:visitor_delete", args=[visitor.pk])
        )
        self.assertRedirects(delete_response, reverse("institute_admin:visitor_list"))
        self.assertFalse(Visitor.objects.filter(pk=visitor.pk).exists())

    def test_list_filters_by_search_and_status(self):
        Visitor.objects.create(
            institute=self.institute,
            visitor_name="Matching Visitor",
            phone_number="9000000021",
            meeting_with="Director",
            visit_date="2026-06-11",
            entry_time="10:00",
            purpose="Meet counsellor",
        )
        Visitor.objects.create(
            institute=self.institute,
            visitor_name="Hidden Visitor",
            phone_number="9000000022",
            meeting_with="Reception",
            visit_date="2026-06-12",
            entry_time="11:00",
            purpose="Other purpose",
        )

        response = self.client.get(
            reverse("institute_admin:visitor_list"),
            {"search": "Matching", "visit_date": "2026-06-11"},
        )

        self.assertContains(response, "Matching")
        self.assertNotContains(response, "9000000022")

    def test_exit_time_cannot_be_before_entry_time(self):
        response = self.client.post(
            reverse("institute_admin:visitor_create"),
            self.visitor_data(entry_time="14:00", exit_time="13:30"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exit time cannot be earlier than entry time.")
        self.assertFalse(Visitor.objects.exists())

    def test_total_person_must_be_at_least_one(self):
        response = self.client.post(
            reverse("institute_admin:visitor_create"),
            self.visitor_data(total_person=0),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ensure this value is greater than or equal to 1.")
        self.assertFalse(Visitor.objects.exists())

    def test_cannot_access_another_institutes_visitor(self):
        visitor = Visitor.objects.create(
            institute=self.other_institute,
            visitor_name="Private Visitor",
            phone_number="9000000031",
            meeting_with="Director",
            visit_date="2026-06-12",
            entry_time="10:00",
            purpose="Private visit",
        )

        edit_response = self.client.get(
            reverse("institute_admin:visitor_update", args=[visitor.pk])
        )
        delete_response = self.client.post(
            reverse("institute_admin:visitor_delete", args=[visitor.pk])
        )

        self.assertEqual(edit_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)
        self.assertTrue(Visitor.objects.filter(pk=visitor.pk).exists())


class AcademicSessionIsolationTests(TestCase):
    def setUp(self):
        cache.clear()
        self.institute = Institute.objects.create(
            name="Saint Monica International School",
            code="smis",
            status=Institute.Status.ACTIVE,
        )
        self.admin_user = User.objects.create_user(
            username="admin",
            password="pass12345",
            first_name="Admin",
        )
        UserProfile.objects.create(
            user=self.admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
            phone="9000000000",
        )
        self.student_user = User.objects.create_user(
            username="student-one",
            password="pass12345",
            first_name="Student",
            last_name="One",
        )
        UserProfile.objects.create(
            user=self.student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
            phone="9111111111",
        )
        self.student = StudentProfile.objects.create(
            institute=self.institute,
            user=self.student_user,
            academic_year=None,
            admission_number="LEGACY-0001",
            is_active=True,
        )
        self.year_2026 = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.year_2027 = AcademicYear.objects.create(
            institute=self.institute,
            name="2027-28",
            start_date=date(2027, 4, 1),
            end_date=date(2028, 3, 31),
        )
        self.session_2026 = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year_2026,
            admission_number="SMIS-2026-27-0001",
            joined_on=date(2026, 4, 5),
            status=StudentAcademicSession.Status.ACTIVE,
            previous_class="11th",
        )
        self.session_2027 = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year_2027,
            admission_number="SMIS-2027-28-0001",
            joined_on=date(2027, 4, 5),
            status=StudentAcademicSession.Status.ACTIVE,
            previous_class="12th",
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Science",
            fee_amount=Decimal("1000.00"),
            is_active=True,
        )
        self.course_2027 = Course.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            name="Science",
            fee_amount=Decimal("1000.00"),
            is_active=True,
        )
        self.batch_2026 = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="11th Batch",
            is_active=True,
        )
        self.batch_2026.courses.add(self.course)
        self.batch_2027 = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            name="12th Batch",
            is_active=True,
        )
        self.batch_2027.courses.add(self.course_2027)
        self.enrollment_2026 = StudentEnrollment.objects.create(
            academic_session=self.session_2026,
            student=self.student,
            batch=self.batch_2026,
            enrolled_on=date(2026, 4, 5),
            custom_fee_amount=Decimal("1000.00"),
        )
        self.enrollment_2026.courses.add(self.course)
        self.enrollment_2027 = StudentEnrollment.objects.create(
            academic_session=self.session_2027,
            student=self.student,
            batch=self.batch_2027,
            enrolled_on=date(2027, 4, 5),
            custom_fee_amount=Decimal("2000.00"),
        )
        self.enrollment_2027.courses.add(self.course_2027)
        self.invoice_2026 = FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            enrollment=self.enrollment_2026,
            batch=self.batch_2026,
            title="2026 Fee",
            amount=Decimal("1000.00"),
            due_date=date(2026, 5, 1),
        )
        self.invoice_2027 = FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2027,
            enrollment=self.enrollment_2027,
            batch=self.batch_2027,
            title="2027 Fee",
            amount=Decimal("2000.00"),
            due_date=date(2027, 5, 1),
        )
        self.payment_2026 = Payment.objects.create(
            invoice=self.invoice_2026,
            amount=Decimal("250.00"),
            paid_on=date(2026, 5, 2),
            method=Payment.Method.CASH,
            received_by=self.admin_user,
        )
        self.payment_2027 = Payment.objects.create(
            invoice=self.invoice_2027,
            amount=Decimal("500.00"),
            paid_on=date(2027, 5, 2),
            method=Payment.Method.CASH,
            received_by=self.admin_user,
        )
        self.attendance_2026 = Attendance.objects.create(
            academic_session=self.session_2026,
            student=self.student,
            batch=self.batch_2026,
            date=date(2026, 5, 3),
            status=Attendance.Status.PRESENT,
            marked_by=self.admin_user,
        )
        self.attendance_2027 = Attendance.objects.create(
            academic_session=self.session_2027,
            student=self.student,
            batch=self.batch_2027,
            date=date(2027, 5, 3),
            status=Attendance.Status.ABSENT,
            marked_by=self.admin_user,
        )
        self.client.force_login(self.admin_user)

    def select_year(self, academic_year):
        session = self.client.session
        session["academic_year_id"] = academic_year.pk
        session.save()

    def test_academic_year_list_is_cached_and_invalidated(self):
        with CaptureQueriesContext(connection) as first_queries:
            first = get_cached_academic_years(self.institute.pk)
        with CaptureQueriesContext(connection) as cached_queries:
            cached = get_cached_academic_years(self.institute.pk)

        self.assertEqual([year.pk for year in first], [year.pk for year in cached])
        self.assertEqual(len(first_queries), 1)
        self.assertEqual(len(cached_queries), 0)

        AcademicYear.objects.create(
            institute=self.institute,
            name="2028-29",
            start_date=date(2028, 4, 1),
            end_date=date(2029, 3, 31),
        )
        with CaptureQueriesContext(connection) as refreshed_queries:
            refreshed = get_cached_academic_years(self.institute.pk)

        self.assertEqual(len(refreshed_queries), 1)
        self.assertEqual(refreshed[0].name, "2028-29")

    def test_batch_course_lookup_is_cached_and_invalidated(self):
        with CaptureQueriesContext(connection) as first_queries:
            first = get_cached_batch_course_data(
                self.institute.pk,
                self.year_2026.pk,
            )
        with CaptureQueriesContext(connection) as cached_queries:
            cached = get_cached_batch_course_data(
                self.institute.pk,
                self.year_2026.pk,
            )

        self.assertEqual(first, cached)
        self.assertGreater(len(first_queries), 0)
        self.assertEqual(len(cached_queries), 0)

        self.course.fee_amount = Decimal("1250.00")
        self.course.save(update_fields=["fee_amount"])
        with CaptureQueriesContext(connection) as refreshed_queries:
            refreshed = get_cached_batch_course_data(
                self.institute.pk,
                self.year_2026.pk,
            )

        self.assertGreater(len(refreshed_queries), 0)
        self.assertEqual(
            refreshed[str(self.batch_2026.pk)][0]["fee"],
            "1250.00",
        )

    def test_student_list_uses_selected_academic_session(self):
        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SMIS-2026-27-0001")
        self.assertContains(response, "11th Batch")
        self.assertNotContains(response, "SMIS-2027-28-0001")
        self.assertNotContains(response, "12th Batch")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:student_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SMIS-2027-28-0001")
        self.assertContains(response, "12th Batch")
        self.assertNotContains(response, "SMIS-2026-27-0001")
        self.assertNotContains(response, "11th Batch")

    def test_notice_list_uses_selected_academic_session(self):
        Notice.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            title="2026 Session Notice",
            message="Only 2026 students should see this.",
            audience=Notice.Audience.EVERYONE,
        )
        Notice.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            title="2027 Session Notice",
            message="Only 2027 students should see this.",
            audience=Notice.Audience.EVERYONE,
        )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:notice_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2026 Session Notice")
        self.assertNotContains(response, "2027 Session Notice")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:notice_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2027 Session Notice")
        self.assertNotContains(response, "2026 Session Notice")

    def test_student_notice_feed_uses_selected_academic_session(self):
        notice_2026 = Notice.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            title="Student 2026 Notice",
            message="Only 2026 feed.",
            audience=Notice.Audience.STUDENTS_PARENTS,
        )
        notice_2027 = Notice.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            title="Student 2027 Notice",
            message="Only 2027 feed.",
            audience=Notice.Audience.STUDENTS_PARENTS,
        )

        visible_2026 = Notice.for_student(self.student, academic_session_id=self.session_2026.pk)
        self.assertIn(notice_2026, visible_2026)
        self.assertNotIn(notice_2027, visible_2026)

        visible_2027 = Notice.for_student(self.student, academic_session_id=self.session_2027.pk)
        self.assertIn(notice_2027, visible_2027)
        self.assertNotIn(notice_2026, visible_2027)

    def test_fee_category_list_uses_selected_academic_session(self):
        FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="2026 Transport",
            default_amount=Decimal("500.00"),
        )
        FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            name="2027 Transport",
            default_amount=Decimal("700.00"),
        )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:fee_category_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2026 Transport")
        self.assertNotContains(response, "2027 Transport")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:fee_category_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2027 Transport")
        self.assertNotContains(response, "2026 Transport")

    def test_expense_list_uses_selected_academic_session(self):
        Expense.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            title="2026 Repair",
            amount=Decimal("125.00"),
            spent_on=date(2026, 7, 1),
            recorded_by=self.admin_user,
        )
        Expense.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            title="2027 Repair",
            amount=Decimal("175.00"),
            spent_on=date(2027, 7, 1),
            recorded_by=self.admin_user,
        )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:expense_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2026 Repair")
        self.assertNotContains(response, "2027 Repair")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:expense_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2027 Repair")
        self.assertNotContains(response, "2026 Repair")

    def test_expense_export_can_include_media_links(self):
        self.select_year(self.year_2026)
        expense = Expense.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            title="Receipt Expense",
            amount=Decimal("125.00"),
            spent_on=date(2026, 7, 1),
            recorded_by=self.admin_user,
        )
        ExpenseDocument.objects.create(
            expense=expense,
            file="expenses/receipts/receipt-one.pdf",
        )

        response = self.client.get(
            reverse("institute_admin:expense_export"),
            {"columns": ["title", "media_links"]},
        )

        self.assertEqual(response.status_code, 200)
        export_text = response.content.decode("utf-8-sig")
        self.assertIn("Title,Media Links", export_text)
        self.assertIn("Receipt Expense", export_text)
        self.assertIn(
            "http://testserver/media/expenses/receipts/receipt-one.pdf",
            export_text,
        )

    def test_lead_list_uses_selected_academic_session(self):
        Lead.objects.create(
            institute=self.institute,
            first_name="Lead",
            last_name="TwentySix",
            mobile_number="9000002026",
            interested_class=self.course,
            interested_batch=self.batch_2026,
        )
        Lead.objects.create(
            institute=self.institute,
            first_name="Lead",
            last_name="TwentySeven",
            mobile_number="9000002027",
            academic_year=self.year_2027,
            interested_class=self.course_2027,
            interested_batch=self.batch_2027,
        )
        Lead.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            first_name="Open",
            last_name="TwentySix",
            mobile_number="9000003026",
        )
        Lead.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            first_name="Open",
            last_name="TwentySeven",
            mobile_number="9000003027",
        )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:lead_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Lead TwentySix")
        self.assertContains(response, "Open TwentySix")
        self.assertNotContains(response, "Lead TwentySeven")
        self.assertNotContains(response, "Open TwentySeven")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:lead_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Lead TwentySeven")
        self.assertContains(response, "Open TwentySeven")
        self.assertNotContains(response, "Lead TwentySix")
        self.assertNotContains(response, "Open TwentySix")

    def test_visitor_list_uses_selected_academic_session_dates(self):
        Visitor.objects.create(
            institute=self.institute,
            visitor_name="Visitor TwentySix",
            phone_number="9000012026",
            meeting_with="Office",
            visit_date=date(2026, 6, 1),
            entry_time="10:00",
        )
        Visitor.objects.create(
            institute=self.institute,
            visitor_name="Visitor TwentySeven",
            phone_number="9000012027",
            meeting_with="Office",
            visit_date=date(2027, 6, 1),
            entry_time="10:00",
        )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:visitor_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Visitor TwentySix")
        self.assertNotContains(response, "Visitor TwentySeven")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:visitor_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Visitor TwentySeven")
        self.assertNotContains(response, "Visitor TwentySix")

    def test_teacher_list_uses_selected_academic_session_assignments(self):
        teacher_2026_user = User.objects.create_user(
            username="teacher-2026",
            password="pass12345",
            first_name="Teacher",
            last_name="TwentySix",
        )
        UserProfile.objects.create(
            user=teacher_2026_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        TeacherProfile.objects.create(
            institute=self.institute,
            user=teacher_2026_user,
            employee_id="T2026",
        )
        self.batch_2026.teachers.add(teacher_2026_user)

        teacher_2027_user = User.objects.create_user(
            username="teacher-2027",
            password="pass12345",
            first_name="Teacher",
            last_name="TwentySeven",
        )
        UserProfile.objects.create(
            user=teacher_2027_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        TeacherProfile.objects.create(
            institute=self.institute,
            user=teacher_2027_user,
            employee_id="T2027",
        )
        self.batch_2027.teachers.add(teacher_2027_user)

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:teacher_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Teacher TwentySix")
        self.assertNotContains(response, "Teacher TwentySeven")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:teacher_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Teacher TwentySeven")
        self.assertNotContains(response, "Teacher TwentySix")

    def test_user_access_list_uses_selected_academic_session(self):
        teacher_2026_user = User.objects.create_user(
            username="access-teacher-2026",
            password="pass12345",
            first_name="Access",
            last_name="Teacher2026",
        )
        UserProfile.objects.create(
            user=teacher_2026_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        TeacherProfile.objects.create(institute=self.institute, user=teacher_2026_user)
        self.batch_2026.teachers.add(teacher_2026_user)

        teacher_2027_user = User.objects.create_user(
            username="access-teacher-2027",
            password="pass12345",
            first_name="Access",
            last_name="Teacher2027",
        )
        UserProfile.objects.create(
            user=teacher_2027_user,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        TeacherProfile.objects.create(institute=self.institute, user=teacher_2027_user)
        self.batch_2027.teachers.add(teacher_2027_user)

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:user_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "access-teacher-2026")
        self.assertContains(response, "student-one")
        self.assertNotContains(response, "access-teacher-2027")

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:user_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "access-teacher-2027")
        self.assertContains(response, "student-one")
        self.assertNotContains(response, "access-teacher-2026")

    def test_student_list_rows_link_to_student_dashboard(self):
        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_list"))
        dashboard_url = reverse("institute_admin:student_dashboard", args=[self.student.pk])

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="student-clickable-row"')
        self.assertContains(response, f'data-dashboard-url="{dashboard_url}"')
        self.assertContains(response, f'href="{dashboard_url}"')

    def test_student_list_search_has_autocomplete_suggestions(self):
        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_list"))
        autocomplete_url = reverse("institute_admin:student_autocomplete")
        dashboard_url_template = reverse("institute_admin:student_dashboard", args=[0])

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="studentListSearchInput"')
        self.assertContains(response, f'data-autocomplete-url="{autocomplete_url}"')
        self.assertContains(response, f'data-dashboard-url-template="{dashboard_url_template}"')
        self.assertContains(response, 'id="studentListSearchSuggestions"')

    def test_student_list_calculates_fee_details_for_current_page_only(self):
        users = [
            User(username=f"page-student-{index}", first_name=f"Page {index}")
            for index in range(1, 25)
        ]
        User.objects.bulk_create(users)
        users = list(User.objects.filter(username__startswith="page-student-").order_by("username"))
        UserProfile.objects.bulk_create(
            [
                UserProfile(
                    user=user,
                    institute=self.institute,
                    role=UserProfile.Role.STUDENT_PARENT,
                )
                for user in users
            ]
        )
        students = [
            StudentProfile(
                institute=self.institute,
                academic_year=self.year_2026,
                user=user,
                admission_number=f"PAGE-{index:04d}",
            )
            for index, user in enumerate(users, start=1)
        ]
        StudentProfile.objects.bulk_create(students)
        students = list(
            StudentProfile.objects.filter(admission_number__startswith="PAGE-").order_by("admission_number")
        )
        StudentAcademicSession.objects.bulk_create(
            [
                StudentAcademicSession(
                    institute=self.institute,
                    student=student,
                    academic_year=self.year_2026,
                    admission_number=student.admission_number,
                )
                for student in students
            ]
        )
        self.select_year(self.year_2026)

        with patch.object(
            views,
            "get_student_session_fee_summaries",
            wraps=views.get_student_session_fee_summaries,
        ) as fee_summary_builder:
            response = self.client.get(reverse("institute_admin:student_list"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["paginator"].count, 25)
        self.assertEqual(len(response.context["students"]), 20)
        self.assertEqual(fee_summary_builder.call_count, 1)
        self.assertEqual(len(list(fee_summary_builder.call_args.args[0])), 20)

    def test_academic_year_switcher_includes_and_switches_inactive_years(self):
        self.year_2027.is_active = False
        self.year_2027.save(update_fields=["is_active"])

        response = self.client.get(reverse("institute_admin:student_list"))
        self.assertContains(response, f'value="{self.year_2026.pk}"')
        self.assertContains(response, f'value="{self.year_2027.pk}"')
        self.assertContains(response, "2027-28 (Inactive)")

        response = self.client.post(
            reverse("institute_admin:academic_year_switch"),
            {"academic_year_id": str(self.year_2027.pk)},
            HTTP_REFERER=reverse("institute_admin:student_list"),
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"), fetch_redirect_response=False)
        self.assertEqual(str(self.client.session["academic_year_id"]), str(self.year_2027.pk))

        response = self.client.get(reverse("institute_admin:student_list"))
        self.assertContains(response, "SMIS-2027-28-0001")
        self.assertNotContains(response, "SMIS-2026-27-0001")

    def test_student_bulk_import_creates_students_in_selected_year(self):
        self.select_year(self.year_2026)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        headers = [
            "First Name *",
            "Last Name",
            "Password",
            "Email",
            "Phone",
            "Date of Birth",
            "Joined On",
            "Address",
            "Current School / College",
            "Current School Address",
            "Previous School / College",
            "Previous Class",
            "Guardian Name",
            "Guardian Relation",
            "Guardian Phone",
            "Guardian Email",
            "Active",
        ]
        sheet.append(["Student Bulk Import Template"])
        sheet.append(["Generated admission numbers"])
        sheet.append(headers)
        sheet.append(
            [
                "Bulk",
                "One",
                "",
                "bulk-one@example.com",
                "9222222221",
                "2012-01-01",
                "2026-04-10",
                "Address one",
                "Current School",
                "Current address",
                "Previous School",
                "8th",
                "Guardian One",
                "Father",
                "9333333331",
                "guardian-one@example.com",
                "Yes",
            ]
        )
        sheet.append(
            [
                "Bulk",
                "Two",
                "Secret123",
                "bulk-two@example.com",
                "9222222222",
                "2012-02-02",
                "2026-04-11",
                "Address two",
                "Current School",
                "Current address",
                "Previous School",
                "9th",
                "",
                "",
                "",
                "",
                "No",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "students.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_import"),
            {"student_file": upload},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        imported_session = StudentAcademicSession.objects.get(admission_number="SMIS26270002")
        self.assertEqual(imported_session.academic_year, self.year_2026)
        self.assertEqual(imported_session.student.user.username, "SMIS26270002")
        self.assertTrue(imported_session.student.user.check_password("Student@123"))
        self.assertTrue(imported_session.student.guardians.filter(name="Guardian One").exists())

        custom_user = User.objects.get(username="SMIS26270003")
        self.assertTrue(custom_user.check_password("Secret123"))
        custom_session = StudentAcademicSession.objects.get(student=custom_user.student_profile)
        self.assertEqual(custom_session.status, StudentAcademicSession.Status.LEFT)

    def test_student_bulk_import_accepts_accidentally_cleared_header(self):
        self.select_year(self.year_2026)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        headers = views.student_import_columns()
        headers[6] = ""
        sheet.append(["Student Bulk Import Template"])
        sheet.append(["Generated admission numbers"])
        sheet.append(headers)
        sheet.append(
            [
                "Bulk",
                "Header",
                "",
                "bulk-header@example.com",
                "9222222299",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Yes",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "students.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_import_validate"),
            {"student_file": upload},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["valid"])
        self.assertEqual(payload["valid_count"], 1)
        self.assertIn("Joined On", payload["warnings"][0])

    def test_student_bulk_import_rejects_renamed_header(self):
        headers = views.student_import_columns()
        headers[6] = "Start Date"

        matches, missing = views.match_student_import_headers(headers)

        self.assertFalse(matches)
        self.assertEqual(missing, [])

    def test_dummy_student_create_builds_complete_unique_institute_records(self):
        self.select_year(self.year_2026)
        other_institute = Institute.objects.create(
            name="Other Institute",
            code="other",
            status=Institute.Status.ACTIVE,
        )
        other_student_count = StudentProfile.objects.filter(institute=other_institute).count()

        response = self.client.post(
            reverse("institute_admin:student_dummy_create"),
            {"count": "12"},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        generated_sessions = StudentAcademicSession.objects.filter(
            institute=self.institute,
            academic_year=self.year_2026,
            admission_number__startswith="SMIS2627",
        ).select_related(
            "student",
            "student__user",
            "student__user__profile",
        ).prefetch_related(
            "student__guardians",
            "enrollments__courses",
        )
        self.assertEqual(generated_sessions.count(), 12)
        self.assertEqual(
            StudentProfile.objects.filter(institute=other_institute).count(),
            other_student_count,
        )

        usernames = []
        emails = []
        phones = []
        guardian_phones = []
        for session in generated_sessions:
            student = session.student
            user = student.user
            profile = user.profile
            guardian = student.guardians.get(is_primary=True)
            enrollment = session.enrollments.get()

            usernames.append(user.username)
            emails.append(user.email)
            phones.append(profile.phone)
            guardian_phones.append(guardian.phone)
            self.assertTrue(user.first_name)
            self.assertTrue(user.last_name)
            self.assertTrue(user.email)
            self.assertTrue(user.check_password("Student@123"))
            self.assertTrue(student.date_of_birth)
            self.assertTrue(student.joined_on)
            self.assertTrue(student.address)
            self.assertTrue(student.current_school_name)
            self.assertTrue(student.current_school_address)
            self.assertTrue(student.previous_school_name)
            self.assertTrue(student.previous_class)
            self.assertTrue(session.joined_on)
            self.assertTrue(session.current_school_name)
            self.assertTrue(session.current_school_address)
            self.assertTrue(session.previous_school_name)
            self.assertTrue(session.previous_class)
            self.assertTrue(guardian.name)
            self.assertTrue(guardian.relation)
            self.assertTrue(guardian.phone)
            self.assertTrue(guardian.email)
            self.assertEqual(enrollment.batch, self.batch_2026)
            self.assertEqual(list(enrollment.courses.all()), [self.course])
            self.assertEqual(enrollment.custom_fee_amount, self.course.fee_amount)

        self.assertEqual(len(usernames), len(set(usernames)))
        self.assertEqual(len(emails), len(set(emails)))
        self.assertEqual(len(phones), len(set(phones)))
        self.assertEqual(len(guardian_phones), len(set(guardian_phones)))
        self.assertFalse(set(phones) & set(guardian_phones))

    def test_dummy_student_create_rejects_invalid_count(self):
        self.select_year(self.year_2026)
        before_count = StudentProfile.objects.filter(institute=self.institute).count()

        page_response = self.client.get(reverse("institute_admin:student_list"))
        self.assertContains(page_response, "Dummy Students")
        self.assertContains(page_response, 'id="dummyStudentModal"')
        self.assertContains(
            page_response,
            reverse("institute_admin:student_dummy_create"),
        )

        response = self.client.post(
            reverse("institute_admin:student_dummy_create"),
            {"count": "0"},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        self.assertEqual(
            StudentProfile.objects.filter(institute=self.institute).count(),
            before_count,
        )

    def test_dummy_student_create_supports_one_thousand_records(self):
        self.select_year(self.year_2026)

        response = self.client.post(
            reverse("institute_admin:student_dummy_create"),
            {"count": "1000"},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        generated_sessions = StudentAcademicSession.objects.filter(
            institute=self.institute,
            academic_year=self.year_2026,
            admission_number__startswith="SMIS2627",
        )
        generated_students = StudentProfile.objects.filter(
            institute=self.institute,
            academic_year=self.year_2026,
            admission_number__startswith="SMIS2627",
        )
        generated_user_ids = generated_students.values_list("user_id", flat=True)
        generated_guardians = GuardianProfile.objects.filter(student__in=generated_students)
        generated_enrollments = StudentEnrollment.objects.filter(
            academic_session__in=generated_sessions,
        )

        self.assertEqual(generated_sessions.count(), 1000)
        self.assertEqual(generated_students.count(), 1000)
        self.assertEqual(generated_guardians.count(), 1000)
        self.assertEqual(generated_enrollments.count(), 1000)
        self.assertEqual(
            User.objects.filter(pk__in=generated_user_ids).values("username").distinct().count(),
            1000,
        )
        self.assertEqual(
            User.objects.filter(pk__in=generated_user_ids).values("email").distinct().count(),
            1000,
        )
        self.assertEqual(
            UserProfile.objects.filter(user_id__in=generated_user_ids).values("phone").distinct().count(),
            1000,
        )
        self.assertEqual(
            generated_guardians.values("phone").distinct().count(),
            1000,
        )

    def test_manual_student_creation_generates_scoped_username_and_default_password(self):
        form = StudentForm(
            data={
                "first_name": "Generated",
                "last_name": "Login",
                "phone": "9444444444",
                "is_active": "on",
            },
            institute=self.institute,
            academic_year=self.year_2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        student = form.save()

        self.assertEqual(student.admission_number, "SMIS26270002")
        self.assertEqual(student.user.username, "SMIS26270002")
        self.assertTrue(student.user.check_password("Student@123"))

    def test_deactivating_student_preserves_existing_custom_and_extra_fees(self):
        self.enrollment_2026.custom_fee_amount = Decimal("750.00")
        self.enrollment_2026.save(update_fields=["custom_fee_amount"])
        extra_invoice = FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            title="Exam Fee",
            amount=Decimal("125.00"),
            due_date=date(2026, 6, 1),
        )

        form = StudentForm(
            data={
                "first_name": self.student.user.first_name,
                "middle_name": self.student.middle_name,
                "last_name": self.student.user.last_name,
                "email": self.student.user.email,
                "phone": self.student.user.profile.phone,
                "joined_on": "2026-04-05",
                "class_course": str(self.course.pk),
                "batch": str(self.batch_2026.pk),
                "fee_discount": "0.00",
            },
            institute=self.institute,
            student=self.student,
            academic_year=self.year_2026,
        )

        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.enrollment_2026.refresh_from_db()
        extra_invoice.refresh_from_db()
        self.session_2026.refresh_from_db()
        self.student.refresh_from_db()
        self.assertFalse(self.student.is_active)
        self.assertEqual(self.session_2026.status, StudentAcademicSession.Status.LEFT)
        self.assertEqual(self.enrollment_2026.custom_fee_amount, Decimal("750.00"))
        self.assertEqual(extra_invoice.amount, Decimal("125.00"))

    def test_profit_loss_report_calculates_collection_expense_and_export(self):
        self.select_year(self.year_2026)
        Expense.objects.create(
            institute=self.institute,
            title="Rent",
            amount=Decimal("75.00"),
            spent_on=date(2026, 5, 4),
            recorded_by=self.admin_user,
        )
        Expense.objects.create(
            institute=self.institute,
            title="Outside Period",
            amount=Decimal("900.00"),
            spent_on=date(2026, 7, 1),
            recorded_by=self.admin_user,
        )

        response = self.client.get(
            reverse("institute_admin:profit_loss_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "course": str(self.course.pk),
                "batch": str(self.batch_2026.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["collection_total"], Decimal("250.00"))
        self.assertEqual(response.context["expense_total"], Decimal("75.00"))
        self.assertEqual(response.context["net_profit"], Decimal("175.00"))
        self.assertContains(response, "Profit & Loss Report")
        self.assertContains(response, "Rent")

        export_response = self.client.get(
            reverse("institute_admin:profit_loss_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "course": str(self.course.pk),
                "batch": str(self.batch_2026.pk),
                "export": "csv",
            },
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "text/csv")
        export_text = export_response.content.decode()
        self.assertIn("Total Collection,250.00", export_text)
        self.assertIn("Total Expenses,75.00", export_text)
        self.assertIn("Net Profit/Loss,175.00", export_text)
        self.assertIn("Rent", export_text)

    def test_profit_loss_report_uses_selected_academic_session_for_payments(self):
        self.select_year(self.year_2026)
        Payment.objects.create(
            invoice=self.invoice_2027,
            amount=Decimal("777.00"),
            paid_on=date(2026, 5, 8),
            method=Payment.Method.CASH,
            receipt_number="RCP-SESSION-LEAK",
            received_by=self.admin_user,
        )

        response = self.client.get(
            reverse("institute_admin:profit_loss_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["collection_total"], Decimal("250.00"))
        self.assertNotContains(response, "RCP-SESSION-LEAK")

        export_response = self.client.get(
            reverse("institute_admin:profit_loss_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "export": "csv",
            },
        )
        export_text = export_response.content.decode()
        self.assertIn("Total Collection,250.00", export_text)
        self.assertNotIn("RCP-SESSION-LEAK", export_text)

    def test_fee_collection_report_filters_groups_and_exports_payments(self):
        self.select_year(self.year_2026)
        matching_payment = Payment.objects.create(
            invoice=self.invoice_2026,
            amount=Decimal("150.00"),
            paid_on=date(2026, 5, 7),
            method=Payment.Method.UPI,
            receipt_number="RCP-UPI-001",
            received_by=self.admin_user,
        )
        Payment.objects.create(
            invoice=self.invoice_2026,
            amount=Decimal("999.00"),
            paid_on=date(2026, 6, 7),
            method=Payment.Method.UPI,
            receipt_number="RCP-OUTSIDE",
            received_by=self.admin_user,
        )

        response = self.client.get(
            reverse("institute_admin:fee_collection_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "course": str(self.course.pk),
                "batch": str(self.batch_2026.pk),
                "method": Payment.Method.UPI,
                "student": "Student",
                "receipt_number": "RCP-UPI",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_collection"], Decimal("150.00"))
        self.assertEqual(response.context["payment_count"], 1)
        self.assertEqual(response.context["method_groups"][0]["label"], "UPI")
        self.assertEqual(response.context["student_groups"][0]["amount"], Decimal("150.00"))
        self.assertContains(response, "Fee Collection Report")
        self.assertContains(response, "RCP-UPI-001")
        self.assertContains(response, reverse("institute_admin:payment_receipt", args=[matching_payment.pk]))
        self.assertNotContains(response, "RCP-OUTSIDE")

        export_response = self.client.get(
            reverse("institute_admin:fee_collection_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "course": str(self.course.pk),
                "batch": str(self.batch_2026.pk),
                "method": Payment.Method.UPI,
                "student": "Student",
                "receipt_number": "RCP-UPI",
                "export": "csv",
            },
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "text/csv")
        export_text = export_response.content.decode()
        self.assertIn("Total Collection,150.00", export_text)
        self.assertIn("Receipt Number,RCP-UPI", export_text)
        self.assertIn("RCP-UPI-001", export_text)
        self.assertIn("Science", export_text)
        self.assertIn("11th Batch", export_text)
        self.assertNotIn("RCP-OUTSIDE", export_text)

    def test_fee_collection_report_uses_selected_academic_session(self):
        self.select_year(self.year_2026)
        Payment.objects.create(
            invoice=self.invoice_2027,
            amount=Decimal("888.00"),
            paid_on=date(2026, 5, 8),
            method=Payment.Method.UPI,
            receipt_number="RCP-FEE-SESSION-LEAK",
            received_by=self.admin_user,
        )

        response = self.client.get(
            reverse("institute_admin:fee_collection_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_collection"], Decimal("250.00"))
        self.assertNotContains(response, "RCP-FEE-SESSION-LEAK")

        export_response = self.client.get(
            reverse("institute_admin:fee_collection_report"),
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "export": "csv",
            },
        )
        export_text = export_response.content.decode()
        self.assertIn("Total Collection,250.00", export_text)
        self.assertNotIn("RCP-FEE-SESSION-LEAK", export_text)

    def test_registration_prefix_uses_normalized_institute_code(self):
        other_institute = Institute.objects.create(
            name="Saint Monica International School",
            code="smis-north",
            status=Institute.Status.ACTIVE,
        )

        first_prefix = get_student_admission_prefix(self.institute, self.year_2026)
        second_prefix = get_student_admission_prefix(other_institute, self.year_2026)
        first_registration = f"{first_prefix}0001"
        second_registration = f"{second_prefix}0001"

        self.assertEqual(first_registration, "SMIS26270001")
        self.assertEqual(second_registration, "SMISNORTH26270001")
        self.assertEqual(build_student_username(self.institute, first_registration), first_registration)
        self.assertEqual(build_student_username(other_institute, second_registration), second_registration)

    def test_student_list_can_filter_by_batch(self):
        matching_sessions = [self.session_2026]
        for index in range(2, 4):
            user = User.objects.create_user(
                username=f"student-{index}",
                password="pass12345",
                first_name="Student",
                last_name=str(index),
            )
            UserProfile.objects.create(
                user=user,
                institute=self.institute,
                role=UserProfile.Role.STUDENT_PARENT,
                phone=f"922222222{index}",
            )
            student = StudentProfile.objects.create(
                institute=self.institute,
                user=user,
                academic_year=self.year_2026,
                admission_number=f"SMIS-2026-27-000{index}",
                is_active=True,
            )
            student_session = StudentAcademicSession.objects.create(
                institute=self.institute,
                student=student,
                academic_year=self.year_2026,
                admission_number=f"SMIS-2026-27-000{index}",
                status=StudentAcademicSession.Status.ACTIVE,
            )
            enrollment = StudentEnrollment.objects.create(
                academic_session=student_session,
                student=student,
                batch=self.batch_2026,
                enrolled_on=date(2026, 4, 8),
                custom_fee_amount=Decimal("1000.00"),
            )
            enrollment.courses.add(self.course)
            matching_sessions.append(student_session)

        other_user = User.objects.create_user(
            username="student-four",
            password="pass12345",
            first_name="Student",
            last_name="Four",
        )
        UserProfile.objects.create(
            user=other_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
            phone="9222222224",
        )
        other_student = StudentProfile.objects.create(
            institute=self.institute,
            user=other_user,
            academic_year=self.year_2026,
            admission_number="SMIS-2026-27-0004",
            is_active=True,
        )
        other_session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=other_student,
            academic_year=self.year_2026,
            admission_number="SMIS-2026-27-0004",
            status=StudentAcademicSession.Status.ACTIVE,
        )
        other_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Evening Batch",
            is_active=True,
        )
        other_batch.courses.add(self.course)
        other_enrollment = StudentEnrollment.objects.create(
            academic_session=other_session,
            student=other_student,
            batch=other_batch,
            enrolled_on=date(2026, 4, 8),
            custom_fee_amount=Decimal("1000.00"),
        )
        other_enrollment.courses.add(self.course)

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_list"), {"batch": self.batch_2026.pk})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SMIS-2026-27-0001")
        self.assertContains(response, "SMIS-2026-27-0002")
        self.assertContains(response, "SMIS-2026-27-0003")
        self.assertNotContains(response, "SMIS-2026-27-0004")
        self.assertEqual(response.context["batch_filter"], str(self.batch_2026.pk))
        self.assertIn(self.batch_2026, list(response.context["batches"]))
        self.assertNotIn(self.batch_2027, list(response.context["batches"]))
        self.assertEqual(response.context["total_students"], 3)
        self.assertEqual(response.context["active_students"], 3)
        self.assertEqual(response.context["inactive_students"], 0)
        self.assertEqual(response.context["total_enrollments"], 3)
        self.assertEqual(response.context["filtered_total_fee_amount"], Decimal("3000.00"))
        self.assertEqual(response.context["filtered_paid_amount"], Decimal("250.00"))
        self.assertEqual(response.context["filtered_due_amount"], Decimal("2750.00"))

    def test_student_export_uses_selected_fields_with_batch_and_fee_columns(self):
        self.select_year(self.year_2026)

        response = self.client.get(
            reverse("institute_admin:student_export"),
            {
                "fields": ["name", "mobile", "batch", "total_fees", "paid_amount", "due_amount"],
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook["Students"]
        headers = [sheet.cell(row=3, column=column).value for column in range(1, 7)]
        values = [sheet.cell(row=4, column=column).value for column in range(1, 7)]

        self.assertEqual(headers, ["Name", "Mobile", "Batch", "Total Fees", "Paid Amount", "Due Amount"])
        self.assertEqual(values[0], "Student One")
        self.assertEqual(values[1], "9111111111")
        self.assertEqual(values[2], "11th Batch")
        self.assertEqual(Decimal(str(values[3])), Decimal("1000"))
        self.assertEqual(Decimal(str(values[4])), Decimal("250"))
        self.assertEqual(Decimal(str(values[5])), Decimal("750"))

    @override_settings(STUDENT_EXPORT_CSV_THRESHOLD=1)
    def test_large_student_export_streams_csv_with_bulk_fee_summary(self):
        self.select_year(self.year_2026)

        with patch.object(
            views,
            "get_student_session_fee_summary",
            wraps=views.get_student_session_fee_summary,
        ) as per_student_fee_summary:
            with CaptureQueriesContext(connection) as queries:
                response = self.client.get(
                    reverse("institute_admin:student_export"),
                    {
                        "fields": [
                            "admission_number",
                            "name",
                            "batch",
                            "total_fees",
                            "paid_amount",
                            "due_amount",
                        ],
                    },
                )
                content = b"".join(response.streaming_content).decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.streaming)
        self.assertEqual(response["X-Export-Mode"], "streamed")
        self.assertIn("Admission Number,Name,Batch,Total Fees,Paid Amount,Due Amount", content)
        self.assertIn("SMIS-2026-27-0001,Student One,11th Batch,1000,250,750", content)
        self.assertEqual(per_student_fee_summary.call_count, 0)
        self.assertLess(len(queries), 25)

    @override_settings(STUDENT_IMPORT_BACKGROUND_THRESHOLD=1)
    def test_large_student_import_is_processed_by_background_worker(self):
        self.select_year(self.year_2026)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(["Student Bulk Import Template"])
        sheet.append(["Generated admission numbers"])
        sheet.append(views.student_import_columns())
        sheet.append(
            [
                "Queued",
                "Student",
                "Student@123",
                "queued@example.com",
                "9333333399",
                "2012-01-01",
                "2026-06-12",
                "Queued address",
                "Current School",
                "School address",
                "Previous School",
                "9th",
                "Queued Guardian",
                "Father",
                "9444444499",
                "guardian.queued@example.com",
                "Yes",
            ]
        )
        upload_buffer = BytesIO()
        workbook.save(upload_buffer)
        upload = SimpleUploadedFile(
            "queued-students.xlsx",
            upload_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_import"),
            {"student_file": upload},
        )

        self.assertEqual(response.status_code, 302)
        job = BackgroundJob.objects.get(job_type=BackgroundJob.JobType.STUDENT_IMPORT)
        self.assertEqual(job.status, BackgroundJob.Status.PENDING)
        self.assertFalse(User.objects.filter(first_name="Queued", last_name="Student").exists())

        call_command("run_background_jobs", "--once")

        job.refresh_from_db()
        self.assertEqual(job.status, BackgroundJob.Status.COMPLETED)
        self.assertEqual(job.result["created_count"], 1)
        self.assertTrue(User.objects.filter(first_name="Queued", last_name="Student").exists())

    def test_notification_fan_out_is_queued_as_durable_jobs(self):
        from student_parent.notifications import (
            enqueue_fee_paid_notification,
            enqueue_notice_published_notification,
        )

        notice = Notice.objects.create(
            institute=self.institute,
            title="Queued notice",
            message="Process this outside the request.",
            push_to_app=True,
            is_published=True,
        )

        fee_job = enqueue_fee_paid_notification(self.payment_2026)
        notice_job = enqueue_notice_published_notification(notice)

        self.assertEqual(fee_job.status, BackgroundJob.Status.PENDING)
        self.assertEqual(fee_job.job_type, BackgroundJob.JobType.FEE_NOTIFICATION)
        self.assertEqual(fee_job.payload["payment_id"], self.payment_2026.pk)
        self.assertEqual(notice_job.status, BackgroundJob.Status.PENDING)
        self.assertEqual(notice_job.job_type, BackgroundJob.JobType.NOTICE_NOTIFICATION)
        self.assertEqual(notice_job.payload["notice_id"], notice.pk)
        self.assertEqual(
            notice_job.payload["notification_version"],
            notice.push_notification_version,
        )

    def test_editing_notice_queues_a_new_notification_version(self):
        self.select_year(self.year_2026)
        notice = Notice.objects.create(
            institute=self.institute,
            created_by=self.admin_user,
            title="Original notice",
            message="Original message.",
            audience=Notice.Audience.STUDENTS_PARENTS,
            category=Notice.Category.GENERAL,
            priority=Notice.Priority.NORMAL,
            push_to_app=True,
            is_published=True,
            push_notification_queued_at=timezone.now(),
        )
        old_job = BackgroundJob.objects.create(
            job_type=BackgroundJob.JobType.NOTICE_NOTIFICATION,
            institute=self.institute,
            created_by=self.admin_user,
            payload={
                "notice_id": notice.pk,
                "notification_version": notice.push_notification_version,
            },
        )

        with (
            patch(
                "institute_admin.background_jobs.dispatch_background_job"
            ) as dispatch,
            self.captureOnCommitCallbacks(execute=True),
        ):
            response = self.client.post(
                reverse("institute_admin:notice_update", args=[notice.pk]),
                {
                    "title": "Updated notice",
                    "message": "Updated message.",
                    "audience": Notice.Audience.STUDENTS_PARENTS,
                    "category": Notice.Category.ACADEMIC,
                    "priority": Notice.Priority.IMPORTANT,
                    "publish_at": "",
                    "expires_at": "",
                    "is_published": "on",
                    "push_to_app": "on",
                },
            )

        self.assertEqual(response.status_code, 200)
        notice.refresh_from_db()
        self.assertEqual(notice.title, "Updated notice")
        self.assertEqual(notice.push_notification_version, 2)
        self.assertIsNotNone(notice.push_notification_queued_at)

        new_job = (
            BackgroundJob.objects.filter(
                job_type=BackgroundJob.JobType.NOTICE_NOTIFICATION,
                payload__notice_id=notice.pk,
            )
            .exclude(pk=old_job.pk)
            .get()
        )
        self.assertEqual(new_job.payload["notification_version"], 2)
        dispatch.assert_called_once_with(new_job.pk)

    def test_student_dashboard_context_is_limited_to_selected_session(self):
        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_dashboard", args=[self.student.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["student_session"], self.session_2026)
        self.assertQuerySetEqual(response.context["enrollments"], [self.enrollment_2026])
        self.assertQuerySetEqual(response.context["payments"], [self.payment_2026])
        self.assertEqual(list(response.context["attendance_records"]), [self.attendance_2026])
        self.assertEqual(response.context["total_fee_amount"], Decimal("1000.00"))
        self.assertEqual(response.context["total_paid_amount"], Decimal("250.00"))

        self.select_year(self.year_2027)
        response = self.client.get(reverse("institute_admin:student_dashboard", args=[self.student.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["student_session"], self.session_2027)
        self.assertQuerySetEqual(response.context["enrollments"], [self.enrollment_2027])
        self.assertQuerySetEqual(response.context["payments"], [self.payment_2027])
        self.assertEqual(list(response.context["attendance_records"]), [self.attendance_2027])
        self.assertEqual(response.context["total_fee_amount"], Decimal("2000.00"))
        self.assertEqual(response.context["total_paid_amount"], Decimal("500.00"))

    def test_student_dashboard_hides_certificate_generate_buttons_for_coaching_classes(self):
        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:student_dashboard", args=[self.student.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["is_school_institute"])
        self.assertNotContains(response, "Generate TC")
        self.assertNotContains(response, "Generate Bonafide")
        self.assertNotContains(response, "ID Card Details")
        self.assertNotContains(response, "Bonafide Certificate Records")
        self.assertNotContains(response, "Transfer Certificate Records")

    def test_student_dashboard_shows_certificate_generate_buttons_for_schools(self):
        self.institute.institute_type = Institute.InstituteType.SCHOOL
        self.institute.save(update_fields=["institute_type"])
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:student_dashboard", args=[self.student.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_school_institute"])
        self.assertContains(response, "Generate TC")
        self.assertContains(response, "Generate Bonafide")

    def test_student_school_only_routes_reject_coaching_classes(self):
        self.select_year(self.year_2026)

        for url_name in [
            "student_tc_generate",
            "student_bonafide_generate",
            "student_id_card_update",
        ]:
            response = self.client.get(reverse(f"institute_admin:{url_name}", args=[self.student.pk]))
            self.assertEqual(response.status_code, 403)

    def test_student_form_uses_limited_fields_for_coaching_classes(self):
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:student_create"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["is_school_institute"])
        self.assertContains(response, "Student Full Name")
        self.assertContains(response, "Class")
        self.assertContains(response, "Batch / Division")
        self.assertContains(response, "Upload Document")
        self.assertContains(response, "Document Type")
        self.assertContains(response, "Document Title")
        self.assertNotContains(response, "PEN No")
        self.assertNotContains(response, "Previous School Name")
        self.assertNotContains(response, "Documents to Collect and Store")
        self.assertNotContains(response, "Birth Certificate")

    def test_student_form_keeps_full_fields_for_schools(self):
        self.institute.institute_type = Institute.InstituteType.SCHOOL
        self.institute.save(update_fields=["institute_type"])
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:student_create"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_school_institute"])
        self.assertContains(response, "PEN No")
        self.assertContains(response, "Previous School Name")
        self.assertContains(response, "Documents to Collect and Store")

    def test_courses_and_batches_are_limited_to_selected_academic_year(self):
        old_only_course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Old Year Only Subject",
            fee_amount=Decimal("300.00"),
            is_active=True,
        )
        old_only_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Old Year Only Batch",
            is_active=True,
        )
        old_only_batch.courses.add(old_only_course)

        self.select_year(self.year_2027)

        response = self.client.get(reverse("institute_admin:course_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Science")
        self.assertNotContains(response, "Old Year Only Subject")

        response = self.client.get(reverse("institute_admin:batch_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "12th Batch")
        self.assertNotContains(response, "Old Year Only Batch")

        response = self.client.get(reverse("institute_admin:enrollment_create"))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(old_only_batch, list(response.context["form"].fields["batch"].queryset))
        self.assertNotIn(old_only_course, list(response.context["form"].fields["courses"].queryset))

    def test_dashboard_fee_totals_include_uninvoiced_enrollment_fees(self):
        extra_course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Extra Subject",
            fee_amount=Decimal("1500.00"),
            is_active=True,
        )
        extra_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Extra Batch",
            is_active=True,
        )
        extra_batch.courses.add(extra_course)
        extra_enrollment = StudentEnrollment.objects.create(
            academic_session=self.session_2026,
            student=self.student,
            batch=extra_batch,
            enrolled_on=date(2026, 4, 8),
            custom_fee_amount=Decimal("1500.00"),
        )
        extra_enrollment.courses.add(extra_course)
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["invoice_amount"], Decimal("2500.00"))
        self.assertEqual(response.context["paid_amount"], Decimal("250.00"))
        self.assertEqual(response.context["due_amount"], Decimal("2250.00"))
        self.assertEqual(response.context["collection_rate"], 10.0)
        self.assertTrue(any(row["due_amount"] == Decimal("2250.00") for row in response.context["due_invoice_rows"]))

    def test_dashboard_summary_is_cached_and_payment_changes_invalidate_it(self):
        self.select_year(self.year_2026)

        with patch.object(
            views,
            "build_dashboard_summary",
            wraps=views.build_dashboard_summary,
        ) as summary_builder:
            first_response = self.client.get(reverse("institute_admin:dashboard"))
            second_response = self.client.get(reverse("institute_admin:dashboard"))

            self.assertEqual(first_response.status_code, 200)
            self.assertEqual(second_response.status_code, 200)
            self.assertEqual(summary_builder.call_count, 1)

            Payment.objects.create(
                invoice=self.invoice_2026,
                amount=Decimal("100.00"),
                paid_on=date(2026, 5, 4),
                method=Payment.Method.CASH,
                received_by=self.admin_user,
            )
            refreshed_response = self.client.get(reverse("institute_admin:dashboard"))

            self.assertEqual(summary_builder.call_count, 2)
            self.assertEqual(refreshed_response.context["paid_amount"], Decimal("350.00"))
            self.assertEqual(refreshed_response.context["due_amount"], Decimal("650.00"))

    def test_dashboard_returns_only_six_highest_dues(self):
        for index in range(1, 8):
            user = User.objects.create(username=f"due-student-{index}", first_name=f"Due {index}")
            UserProfile.objects.create(
                user=user,
                institute=self.institute,
                role=UserProfile.Role.STUDENT_PARENT,
            )
            student = StudentProfile.objects.create(
                institute=self.institute,
                academic_year=self.year_2026,
                user=user,
                admission_number=f"DUE-{index:04d}",
            )
            student_session = StudentAcademicSession.objects.create(
                institute=self.institute,
                student=student,
                academic_year=self.year_2026,
                admission_number=f"DUE-{index:04d}",
                joined_on=date(2026, 4, index),
            )
            StudentEnrollment.objects.create(
                student=student,
                academic_session=student_session,
                batch=self.batch_2026,
                enrolled_on=date(2026, 4, index),
                custom_fee_amount=Decimal(index * 1000),
            )

        self.select_year(self.year_2026)
        response = self.client.get(reverse("institute_admin:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["due_invoice_rows"]), 6)
        self.assertEqual(
            [row["due_amount"] for row in response.context["due_invoice_rows"]],
            [Decimal(value) for value in (7000, 6000, 5000, 4000, 3000, 2000)],
        )

    def test_session_link_is_required_for_operational_records(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                StudentEnrollment.objects.create(
                    student=self.student,
                    batch=self.batch_2026,
                    enrolled_on=date(2026, 4, 6),
                )
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                FeeInvoice.objects.create(
                    institute=self.institute,
                    student=self.student,
                    title="Broken Invoice",
                    amount=Decimal("100.00"),
                    due_date=date(2026, 6, 1),
                )
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Attendance.objects.create(
                    student=self.student,
                    batch=self.batch_2026,
                    date=date(2026, 6, 1),
                )

    def test_models_align_student_from_academic_session(self):
        other_user = User.objects.create_user(username="other-student", password="pass12345")
        other_student = StudentProfile.objects.create(
            institute=self.institute,
            user=other_user,
            academic_year=None,
            admission_number="LEGACY-0002",
            is_active=True,
        )
        batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2027,
            name="Alignment Batch",
            is_active=True,
        )
        enrollment = StudentEnrollment.objects.create(
            academic_session=self.session_2027,
            student=other_student,
            batch=batch,
        )
        self.assertEqual(enrollment.student, self.student)

        invoice = FeeInvoice.objects.create(
            institute=self.institute,
            student=other_student,
            academic_session=self.session_2027,
            title="Aligned Invoice",
            amount=Decimal("50.00"),
            due_date=date(2027, 7, 1),
        )
        self.assertEqual(invoice.student, self.student)
        self.assertEqual(invoice.institute, self.institute)

        attendance = Attendance.objects.create(
            academic_session=self.session_2027,
            student=other_student,
            batch=batch,
            date=date(2027, 7, 2),
        )
        self.assertEqual(attendance.student, self.student)

    def test_receive_fee_rejects_more_than_enrollment_due(self):
        form = ReceiveFeeForm(
            data={
                "enrollment": self.enrollment_2026.pk,
                "category": "",
                "title": "2026 Fee",
                "invoice_amount": "1000.00",
                "payment_amount": "800.00",
                "paid_on": "2026-05-10",
                "method": Payment.Method.CASH,
                "receipt_number": "",
                "note": "",
            },
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("Payment amount cannot be greater", str(form.errors))

    def test_receive_fee_rejects_more_than_category_due(self):
        category = FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Books",
            default_amount=Decimal("500.00"),
        )
        invoice = FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            category=category,
            title="Books",
            amount=Decimal("500.00"),
            due_date=date(2026, 5, 10),
        )
        Payment.objects.create(
            invoice=invoice,
            amount=Decimal("200.00"),
            paid_on=date(2026, 5, 11),
            method=Payment.Method.CASH,
            received_by=self.admin_user,
        )
        form = ReceiveFeeForm(
            data={
                "enrollment": "",
                "category": category.pk,
                "title": "Books",
                "invoice_amount": "500.00",
                "payment_amount": "400.00",
                "paid_on": "2026-05-12",
                "method": Payment.Method.CASH,
                "receipt_number": "",
                "note": "",
            },
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("Payment amount cannot be greater", str(form.errors))

    def test_receive_fee_category_list_only_includes_student_extra_fee_categories(self):
        books = FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Books",
            default_amount=Decimal("500.00"),
        )
        uniform = FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Uniform",
            default_amount=Decimal("300.00"),
        )
        enrollment_category = FeeCategory.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Enrollment Category",
            default_amount=Decimal("100.00"),
        )
        FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            category=books,
            title="Books",
            amount=Decimal("500.00"),
            due_date=date(2026, 5, 10),
        )
        FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            enrollment=self.enrollment_2026,
            category=books,
            title="Enrollment Books",
            amount=Decimal("100.00"),
            due_date=date(2026, 5, 10),
        )
        FeeInvoice.objects.create(
            institute=self.institute,
            student=self.student,
            academic_session=self.session_2026,
            enrollment=self.enrollment_2026,
            category=enrollment_category,
            title="Enrollment Category",
            amount=Decimal("100.00"),
            due_date=date(2026, 5, 10),
        )

        form = ReceiveFeeForm(institute=self.institute, student=self.student, academic_session=self.session_2026)
        category_ids = set(form.fields["category"].queryset.values_list("pk", flat=True))
        due_data = views.get_student_category_due_data(self.session_2026)
        due_category_ids = set(due_data["categories"].keys())

        self.assertEqual(category_ids, {books.pk})
        self.assertEqual(due_category_ids, {str(books.pk)})
        self.assertEqual(due_data["categories"][str(books.pk)]["total"], "500.00")
        self.assertNotIn(uniform.pk, category_ids)
        self.assertNotIn(enrollment_category.pk, category_ids)

    def test_receive_fee_reuses_existing_enrollment_invoice(self):
        self.select_year(self.year_2026)
        invoice_count = FeeInvoice.objects.filter(academic_session=self.session_2026).count()

        response = self.client.post(
            reverse("institute_admin:student_receive_fee", args=[self.student.pk]),
            data={
                "enrollment": self.enrollment_2026.pk,
                "category": "",
                "title": "2026 Fee",
                "invoice_amount": "1000.00",
                "payment_amount": "300.00",
                "paid_on": "2026-05-12",
                "method": Payment.Method.CASH,
                "receipt_number": "",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(FeeInvoice.objects.filter(academic_session=self.session_2026).count(), invoice_count)
        active_paid = sum(
            payment.amount
            for payment in Payment.objects.filter(invoice=self.invoice_2026, status=Payment.Status.ACTIVE)
        )
        self.assertEqual(active_paid, Decimal("550.00"))

    def test_payment_correction_rejects_overpayment(self):
        form = PaymentUpdateForm(
            data={
                "amount": "1200.00",
                "paid_on": "2026-05-12",
                "method": Payment.Method.CASH,
                "receipt_number": "CORRECTED",
                "note": "",
                "correction_reason": "Testing overpayment guard.",
            },
            payment=self.payment_2026,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("Corrected amount cannot be greater", str(form.errors))

    def test_voided_payment_is_excluded_from_dashboard_totals(self):
        extra_payment = Payment.objects.create(
            invoice=self.invoice_2026,
            amount=Decimal("100.00"),
            paid_on=date(2026, 5, 13),
            method=Payment.Method.CASH,
            received_by=self.admin_user,
        )
        extra_payment.void(self.admin_user, "Mistaken payment")
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:student_dashboard", args=[self.student.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_paid_amount"], Decimal("250.00"))

    def test_receipt_uses_academic_session_admission_number(self):
        self.select_year(self.year_2027)

        response = self.client.get(reverse("institute_admin:payment_receipt", args=[self.payment_2027.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SMIS-2027-28-0001")
        self.assertNotContains(response, "LEGACY-0001")
        self.assertNotContains(response, "SMIS-2026-27-0001")

    def test_attendance_post_creates_record_for_selected_academic_session(self):
        self.select_year(self.year_2026)
        target_date = date(2026, 6, 10)

        response = self.client.post(
            reverse("institute_admin:attendance_list")
            + f"?batch={self.batch_2026.pk}&date={target_date.isoformat()}",
            data={
                "student_ids": [str(self.student.pk)],
                f"status_{self.student.pk}": Attendance.Status.LATE,
                f"note_{self.student.pk}": "Reached after first lecture",
            },
        )

        self.assertEqual(response.status_code, 302)
        attendance = Attendance.objects.get(
            academic_session=self.session_2026,
            batch=self.batch_2026,
            date=target_date,
        )
        self.assertEqual(attendance.student, self.student)
        self.assertEqual(attendance.status, Attendance.Status.LATE)
        self.assertEqual(attendance.note, "Reached after first lecture")
        self.assertFalse(
            Attendance.objects.filter(
                academic_session=self.session_2027,
                batch=self.batch_2026,
                date=target_date,
            ).exists()
        )

    def test_attendance_waits_for_batch_selection(self):
        self.select_year(self.year_2026)

        response = self.client.get(reverse("institute_admin:attendance_list"))

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context["selected_batch"])
        self.assertEqual(response.context["attendance_rows"], [])
        self.assertContains(response, "Select a batch to load students for attendance.")

    def test_attendance_uses_bulk_create_and_bulk_update(self):
        self.select_year(self.year_2026)
        target_date = date(2026, 6, 12)
        Attendance.objects.create(
            student=self.student,
            academic_session=self.session_2026,
            batch=self.batch_2026,
            date=target_date,
            status=Attendance.Status.PRESENT,
            marked_by=self.admin_user,
        )

        with patch.object(
            Attendance.objects,
            "bulk_update",
            wraps=Attendance.objects.bulk_update,
        ) as bulk_update:
            response = self.client.post(
                reverse("institute_admin:attendance_list")
                + f"?batch={self.batch_2026.pk}&date={target_date.isoformat()}",
                data={
                    "student_ids": [str(self.student.pk)],
                    f"status_{self.student.pk}": Attendance.Status.ABSENT,
                    f"note_{self.student.pk}": "Updated in bulk",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(bulk_update.call_count, 1)
        record = Attendance.objects.get(
            academic_session=self.session_2026,
            batch=self.batch_2026,
            date=target_date,
        )
        self.assertEqual(record.status, Attendance.Status.ABSENT)
        self.assertEqual(record.note, "Updated in bulk")

    def test_attendance_large_batch_is_paginated_below_response_limit(self):
        users = [
            User(username=f"attendance-page-{index}", first_name=f"Student {index}")
            for index in range(1, 101)
        ]
        User.objects.bulk_create(users)
        users = list(User.objects.filter(username__startswith="attendance-page-").order_by("username"))
        UserProfile.objects.bulk_create(
            [
                UserProfile(
                    user=user,
                    institute=self.institute,
                    role=UserProfile.Role.STUDENT_PARENT,
                )
                for user in users
            ]
        )
        students = [
            StudentProfile(
                institute=self.institute,
                academic_year=self.year_2026,
                user=user,
                admission_number=f"ATT-{index:04d}",
            )
            for index, user in enumerate(users, start=1)
        ]
        StudentProfile.objects.bulk_create(students)
        students = list(
            StudentProfile.objects.filter(admission_number__startswith="ATT-").order_by("admission_number")
        )
        sessions = [
            StudentAcademicSession(
                institute=self.institute,
                student=student,
                academic_year=self.year_2026,
                admission_number=student.admission_number,
            )
            for student in students
        ]
        StudentAcademicSession.objects.bulk_create(sessions)
        sessions = list(
            StudentAcademicSession.objects.filter(admission_number__startswith="ATT-").order_by(
                "admission_number"
            )
        )
        StudentEnrollment.objects.bulk_create(
            [
                StudentEnrollment(
                    student_id=session.student_id,
                    academic_session=session,
                    batch=self.batch_2026,
                )
                for session in sessions
            ]
        )
        self.select_year(self.year_2026)

        response = self.client.get(
            reverse("institute_admin:attendance_list"),
            {"batch": self.batch_2026.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["attendance_page"].paginator.per_page, 100)
        self.assertEqual(len(response.context["attendance_rows"]), 100)
        self.assertGreater(response.context["attendance_page"].paginator.count, 100)
        self.assertLess(len(response.content), 300 * 1024)

    def test_attendance_post_ignores_student_outside_selected_year_batch(self):
        self.select_year(self.year_2026)
        target_date = date(2026, 6, 11)

        response = self.client.post(
            reverse("institute_admin:attendance_list")
            + f"?batch={self.batch_2027.pk}&date={target_date.isoformat()}",
            data={
                "student_ids": [str(self.student.pk)],
                f"status_{self.student.pk}": Attendance.Status.ABSENT,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Attendance.objects.filter(date=target_date).exists())

    def test_enrollment_create_attaches_selected_academic_session(self):
        self.select_year(self.year_2026)
        new_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="2026 Revision",
            is_active=True,
        )
        new_batch.courses.add(self.course)

        response = self.client.post(
            reverse("institute_admin:enrollment_create"),
            data={
                "student": self.student.pk,
                "batch": new_batch.pk,
                "courses": [self.course.pk],
                "enrolled_on": "2026-06-15",
                "status": StudentEnrollment.Status.ACTIVE,
                "custom_fee_amount": "750.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        enrollment = StudentEnrollment.objects.get(batch=new_batch)
        self.assertEqual(enrollment.academic_session, self.session_2026)
        self.assertEqual(enrollment.student, self.student)

    def test_student_autocomplete_requires_two_characters_and_scopes_academic_year(self):
        self.select_year(self.year_2026)

        short_response = self.client.get(
            reverse("institute_admin:student_autocomplete"),
            {"q": "S"},
        )
        response = self.client.get(
            reverse("institute_admin:student_autocomplete"),
            {"q": "SMIS", "academic_year": self.year_2026.pk},
        )

        self.assertEqual(short_response.status_code, 200)
        self.assertEqual(short_response.json()["results"], [])
        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(response.json()["results"]), 20)
        self.assertEqual(response.json()["results"][0]["id"], self.student.pk)
        self.assertIn(self.session_2026.admission_number, response.json()["results"][0]["text"])
        self.assertNotIn(self.session_2027.admission_number, response.content.decode())

    def test_large_student_forms_use_ajax_without_rendering_full_student_choices(self):
        self.select_year(self.year_2026)

        enrollment_response = self.client.get(reverse("institute_admin:enrollment_create"))
        selected_enrollment_response = self.client.get(
            reverse("institute_admin:enrollment_create"),
            {"student": self.student.pk},
        )
        notice_response = self.client.get(reverse("institute_admin:notice_create"))
        attendance_response = self.client.get(reverse("institute_admin:attendance_list"))

        self.assertContains(enrollment_response, 'data-student-autocomplete="true"')
        self.assertNotContains(enrollment_response, self.session_2026.admission_number)
        self.assertContains(selected_enrollment_response, f'value="{self.student.pk}" selected')
        self.assertContains(notice_response, 'data-student-autocomplete="true"')
        self.assertNotContains(notice_response, self.session_2026.admission_number)
        self.assertContains(attendance_response, 'data-student-autocomplete="true"')
        self.assertNotContains(
            attendance_response,
            f'<option value="{self.student.pk}">{self.session_2026.admission_number}',
            html=False,
        )

    def test_enrollment_forms_stay_below_query_and_response_size_targets(self):
        self.select_year(self.year_2026)

        with CaptureQueriesContext(connection) as create_queries:
            create_response = self.client.get(reverse("institute_admin:enrollment_create"))
            create_size = len(create_response.content)
        with CaptureQueriesContext(connection) as update_queries:
            update_response = self.client.get(
                reverse("institute_admin:enrollment_update", args=[self.enrollment_2026.pk])
            )
            update_size = len(update_response.content)

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(update_response.status_code, 200)
        self.assertLess(len(create_queries), 25)
        self.assertLess(len(update_queries), 25)
        self.assertLess(create_size, 150 * 1024)
        self.assertLess(update_size, 150 * 1024)
        self.assertContains(create_response, 'data-student-autocomplete="true"')
        self.assertContains(update_response, f'value="{self.student.pk}" selected')

    def test_enrollment_create_rejects_student_without_selected_year_session(self):
        self.select_year(self.year_2026)
        other_user = User.objects.create_user(username="future-only", password="pass12345")
        UserProfile.objects.create(
            user=other_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        other_student = StudentProfile.objects.create(
            institute=self.institute,
            user=other_user,
            academic_year=None,
            admission_number="LEGACY-0003",
            is_active=True,
        )
        StudentAcademicSession.objects.create(
            institute=self.institute,
            student=other_student,
            academic_year=self.year_2027,
            admission_number="SMIS-2027-28-0003",
            status=StudentAcademicSession.Status.ACTIVE,
        )
        new_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year_2026,
            name="Rejected Batch",
            is_active=True,
        )
        new_batch.courses.add(self.course)

        response = self.client.post(
            reverse("institute_admin:enrollment_create"),
            data={
                "student": other_student.pk,
                "batch": new_batch.pk,
                "courses": [self.course.pk],
                "enrolled_on": "2026-06-16",
                "status": StudentEnrollment.Status.ACTIVE,
                "custom_fee_amount": "750.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(StudentEnrollment.objects.filter(batch=new_batch).exists())
        self.assertContains(response, "Select a valid choice", status_code=200)

    def test_student_promotion_choices_are_scoped_to_selected_sessions(self):
        self.select_year(self.year_2026)

        response = self.client.get(
            reverse("institute_admin:student_promote"),
            {
                "source_year": self.year_2026.pk,
                "target_year": self.year_2027.pk,
                "source_course": self.course.pk,
                "source_batch": self.batch_2026.pk,
                "target_course": self.course_2027.pk,
                "target_batch": self.batch_2027.pk,
                "load_students": "1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.course, list(response.context["source_courses"]))
        self.assertNotIn(self.course_2027, list(response.context["source_courses"]))
        self.assertIn(self.course_2027, list(response.context["target_courses"]))
        self.assertNotIn(self.course, list(response.context["target_courses"]))
        self.assertIn(self.batch_2026, list(response.context["source_batches"]))
        self.assertNotIn(self.batch_2027, list(response.context["source_batches"]))
        self.assertIn(self.batch_2027, list(response.context["target_batches"]))
        self.assertNotIn(self.batch_2026, list(response.context["target_batches"]))
        self.assertContains(response, 'data-student-autocomplete="true"')
        self.assertNotContains(response, self.session_2026.admission_number)

    def test_student_promotion_creates_target_session_and_selected_enrollment(self):
        self.session_2027.delete()
        self.select_year(self.year_2026)
        invoice_count = FeeInvoice.objects.count()
        payment_count = Payment.objects.count()
        attendance_count = Attendance.objects.count()

        response = self.client.post(
            reverse("institute_admin:student_promote"),
            data={
                "source_year": self.year_2026.pk,
                "target_year": self.year_2027.pk,
                "source_course": self.course.pk,
                "source_batch": self.batch_2026.pk,
                "target_course": self.course_2027.pk,
                "target_batch": self.batch_2027.pk,
                "students": [str(self.student.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        promoted_session = StudentAcademicSession.objects.get(
            student=self.student,
            academic_year=self.year_2027,
        )
        promoted_enrollment = StudentEnrollment.objects.get(academic_session=promoted_session)
        self.assertEqual(promoted_enrollment.batch, self.batch_2027)
        self.assertQuerySetEqual(promoted_enrollment.courses.all(), [self.course_2027])
        self.assertEqual(FeeInvoice.objects.count(), invoice_count)
        self.assertEqual(Payment.objects.count(), payment_count)
        self.assertEqual(Attendance.objects.count(), attendance_count)
        self.assertFalse(FeeInvoice.objects.filter(academic_session=promoted_session).exists())
        self.assertFalse(Attendance.objects.filter(academic_session=promoted_session).exists())

    def test_student_promotion_requires_target_batch(self):
        self.session_2027.delete()
        self.select_year(self.year_2026)

        response = self.client.post(
            reverse("institute_admin:student_promote"),
            data={
                "source_year": self.year_2026.pk,
                "target_year": self.year_2027.pk,
                "source_course": self.course.pk,
                "source_batch": self.batch_2026.pk,
                "target_course": self.course_2027.pk,
                "students": [str(self.student.pk)],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select a valid target batch")
        self.assertFalse(
            StudentAcademicSession.objects.filter(
                student=self.student,
                academic_year=self.year_2027,
            ).exists()
        )

    def test_student_promotion_repairs_existing_target_session_allocation(self):
        self.enrollment_2027.delete()
        self.select_year(self.year_2026)

        response = self.client.post(
            reverse("institute_admin:student_promote"),
            data={
                "source_year": self.year_2026.pk,
                "target_year": self.year_2027.pk,
                "source_course": self.course.pk,
                "source_batch": self.batch_2026.pk,
                "target_course": self.course_2027.pk,
                "target_batch": self.batch_2027.pk,
                "students": [str(self.student.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        repaired_enrollment = StudentEnrollment.objects.get(
            academic_session=self.session_2027,
            batch=self.batch_2027,
        )
        self.assertQuerySetEqual(repaired_enrollment.courses.all(), [self.course_2027])


class TenantIsolationTests(TestCase):
    def setUp(self):
        self.institute_a = Institute.objects.create(
            name="Arohan Academy",
            code="aa",
            status=Institute.Status.ACTIVE,
        )
        self.institute_b = Institute.objects.create(
            name="Saint Monica International School",
            code="smis",
            status=Institute.Status.ACTIVE,
        )
        self.admin_a = User.objects.create_user(username="admin-a", password="pass12345")
        self.admin_b = User.objects.create_user(username="admin-b", password="pass12345")
        UserProfile.objects.create(
            user=self.admin_a,
            institute=self.institute_a,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        UserProfile.objects.create(
            user=self.admin_b,
            institute=self.institute_b,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.year_a = AcademicYear.objects.create(
            institute=self.institute_a,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.year_b = AcademicYear.objects.create(
            institute=self.institute_b,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course_a = Course.objects.create(
            institute=self.institute_a,
            academic_year=self.year_a,
            name="Math A",
            fee_amount=Decimal("1000.00"),
            is_active=True,
        )
        self.course_b = Course.objects.create(
            institute=self.institute_b,
            academic_year=self.year_b,
            name="Math B",
            fee_amount=Decimal("2000.00"),
            is_active=True,
        )
        self.batch_a = Batch.objects.create(
            institute=self.institute_a,
            academic_year=self.year_a,
            name="Batch A",
            is_active=True,
        )
        self.batch_a.courses.add(self.course_a)
        self.batch_b = Batch.objects.create(
            institute=self.institute_b,
            academic_year=self.year_b,
            name="Batch B",
            is_active=True,
        )
        self.batch_b.courses.add(self.course_b)
        self.student_a = self.create_student(
            institute=self.institute_a,
            academic_year=self.year_a,
            username="student-a",
            admission_number="AA-2026-27-0001",
        )
        self.student_b = self.create_student(
            institute=self.institute_b,
            academic_year=self.year_b,
            username="student-b",
            admission_number="SMIS-2026-27-0001",
        )
        self.enrollment_a = StudentEnrollment.objects.create(
            academic_session=self.student_a["session"],
            student=self.student_a["profile"],
            batch=self.batch_a,
            custom_fee_amount=Decimal("1000.00"),
        )
        self.enrollment_b = StudentEnrollment.objects.create(
            academic_session=self.student_b["session"],
            student=self.student_b["profile"],
            batch=self.batch_b,
            custom_fee_amount=Decimal("2000.00"),
        )
        self.invoice_a = FeeInvoice.objects.create(
            institute=self.institute_a,
            academic_session=self.student_a["session"],
            student=self.student_a["profile"],
            enrollment=self.enrollment_a,
            batch=self.batch_a,
            title="Institute A Fee",
            amount=Decimal("1000.00"),
            due_date=date(2026, 5, 1),
        )
        self.invoice_b = FeeInvoice.objects.create(
            institute=self.institute_b,
            academic_session=self.student_b["session"],
            student=self.student_b["profile"],
            enrollment=self.enrollment_b,
            batch=self.batch_b,
            title="Institute B Fee",
            amount=Decimal("2000.00"),
            due_date=date(2026, 5, 1),
        )
        self.payment_b = Payment.objects.create(
            invoice=self.invoice_b,
            amount=Decimal("500.00"),
            paid_on=date(2026, 5, 2),
            method=Payment.Method.CASH,
            received_by=self.admin_b,
            receipt_number="B-RECEIPT",
        )
        self.attendance_b = Attendance.objects.create(
            academic_session=self.student_b["session"],
            student=self.student_b["profile"],
            batch=self.batch_b,
            date=date(2026, 5, 3),
            status=Attendance.Status.ABSENT,
            marked_by=self.admin_b,
        )
        self.homework_b = Homework.objects.create(
            batch=self.batch_b,
            course=self.course_b,
            title="Institute B Homework",
            instructions="Private homework",
            due_date=date(2026, 5, 4),
            created_by=self.admin_b,
        )
        self.notice_b = Notice.objects.create(
            institute=self.institute_b,
            title="Institute B Notice",
            message="Private notice",
            audience=Notice.Audience.EVERYONE,
            created_by=self.admin_b,
        )
        self.exam_a = Exam.objects.create(
            academic_year=self.year_a,
            batch=self.batch_a,
            course=self.course_a,
            title="Institute A Exam",
            exam_date=date(2026, 6, 10),
            is_published=False,
        )
        self.exam_b = Exam.objects.create(
            academic_year=self.year_b,
            batch=self.batch_b,
            course=self.course_b,
            title="Institute B Exam",
            exam_date=date(2026, 6, 10),
            is_published=False,
        )
        self.client.force_login(self.admin_a)
        session = self.client.session
        session["academic_year_id"] = self.year_a.pk
        session.save()

    def create_student(self, *, institute, academic_year, username, admission_number):
        user = User.objects.create_user(username=username, password="pass12345", first_name=username)
        UserProfile.objects.create(
            user=user,
            institute=institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        profile = StudentProfile.objects.create(
            institute=institute,
            user=user,
            academic_year=academic_year,
            admission_number=admission_number,
            is_active=True,
        )
        academic_session = StudentAcademicSession.objects.create(
            institute=institute,
            student=profile,
            academic_year=academic_year,
            admission_number=admission_number,
            status=StudentAcademicSession.Status.ACTIVE,
        )
        return {"user": user, "profile": profile, "session": academic_session}

    def assert_not_found(self, url_name, *args):
        response = self.client.get(reverse(url_name, args=args))
        self.assertEqual(response.status_code, 404)

    def test_institute_admin_cannot_open_other_institute_records_by_url(self):
        other_student_pk = self.student_b["profile"].pk
        protected_urls = [
            ("institute_admin:student_dashboard", other_student_pk),
            ("institute_admin:student_add_fee", other_student_pk),
            ("institute_admin:student_receive_fee", other_student_pk),
            ("institute_admin:student_update", other_student_pk),
            ("institute_admin:student_basic_update", other_student_pk),
            ("institute_admin:student_education_update", other_student_pk),
            ("institute_admin:student_guardian_update", other_student_pk),
            ("institute_admin:student_document_upload", other_student_pk),
            ("institute_admin:payment_receipt", self.payment_b.pk),
            ("institute_admin:payment_update", self.payment_b.pk),
            ("institute_admin:payment_void", self.payment_b.pk),
            ("institute_admin:enrollment_update", self.enrollment_b.pk),
            ("institute_admin:enrollment_delete", self.enrollment_b.pk),
            ("institute_admin:homework_update", self.homework_b.pk),
            ("institute_admin:homework_delete", self.homework_b.pk),
            ("institute_admin:notice_update", self.notice_b.pk),
            ("institute_admin:notice_delete", self.notice_b.pk),
        ]
        for url_name, pk in protected_urls:
            with self.subTest(url_name=url_name):
                self.assert_not_found(url_name, pk)

    def test_institute_admin_lists_do_not_show_other_institute_records(self):
        list_urls = [
            reverse("institute_admin:student_list"),
            reverse("institute_admin:enrollment_list"),
            reverse("institute_admin:homework_list"),
            reverse("institute_admin:notice_list"),
            reverse("institute_admin:attendance_list"),
        ]
        forbidden_text = [
            "SMIS-2026-27-0001",
            "student-b",
            "Batch B",
            "Institute B Homework",
            "Institute B Notice",
        ]
        for url in list_urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            for text in forbidden_text:
                with self.subTest(url=url, text=text):
                    self.assertNotContains(response, text)

    def test_bulk_student_delete_deletes_selected_students_without_invoices(self):
        deletable_student = self.create_student(
            institute=self.institute_a,
            academic_year=self.year_a,
            username="student-delete",
            admission_number="AA-2026-27-0002",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_delete"),
            data={"student_ids": [str(deletable_student["profile"].pk)]},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        self.assertFalse(User.objects.filter(pk=deletable_student["user"].pk).exists())
        self.assertFalse(StudentProfile.objects.filter(pk=deletable_student["profile"].pk).exists())

    def test_bulk_student_delete_skips_students_with_fee_invoices(self):
        deletable_student = self.create_student(
            institute=self.institute_a,
            academic_year=self.year_a,
            username="student-delete",
            admission_number="AA-2026-27-0002",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_delete"),
            data={"student_ids": [str(self.student_a["profile"].pk), str(deletable_student["profile"].pk)]},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        self.assertTrue(User.objects.filter(pk=self.student_a["user"].pk).exists())
        self.assertTrue(StudentProfile.objects.filter(pk=self.student_a["profile"].pk).exists())
        self.assertFalse(User.objects.filter(pk=deletable_student["user"].pk).exists())

    def test_bulk_student_delete_ignores_other_institute_students(self):
        deletable_student = self.create_student(
            institute=self.institute_a,
            academic_year=self.year_a,
            username="student-delete",
            admission_number="AA-2026-27-0002",
        )

        response = self.client.post(
            reverse("institute_admin:student_bulk_delete"),
            data={"student_ids": [str(deletable_student["profile"].pk), str(self.student_b["profile"].pk)]},
        )

        self.assertRedirects(response, reverse("institute_admin:student_list"))
        self.assertFalse(User.objects.filter(pk=deletable_student["user"].pk).exists())
        self.assertTrue(User.objects.filter(pk=self.student_b["user"].pk).exists())
        self.assertTrue(StudentProfile.objects.filter(pk=self.student_b["profile"].pk).exists())

    def test_attendance_export_does_not_include_other_institute_records(self):
        response = self.client.get(reverse("institute_admin:attendance_export"), {"format": "excel"})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(b"SMIS-2026-27-0001", response.content)
        self.assertNotIn(b"Batch B", response.content)

        response = self.client.get(reverse("institute_admin:attendance_export"), {"format": "pdf"})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(b"SMIS-2026-27-0001", response.content)
        self.assertNotIn(b"Batch B", response.content)

    def test_institute_admin_can_publish_exam_from_submissions(self):
        response = self.client.get(reverse("institute_admin:institute_exam_submissions", args=[self.exam_a.pk]))

        self.assertContains(response, "Exam Not Published")
        self.assertContains(response, "Publish Exam")

        response = self.client.post(reverse("institute_admin:institute_exam_publish", args=[self.exam_a.pk]))

        self.assertRedirects(response, reverse("institute_admin:institute_exam_submissions", args=[self.exam_a.pk]))
        self.exam_a.refresh_from_db()
        self.assertTrue(self.exam_a.is_published)

    def test_institute_admin_publishing_results_notifies_each_result_student(self):
        self.exam_a.show_result_after_submit = False
        self.exam_a.save(update_fields=["show_result_after_submit"])
        result = ExamResult.objects.create(
            exam=self.exam_a,
            student=self.student_a["profile"],
            marks_obtained="91.00",
        )

        with (
            patch(
                "student_parent.notifications.notify_result_declared"
            ) as notify_result,
            self.captureOnCommitCallbacks(execute=True),
        ):
            response = self.client.post(
                reverse(
                    "institute_admin:institute_exam_toggle_result_publish",
                    args=[self.exam_a.pk],
                ),
                {"action": "publish"},
            )

        self.assertRedirects(
            response,
            reverse(
                "institute_admin:institute_exam_submissions",
                args=[self.exam_a.pk],
            ),
        )
        self.exam_a.refresh_from_db()
        self.assertTrue(self.exam_a.show_result_after_submit)
        notify_result.assert_called_once()
        self.assertEqual(notify_result.call_args.args[0].pk, result.pk)

        with (
            patch(
                "student_parent.notifications.notify_result_declared"
            ) as notify_result,
            self.captureOnCommitCallbacks(execute=True),
        ):
            self.client.post(
                reverse(
                    "institute_admin:institute_exam_toggle_result_publish",
                    args=[self.exam_a.pk],
                ),
                {"action": "publish"},
            )
        notify_result.assert_not_called()

    def test_exam_submissions_shows_upload_count_and_uniform_actions(self):
        attempt = ExamAttempt.objects.create(
            exam=self.exam_a,
            academic_session=self.student_a["session"],
            student=self.student_a["profile"],
            total_marks=Decimal("20.00"),
        )
        ExamAttemptUpload.objects.create(
            attempt=attempt,
            image="exams/rough-work/private-upload-one.png",
        )
        ExamAttemptUpload.objects.create(
            attempt=attempt,
            image="exams/rough-work/private-upload-two.png",
        )

        response = self.client.get(
            reverse("institute_admin:institute_exam_submissions", args=[self.exam_a.pk])
        )

        self.assertEqual(response.status_code, 200)
        attempt_row = next(row for row in response.context["rows"] if row["attempt"] == attempt)
        self.assertEqual(attempt_row["upload_count"], 2)
        self.assertNotContains(response, "private-upload-one.png")
        self.assertNotContains(response, "private-upload-two.png")
        self.assertContains(response, 'class="submission-action manage"')
        self.assertContains(response, 'class="submission-action reset"')

    def test_institute_admin_cannot_publish_other_institute_exam(self):
        response = self.client.post(reverse("institute_admin:institute_exam_publish", args=[self.exam_b.pk]))

        self.assertEqual(response.status_code, 404)
        self.exam_b.refresh_from_db()
        self.assertFalse(self.exam_b.is_published)


class ExamResultsReportTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Result Institute",
            code="result-institute",
            status=Institute.Status.ACTIVE,
        )
        self.year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.course = Course.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Mathematics",
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Batch Result",
        )
        self.batch.courses.add(self.course)
        self.exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            title="Ranked Mathematics Exam",
            exam_date=date(2026, 6, 5),
            total_marks=100,
        )
        self.admin = User.objects.create_user(username="result-admin", password="pass12345")
        UserProfile.objects.create(
            user=self.admin,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(self.admin)
        session = self.client.session
        session["academic_year_id"] = self.year.pk
        session.save()

        self.attempts = [
            self.create_attempt("rank-one", "RI-001", Decimal("90.00")),
            self.create_attempt("rank-two-a", "RI-002", Decimal("75.00")),
            self.create_attempt("rank-two-b", "RI-003", Decimal("75.00")),
            self.create_attempt("needs-help", "RI-004", Decimal("35.00")),
        ]

    def create_attempt(self, username, admission_number, score):
        user = User.objects.create_user(username=username, password="pass12345", first_name=username)
        UserProfile.objects.create(
            user=user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=user,
            admission_number=admission_number,
            is_active=True,
        )
        academic_session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=student,
            academic_year=self.year,
            admission_number=admission_number,
            status=StudentAcademicSession.Status.ACTIVE,
        )
        StudentEnrollment.objects.create(
            student=student,
            academic_session=academic_session,
            batch=self.batch,
            status=StudentEnrollment.Status.ACTIVE,
        )
        return ExamAttempt.objects.create(
            exam=self.exam,
            academic_session=academic_session,
            student=student,
            submitted_at=timezone.now(),
            score=score,
            total_marks=Decimal("100.00"),
            correct_count=int(score),
            wrong_count=100 - int(score),
            unattempted_count=0,
        )

    def test_results_are_generated_from_submitted_attempts_with_exam_ranks(self):
        manual_user = User.objects.create_user(username="manual-only", password="pass12345")
        UserProfile.objects.create(
            user=manual_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        manual_student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=manual_user,
            admission_number="RI-MANUAL",
        )
        ExamResult.objects.create(
            exam=self.exam,
            student=manual_student,
            marks_obtained=Decimal("99.00"),
            remark="Manual result must not appear",
        )

        response = self.client.get(reverse("institute_admin:institute_results"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Add Result")
        self.assertNotContains(response, "manual-only")
        ranks = {
            row["attempt"].student.user.username: row["rank"]
            for row in response.context["rows"]
        }
        self.assertEqual(ranks["rank-one"], 1)
        self.assertEqual(ranks["rank-two-a"], 2)
        self.assertEqual(ranks["rank-two-b"], 2)
        self.assertEqual(ranks["needs-help"], 4)

    def test_result_filters_apply_to_page_and_exports(self):
        query = {
            "exam": str(self.exam.pk),
            "performance": "failed",
            "min_percentage": "30",
            "max_percentage": "40",
            "student": "needs-help",
        }
        response = self.client.get(reverse("institute_admin:institute_results"), query)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["result_count"], 1)
        self.assertContains(response, "needs-help")
        self.assertNotContains(response, "rank-one")

        excel_response = self.client.get(
            reverse("institute_admin:institute_results_export"),
            {**query, "format": "excel"},
        )
        self.assertEqual(excel_response.status_code, 200)
        workbook = load_workbook(BytesIO(excel_response.content))
        sheet = workbook["Exam Results"]
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("needs-help", values)
        self.assertNotIn("rank-one", values)

        pdf_response = self.client.get(
            reverse("institute_admin:institute_results_export"),
            {**query, "format": "pdf"},
        )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        self.assertIn(b"needs-help", pdf_response.content)
        self.assertNotIn(b"rank-one", pdf_response.content)

    def test_results_do_not_include_other_institute_attempts(self):
        other_institute = Institute.objects.create(name="Other Result Institute", code="other-results")
        other_year = AcademicYear.objects.create(
            institute=other_institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        other_course = Course.objects.create(
            institute=other_institute,
            academic_year=other_year,
            name="Private Course",
        )
        other_batch = Batch.objects.create(
            institute=other_institute,
            academic_year=other_year,
            name="Private Batch",
        )
        other_batch.courses.add(other_course)
        other_exam = Exam.objects.create(
            academic_year=other_year,
            batch=other_batch,
            course=other_course,
            title="PRIVATE EXAM RESULT",
            exam_date=date(2026, 6, 5),
        )
        other_user = User.objects.create_user(username="private-result-user", password="pass12345")
        other_student = StudentProfile.objects.create(
            institute=other_institute,
            academic_year=other_year,
            user=other_user,
            admission_number="PRIVATE-001",
        )
        other_session = StudentAcademicSession.objects.create(
            institute=other_institute,
            student=other_student,
            academic_year=other_year,
            admission_number="PRIVATE-001",
        )
        ExamAttempt.objects.create(
            exam=other_exam,
            academic_session=other_session,
            student=other_student,
            submitted_at=timezone.now(),
            score=Decimal("100.00"),
            total_marks=Decimal("100.00"),
        )

        response = self.client.get(reverse("institute_admin:institute_results"))

        self.assertNotContains(response, "PRIVATE EXAM RESULT")
        self.assertNotContains(response, "private-result-user")


class SessionAuditCommandTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(
            name="Arohan Academy",
            code="aa",
            status=Institute.Status.ACTIVE,
        )
        self.year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.user = User.objects.create_user(username="audit-student", password="pass12345")
        UserProfile.objects.create(
            user=self.user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
            phone="9000000001",
        )
        self.student = StudentProfile.objects.create(
            institute=self.institute,
            user=self.user,
            academic_year=self.year,
            admission_number="AA-2026-27-0001",
            is_active=True,
        )
        self.session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=self.student,
            academic_year=self.year,
            admission_number="AA-2026-27-0001",
            status=StudentAcademicSession.Status.ACTIVE,
        )
        self.batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="Audit Batch",
            is_active=True,
        )
        self.invoice = FeeInvoice.objects.create(
            institute=self.institute,
            academic_session=self.session,
            student=self.student,
            batch=self.batch,
            title="Audit Fee",
            amount=Decimal("1000.00"),
            due_date=date(2026, 5, 1),
            status=FeeInvoice.Status.UNPAID,
        )

    def test_audit_sessions_reports_clean_data_without_errors(self):
        output = StringIO()

        call_command("audit_sessions", institute_code="aa", stdout=output)

        self.assertIn("Audit completed with 0 errors", output.getvalue())
        self.assertIn("Invoices with active payment total greater than invoice amount", output.getvalue())

    def test_audit_sessions_can_fail_on_error_level_issues(self):
        Payment.objects.create(
            invoice=self.invoice,
            amount=Decimal("1200.00"),
            paid_on=date(2026, 5, 2),
            method=Payment.Method.CASH,
        )
        output = StringIO()

        with self.assertRaises(CommandError):
            call_command("audit_sessions", institute_code="aa", fail_on_issues=True, stdout=output)

        self.assertIn("Invoices with active payment total greater than invoice amount", output.getvalue())
        self.assertIn("Audit completed with", output.getvalue())
