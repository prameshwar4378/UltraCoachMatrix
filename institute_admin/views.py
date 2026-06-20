from datetime import date, datetime, timedelta
from collections import defaultdict
from decimal import Decimal
from io import BytesIO
import csv
import json
import random

from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.contrib.sessions.models import Session
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, connection, transaction
from django.db.models import Avg, Case, Count, DecimalField, ExpressionWrapper, F, IntegerField, Max, OuterRef, Prefetch, Q, Subquery, Sum, Value, When
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.urls import reverse
from django.template import Context, Template, TemplateSyntaxError
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from accountant.models import Expense, ExpenseDocument, FeeCategory, FeeInvoice, Payment, PaymentActivity
from student_parent.models import (
    GuardianProfile,
    StudentBonafideCertificate,
    StudentAcademicSession,
    StudentDocument,
    StudentEnrollment,
    StudentProfile,
    StudentTransferCertificate,
)
from student_parent.notifications import (
    enqueue_fee_paid_notification,
    enqueue_notice_published_notification,
    notify_exam_results_declared,
)
from super_admin.models import Institute, SubscriptionPayment, UserProfile
from super_admin.decorators import institute_admin_required
from super_admin.session_security import user_web_sessions
from teacher.models import (
    Attendance,
    Exam,
    ExamAttempt,
    ExamAttemptUpload,
    ExamQuestion,
    ExamQuestionAttempt,
    ExamQuestionOption,
    ExamResult,
    Homework,
    TeacherProfile,
)
from teacher.forms import ExamQuestionForm, ExamQuestionOptionFormSet, TeacherExamForm


from .forms import (
    AddStudentFeeForm,
    AcademicYearForm,
    BatchForm,
    build_student_username,
    CourseForm,
    DummyStudentCreateForm,
    ExpenseForm,
    LeadForm,
    FeeCategoryForm,
    generate_student_admission_number,
    generate_student_login_credentials,
    get_academic_year_label,
    get_last_student_admission_sequence,
    get_or_create_academic_year,
    get_student_admission_prefix,
    HomeworkForm,
    InstituteProfileForm,
    InstitutePrintTemplateForm,
    InstituteUserForm,
    NoticeForm,
    PaymentUpdateForm,
    PaymentVoidForm,
    ReceiveFeeForm,
    SecurityPasswordChangeForm,
    SupportTicketForm,
    StudentBasicForm,
    StudentBonafideCertificateForm,
    StudentDocumentUploadForm,
    StudentEducationForm,
    StudentEnrollmentForm,
    StudentForm,
    StudentGuardianForm,
    StudentIdCardForm,
    StudentTransferCertificateForm,
    SubjectForm,
    TeacherForm,
    VisitorForm,
)
from .dashboard_cache import (
    get_dashboard_summary,
    invalidate_dashboard_summary,
    set_dashboard_summary,
)
from .attendance_service import bulk_save_attendance
from .background_jobs import enqueue_background_job
from .lookup_cache import (
    get_cached_batch_course_data,
    get_cached_course_batch_data,
    invalidate_academic_years_cache,
)
from .models import (
    AcademicYear,
    BackgroundJob,
    Batch,
    Course,
    InstituteGlobalPrintTemplate,
    InstitutePrintTemplate,
    Lead,
    Notice,
    PrintDocumentType,
    Subject,
    SupportTicket,
    Visitor,
)
from UltraCoachMatrix.email_notifications import (
    on_commit_email,
    send_institute_welcome,
    send_payment_confirmation,
    send_payment_update,
    send_bulk_student_welcomes,
    send_student_welcome,
    send_teacher_welcome,
)


def sync_students_to_academic_session(institute, academic_year):
    with transaction.atomic():
        locked_year = AcademicYear.objects.select_for_update().get(
            pk=academic_year.pk,
            institute=institute,
        )
        students = list(
            StudentProfile.objects.filter(
                institute=institute,
                is_active=True,
            )
            .exclude(academic_sessions__academic_year=locked_year)
            .order_by("pk")
        )
        if not students:
            return 0

        prefix = get_student_admission_prefix(institute, locked_year)
        sequence = get_last_student_admission_sequence(institute, locked_year)
        sessions = []
        for student in students:
            sequence += 1
            sessions.append(
                StudentAcademicSession(
                    institute=institute,
                    student=student,
                    academic_year=locked_year,
                    admission_number=f"{prefix}{sequence:04d}",
                    joined_on=timezone.localdate(),
                    status=StudentAcademicSession.Status.ACTIVE,
                    current_school_name=student.current_school_name,
                    current_school_address=student.current_school_address,
                    previous_school_name=student.previous_school_name,
                    previous_class=student.previous_class,
                )
            )
        StudentAcademicSession.objects.bulk_create(sessions, batch_size=500)
        return len(sessions)


def close_popup_response(receipt_url=None):
    receipt_script = ""
    if receipt_url:
        receipt_script = (
            f"window.open({json.dumps(receipt_url)}, "
            "'feeReceiptWindow', "
            "'width=980,height=900,scrollbars=yes,resizable=yes');"
        )
    return HttpResponse(
        """
        <script>
            __RECEIPT_SCRIPT__
            if (window.opener) {
                window.opener.location.reload();
                window.close();
            } else {
                window.location.href = '/institute/';
            }
        </script>
        """.replace("__RECEIPT_SCRIPT__", receipt_script)
    )


def render_print_document(request, document_type, default_template_name, context):
    institute = get_current_institute(request)
    custom_template = None
    if institute:
        custom_template = (
            InstitutePrintTemplate.objects.filter(
                institute=institute,
                document_type=document_type,
                is_active=True,
            )
            .order_by("-updated_at")
            .first()
        )
    if not custom_template:
        return render(request, default_template_name, context)

    try:
        html_file = custom_template.effective_html_file
        if not html_file:
            return render(request, default_template_name, context)
        with html_file.open("rb") as template_file:
            template_source = template_file.read().decode("utf-8-sig")
        rendered_html = Template(template_source).render(
            Context(
                {
                    **context,
                    "request": request,
                    "institute": institute,
                    "print_template": custom_template,
                    "generated_at": context.get("generated_at") or timezone.localtime(),
                }
            )
        )
    except (OSError, UnicodeDecodeError, TemplateSyntaxError) as exc:
        messages.warning(
            request,
            f"Custom {custom_template.get_document_type_display()} template could not be rendered. Default template was used. {exc}",
        )
        return render(request, default_template_name, context)

    return HttpResponse(rendered_html)


def paginate_queryset(request, queryset, per_page=20):
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))
    query_params = request.GET.copy()
    query_params.pop("page", None)
    return page_obj, paginator, query_params.urlencode()


@institute_admin_required
def print_template_list(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before managing print templates.")
        return redirect("institute_admin:dashboard")
    templates = {
        template.document_type: template
        for template in InstitutePrintTemplate.objects.filter(institute=institute)
    }
    library_templates = (
        InstituteGlobalPrintTemplate.objects.filter(is_active=True)
        .order_by("document_type", "title")
    )
    library_by_type = defaultdict(list)
    for template in library_templates:
        library_by_type[template.document_type].append(template)
    groups = [
        {
            "document_type": document_type,
            "label": label,
            "selected_template": templates.get(document_type),
            "library_templates": library_by_type.get(document_type, []),
        }
        for document_type, label in PrintDocumentType.choices
    ]
    return render(
        request,
        "institute_admin/print_template_list.html",
        {"groups": groups, "institute": institute},
    )


@institute_admin_required
def print_template_create(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before uploading a print template.")
        return redirect("institute_admin:dashboard")
    if request.method == "POST":
        form = InstitutePrintTemplateForm(request.POST, request.FILES)
        if form.is_valid():
            document_type = form.cleaned_data["document_type"]
            template = InstitutePrintTemplate.objects.filter(
                institute=institute,
                document_type=document_type,
            ).first()
            if template is None:
                template = form.save(commit=False)
                template.institute = institute
            else:
                template.title = form.cleaned_data["title"]
                template.html_file = form.cleaned_data["html_file"]
                template.is_active = form.cleaned_data["is_active"]
                template.library_template = None
            template.uploaded_by = request.user
            template.save()
            messages.success(request, "Print template saved successfully.")
            return close_popup_response()
    else:
        form = InstitutePrintTemplateForm(
            initial={"document_type": request.GET.get("type", "")}
        )
    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Upload Print Template",
            "subtitle": "Select TC, Admission Form, or Bonafide, then upload that institute's HTML print template.",
            "button_text": "Save Template",
            "icon": "bi-filetype-html",
        },
    )


@institute_admin_required
def print_template_update(request, pk):
    institute = get_current_institute(request)
    template = get_object_or_404(InstitutePrintTemplate, pk=pk, institute=institute)
    if request.method == "POST":
        form = InstitutePrintTemplateForm(request.POST, request.FILES, instance=template)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.institute = institute
            updated.uploaded_by = request.user
            updated.save()
            messages.success(request, "Print template updated successfully.")
            return close_popup_response()
    else:
        form = InstitutePrintTemplateForm(instance=template)
    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Update Print Template",
            "subtitle": f"Update the {template.get_document_type_display()} HTML template.",
            "button_text": "Update Template",
            "icon": "bi-filetype-html",
        },
    )


@require_POST
@institute_admin_required
def print_template_delete(request, pk):
    institute = get_current_institute(request)
    template = get_object_or_404(InstitutePrintTemplate, pk=pk, institute=institute)
    template.delete()
    messages.success(request, "Custom print template deleted. Default template will be used.")
    return redirect("institute_admin:print_template_list")


def _global_template_for_request(request, pk):
    return get_object_or_404(InstituteGlobalPrintTemplate, pk=pk, is_active=True)


@institute_admin_required
def print_template_library_view(request, pk):
    template = _global_template_for_request(request, pk)
    try:
        with template.html_file.open("rb") as html_file:
            html = html_file.read().decode("utf-8-sig")
    except (OSError, UnicodeDecodeError) as exc:
        return HttpResponse(f"Template could not be opened: {exc}", status=500)
    return HttpResponse(html)


@institute_admin_required
def print_template_library_download(request, pk):
    template = _global_template_for_request(request, pk)
    try:
        with template.html_file.open("rb") as html_file:
            content = html_file.read()
    except OSError:
        return HttpResponse("Template file not found.", status=404)
    filename = template.html_file.name.rsplit("/", 1)[-1] or "template.html"
    response = HttpResponse(content, content_type="text/html")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_POST
@institute_admin_required
def print_template_set_library(request, pk):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before setting a print template.")
        return redirect("institute_admin:print_template_list")
    library_template = _global_template_for_request(request, pk)
    InstitutePrintTemplate.objects.update_or_create(
        institute=institute,
        document_type=library_template.document_type,
        defaults={
            "title": library_template.title,
            "html_file": "",
            "library_template": library_template,
            "is_active": True,
            "uploaded_by": request.user,
        },
    )
    messages.success(
        request,
        f"{library_template.get_document_type_display()} template set for this institute.",
    )
    return redirect("institute_admin:print_template_list")


def get_current_institute(request):
    profile = getattr(request.user, "profile", None)
    if request.user.is_superuser:
        return None
    if not profile or profile.role != UserProfile.Role.INSTITUTE_ADMIN:
        raise PermissionDenied("Only institute admins can access this page.")
    return profile.institute


@login_required
def student_autocomplete(request):
    profile = getattr(request.user, "profile", None)
    allowed_roles = {UserProfile.Role.INSTITUTE_ADMIN, UserProfile.Role.TEACHER}
    if not profile or profile.role not in allowed_roles or not profile.institute_id:
        return JsonResponse({"detail": "You are not allowed to search students."}, status=403)

    query = request.GET.get("q", "").strip()
    if len(query) < 2:
        return JsonResponse({"results": [], "pagination": {"more": False}})

    academic_year_id = request.GET.get("academic_year", "").strip()
    if not academic_year_id:
        academic_year_id = request.session.get("academic_year_id")

    sessions = StudentAcademicSession.objects.filter(
        institute=profile.institute,
        student__is_active=True,
    ).select_related("student", "student__user", "student__user__profile")
    if academic_year_id:
        sessions = sessions.filter(academic_year_id=academic_year_id)

    batch_id = request.GET.get("batch", "").strip()
    course_id = request.GET.get("course", "").strip()
    if batch_id:
        sessions = sessions.filter(enrollments__batch_id=batch_id)
    if course_id:
        sessions = sessions.filter(enrollments__courses__id=course_id)
    if profile.role == UserProfile.Role.TEACHER:
        sessions = sessions.filter(enrollments__batch__teachers=request.user)

    sessions = (
        sessions.filter(
            Q(admission_number__icontains=query)
            | Q(student__user__first_name__icontains=query)
            | Q(student__user__last_name__icontains=query)
            | Q(student__user__username__icontains=query)
            | Q(student__user__email__icontains=query)
            | Q(student__user__profile__phone__icontains=query)
            | Q(student__guardians__name__icontains=query)
            | Q(student__guardians__phone__icontains=query)
        )
        .distinct()
        .order_by("admission_number", "student__user__first_name", "student__user__username")
    )
    matches = list(sessions[:21])
    results = []
    for session in matches[:20]:
        user = session.student.user
        name = user.get_full_name() or user.username
        phone = getattr(getattr(user, "profile", None), "phone", "")
        meta_parts = [f"Admission: {session.admission_number}", f"Username: {user.username}"]
        if phone:
            meta_parts.append(f"Phone: {phone}")
        results.append(
            {
                "id": session.student_id,
                "text": f"{session.admission_number} - {name}",
                "meta": " | ".join(meta_parts),
            }
        )
    return JsonResponse({"results": results, "pagination": {"more": len(matches) > 20}})


@institute_admin_required
def software_tour(request):
    profile = request.user.profile
    subscription = getattr(profile.institute, "subscription", None)

    if profile.onboarding_completed_at:
        return redirect("institute_admin:dashboard")

    if request.method == "POST":
        profile.onboarding_completed_at = timezone.now()
        profile.save(update_fields=["onboarding_completed_at"])
        messages.success(request, "Setup tour completed. Welcome to UltraCoachMatrix.")
        return redirect("institute_admin:dashboard")

    return render(
        request,
        "institute_admin/software_tour.html",
        {
            "institute": profile.institute,
            "subscription": subscription,
        },
    )


@institute_admin_required
def institute_profile(request):
    institute = get_current_institute(request)
    form = InstituteProfileForm(request.POST or None, request.FILES or None, instance=institute)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Institute profile updated successfully.")
        return redirect("institute_profile")

    return render(
        request,
        "institute_admin/institute_profile.html",
        {
            "form": form,
            "institute": institute,
        },
    )


@institute_admin_required
def academic_session_settings(request):
    institute = get_current_institute(request)
    edit_id = request.GET.get("edit", "").strip()
    editing_session = None
    if edit_id:
        editing_session = get_object_or_404(AcademicYear, pk=edit_id, institute=institute)

    form = AcademicYearForm(instance=editing_session, institute=institute)
    if request.method == "POST":
        action = request.POST.get("action", "save")
        session_id = request.POST.get("session_id", "").strip()

        if action == "save":
            instance = None
            if session_id:
                instance = get_object_or_404(AcademicYear, pk=session_id, institute=institute)
            form = AcademicYearForm(request.POST, instance=instance, institute=institute)
            if form.is_valid():
                academic_session = form.save(commit=False)
                academic_session.institute = institute
                academic_session.save()
                synced_count = sync_students_to_academic_session(institute, academic_session)
                invalidate_academic_years_cache(institute.pk)
                messages.success(
                    request,
                    f"Academic session {academic_session.name} "
                    f"{'updated' if instance else 'created'} successfully. "
                    f"{synced_count} active student session(s) synced for mobile access.",
                )
                return redirect("institute_admin:academic_session_settings")
            editing_session = instance

        elif action == "sync_students":
            academic_session = get_object_or_404(AcademicYear, pk=session_id, institute=institute)
            synced_count = sync_students_to_academic_session(institute, academic_session)
            if synced_count:
                messages.success(
                    request,
                    f"{synced_count} active student session(s) added to {academic_session.name}.",
                )
            else:
                messages.info(
                    request,
                    f"All active students are already synced with {academic_session.name}.",
                )
            return redirect("institute_admin:academic_session_settings")

        elif action == "set_current":
            academic_session = get_object_or_404(AcademicYear, pk=session_id, institute=institute)
            request.session["academic_year_id"] = academic_session.pk
            request._selected_academic_year = academic_session
            request._academic_year_context = None
            messages.success(request, f"{academic_session.name} is now your current session.")
            return redirect("institute_admin:academic_session_settings")

        elif action == "toggle_active":
            academic_session = get_object_or_404(AcademicYear, pk=session_id, institute=institute)
            academic_session.is_active = not academic_session.is_active
            academic_session.save(update_fields=["is_active"])
            messages.success(
                request,
                f"{academic_session.name} marked "
                f"{'active' if academic_session.is_active else 'inactive'}.",
            )
            return redirect("institute_admin:academic_session_settings")

        elif action == "delete":
            academic_session = get_object_or_404(AcademicYear, pk=session_id, institute=institute)
            session_name = academic_session.name
            try:
                academic_session.delete()
            except ProtectedError:
                messages.error(
                    request,
                    f"{session_name} cannot be deleted because academic records are linked to it. "
                    "Mark it inactive instead.",
                )
            else:
                if str(request.session.get("academic_year_id", "")) == str(session_id):
                    request.session.pop("academic_year_id", None)
                messages.success(request, f"Academic session {session_name} deleted.")
            return redirect("institute_admin:academic_session_settings")

        else:
            messages.error(request, "Invalid academic session action.")
            return redirect("institute_admin:academic_session_settings")

    academic_sessions = AcademicYear.objects.filter(institute=institute).annotate(
        course_count=Count("courses", distinct=True),
        batch_count=Count("batches", distinct=True),
        student_count=Count("student_sessions", distinct=True),
    )
    return render(
        request,
        "institute_admin/academic_session_settings.html",
        {
            "form": form,
            "editing_session": editing_session,
            "academic_sessions": academic_sessions,
            "institute": institute,
        },
    )


@institute_admin_required
def subscription_billing(request):
    institute = get_current_institute(request)
    subscription = getattr(institute, "subscription", None)
    payments = SubscriptionPayment.objects.filter(institute=institute)

    days_remaining = None
    if subscription and subscription.ends_on:
        days_remaining = max((subscription.ends_on - timezone.localdate()).days, 0)

    if institute.status not in {institute.Status.ACTIVE, institute.Status.TRIAL}:
        access_status = institute.get_status_display()
    elif subscription and subscription.is_expired:
        access_status = "Expired"
    elif subscription and not subscription.is_active:
        access_status = "Not started"
    else:
        access_status = "Active"

    return render(
        request,
        "institute_admin/subscription_billing.html",
        {
            "institute": institute,
            "subscription": subscription,
            "payments": payments,
            "days_remaining": days_remaining,
            "access_status": access_status,
        },
    )


@institute_admin_required
def subscription_payment_bill(request, pk):
    institute = get_current_institute(request)
    payment = get_object_or_404(
        SubscriptionPayment,
        pk=pk,
        institute=institute,
    )
    subscription = getattr(institute, "subscription", None)

    return render(
        request,
        "institute_admin/subscription_payment_bill.html",
        {
            "institute": institute,
            "subscription": subscription,
            "payment": payment,
            "bill_number": f"UCM-{payment.paid_on:%Y%m%d}-{payment.pk:06d}",
        },
    )


@institute_admin_required
def security_settings(request):
    institute = get_current_institute(request)
    password_form = SecurityPasswordChangeForm(user=request.user)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "change_password":
            password_form = SecurityPasswordChangeForm(request.POST, user=request.user)
            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, "Your password has been changed successfully.")
                return redirect("security_settings")

        elif action == "logout_other_sessions":
            current_session_key = request.session.session_key
            other_session_keys = [
                session["session_key"]
                for session in user_web_sessions(request.user, request.session)
                if session["session_key"] != current_session_key
            ]
            deleted_count = Session.objects.filter(session_key__in=other_session_keys).delete()[0]
            if deleted_count:
                messages.success(request, f"Signed out {deleted_count} other session(s).")
            else:
                messages.info(request, "No other active sessions were found.")
            return redirect("security_settings")

        elif action == "logout_session":
            requested_identifier = request.POST.get("session_identifier", "")
            matching_session = next(
                (
                    session
                    for session in user_web_sessions(request.user, request.session)
                    if session["identifier"] == requested_identifier
                ),
                None,
            )
            if not matching_session:
                messages.error(request, "That session is no longer active.")
            elif matching_session["session_key"] == request.session.session_key:
                messages.warning(request, "Use the main logout option to sign out this device.")
            else:
                Session.objects.filter(session_key=matching_session["session_key"]).delete()
                messages.success(request, "The selected device has been signed out.")
            return redirect("security_settings")

        else:
            messages.error(request, "Invalid security action.")
            return redirect("security_settings")

    institute_users = (
        UserProfile.objects.filter(institute=institute)
        .select_related("user")
        .order_by("role", "user__username")
    )
    active_user_count = institute_users.filter(user__is_active=True).count()
    inactive_user_count = institute_users.filter(user__is_active=False).count()
    web_sessions = user_web_sessions(request.user, request.session)
    for web_session in web_sessions:
        web_session["is_current"] = web_session["session_key"] == request.session.session_key
    session_count = len(web_sessions)

    return render(
        request,
        "institute_admin/security_settings.html",
        {
            "institute": institute,
            "password_form": password_form,
            "institute_users": institute_users,
            "active_user_count": active_user_count,
            "inactive_user_count": inactive_user_count,
            "session_count": session_count,
            "web_sessions": web_sessions,
        },
    )


@institute_admin_required
def help_support(request):
    institute = get_current_institute(request)
    ticket_form = SupportTicketForm(request.POST or None)

    if request.method == "POST" and ticket_form.is_valid():
        ticket = ticket_form.save(commit=False)
        ticket.institute = institute
        ticket.created_by = request.user
        ticket.save()
        messages.success(
            request,
            f"Support request #{ticket.pk} submitted successfully.",
        )
        return redirect("help_support")

    tickets = SupportTicket.objects.filter(institute=institute).select_related("created_by")
    open_ticket_count = tickets.filter(
        status__in=[SupportTicket.Status.NEW, SupportTicket.Status.IN_PROGRESS]
    ).count()

    return render(
        request,
        "institute_admin/help_support.html",
        {
            "institute": institute,
            "ticket_form": ticket_form,
            "tickets": tickets[:20],
            "open_ticket_count": open_ticket_count,
        },
    )


def get_current_academic_year(request, institute=None):
    institute = institute or get_current_institute(request)
    if not institute:
        return None

    cached_year = getattr(request, "_selected_academic_year", None)
    if cached_year and cached_year.institute_id == institute.pk:
        return cached_year

    year_id = request.session.get("academic_year_id")
    if year_id:
        academic_year = AcademicYear.objects.filter(pk=year_id, institute=institute).first()
        if academic_year:
            request._selected_academic_year = academic_year
            return academic_year
    academic_year = get_or_create_academic_year(institute)
    request.session["academic_year_id"] = academic_year.pk
    request._selected_academic_year = academic_year
    return academic_year


def get_batch_course_data(institute, academic_year=None):
    if not institute:
        return {}
    return get_cached_batch_course_data(
        institute.pk,
        academic_year.pk if academic_year else None,
    )


def get_course_batch_data(institute, academic_year=None):
    if not institute:
        return {}
    return get_cached_course_batch_data(
        institute.pk,
        academic_year.pk if academic_year else None,
    )


LEAD_EXPORT_FIELD_OPTIONS = [
    ("name", "Name"),
    ("mobile_number", "Mobile Number"),
    ("email", "Email"),
    ("status", "Status"),
    ("source", "Source"),
    ("interested_class", "Interested Class"),
    ("interested_batch", "Interested Batch"),
    ("follow_up_on", "Follow Up Date"),
    ("message", "Message / Notes"),
    ("created_by", "Created By"),
    ("created_at", "Created Date"),
    ("updated_at", "Updated Date"),
    ("converted_at", "Converted Date"),
]

LEAD_EXPORT_DEFAULT_FIELDS = [
    "name",
    "mobile_number",
    "email",
    "status",
    "interested_class",
    "interested_batch",
    "follow_up_on",
    "created_at",
]


def get_lead_export_field_keys(request):
    allowed_fields = {field for field, _label in LEAD_EXPORT_FIELD_OPTIONS}
    selected_fields = [
        field
        for field in request.GET.getlist("columns")
        if field in allowed_fields
    ]
    return selected_fields or LEAD_EXPORT_DEFAULT_FIELDS


def get_lead_export_value(lead, field):
    if field == "name":
        return lead.full_name
    if field == "mobile_number":
        return lead.mobile_number
    if field == "email":
        return lead.email
    if field == "status":
        return lead.get_status_display()
    if field == "source":
        return lead.get_source_display()
    if field == "interested_class":
        return lead.interested_class.name if lead.interested_class else ""
    if field == "interested_batch":
        return lead.interested_batch.name if lead.interested_batch else ""
    if field == "follow_up_on":
        return lead.follow_up_on.strftime("%Y-%m-%d") if lead.follow_up_on else ""
    if field == "message":
        return lead.message
    if field == "created_by":
        if not lead.created_by:
            return ""
        return lead.created_by.get_full_name() or lead.created_by.username
    if field == "created_at":
        return timezone.localtime(lead.created_at).strftime("%Y-%m-%d %H:%M") if lead.created_at else ""
    if field == "updated_at":
        return timezone.localtime(lead.updated_at).strftime("%Y-%m-%d %H:%M") if lead.updated_at else ""
    if field == "converted_at":
        return timezone.localtime(lead.converted_at).strftime("%Y-%m-%d %H:%M") if lead.converted_at else ""
    return ""


@institute_admin_required
def lead_list(request):
    institute = get_current_institute(request)
    leads = Lead.objects.select_related(
        "interested_class",
        "interested_batch",
        "created_by",
    )
    if institute:
        leads = leads.filter(institute=institute)

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        leads = leads.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(mobile_number__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(interested_class__name__icontains=search_query)
            | Q(interested_batch__name__icontains=search_query)
        )
    if status_filter in Lead.Status.values:
        leads = leads.filter(status=status_filter)

    base_queryset = Lead.objects.filter(institute=institute)
    paginator = Paginator(leads, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "institute_admin/lead_list.html",
        {
            "leads": page_obj.object_list,
            "page_obj": page_obj,
            "search_query": search_query,
            "status_filter": status_filter,
            "status_choices": Lead.Status.choices,
            "lead_export_field_options": LEAD_EXPORT_FIELD_OPTIONS,
            "lead_export_default_fields": LEAD_EXPORT_DEFAULT_FIELDS,
            "total_leads": base_queryset.count(),
            "new_leads": base_queryset.filter(status=Lead.Status.NEW).count(),
            "follow_up_leads": base_queryset.filter(status=Lead.Status.FOLLOW_UP).count(),
            "converted_leads": base_queryset.filter(status=Lead.Status.CONVERTED).count(),
        },
    )


@institute_admin_required
def lead_export(request):
    institute = get_current_institute(request)
    leads = Lead.objects.select_related(
        "interested_class",
        "interested_batch",
        "created_by",
    )
    if institute:
        leads = leads.filter(institute=institute)
    else:
        leads = leads.none()

    status_filter = request.GET.get("status", "").strip()
    start_date = parse_date(request.GET.get("start_date", "").strip())
    end_date = parse_date(request.GET.get("end_date", "").strip())

    if status_filter in Lead.Status.values:
        leads = leads.filter(status=status_filter)
    if start_date:
        leads = leads.filter(created_at__date__gte=start_date)
    if end_date:
        leads = leads.filter(created_at__date__lte=end_date)

    selected_fields = get_lead_export_field_keys(request)
    field_labels = dict(LEAD_EXPORT_FIELD_OPTIONS)
    columns = [field_labels[field] for field in selected_fields]

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="lead_export_{stamp}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(columns)
    for lead in leads.order_by("-created_at"):
        writer.writerow([get_lead_export_value(lead, field) for field in selected_fields])

    return response


@institute_admin_required
def lead_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if request.method == "POST":
        form = LeadForm(
            request.POST,
            institute=institute,
            academic_year=academic_year,
        )
        if form.is_valid():
            lead = form.save(commit=False)
            lead.institute = institute
            lead.created_by = request.user
            lead.save()
            messages.success(request, "Lead created successfully.")
            return close_popup_response()
    else:
        form = LeadForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/lead_form.html",
        {
            "form": form,
            "title": "Create Lead",
            "button_text": "Save Lead",
            "course_batch_data": get_course_batch_data(institute, academic_year),
        },
    )


@institute_admin_required
def lead_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    lead = get_object_or_404(Lead, pk=pk, institute=institute)

    if request.method == "POST":
        form = LeadForm(
            request.POST,
            instance=lead,
            institute=institute,
            academic_year=academic_year,
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Lead updated successfully.")
            return close_popup_response()
    else:
        form = LeadForm(
            instance=lead,
            institute=institute,
            academic_year=academic_year,
        )

    return render(
        request,
        "institute_admin/lead_form.html",
        {
            "form": form,
            "title": "Edit Lead",
            "button_text": "Update Lead",
            "course_batch_data": get_course_batch_data(institute, academic_year),
        },
    )


@institute_admin_required
@require_POST
def lead_delete(request, pk):
    institute = get_current_institute(request)
    lead = get_object_or_404(Lead, pk=pk, institute=institute)
    lead.delete()
    messages.success(request, "Lead deleted successfully.")
    return redirect("institute_admin:lead_list")


EXPENSE_EXPORT_FIELD_OPTIONS = [
    ("title", "Title"),
    ("amount", "Amount"),
    ("spent_on", "Spent Date"),
    ("note", "Note"),
    ("recorded_by", "Recorded By"),
]

EXPENSE_EXPORT_DEFAULT_FIELDS = ["title", "amount", "spent_on", "recorded_by"]


def get_expense_export_field_keys(request):
    allowed = {field for field, _ in EXPENSE_EXPORT_FIELD_OPTIONS}
    selected = [field for field in request.GET.getlist("columns") if field in allowed]
    return selected or EXPENSE_EXPORT_DEFAULT_FIELDS


def get_expense_export_value(expense, field):
    if field == "title":
        return expense.title
    if field == "amount":
        return expense.amount
    if field == "spent_on":
        return expense.spent_on.strftime("%Y-%m-%d") if expense.spent_on else ""
    if field == "note":
        return expense.note
    if field == "recorded_by":
        if not expense.recorded_by:
            return ""
        return expense.recorded_by.get_full_name() or expense.recorded_by.username
    return ""


def save_expense_documents(expense, uploaded_files):
    for uploaded_file in uploaded_files or []:
        if uploaded_file:
            ExpenseDocument.objects.create(expense=expense, file=uploaded_file)


@institute_admin_required
def expense_list(request):
    institute = get_current_institute(request)
    expenses = Expense.objects.select_related("recorded_by").prefetch_related("documents")
    if institute:
        expenses = expenses.filter(institute=institute)
    else:
        expenses = expenses.none()

    search_query = request.GET.get("search", "").strip()
    start_date_value = request.GET.get("start_date", "").strip()
    end_date_value = request.GET.get("end_date", "").strip()
    start_date = parse_date(start_date_value)
    end_date = parse_date(end_date_value)

    if search_query:
        expenses = expenses.filter(
            Q(title__icontains=search_query)
            | Q(note__icontains=search_query)
            | Q(recorded_by__username__icontains=search_query)
            | Q(recorded_by__first_name__icontains=search_query)
            | Q(recorded_by__last_name__icontains=search_query)
        )
    if start_date:
        expenses = expenses.filter(spent_on__gte=start_date)
    if end_date:
        expenses = expenses.filter(spent_on__lte=end_date)

    base_queryset = Expense.objects.filter(institute=institute) if institute else Expense.objects.none()
    today = timezone.localdate()
    month_start = today.replace(day=1)
    filtered_total = expenses.aggregate(total=Coalesce(Sum("amount"), Decimal("0")))["total"]
    paginator = Paginator(expenses, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "institute_admin/expense_list.html",
        {
            "expenses": page_obj.object_list,
            "page_obj": page_obj,
            "search_query": search_query,
            "start_date": start_date_value,
            "end_date": end_date_value,
            "expense_export_field_options": EXPENSE_EXPORT_FIELD_OPTIONS,
            "expense_export_default_fields": EXPENSE_EXPORT_DEFAULT_FIELDS,
            "total_expenses": base_queryset.count(),
            "total_amount": base_queryset.aggregate(total=Coalesce(Sum("amount"), Decimal("0")))["total"],
            "month_amount": base_queryset.filter(spent_on__gte=month_start).aggregate(
                total=Coalesce(Sum("amount"), Decimal("0"))
            )["total"],
            "filtered_amount": filtered_total,
        },
    )


@institute_admin_required
def expense_export(request):
    institute = get_current_institute(request)
    expenses = Expense.objects.select_related("recorded_by").prefetch_related("documents")
    if institute:
        expenses = expenses.filter(institute=institute)
    else:
        expenses = expenses.none()

    search_query = request.GET.get("search", "").strip()
    start_date = parse_date(request.GET.get("start_date", "").strip())
    end_date = parse_date(request.GET.get("end_date", "").strip())

    if search_query:
        expenses = expenses.filter(
            Q(title__icontains=search_query)
            | Q(note__icontains=search_query)
            | Q(recorded_by__username__icontains=search_query)
            | Q(recorded_by__first_name__icontains=search_query)
            | Q(recorded_by__last_name__icontains=search_query)
        )
    if start_date:
        expenses = expenses.filter(spent_on__gte=start_date)
    if end_date:
        expenses = expenses.filter(spent_on__lte=end_date)

    selected_fields = get_expense_export_field_keys(request)
    field_labels = dict(EXPENSE_EXPORT_FIELD_OPTIONS)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="expense_export_{stamp}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow([field_labels[field] for field in selected_fields])
    for expense in expenses.order_by("-spent_on", "-pk"):
        writer.writerow([get_expense_export_value(expense, field) for field in selected_fields])
    return response


@institute_admin_required
def expense_create(request):
    institute = get_current_institute(request)
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.institute = institute
            expense.recorded_by = request.user
            expense.save()
            save_expense_documents(expense, form.cleaned_data.get("files"))
            messages.success(request, "Expense created successfully.")
            return close_popup_response()
    else:
        form = ExpenseForm()

    return render(
        request,
        "institute_admin/expense_form.html",
        {"form": form, "title": "Create Expense", "button_text": "Save Expense"},
    )


@institute_admin_required
def expense_update(request, pk):
    institute = get_current_institute(request)
    expense = get_object_or_404(
        Expense.objects.prefetch_related("documents"),
        pk=pk,
        institute=institute,
    )
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            expense = form.save()
            save_expense_documents(expense, form.cleaned_data.get("files"))
            messages.success(request, "Expense updated successfully.")
            return close_popup_response()
    else:
        form = ExpenseForm(instance=expense)

    return render(
        request,
        "institute_admin/expense_form.html",
        {
            "form": form,
            "title": "Edit Expense",
            "button_text": "Update Expense",
            "expense": expense,
        },
    )


@institute_admin_required
@require_POST
def expense_delete(request, pk):
    institute = get_current_institute(request)
    expense = get_object_or_404(Expense, pk=pk, institute=institute)
    expense.delete()
    messages.success(request, "Expense deleted successfully.")
    return redirect("institute_admin:expense_list")


@institute_admin_required
@require_POST
def expense_document_delete(request, pk):
    institute = get_current_institute(request)
    document = get_object_or_404(
        ExpenseDocument.objects.select_related("expense"),
        pk=pk,
        expense__institute=institute,
    )
    if document.file:
        document.file.delete(save=False)
    document.delete()
    messages.success(request, "Expense document deleted successfully.")
    return close_popup_response()


VISITOR_EXPORT_FIELD_OPTIONS = [
    ("visitor_name", "Visitor Name"),
    ("phone_number", "Phone Number"),
    ("id_card_number", "ID Card / Pass No"),
    ("meeting_with", "Meeting With"),
    ("total_person", "Total Persons"),
    ("visit_date", "Visit Date"),
    ("entry_time", "Entry Time"),
    ("exit_time", "Exit Time"),
    ("purpose", "Purpose"),
    ("attachment", "Attachment"),
    ("created_by", "Created By"),
    ("created_at", "Created Date"),
    ("updated_at", "Updated Date"),
]

VISITOR_EXPORT_DEFAULT_FIELDS = [
    "visitor_name",
    "phone_number",
    "meeting_with",
    "total_person",
    "visit_date",
    "entry_time",
    "exit_time",
    "purpose",
]


def get_visitor_export_field_keys(request):
    allowed_fields = {field for field, _label in VISITOR_EXPORT_FIELD_OPTIONS}
    selected_fields = [
        field
        for field in request.GET.getlist("columns")
        if field in allowed_fields
    ]
    return selected_fields or VISITOR_EXPORT_DEFAULT_FIELDS


def get_visitor_export_value(visitor, field, request=None):
    if field == "visitor_name":
        return visitor.visitor_name
    if field == "phone_number":
        return visitor.phone_number
    if field == "id_card_number":
        return visitor.id_card_number
    if field == "meeting_with":
        return visitor.meeting_with
    if field == "total_person":
        return visitor.total_person
    if field == "visit_date":
        return visitor.visit_date.strftime("%Y-%m-%d") if visitor.visit_date else ""
    if field == "entry_time":
        return visitor.entry_time.strftime("%H:%M") if visitor.entry_time else ""
    if field == "exit_time":
        return visitor.exit_time.strftime("%H:%M") if visitor.exit_time else ""
    if field == "purpose":
        return visitor.purpose
    if field == "attachment":
        if not visitor.attachment:
            return ""
        return request.build_absolute_uri(visitor.attachment.url) if request else visitor.attachment.url
    if field == "created_by":
        if not visitor.created_by:
            return ""
        return visitor.created_by.get_full_name() or visitor.created_by.username
    if field == "created_at":
        return timezone.localtime(visitor.created_at).strftime("%Y-%m-%d %H:%M") if visitor.created_at else ""
    if field == "updated_at":
        return timezone.localtime(visitor.updated_at).strftime("%Y-%m-%d %H:%M") if visitor.updated_at else ""
    return ""


@institute_admin_required
def visitor_list(request):
    institute = get_current_institute(request)
    visitors = Visitor.objects.select_related("created_by").filter(institute=institute)

    search_query = request.GET.get("search", "").strip()
    visit_date_filter = request.GET.get("visit_date", "").strip()

    if search_query:
        visitors = visitors.filter(
            Q(visitor_name__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(id_card_number__icontains=search_query)
            | Q(meeting_with__icontains=search_query)
            | Q(purpose__icontains=search_query)
        )
    if visit_date_filter:
        visit_date = parse_date(visit_date_filter)
        if visit_date:
            visitors = visitors.filter(visit_date=visit_date)

    paginator = Paginator(visitors, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "institute_admin/visitor_list.html",
        {
            "visitors": page_obj.object_list,
            "page_obj": page_obj,
            "search_query": search_query,
            "visit_date_filter": visit_date_filter,
            "visitor_export_field_options": VISITOR_EXPORT_FIELD_OPTIONS,
            "visitor_export_default_fields": VISITOR_EXPORT_DEFAULT_FIELDS,
        },
    )


@institute_admin_required
def visitor_export(request):
    institute = get_current_institute(request)
    visitors = Visitor.objects.select_related("created_by")
    if institute:
        visitors = visitors.filter(institute=institute)
    else:
        visitors = visitors.none()

    search_query = request.GET.get("search", "").strip()
    start_date = parse_date(request.GET.get("start_date", "").strip())
    end_date = parse_date(request.GET.get("end_date", "").strip())

    if search_query:
        visitors = visitors.filter(
            Q(visitor_name__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(id_card_number__icontains=search_query)
            | Q(meeting_with__icontains=search_query)
            | Q(purpose__icontains=search_query)
        )
    if start_date:
        visitors = visitors.filter(visit_date__gte=start_date)
    if end_date:
        visitors = visitors.filter(visit_date__lte=end_date)

    selected_fields = get_visitor_export_field_keys(request)
    field_labels = dict(VISITOR_EXPORT_FIELD_OPTIONS)
    columns = [field_labels[field] for field in selected_fields]

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="visitor_export_{stamp}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(columns)
    for visitor in visitors.order_by("-visit_date", "-entry_time", "-created_at"):
        writer.writerow([get_visitor_export_value(visitor, field, request) for field in selected_fields])

    return response


@institute_admin_required
def visitor_create(request):
    institute = get_current_institute(request)

    if request.method == "POST":
        form = VisitorForm(request.POST, request.FILES)
        if form.is_valid():
            visitor = form.save(commit=False)
            visitor.institute = institute
            visitor.created_by = request.user
            visitor.save()
            messages.success(request, "Visitor created successfully.")
            return close_popup_response()
    else:
        form = VisitorForm(
            initial={
                "visit_date": timezone.localdate().isoformat(),
                "entry_time": timezone.localtime().strftime("%H:%M"),
                "total_person": 1,
            },
        )

    return render(
        request,
        "institute_admin/visitor_form.html",
        {
            "form": form,
            "title": "Create Visitor",
            "button_text": "Save Entry",
        },
    )


@institute_admin_required
def visitor_update(request, pk):
    institute = get_current_institute(request)
    visitor = get_object_or_404(Visitor, pk=pk, institute=institute)

    if request.method == "POST":
        form = VisitorForm(
            request.POST,
            request.FILES,
            instance=visitor,
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Visitor updated successfully.")
            return close_popup_response()
    else:
        form = VisitorForm(instance=visitor)

    return render(
        request,
        "institute_admin/visitor_form.html",
        {
            "form": form,
            "title": "Edit Visitor",
            "button_text": "Update Entry",
            "visitor": visitor,
        },
    )


@institute_admin_required
@require_POST
def visitor_delete(request, pk):
    institute = get_current_institute(request)
    visitor = get_object_or_404(Visitor, pk=pk, institute=institute)
    visitor.delete()
    messages.success(request, "Visitor deleted successfully.")
    return redirect("institute_admin:visitor_list")


@institute_admin_required
@require_POST
def lead_convert(request, pk):
    institute = get_current_institute(request)
    lead = get_object_or_404(
        Lead.objects.select_related(
            "interested_class",
            "interested_batch",
            "converted_student",
        ),
        pk=pk,
        institute=institute,
    )

    if lead.status == Lead.Status.CONVERTED or lead.converted_student_id:
        messages.warning(request, "This lead has already been converted.")
        return redirect("institute_admin:lead_list")

    course = lead.interested_class
    batch = lead.interested_batch
    if not course or not batch:
        messages.error(request, "Select an interested class and batch before conversion.")
        return redirect("institute_admin:lead_list")
    if (
        course.institute_id != institute.pk
        or batch.institute_id != institute.pk
        or course.academic_year_id != batch.academic_year_id
        or not batch.courses.filter(pk=course.pk).exists()
    ):
        messages.error(request, "The selected class and batch are not valid for conversion.")
        return redirect("institute_admin:lead_list")

    academic_year = course.academic_year
    joined_on = timezone.localdate()
    try:
        with transaction.atomic():
            admission_number, username = generate_student_login_credentials(
                institute,
                academic_year,
            )
            user = User(
                username=username,
                first_name=lead.first_name,
                last_name=lead.last_name,
                email=lead.email,
                is_active=True,
            )
            user.set_password("Student@123")
            user.save()
            UserProfile.objects.create(
                user=user,
                institute=institute,
                role=UserProfile.Role.STUDENT_PARENT,
                phone=lead.mobile_number,
            )

            student = StudentProfile.objects.create(
                institute=institute,
                academic_year=academic_year,
                user=user,
                admission_number=admission_number,
                joined_on=joined_on,
                is_active=True,
            )
            academic_session = StudentAcademicSession.objects.create(
                institute=institute,
                student=student,
                academic_year=academic_year,
                admission_number=admission_number,
                joined_on=joined_on,
                status=StudentAcademicSession.Status.ACTIVE,
            )
            enrollment = StudentEnrollment.objects.create(
                student=student,
                academic_session=academic_session,
                batch=batch,
                enrolled_on=joined_on,
                status=StudentEnrollment.Status.ACTIVE,
            )
            enrollment.courses.add(course)

            lead.status = Lead.Status.CONVERTED
            lead.converted_student = student
            lead.converted_at = timezone.now()
            lead.save(
                update_fields=[
                    "status",
                    "converted_student",
                    "converted_at",
                    "updated_at",
                ]
            )
            on_commit_email(send_student_welcome, student.pk, "Student@123")
    except IntegrityError:
        messages.error(
            request,
            "The lead could not be converted because the student account already exists.",
        )
        return redirect("institute_admin:lead_list")

    messages.success(
        request,
        f"{lead.full_name} converted successfully. Username: {username}",
    )
    return redirect("institute_admin:lead_list")


def refresh_invoice_status(invoice):
    paid_amount = sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
    if paid_amount >= invoice.amount:
        invoice.status = FeeInvoice.Status.PAID
    elif paid_amount > 0:
        invoice.status = FeeInvoice.Status.PARTIAL
    else:
        invoice.status = FeeInvoice.Status.UNPAID
    invoice.save(update_fields=["status"])
    return invoice


def get_student_enrollment_due_data(student_or_session):
    if isinstance(student_or_session, StudentAcademicSession):
        student_session = student_or_session
        student = student_session.student
        enrollments = student_session.enrollments.exclude(status=StudentEnrollment.Status.CANCELLED).select_related(
            "batch"
        ).prefetch_related("courses")
    else:
        student_session = None
        student = student_or_session
        enrollments = student.enrollments.exclude(status=StudentEnrollment.Status.CANCELLED).select_related(
            "batch"
        ).prefetch_related("courses")
    data = {}
    for enrollment in enrollments:
        payment_filter = {
            "invoice__student": student,
            "invoice__enrollment": enrollment,
            "status": Payment.Status.ACTIVE,
        }
        if student_session:
            payment_filter["invoice__academic_session"] = student_session
        paid_amount = Payment.objects.filter(
            **payment_filter,
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        due_amount = enrollment.total_course_fee - paid_amount
        if due_amount < 0:
            due_amount = Decimal("0.00")
        first_due_invoice = None
        pending_filter = {
            "student": student,
            "enrollment": enrollment,
            "status__in": [FeeInvoice.Status.UNPAID, FeeInvoice.Status.PARTIAL],
        }
        if student_session:
            pending_filter["academic_session"] = student_session
        pending_invoices = (
            FeeInvoice.objects.filter(
                **pending_filter,
            )
            .prefetch_related("payments")
            .order_by("-created_at", "-pk")
        )
        for invoice in pending_invoices:
            invoice_paid = sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
            invoice_due = invoice.amount - invoice_paid
            if invoice_due <= 0 or due_amount <= 0:
                continue
            collectable_due = min(invoice_due, due_amount)
            first_due_invoice = {
                "id": str(invoice.pk),
                "title": invoice.title,
                "amount": str(invoice.amount),
                "paid": str(invoice_paid),
                "due": str(collectable_due),
            }
            break
        data[str(enrollment.pk)] = {
            "title": f"{enrollment.batch.name} Fee",
            "total_fee": str(enrollment.total_course_fee),
            "paid": str(paid_amount),
            "due": str(due_amount),
            "invoice": first_due_invoice,
        }
    return data


def get_student_category_due_data(student_or_session):
    if isinstance(student_or_session, StudentAcademicSession):
        student_session = student_or_session
        student = student_session.student
    else:
        student_session = None
        student = student_or_session
    data = {}
    total_due = Decimal("0.00")
    categories = FeeCategory.objects.filter(institute=student.institute, is_active=True)
    for category in categories:
        invoice_filter = {
            "student": student,
            "category": category,
            "status__in": [FeeInvoice.Status.UNPAID, FeeInvoice.Status.PARTIAL],
        }
        if student_session:
            invoice_filter["academic_session"] = student_session
        invoices = (
            FeeInvoice.objects.filter(
                **invoice_filter,
            )
            .select_related("category")
            .prefetch_related("payments")
        )
        category_total = Decimal("0.00")
        category_paid = Decimal("0.00")
        first_due_invoice = None
        for invoice in invoices:
            paid_amount = sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
            due_amount = invoice.amount - paid_amount
            if due_amount < 0:
                due_amount = Decimal("0.00")
            if due_amount <= 0:
                continue
            category_total += invoice.amount
            category_paid += paid_amount
            total_due += due_amount
            if first_due_invoice is None:
                first_due_invoice = {
                    "id": str(invoice.pk),
                    "title": invoice.title,
                    "amount": str(invoice.amount),
                    "paid": str(paid_amount),
                    "due": str(due_amount),
                }

        category_due = category_total - category_paid
        if category_due < 0:
            category_due = Decimal("0.00")
        data[str(category.pk)] = {
            "name": category.name,
            "default_amount": str(category.default_amount),
            "total": str(category_total),
            "paid": str(category_paid),
            "due": str(category_due),
            "invoice": first_due_invoice,
        }

    return {
        "categories": data,
        "total_due": str(total_due),
    }


@login_required
def academic_year_switch(request):
    profile = getattr(request.user, "profile", None)
    if not profile or profile.role not in {UserProfile.Role.INSTITUTE_ADMIN, UserProfile.Role.TEACHER} or not profile.institute_id:
        return redirect(reverse("school_dashboard"))
    institute = profile.institute
    if request.method == "POST":
        year_id = request.POST.get("academic_year_id", "").strip()
        academic_year = AcademicYear.objects.filter(pk=year_id, institute=institute).first() if year_id else None
        if academic_year:
            request.session["academic_year_id"] = academic_year.pk
            request._selected_academic_year = academic_year
            request._academic_year_context = None
            messages.success(request, f"Academic year changed to {academic_year.name}.")
    return redirect(request.META.get("HTTP_REFERER") or reverse("institute_admin:dashboard"))


def dashboard_financial_summary(institute, academic_year):
    session_table = connection.ops.quote_name(StudentAcademicSession._meta.db_table)
    enrollment_table = connection.ops.quote_name(StudentEnrollment._meta.db_table)
    enrollment_course_table = connection.ops.quote_name(
        StudentEnrollment.courses.through._meta.db_table
    )
    course_table = connection.ops.quote_name(Course._meta.db_table)
    batch_table = connection.ops.quote_name(Batch._meta.db_table)
    invoice_table = connection.ops.quote_name(FeeInvoice._meta.db_table)
    payment_table = connection.ops.quote_name(Payment._meta.db_table)
    sql = f"""
        WITH target_sessions AS (
            SELECT id
            FROM {session_table}
            WHERE institute_id = %s AND academic_year_id = %s
        ),
        enrollment_course_fees AS (
            SELECT sec.studentenrollment_id AS enrollment_id, SUM(c.fee_amount) AS total
            FROM {enrollment_course_table} sec
            INNER JOIN {course_table} c ON c.id = sec.course_id
            GROUP BY sec.studentenrollment_id
        ),
        enrollment_summary AS (
            SELECT
                se.academic_session_id AS session_id,
                SUM(
                    CASE
                        WHEN se.custom_fee_amount IS NOT NULL THEN se.custom_fee_amount
                        ELSE COALESCE(ecf.total, 0)
                    END
                ) AS enrollment_total,
                MIN(b.name) AS batch_name,
                MIN(se.enrolled_on) AS due_date
            FROM {enrollment_table} se
            INNER JOIN target_sessions ts ON ts.id = se.academic_session_id
            INNER JOIN {batch_table} b ON b.id = se.batch_id
            LEFT JOIN enrollment_course_fees ecf ON ecf.enrollment_id = se.id
            WHERE se.status <> %s
            GROUP BY se.academic_session_id
        ),
        invoice_summary AS (
            SELECT
                fi.academic_session_id AS session_id,
                SUM(fi.amount) AS invoiced_total,
                SUM(CASE WHEN fi.enrollment_id IS NULL THEN fi.amount ELSE 0 END) AS additional_total
            FROM {invoice_table} fi
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE fi.status <> %s
            GROUP BY fi.academic_session_id
        ),
        payment_summary AS (
            SELECT fi.academic_session_id AS session_id, SUM(p.amount) AS paid_total
            FROM {payment_table} p
            INNER JOIN {invoice_table} fi ON fi.id = p.invoice_id
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE p.status = %s AND fi.status <> %s
            GROUP BY fi.academic_session_id
        ),
        session_financial AS (
            SELECT
                ts.id AS session_id,
                CASE
                    WHEN COALESCE(es.enrollment_total, 0) > 0
                    THEN COALESCE(es.enrollment_total, 0) + COALESCE(ins.additional_total, 0)
                    ELSE COALESCE(ins.invoiced_total, 0)
                END AS total_fee,
                COALESCE(ps.paid_total, 0) AS paid_amount,
                es.batch_name,
                es.due_date
            FROM target_sessions ts
            LEFT JOIN enrollment_summary es ON es.session_id = ts.id
            LEFT JOIN invoice_summary ins ON ins.session_id = ts.id
            LEFT JOIN payment_summary ps ON ps.session_id = ts.id
        ),
        financial AS (
            SELECT
                session_id,
                total_fee,
                paid_amount,
                CASE WHEN total_fee > paid_amount THEN total_fee - paid_amount ELSE 0 END AS due_amount,
                batch_name,
                due_date
            FROM session_financial
        )
        SELECT
            0 AS row_type,
            NULL AS session_id,
            COALESCE(SUM(total_fee), 0) AS total_fee,
            COALESCE(SUM(paid_amount), 0) AS paid_amount,
            COALESCE(SUM(due_amount), 0) AS due_amount,
            NULL AS batch_name,
            NULL AS due_date
        FROM financial
        UNION ALL
        SELECT row_type, session_id, total_fee, paid_amount, due_amount, batch_name, due_date
        FROM (
            SELECT
                1 AS row_type,
                session_id,
                total_fee,
                paid_amount,
                due_amount,
                batch_name,
                due_date
            FROM financial
            WHERE due_amount > 0
            ORDER BY due_amount DESC, session_id
            LIMIT 6
        ) top_dues
        ORDER BY row_type, due_amount DESC
    """
    params = [
        institute.pk,
        academic_year.pk,
        StudentEnrollment.Status.CANCELLED,
        FeeInvoice.Status.CANCELLED,
        Payment.Status.ACTIVE,
        FeeInvoice.Status.CANCELLED,
    ]
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()

    total_row = rows[0]
    totals = {
        "invoice_amount": Decimal(str(total_row[2] or 0)),
        "paid_amount": Decimal(str(total_row[3] or 0)),
        "due_amount": Decimal(str(total_row[4] or 0)),
    }
    due_rows = [
        {
            "session_id": row[1],
            "due_amount": Decimal(str(row[4] or 0)),
            "batch_name": row[5],
            "due_date": date.fromisoformat(row[6]) if isinstance(row[6], str) else row[6],
        }
        for row in rows[1:]
    ]
    return totals, due_rows


def build_dashboard_summary(institute, academic_year, today):
    students = StudentAcademicSession.objects.filter(institute=institute, academic_year=academic_year)
    batches = Batch.objects.filter(institute=institute, academic_year=academic_year)
    courses = Course.objects.filter(institute=institute, academic_year=academic_year)
    invoices = FeeInvoice.objects.filter(
        institute=institute,
        academic_session__academic_year=academic_year,
    ).exclude(status=FeeInvoice.Status.CANCELLED)
    payments = Payment.objects.filter(
        status=Payment.Status.ACTIVE,
        invoice__institute=institute,
        invoice__academic_session__academic_year=academic_year,
    ).exclude(invoice__status=FeeInvoice.Status.CANCELLED)
    attendance = Attendance.objects.filter(
        date=today,
        batch__institute=institute,
        academic_session__academic_year=academic_year,
    )

    attendance_counts = attendance.aggregate(
        total=Count("pk"),
        present=Count("pk", filter=Q(status=Attendance.Status.PRESENT)),
        absent=Count("pk", filter=Q(status=Attendance.Status.ABSENT)),
        late=Count("pk", filter=Q(status=Attendance.Status.LATE)),
    )
    payment_totals = payments.aggregate(
        today_collection=Coalesce(
            Sum("amount", filter=Q(paid_on=today)),
            Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2)),
        ),
        month_collection=Coalesce(
            Sum("amount", filter=Q(paid_on__year=today.year, paid_on__month=today.month)),
            Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2)),
        ),
    )
    student_counts = students.aggregate(
        total=Count("pk"),
        active=Count(
            "pk",
            filter=Q(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True),
        ),
    )
    batch_counts = batches.aggregate(
        total=Count("pk"),
        active=Count("pk", filter=Q(is_active=True)),
    )
    staff_counts = UserProfile.objects.filter(institute=institute).aggregate(
        teachers=Count("pk", filter=Q(role=UserProfile.Role.TEACHER)),
        accountants=Count("pk", filter=Q(role=UserProfile.Role.ACCOUNTANT)),
    )

    financial_totals, top_due_rows = dashboard_financial_summary(institute, academic_year)
    due_sessions = StudentAcademicSession.objects.select_related(
        "student",
        "student__user",
    ).in_bulk(row["session_id"] for row in top_due_rows)
    due_invoice_rows = []
    for row in top_due_rows:
        student_session = due_sessions[row["session_id"]]
        user = student_session.student.user
        due_invoice_rows.append(
            {
                "student_name": user.get_full_name() or user.username,
                "title": row["batch_name"] or "Pending fees",
                "due_date": row["due_date"],
                "due_amount": row["due_amount"],
            }
        )

    invoice_amount = financial_totals["invoice_amount"]
    paid_amount = financial_totals["paid_amount"]
    attendance_total = attendance_counts["total"]
    return {
        "course_count": courses.count(),
        "batch_count": batch_counts["total"],
        "active_batch_count": batch_counts["active"],
        "student_count": student_counts["total"],
        "active_student_count": student_counts["active"],
        "teacher_count": staff_counts["teachers"],
        "accountant_count": staff_counts["accountants"],
        "invoice_count": invoices.count(),
        "invoice_amount": invoice_amount,
        "due_amount": financial_totals["due_amount"],
        "paid_amount": paid_amount,
        "collection_rate": round((paid_amount / invoice_amount) * 100, 1) if invoice_amount else 0,
        "today_attendance_count": attendance_total,
        "today_present_count": attendance_counts["present"],
        "today_absent_count": attendance_counts["absent"],
        "today_late_count": attendance_counts["late"],
        "attendance_rate": round((attendance_counts["present"] / attendance_total) * 100, 1)
        if attendance_total
        else 0,
        "today_collection": payment_totals["today_collection"],
        "month_collection": payment_totals["month_collection"],
        "due_invoice_rows": due_invoice_rows,
    }


@institute_admin_required
def dashboard(request):
    profile = getattr(request.user, "profile", None)
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    institute_filter = {"institute": institute} if institute else {}
    today = timezone.localdate()

    summary = get_dashboard_summary(institute.pk, academic_year.pk)
    if summary is None:
        summary = build_dashboard_summary(institute, academic_year, today)
        set_dashboard_summary(institute.pk, academic_year.pk, summary)

    students = StudentAcademicSession.objects.filter(institute=institute, academic_year=academic_year)
    recent_payments = Payment.objects.filter(
        status=Payment.Status.ACTIVE,
        invoice__institute=institute,
        invoice__academic_session__academic_year=academic_year,
    ).select_related(
        "invoice",
        "invoice__student",
        "invoice__student__user",
    )

    latest_students = students.select_related("student", "student__user").order_by("-id")[:5]

    context = {
        "profile": profile,
        "institute": institute,
        "today": today,
        "recent_payments": recent_payments.order_by("-created_at", "-pk")[:5],
        "latest_students": latest_students,
        "recent_notices": Notice.objects.filter(**institute_filter)[:5],
        "recent_homework": Homework.objects.filter(
            batch__institute=institute,
            batch__academic_year=academic_year,
        )[:5] if institute and academic_year else [],
        "recent_attendance": Attendance.objects.filter(
            batch__institute=institute,
            academic_session__academic_year=academic_year,
            batch__academic_year=academic_year,
        )[:5] if institute and academic_year else [],
    }
    context.update(summary)
    return render(request, "institute_admin/dashboard.html", context)


@institute_admin_required
def course_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    courses = Course.objects.select_related("institute").annotate(batch_count=Count("batches")).order_by("name", "pk")
    if institute:
        courses = courses.filter(institute=institute)
    if academic_year:
        courses = courses.filter(academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        courses = courses.filter(name__icontains=search_query)

    if status_filter == "active":
        courses = courses.filter(is_active=True)
    elif status_filter == "inactive":
        courses = courses.filter(is_active=False)

    base_queryset = Course.objects.filter(institute=institute) if institute else Course.objects.all()
    batch_queryset = Batch.objects.filter(institute=institute) if institute else Batch.objects.all()
    if academic_year:
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
    if academic_year:
        base_queryset = base_queryset.filter(academic_year=academic_year)
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
    page_obj, paginator, pagination_query = paginate_queryset(request, courses)
    context = {
        "courses": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "courses",
        "search_query": search_query,
        "status_filter": status_filter,
        "total_courses": base_queryset.count(),
        "active_courses": base_queryset.filter(is_active=True).count(),
        "inactive_courses": base_queryset.filter(is_active=False).count(),
        "total_batches": batch_queryset.count(),
    }
    return render(request, "institute_admin/course_list.html", context)


@institute_admin_required
def course_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating a course.")
        return redirect("institute_admin:course_list")

    if request.method == "POST":
        form = CourseForm(request.POST, institute=institute, academic_year=academic_year)
        if form.is_valid():
            course = form.save(commit=False)
            course.institute = institute
            course.academic_year = academic_year
            course.save()
            messages.success(request, "Course created successfully.")
            return close_popup_response()
    else:
        form = CourseForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/course_form.html",
        {
            "form": form,
            "title": "Create Course",
            "subtitle": "Add a course offered by your institute.",
            "button_text": "Save Course",
        },
    )


@institute_admin_required
def course_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Course.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    course = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = CourseForm(request.POST, instance=course, institute=course.institute, academic_year=course.academic_year)
        if form.is_valid():
            form.save()
            messages.success(request, "Course updated successfully.")
            return close_popup_response()
    else:
        form = CourseForm(instance=course, institute=course.institute, academic_year=course.academic_year)

    return render(
        request,
        "institute_admin/course_form.html",
        {
            "form": form,
            "title": "Edit Course",
            "subtitle": "Update course details and active status.",
            "button_text": "Update Course",
        },
    )


@institute_admin_required
def course_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Course.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    course = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if course.batches.exists():
            messages.error(request, "This course is used in batches. Remove it from those batches before deleting.")
        else:
            course.delete()
            messages.success(request, "Course deleted successfully.")

    return redirect("institute_admin:course_list")


@institute_admin_required
def subject_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    subjects = Subject.objects.select_related("institute", "academic_year")
    if institute:
        subjects = subjects.filter(institute=institute)
    if academic_year:
        subjects = subjects.filter(academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        subjects = subjects.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))

    if status_filter == "active":
        subjects = subjects.filter(is_active=True)
    elif status_filter == "inactive":
        subjects = subjects.filter(is_active=False)

    base_queryset = Subject.objects.filter(institute=institute) if institute else Subject.objects.all()
    if academic_year:
        base_queryset = base_queryset.filter(academic_year=academic_year)

    context = {
        "subjects": subjects,
        "search_query": search_query,
        "status_filter": status_filter,
        "total_subjects": base_queryset.count(),
        "active_subjects": base_queryset.filter(is_active=True).count(),
        "inactive_subjects": base_queryset.filter(is_active=False).count(),
    }
    return render(request, "institute_admin/subject_list.html", context)


@institute_admin_required
def subject_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating a subject.")
        return redirect("institute_admin:subject_list")

    if request.method == "POST":
        form = SubjectForm(request.POST, institute=institute, academic_year=academic_year)
        if form.is_valid():
            subject = form.save(commit=False)
            subject.institute = institute
            subject.academic_year = academic_year
            subject.save()
            messages.success(request, "Subject created successfully.")
            return close_popup_response()
    else:
        form = SubjectForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/subject_form.html",
        {
            "form": form,
            "title": "Create Subject",
            "subtitle": "Add a subject for exams and academic work.",
            "button_text": "Save Subject",
        },
    )


@institute_admin_required
def subject_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Subject.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    subject = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject, institute=subject.institute, academic_year=subject.academic_year)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject updated successfully.")
            return close_popup_response()
    else:
        form = SubjectForm(instance=subject, institute=subject.institute, academic_year=subject.academic_year)

    return render(
        request,
        "institute_admin/subject_form.html",
        {
            "form": form,
            "title": "Edit Subject",
            "subtitle": "Update subject details and active status.",
            "button_text": "Update Subject",
        },
    )


@institute_admin_required
def subject_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Subject.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    subject = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if subject.exams.exists():
            messages.error(request, "This subject is used in exams. Remove it from exams before deleting.")
        else:
            subject.delete()
            messages.success(request, "Subject deleted successfully.")

    return redirect("institute_admin:subject_list")


@institute_admin_required
def fee_category_list(request):
    institute = get_current_institute(request)
    categories = FeeCategory.objects.select_related("institute").annotate(invoice_count=Count("invoices"))
    if institute:
        categories = categories.filter(institute=institute)

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        categories = categories.filter(name__icontains=search_query)

    if status_filter == "active":
        categories = categories.filter(is_active=True)
    elif status_filter == "inactive":
        categories = categories.filter(is_active=False)

    base_queryset = FeeCategory.objects.filter(institute=institute) if institute else FeeCategory.objects.all()
    context = {
        "categories": categories,
        "search_query": search_query,
        "status_filter": status_filter,
        "total_categories": base_queryset.count(),
        "active_categories": base_queryset.filter(is_active=True).count(),
        "inactive_categories": base_queryset.filter(is_active=False).count(),
        "used_categories": base_queryset.filter(invoices__isnull=False).distinct().count(),
    }
    return render(request, "institute_admin/fee_category_list.html", context)


@institute_admin_required
def fee_category_create(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before creating a fee category.")
        return redirect("institute_admin:fee_category_list")

    if request.method == "POST":
        form = FeeCategoryForm(request.POST, institute=institute)
        if form.is_valid():
            category = form.save(commit=False)
            category.institute = institute
            category.save()
            messages.success(request, "Fee category created successfully.")
            return close_popup_response()
    else:
        form = FeeCategoryForm(institute=institute)

    return render(
        request,
        "institute_admin/fee_category_form.html",
        {
            "form": form,
            "title": "Create Fee Category",
            "subtitle": "Add a dynamic fee head for your institute services.",
            "button_text": "Save Category",
        },
    )


@institute_admin_required
def fee_category_update(request, pk):
    institute = get_current_institute(request)
    queryset = FeeCategory.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    category = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = FeeCategoryForm(request.POST, institute=category.institute, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Fee category updated successfully.")
            return close_popup_response()
    else:
        form = FeeCategoryForm(institute=category.institute, instance=category)

    return render(
        request,
        "institute_admin/fee_category_form.html",
        {
            "form": form,
            "title": "Edit Fee Category",
            "subtitle": "Update category name, default amount and status.",
            "button_text": "Update Category",
        },
    )


@institute_admin_required
def fee_category_delete(request, pk):
    institute = get_current_institute(request)
    queryset = FeeCategory.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    category = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if category.invoices.exists():
            messages.error(request, "This category is used in invoices. Mark it inactive instead of deleting.")
        else:
            category.delete()
            messages.success(request, "Fee category deleted successfully.")

    return redirect("institute_admin:fee_category_list")


@institute_admin_required
def batch_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    batches = Batch.objects.select_related("institute").prefetch_related("courses", "teachers").order_by("name", "pk")
    if institute:
        batches = batches.filter(institute=institute)
    if academic_year:
        batches = batches.filter(academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    course_filter = request.GET.get("course", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        batches = batches.filter(
            Q(name__icontains=search_query)
            | Q(courses__name__icontains=search_query)
            | Q(timing__icontains=search_query)
        ).distinct()

    if course_filter:
        batches = batches.filter(courses__id=course_filter)

    if status_filter == "active":
        batches = batches.filter(is_active=True)
    elif status_filter == "inactive":
        batches = batches.filter(is_active=False)

    base_queryset = Batch.objects.filter(institute=institute) if institute else Batch.objects.all()
    course_queryset = Course.objects.filter(institute=institute) if institute else Course.objects.all()
    if academic_year:
        base_queryset = base_queryset.filter(academic_year=academic_year)
        course_queryset = course_queryset.filter(academic_year=academic_year)

    page_obj, paginator, pagination_query = paginate_queryset(request, batches)
    context = {
        "batches": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "batches",
        "courses": course_queryset,
        "search_query": search_query,
        "course_filter": course_filter,
        "status_filter": status_filter,
        "total_batches": base_queryset.count(),
        "active_batches": base_queryset.filter(is_active=True).count(),
        "inactive_batches": base_queryset.filter(is_active=False).count(),
        "total_courses": course_queryset.count(),
    }
    return render(request, "institute_admin/batch_list.html", context)


@institute_admin_required
def batch_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating a batch.")
        return redirect("institute_admin:batch_list")

    if request.method == "POST":
        form = BatchForm(request.POST, institute=institute, academic_year=academic_year)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.institute = institute
            batch.academic_year = academic_year
            batch.save()
            batch.courses.set(form.cleaned_data["courses"])
            teacher_profiles = form.cleaned_data["teachers"]
            batch.teachers.set([profile.user for profile in teacher_profiles])
            messages.success(request, "Batch created successfully.")
            return close_popup_response()
    else:
        form = BatchForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/batch_form.html",
        {
            "form": form,
            "title": "Create Batch",
            "subtitle": "Add a batch, select one or more courses, and assign teachers.",
            "button_text": "Save Batch",
            "show_timetable_builder": True,
        },
    )


@institute_admin_required
def batch_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Batch.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    batch = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = BatchForm(request.POST, instance=batch, institute=batch.institute, academic_year=batch.academic_year)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.save()
            batch.courses.set(form.cleaned_data["courses"])
            teacher_profiles = form.cleaned_data["teachers"]
            batch.teachers.set([profile.user for profile in teacher_profiles])
            messages.success(request, "Batch updated successfully.")
            return close_popup_response()
    else:
        initial_profiles = UserProfile.objects.filter(user__in=batch.teachers.all())
        form = BatchForm(instance=batch, institute=batch.institute, academic_year=batch.academic_year, initial={"teachers": initial_profiles})

    return render(
        request,
        "institute_admin/batch_form.html",
        {
            "form": form,
            "title": "Edit Batch",
            "subtitle": "Update batch details, timing, timetable and assigned teachers.",
            "button_text": "Update Batch",
            "show_timetable_builder": True,
        },
    )


@institute_admin_required
def batch_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Batch.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    batch = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        try:
            batch.delete()
            messages.success(request, "Batch deleted successfully.")
        except ProtectedError:
            messages.error(request, "This batch has related records and cannot be deleted.")

    return redirect("institute_admin:batch_list")


@institute_admin_required
def user_list(request):
    institute = get_current_institute(request)
    profiles = UserProfile.objects.select_related("user", "institute").exclude(
        role=UserProfile.Role.SUPER_ADMIN
    ).order_by("user__first_name", "user__username", "pk")
    if institute:
        profiles = profiles.filter(institute=institute)

    search_query = request.GET.get("search", "").strip()
    role_filter = request.GET.get("role", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        profiles = profiles.filter(
            Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(phone__icontains=search_query)
        )

    if role_filter:
        profiles = profiles.filter(role=role_filter)

    if status_filter == "active":
        profiles = profiles.filter(user__is_active=True)
    elif status_filter == "inactive":
        profiles = profiles.filter(user__is_active=False)

    base_queryset = UserProfile.objects.exclude(role=UserProfile.Role.SUPER_ADMIN)
    if institute:
        base_queryset = base_queryset.filter(institute=institute)

    page_obj, paginator, pagination_query = paginate_queryset(request, profiles)
    context = {
        "profiles": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "users",
        "role_choices": InstituteUserForm.ROLE_CHOICES,
        "search_query": search_query,
        "role_filter": role_filter,
        "status_filter": status_filter,
        "total_users": base_queryset.count(),
        "active_users": base_queryset.filter(user__is_active=True).count(),
        "teacher_users": base_queryset.filter(role=UserProfile.Role.TEACHER).count(),
        "accountant_users": base_queryset.filter(role=UserProfile.Role.ACCOUNTANT).count(),
        "student_parent_users": base_queryset.filter(role=UserProfile.Role.STUDENT_PARENT).count(),
    }
    return render(request, "institute_admin/user_list.html", context)


@institute_admin_required
def user_create(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before creating a user.")
        return redirect("institute_admin:user_list")

    if request.method == "POST":
        form = InstituteUserForm(request.POST, institute=institute)
        if form.is_valid():
            profile = form.save()
            if profile.role == UserProfile.Role.TEACHER and hasattr(profile.user, "teacher_profile"):
                on_commit_email(
                    send_teacher_welcome,
                    profile.user.teacher_profile.pk,
                    form.cleaned_data["password"],
                )
            elif profile.role == UserProfile.Role.INSTITUTE_ADMIN:
                on_commit_email(
                    send_institute_welcome,
                    profile.user.pk,
                    form.cleaned_data["password"],
                )
            messages.success(request, "User account created successfully.")
            return close_popup_response()
    else:
        form = InstituteUserForm(institute=institute)

    return render(
        request,
        "institute_admin/user_form.html",
        {
            "form": form,
            "title": "Create User",
            "subtitle": "Create login access and assign a role.",
            "button_text": "Save User",
        },
    )


@institute_admin_required
def user_update(request, pk):
    institute = get_current_institute(request)
    queryset = UserProfile.objects.select_related("user").exclude(role=UserProfile.Role.SUPER_ADMIN)
    if institute:
        queryset = queryset.filter(institute=institute)
    profile = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = InstituteUserForm(request.POST, institute=profile.institute, profile=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "User account updated successfully.")
            return close_popup_response()
    else:
        form = InstituteUserForm(institute=profile.institute, profile=profile)

    return render(
        request,
        "institute_admin/user_form.html",
        {
            "form": form,
            "title": "Edit User",
            "subtitle": "Update login access, role and account status.",
            "button_text": "Update User",
        },
    )


@institute_admin_required
def user_delete(request, pk):
    institute = get_current_institute(request)
    queryset = UserProfile.objects.select_related("user").exclude(role=UserProfile.Role.SUPER_ADMIN)
    if institute:
        queryset = queryset.filter(institute=institute)
    profile = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if profile.user_id == request.user.id:
            messages.error(request, "You cannot delete your own account.")
        elif hasattr(profile.user, "student_profile"):
            messages.error(request, "This user is linked with a student profile. Manage it from Student section.")
        elif hasattr(profile.user, "teacher_profile") and profile.user.assigned_batches.exists():
            messages.error(request, "This teacher is assigned to batches. Remove them from batches before deleting.")
        else:
            profile.user.delete()
            messages.success(request, "User account deleted successfully.")

    return redirect("institute_admin:user_list")


def get_student_session_fee_summaries(session_ids, display_session_ids=None):
    session_ids = list(session_ids)
    if display_session_ids is None:
        display_session_ids = session_ids
    display_session_ids = set(display_session_ids)
    money_unit = Decimal("0.01")
    empty_summary = {
        "total_fee_amount": Decimal("0.00"),
        "paid_amount": Decimal("0.00"),
        "due_amount": Decimal("0.00"),
        "display_enrollments": [],
        "display_enrollment_count": 0,
    }
    if not session_ids:
        return {}

    enrollment_fee_by_session = defaultdict(lambda: Decimal("0.00"))
    invoiced_amount_by_session = defaultdict(lambda: Decimal("0.00"))
    additional_fee_by_session = defaultdict(lambda: Decimal("0.00"))
    paid_amount_by_session = defaultdict(lambda: Decimal("0.00"))
    display_enrollments_by_session = defaultdict(list)
    display_enrollment_count_by_session = defaultdict(int)

    active_enrollments = StudentEnrollment.objects.filter(
        academic_session_id__in=session_ids,
    ).exclude(status=StudentEnrollment.Status.CANCELLED)

    custom_fee_rows = (
        active_enrollments.filter(custom_fee_amount__isnull=False)
        .values("academic_session_id")
        .annotate(total=Sum("custom_fee_amount"))
    )
    for row in custom_fee_rows:
        enrollment_fee_by_session[row["academic_session_id"]] += row["total"] or Decimal("0.00")

    course_fee_rows = (
        Course.objects.filter(
            student_enrollments__academic_session_id__in=session_ids,
            student_enrollments__custom_fee_amount__isnull=True,
        )
        .exclude(student_enrollments__status=StudentEnrollment.Status.CANCELLED)
        .values("student_enrollments__academic_session_id")
        .annotate(total=Sum("fee_amount"))
    )
    for row in course_fee_rows:
        enrollment_fee_by_session[row["student_enrollments__academic_session_id"]] += row["total"] or Decimal("0.00")

    enrollment_rows = active_enrollments.filter(academic_session_id__in=display_session_ids).select_related("batch").order_by(
        "academic_session__admission_number",
        "batch__name",
    )
    for enrollment in enrollment_rows:
        session_id = enrollment.academic_session_id
        display_enrollment_count_by_session[session_id] += 1
        if len(display_enrollments_by_session[session_id]) < 2:
            display_enrollments_by_session[session_id].append(enrollment)

    invoice_queryset = FeeInvoice.objects.filter(
        academic_session_id__in=session_ids,
    ).exclude(status=FeeInvoice.Status.CANCELLED)
    invoice_amount_rows = invoice_queryset.values("academic_session_id").annotate(total=Sum("amount"))
    for row in invoice_amount_rows:
        invoiced_amount_by_session[row["academic_session_id"]] = row["total"] or Decimal("0.00")

    additional_fee_rows = (
        invoice_queryset.filter(enrollment__isnull=True)
        .values("academic_session_id")
        .annotate(total=Sum("amount"))
    )
    for row in additional_fee_rows:
        additional_fee_by_session[row["academic_session_id"]] = row["total"] or Decimal("0.00")

    payment_rows = (
        Payment.objects.filter(
            invoice__academic_session_id__in=session_ids,
            status=Payment.Status.ACTIVE,
        )
        .exclude(invoice__status=FeeInvoice.Status.CANCELLED)
        .values("invoice__academic_session_id")
        .annotate(total=Sum("amount"))
    )
    for row in payment_rows:
        paid_amount_by_session[row["invoice__academic_session_id"]] = row["total"] or Decimal("0.00")

    summaries = {}
    for session_id in session_ids:
        enrollment_fee_amount = enrollment_fee_by_session[session_id]
        invoiced_amount = invoiced_amount_by_session[session_id]
        additional_fee_amount = additional_fee_by_session[session_id]
        paid_amount = paid_amount_by_session[session_id]
        total_fee_amount = enrollment_fee_amount + additional_fee_amount if enrollment_fee_amount > 0 else invoiced_amount
        due_amount = total_fee_amount - paid_amount
        if due_amount < 0:
            due_amount = Decimal("0.00")
        summaries[session_id] = {
            **empty_summary,
            "total_fee_amount": total_fee_amount.quantize(money_unit),
            "paid_amount": paid_amount.quantize(money_unit),
            "due_amount": due_amount.quantize(money_unit),
            "display_enrollments": display_enrollments_by_session[session_id],
            "display_enrollment_count": display_enrollment_count_by_session[session_id],
        }
    return summaries


def get_student_list_financial_totals(sessions):
    target_sql, target_params = sessions.order_by().values("pk").query.sql_with_params()
    enrollment_table = connection.ops.quote_name(StudentEnrollment._meta.db_table)
    enrollment_course_table = connection.ops.quote_name(
        StudentEnrollment.courses.through._meta.db_table
    )
    course_table = connection.ops.quote_name(Course._meta.db_table)
    invoice_table = connection.ops.quote_name(FeeInvoice._meta.db_table)
    payment_table = connection.ops.quote_name(Payment._meta.db_table)
    sql = f"""
        WITH target_session_ids AS (
            {target_sql}
        ),
        target_sessions AS (
            SELECT pk AS id FROM target_session_ids
        ),
        enrollment_course_fees AS (
            SELECT sec.studentenrollment_id AS enrollment_id, SUM(c.fee_amount) AS total
            FROM {enrollment_course_table} sec
            INNER JOIN {course_table} c ON c.id = sec.course_id
            GROUP BY sec.studentenrollment_id
        ),
        enrollment_summary AS (
            SELECT
                se.academic_session_id AS session_id,
                SUM(
                    CASE
                        WHEN se.custom_fee_amount IS NOT NULL THEN se.custom_fee_amount
                        ELSE COALESCE(ecf.total, 0)
                    END
                ) AS enrollment_total
            FROM {enrollment_table} se
            INNER JOIN target_sessions ts ON ts.id = se.academic_session_id
            LEFT JOIN enrollment_course_fees ecf ON ecf.enrollment_id = se.id
            WHERE se.status <> %s
            GROUP BY se.academic_session_id
        ),
        invoice_summary AS (
            SELECT
                fi.academic_session_id AS session_id,
                SUM(fi.amount) AS invoiced_total,
                SUM(CASE WHEN fi.enrollment_id IS NULL THEN fi.amount ELSE 0 END) AS additional_total
            FROM {invoice_table} fi
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE fi.status <> %s
            GROUP BY fi.academic_session_id
        ),
        payment_summary AS (
            SELECT fi.academic_session_id AS session_id, SUM(p.amount) AS paid_total
            FROM {payment_table} p
            INNER JOIN {invoice_table} fi ON fi.id = p.invoice_id
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE p.status = %s AND fi.status <> %s
            GROUP BY fi.academic_session_id
        ),
        financial AS (
            SELECT
                CASE
                    WHEN COALESCE(es.enrollment_total, 0) > 0
                    THEN COALESCE(es.enrollment_total, 0) + COALESCE(ins.additional_total, 0)
                    ELSE COALESCE(ins.invoiced_total, 0)
                END AS total_fee,
                COALESCE(ps.paid_total, 0) AS paid_amount
            FROM target_sessions ts
            LEFT JOIN enrollment_summary es ON es.session_id = ts.id
            LEFT JOIN invoice_summary ins ON ins.session_id = ts.id
            LEFT JOIN payment_summary ps ON ps.session_id = ts.id
        )
        SELECT
            COALESCE(SUM(total_fee), 0),
            COALESCE(SUM(paid_amount), 0),
            COALESCE(SUM(CASE WHEN total_fee > paid_amount THEN total_fee - paid_amount ELSE 0 END), 0)
        FROM financial
    """
    params = [
        *target_params,
        StudentEnrollment.Status.CANCELLED,
        FeeInvoice.Status.CANCELLED,
        Payment.Status.ACTIVE,
        FeeInvoice.Status.CANCELLED,
    ]
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        total_fee, paid_amount, due_amount = cursor.fetchone()
    return {
        "filtered_total_fee_amount": Decimal(str(total_fee or 0)),
        "filtered_paid_amount": Decimal(str(paid_amount or 0)),
        "filtered_due_amount": Decimal(str(due_amount or 0)),
    }


@institute_admin_required
def student_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    sessions = StudentAcademicSession.objects.select_related(
        "institute",
        "academic_year",
        "student",
        "student__user",
        "student__user__profile",
    )
    if institute:
        sessions = sessions.filter(institute=institute)
    if academic_year:
        sessions = sessions.filter(academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    batch_queryset = Batch.objects.filter(institute=institute, is_active=True) if institute else Batch.objects.none()
    if academic_year:
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
    selected_batch = batch_queryset.filter(pk=batch_filter).first() if batch_filter else None

    if batch_filter:
        if selected_batch:
            sessions = sessions.filter(enrollments__batch=selected_batch).distinct()
        else:
            sessions = sessions.none()

    if search_query:
        sessions = sessions.filter(
            Q(admission_number__icontains=search_query)
            | Q(student__user__first_name__icontains=search_query)
            | Q(student__user__last_name__icontains=search_query)
            | Q(student__user__username__icontains=search_query)
            | Q(student__user__email__icontains=search_query)
            | Q(student__user__profile__phone__icontains=search_query)
            | Q(student__guardians__name__icontains=search_query)
            | Q(student__guardians__phone__icontains=search_query)
        ).distinct()

    if status_filter == "active":
        sessions = sessions.filter(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True)
    elif status_filter == "inactive":
        sessions = sessions.exclude(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True)

    paginator = Paginator(sessions, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    sessions_page = list(
        page_obj.object_list.prefetch_related(
            "student__guardians",
        )
    )
    page_session_ids = [session.pk for session in sessions_page]
    fee_summaries = get_student_session_fee_summaries(page_session_ids)

    filtered_session_ids = sessions.order_by().values("pk")
    student_counts = StudentAcademicSession.objects.filter(
        pk__in=Subquery(filtered_session_ids),
    ).aggregate(
        total=Count("pk"),
        active=Count(
            "pk",
            filter=Q(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True),
        ),
    )
    total_enrollments = StudentEnrollment.objects.filter(
        academic_session_id__in=Subquery(filtered_session_ids),
    ).count()
    financial_totals = get_student_list_financial_totals(sessions)
    query_params = request.GET.copy()
    query_params.pop("page", None)
    pagination_query = query_params.urlencode()

    for session in sessions_page:
        summary = fee_summaries.get(session.pk, {})
        session.total_fee_amount = summary.get("total_fee_amount", Decimal("0.00"))
        session.paid_amount = summary.get("paid_amount", Decimal("0.00"))
        session.due_amount = summary.get("due_amount", Decimal("0.00"))
        session.display_enrollments = summary.get("display_enrollments", [])
        session.display_enrollment_count = summary.get("display_enrollment_count", 0)

    context = {
        "students": sessions_page,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "students",
        "search_query": search_query,
        "status_filter": status_filter,
        "batch_filter": batch_filter,
        "batches": batch_queryset,
        "selected_batch": selected_batch,
        "student_export_field_options": STUDENT_EXPORT_FIELD_OPTIONS,
        "student_export_default_fields": STUDENT_EXPORT_DEFAULT_FIELDS,
        "total_students": student_counts["total"],
        "active_students": student_counts["active"],
        "inactive_students": student_counts["total"] - student_counts["active"],
        "total_enrollments": total_enrollments,
        **financial_totals,
    }
    return render(request, "institute_admin/student_list.html", context)


@institute_admin_required
def student_promote(request):
    institute = get_current_institute(request)
    current_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before promoting students.")
        return redirect("institute_admin:student_list")

    academic_years = AcademicYear.objects.filter(institute=institute, is_active=True).order_by("start_date")
    source_year_id = request.POST.get("source_year") or request.GET.get("source_year")
    source_year = academic_years.filter(pk=source_year_id).first() if source_year_id else current_year
    source_year = source_year or academic_years.first()

    target_year_id = request.POST.get("target_year") or request.GET.get("target_year")
    target_year = academic_years.filter(pk=target_year_id).first() if target_year_id else None
    if not target_year and source_year:
        target_year = academic_years.filter(start_date__gt=source_year.start_date).first()
    if not target_year and source_year:
        target_year = academic_years.exclude(pk=source_year.pk).first()

    all_courses = Course.objects.filter(institute=institute, is_active=True).select_related("academic_year")
    all_batches = Batch.objects.filter(institute=institute, is_active=True).prefetch_related("courses")
    source_courses = all_courses.filter(academic_year=source_year) if source_year else Course.objects.none()
    target_courses = all_courses.filter(academic_year=target_year) if target_year else Course.objects.none()

    source_course_id = request.POST.get("source_course") or request.GET.get("source_course")
    target_course_id = request.POST.get("target_course") or request.GET.get("target_course")
    source_course = source_courses.filter(pk=source_course_id).first() if source_course_id else None
    target_course = target_courses.filter(pk=target_course_id).first() if target_course_id else None

    source_batches = Batch.objects.none()
    if source_year:
        source_batches = all_batches.filter(academic_year=source_year)
        if source_course:
            source_batches = source_batches.filter(courses=source_course)
    target_batches = Batch.objects.none()
    if target_year:
        target_batches = all_batches.filter(academic_year=target_year)
        if target_course:
            target_batches = target_batches.filter(courses=target_course)

    source_batch_id = request.POST.get("source_batch") or request.GET.get("source_batch")
    target_batch_id = request.POST.get("target_batch") or request.GET.get("target_batch")
    source_batch = source_batches.filter(pk=source_batch_id).first() if source_batch_id else None
    target_batch = target_batches.filter(pk=target_batch_id).first() if target_batch_id else None
    promotion_loaded = request.method == "POST" or request.GET.get("load_students") == "1"

    source_sessions = StudentAcademicSession.objects.none()
    if source_year and source_course and source_year != target_year:
        enrollment_filters = {
            "enrollments__courses": source_course,
            "enrollments__status": StudentEnrollment.Status.ACTIVE,
        }
        if source_batch:
            enrollment_filters["enrollments__batch"] = source_batch
        source_sessions = (
            StudentAcademicSession.objects.filter(
                institute=institute,
                academic_year=source_year,
                status=StudentAcademicSession.Status.ACTIVE,
                student__is_active=True,
                **enrollment_filters,
            )
            .select_related("student", "student__user", "student__user__profile")
            .prefetch_related("student__guardians", "enrollments__batch", "enrollments__courses")
            .distinct()
            .order_by("admission_number")
        )

    fully_promoted_student_ids = set()
    if target_year and target_course and target_batch:
        fully_promoted_student_ids = set(
            StudentAcademicSession.objects.filter(
                institute=institute,
                academic_year=target_year,
                enrollments__batch=target_batch,
                enrollments__courses=target_course,
                enrollments__status=StudentEnrollment.Status.ACTIVE,
            ).values_list("student_id", flat=True)
        )

    if request.method == "POST":
        selected_ids = [student_id for student_id in request.POST.getlist("students") if student_id.isdigit()]
        redirect_params = (
            f"?source_year={source_year.pk if source_year else ''}"
            f"&target_year={target_year.pk if target_year else ''}"
            f"&source_course={source_course.pk if source_course else ''}"
            f"&target_course={target_course.pk if target_course else ''}"
            f"&source_batch={source_batch.pk if source_batch else ''}"
            f"&target_batch={target_batch.pk if target_batch else ''}"
            "&load_students=1"
        )
        error_message = None
        if not source_year or not target_year:
            error_message = "Select source and target academic sessions."
        elif source_year == target_year:
            error_message = "Target academic session must be different from source academic session."
        elif not source_course:
            error_message = "Select a valid source course from the source academic session."
        elif not target_course:
            error_message = "Select a valid target course from the target academic session."
        elif not source_batch:
            error_message = "Select a valid source batch for the chosen source course."
        elif not target_batch:
            error_message = "Select a valid target batch for the chosen target course."
        elif not selected_ids:
            error_message = "Select at least one student to promote."
        if error_message:
            messages.error(request, error_message)
            return redirect(reverse("institute_admin:student_promote") + redirect_params)

        selected_sessions = source_sessions.filter(student_id__in=selected_ids)
        created_count = 0
        enrollment_count = 0
        skipped_count = 0

        with transaction.atomic():
            for source_session in selected_sessions:
                promoted_session = StudentAcademicSession.objects.filter(
                    student=source_session.student,
                    academic_year=target_year,
                ).first()
                if not promoted_session:
                    promoted_session = StudentAcademicSession.objects.create(
                        institute=institute,
                        student=source_session.student,
                        academic_year=target_year,
                        admission_number=generate_student_admission_number(institute, target_year),
                        joined_on=timezone.localdate(),
                        status=StudentAcademicSession.Status.ACTIVE,
                    )
                    created_count += 1

                enrollment, enrollment_created = StudentEnrollment.objects.get_or_create(
                    student=source_session.student,
                    academic_session=promoted_session,
                    batch=target_batch,
                    defaults={
                        "enrolled_on": timezone.localdate(),
                        "status": StudentEnrollment.Status.ACTIVE,
                    },
                )
                course_already_assigned = enrollment.courses.filter(pk=target_course.pk).exists()
                if enrollment.status != StudentEnrollment.Status.ACTIVE:
                    enrollment.status = StudentEnrollment.Status.ACTIVE
                    enrollment.save(update_fields=["status"])
                enrollment.courses.add(target_course)
                if enrollment_created or not course_already_assigned:
                    enrollment_count += 1
                else:
                    skipped_count += 1

        request.session["academic_year_id"] = target_year.pk
        if created_count or enrollment_count:
            messages.success(
                request,
                f"{created_count} target admission(s) created and {enrollment_count} student(s) allocated to "
                f"{target_batch.name} / {target_course.name} in {target_year.name}.",
            )
        if skipped_count:
            messages.warning(request, f"{skipped_count} student(s) were already allocated to the selected target batch and course.")
        return redirect("institute_admin:student_list")

    choice_data = {}
    for year in academic_years:
        year_courses = all_courses.filter(academic_year=year)
        year_batches = all_batches.filter(academic_year=year)
        choice_data[str(year.pk)] = {
            "courses": [{"id": course.pk, "name": course.name} for course in year_courses],
            "batches": [
                {
                    "id": batch.pk,
                    "name": batch.name,
                    "course_ids": [course.pk for course in batch.courses.all()],
                }
                for batch in year_batches
            ],
        }

    promotion_ready = bool(
        promotion_loaded
        and source_year
        and target_year
        and source_course
        and target_course
        and source_batch
        and target_batch
        and source_year != target_year
    )
    context = {
        "academic_years": academic_years,
        "source_year": source_year,
        "target_year": target_year,
        "source_courses": source_courses,
        "target_courses": target_courses,
        "source_course": source_course,
        "target_course": target_course,
        "source_batches": source_batches,
        "target_batches": target_batches,
        "source_batch": source_batch,
        "target_batch": target_batch,
        "promotion_loaded": promotion_loaded,
        "promotion_ready": promotion_ready,
        "already_promoted_student_ids": fully_promoted_student_ids,
        "promotion_students": source_sessions,
        "available_count": source_sessions.exclude(student_id__in=fully_promoted_student_ids).count(),
        "choice_data": choice_data,
    }
    return render(request, "institute_admin/student_promote.html", context)


def style_student_workbook_header(sheet, title, max_col):
    primary = "0F766E"
    dark = "0F172A"
    border_color = "CBD5E1"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(size=18, bold=True, color="FFFFFF")
    sheet.cell(row=1, column=1).fill = PatternFill("solid", fgColor=primary)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=3, column=col)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    return border


def student_import_columns():
    return [
        "First Name *",
        "Last Name",
        "Password",
        "Email",
        "Phone",
        "Date of Birth",
        "Joined On",
        "Address",
        "Current School / College",
        "Current School Address",
        "Previous School / College",
        "Previous Class",
        "Guardian Name",
        "Guardian Relation",
        "Guardian Phone",
        "Guardian Email",
        "Active",
    ]


STUDENT_EXPORT_FIELD_OPTIONS = [
    ("student_id", "Student ID"),
    ("admission_number", "Admission Number"),
    ("name", "Name"),
    ("institute_name", "Institute"),
    ("academic_year", "Academic Year"),
    ("mobile", "Mobile"),
    ("batch", "Batch"),
    ("total_fees", "Total Fees"),
    ("paid_amount", "Paid Amount"),
    ("due_amount", "Due Amount"),
    ("status", "Status"),
    ("session_status", "Session Status"),
    ("student_status", "Student Status"),
    ("first_name", "First Name"),
    ("middle_name", "Middle Name"),
    ("last_name", "Last Name"),
    ("username", "Username"),
    ("email", "Email"),
    ("profile_image", "Profile Photo"),
    ("pen_no", "PEN No"),
    ("appar_id", "Appar ID"),
    ("gr_number", "GR Number"),
    ("udise_number", "UDISE Number"),
    ("roll_number", "Roll Number"),
    ("gender", "Gender"),
    ("date_of_birth", "Date of Birth"),
    ("blood_group", "Blood Group"),
    ("religion", "Religion"),
    ("cast", "Cast"),
    ("caste_category", "Caste Category"),
    ("nationality", "Nationality"),
    ("aadhaar_number", "Aadhaar Number"),
    ("birth_certificate_number", "Birth Certificate Number"),
    ("place_of_birth", "Place of Birth"),
    ("mother_tongue", "Mother Tongue"),
    ("joined_on", "Joined On"),
    ("admission_class", "Admission Class"),
    ("current_class", "Current Class"),
    ("division", "Division"),
    ("medium", "Medium"),
    ("guardian_name", "Guardian Name"),
    ("guardian_relation", "Guardian Relation"),
    ("guardian_phone", "Guardian Phone"),
    ("guardian_email", "Guardian Email"),
    ("father_name", "Father Name"),
    ("father_occupation", "Father Occupation"),
    ("father_qualification", "Father Qualification"),
    ("father_mobile_number", "Father Mobile Number"),
    ("father_email", "Father Email"),
    ("father_aadhaar_number", "Father Aadhaar Number"),
    ("father_annual_income", "Father Annual Income"),
    ("mother_name", "Mother Name"),
    ("mother_occupation", "Mother Occupation"),
    ("mother_qualification", "Mother Qualification"),
    ("mother_mobile_number", "Mother Mobile Number"),
    ("mother_aadhaar_number", "Mother Aadhaar Number"),
    ("mother_annual_income", "Mother Annual Income"),
    ("guardian_address", "Guardian Address"),
    ("address", "Address"),
    ("current_house_number", "Current House Number"),
    ("current_street_area", "Current Street / Area"),
    ("current_village_city", "Current Village / City"),
    ("current_taluka", "Current Taluka"),
    ("current_district", "Current District"),
    ("current_state", "Current State"),
    ("current_pin_code", "Current PIN Code"),
    ("permanent_house_number", "Permanent House Number"),
    ("permanent_street_area", "Permanent Street / Area"),
    ("permanent_village_city", "Permanent Village / City"),
    ("permanent_taluka", "Permanent Taluka"),
    ("permanent_district", "Permanent District"),
    ("permanent_state", "Permanent State"),
    ("permanent_pin_code", "Permanent PIN Code"),
    ("current_school_name", "Current School / College"),
    ("current_school_address", "Current School Address"),
    ("previous_school_name", "Previous School / College"),
    ("previous_school_address", "Previous School Address"),
    ("previous_school_udise_code", "Previous School UDISE Code"),
    ("previous_class", "Previous Class"),
    ("previous_class_passed", "Previous Class Passed"),
    ("last_exam_result", "Last Exam Result"),
    ("result", "Result"),
    ("conduct", "Conduct"),
    ("reason_for_leaving", "Reason For Leaving"),
    ("date_of_leaving_school", "Date Of Leaving School"),
    ("tc_issue_date", "TC Issue Date"),
    ("bonafide_purpose", "Bonafide Purpose"),
    ("emergency_contact_number", "Emergency Contact Number"),
    ("is_active", "Active Login"),
]

STUDENT_EXPORT_DEFAULT_FIELDS = [
    "admission_number",
    "name",
    "mobile",
    "batch",
    "total_fees",
    "paid_amount",
    "due_amount",
    "status",
]


def get_student_session_fee_summary(session):
    fee_enrollments = session.enrollments.exclude(status=StudentEnrollment.Status.CANCELLED)
    invoices = session.fee_invoices.exclude(status=FeeInvoice.Status.CANCELLED)
    enrollment_fee_amount = sum(enrollment.total_course_fee for enrollment in fee_enrollments)
    invoiced_amount = Decimal("0.00")
    additional_fee_amount = Decimal("0.00")
    paid_amount = Decimal("0.00")
    for invoice in invoices:
        invoiced_amount += invoice.amount
        if not invoice.enrollment_id:
            additional_fee_amount += invoice.amount
        paid_amount += sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
    total_fee_amount = enrollment_fee_amount + additional_fee_amount if enrollment_fee_amount > 0 else invoiced_amount
    due_amount = total_fee_amount - paid_amount
    if due_amount < 0:
        due_amount = Decimal("0.00")
    return total_fee_amount, paid_amount, due_amount


def get_student_export_fee_summaries(sessions):
    target_sql, target_params = sessions.order_by().values("pk").query.sql_with_params()
    enrollment_table = connection.ops.quote_name(StudentEnrollment._meta.db_table)
    enrollment_course_table = connection.ops.quote_name(
        StudentEnrollment.courses.through._meta.db_table
    )
    course_table = connection.ops.quote_name(Course._meta.db_table)
    invoice_table = connection.ops.quote_name(FeeInvoice._meta.db_table)
    payment_table = connection.ops.quote_name(Payment._meta.db_table)
    sql = f"""
        WITH target_session_ids AS (
            {target_sql}
        ),
        target_sessions AS (
            SELECT DISTINCT pk AS id FROM target_session_ids
        ),
        enrollment_course_fees AS (
            SELECT sec.studentenrollment_id AS enrollment_id, SUM(c.fee_amount) AS total
            FROM {enrollment_course_table} sec
            INNER JOIN {course_table} c ON c.id = sec.course_id
            GROUP BY sec.studentenrollment_id
        ),
        enrollment_summary AS (
            SELECT
                se.academic_session_id AS session_id,
                SUM(
                    CASE
                        WHEN se.custom_fee_amount IS NOT NULL THEN se.custom_fee_amount
                        ELSE COALESCE(ecf.total, 0)
                    END
                ) AS enrollment_total
            FROM {enrollment_table} se
            INNER JOIN target_sessions ts ON ts.id = se.academic_session_id
            LEFT JOIN enrollment_course_fees ecf ON ecf.enrollment_id = se.id
            WHERE se.status <> %s
            GROUP BY se.academic_session_id
        ),
        invoice_summary AS (
            SELECT
                fi.academic_session_id AS session_id,
                SUM(fi.amount) AS invoiced_total,
                SUM(CASE WHEN fi.enrollment_id IS NULL THEN fi.amount ELSE 0 END) AS additional_total
            FROM {invoice_table} fi
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE fi.status <> %s
            GROUP BY fi.academic_session_id
        ),
        payment_summary AS (
            SELECT fi.academic_session_id AS session_id, SUM(p.amount) AS paid_total
            FROM {payment_table} p
            INNER JOIN {invoice_table} fi ON fi.id = p.invoice_id
            INNER JOIN target_sessions ts ON ts.id = fi.academic_session_id
            WHERE p.status = %s AND fi.status <> %s
            GROUP BY fi.academic_session_id
        )
        SELECT
            ts.id,
            CASE
                WHEN COALESCE(es.enrollment_total, 0) > 0
                THEN COALESCE(es.enrollment_total, 0) + COALESCE(ins.additional_total, 0)
                ELSE COALESCE(ins.invoiced_total, 0)
            END AS total_fee,
            COALESCE(ps.paid_total, 0) AS paid_amount
        FROM target_sessions ts
        LEFT JOIN enrollment_summary es ON es.session_id = ts.id
        LEFT JOIN invoice_summary ins ON ins.session_id = ts.id
        LEFT JOIN payment_summary ps ON ps.session_id = ts.id
    """
    params = [
        *target_params,
        StudentEnrollment.Status.CANCELLED,
        FeeInvoice.Status.CANCELLED,
        Payment.Status.ACTIVE,
        FeeInvoice.Status.CANCELLED,
    ]
    summaries = {}
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        for session_id, total_fee, paid_amount in cursor.fetchall():
            total_fee = Decimal(str(total_fee or 0))
            paid_amount = Decimal(str(paid_amount or 0))
            summaries[session_id] = (
                total_fee,
                paid_amount,
                max(total_fee - paid_amount, Decimal("0.00")),
            )
    return summaries


def get_student_export_field_keys(request):
    allowed_fields = {key for key, _label in STUDENT_EXPORT_FIELD_OPTIONS}
    requested_fields = [field for field in request.GET.getlist("fields") if field in allowed_fields]
    return requested_fields or STUDENT_EXPORT_DEFAULT_FIELDS


def get_student_export_related_summaries(sessions):
    session_ids = sessions.order_by().values("pk")
    student_ids = sessions.order_by().values("student_id")
    guardian_summaries = {}
    for guardian in (
        GuardianProfile.objects.filter(student_id__in=Subquery(student_ids))
        .order_by("student_id", "-is_primary", "pk")
        .iterator(chunk_size=1000)
    ):
        guardian_summaries.setdefault(guardian.student_id, guardian)

    batch_summaries = defaultdict(list)
    batch_rows = (
        StudentEnrollment.objects.filter(academic_session_id__in=Subquery(session_ids))
        .exclude(status=StudentEnrollment.Status.CANCELLED)
        .values_list("academic_session_id", "batch__name")
        .order_by("academic_session_id", "batch__name")
    )
    for session_id, batch_name in batch_rows.iterator(chunk_size=1000):
        if batch_name and batch_name not in batch_summaries[session_id]:
            batch_summaries[session_id].append(batch_name)
    return guardian_summaries, batch_summaries


def student_export_row(
    session,
    selected_fields,
    fee_summaries,
    guardian_summaries=None,
    batch_summaries=None,
):
    student = session.student
    profile = getattr(student.user, "profile", None)
    if guardian_summaries is None:
        guardians = list(student.guardians.all())
        guardian = next((item for item in guardians if item.is_primary), guardians[0] if guardians else None)
    else:
        guardian = guardian_summaries.get(student.pk)
    if batch_summaries is None:
        batch_names = ", ".join(
            enrollment.batch.name
            for enrollment in session.enrollments.all()
            if enrollment.batch_id
        )
    else:
        batch_names = ", ".join(batch_summaries.get(session.pk, []))
    total_fee_amount, paid_amount, due_amount = fee_summaries.get(
        session.pk,
        (Decimal("0.00"), Decimal("0.00"), Decimal("0.00")),
    )
    full_name = student.user.get_full_name() or student.user.username
    field_values = {
        "student_id": student.pk,
        "admission_number": session.admission_number,
        "name": full_name,
        "institute_name": student.institute.name if student.institute_id else "",
        "academic_year": session.academic_year.name if session.academic_year_id else "",
        "mobile": profile.phone if profile else "",
        "batch": batch_names,
        "total_fees": total_fee_amount,
        "paid_amount": paid_amount,
        "due_amount": due_amount,
        "status": "Active"
        if session.status == StudentAcademicSession.Status.ACTIVE and student.is_active
        else "Inactive",
        "session_status": session.get_status_display(),
        "student_status": student.get_student_status_display(),
        "first_name": student.user.first_name,
        "middle_name": student.middle_name,
        "last_name": student.user.last_name,
        "username": student.user.username,
        "email": student.user.email,
        "profile_image": student.profile_image.url if student.profile_image else "",
        "pen_no": student.pen_no,
        "appar_id": student.appar_id,
        "gr_number": student.gr_number_udise,
        "udise_number": student.udise_number,
        "roll_number": student.roll_number,
        "gender": student.get_gender_display() if student.gender else "",
        "date_of_birth": student.date_of_birth.isoformat() if student.date_of_birth else "",
        "blood_group": student.blood_group,
        "religion": student.religion,
        "cast": student.cast,
        "caste_category": student.caste_category,
        "nationality": student.nationality,
        "aadhaar_number": student.aadhaar_number,
        "birth_certificate_number": student.birth_certificate_number,
        "place_of_birth": student.place_of_birth,
        "mother_tongue": student.mother_tongue,
        "joined_on": session.joined_on.isoformat() if session.joined_on else "",
        "admission_class": student.admission_class,
        "current_class": student.current_class,
        "division": student.division,
        "medium": student.medium,
        "guardian_name": guardian.name if guardian else "",
        "guardian_relation": guardian.relation if guardian else "",
        "guardian_phone": guardian.phone if guardian else "",
        "guardian_email": guardian.email if guardian else "",
        "father_name": student.father_name,
        "father_occupation": student.father_occupation,
        "father_qualification": student.father_qualification,
        "father_mobile_number": student.father_mobile_number,
        "father_email": student.father_email,
        "father_aadhaar_number": student.father_aadhaar_number,
        "father_annual_income": student.father_annual_income,
        "mother_name": student.mother_name,
        "mother_occupation": student.mother_occupation,
        "mother_qualification": student.mother_qualification,
        "mother_mobile_number": student.mother_mobile_number,
        "mother_aadhaar_number": student.mother_aadhaar_number,
        "mother_annual_income": student.mother_annual_income,
        "guardian_address": student.guardian_address,
        "address": student.address,
        "current_house_number": student.current_house_number,
        "current_street_area": student.current_street_area,
        "current_village_city": student.current_village_city,
        "current_taluka": student.current_taluka,
        "current_district": student.current_district,
        "current_state": student.current_state,
        "current_pin_code": student.current_pin_code,
        "permanent_house_number": student.permanent_house_number,
        "permanent_street_area": student.permanent_street_area,
        "permanent_village_city": student.permanent_village_city,
        "permanent_taluka": student.permanent_taluka,
        "permanent_district": student.permanent_district,
        "permanent_state": student.permanent_state,
        "permanent_pin_code": student.permanent_pin_code,
        "current_school_name": session.current_school_name,
        "current_school_address": session.current_school_address,
        "previous_school_name": session.previous_school_name,
        "previous_school_address": student.previous_school_address,
        "previous_school_udise_code": student.previous_school_udise_code,
        "previous_class": session.previous_class,
        "previous_class_passed": student.previous_class_passed,
        "last_exam_result": student.last_exam_result,
        "result": student.result,
        "conduct": student.conduct,
        "reason_for_leaving": student.reason_for_leaving,
        "date_of_leaving_school": student.date_of_leaving_school.isoformat()
        if student.date_of_leaving_school
        else "",
        "tc_issue_date": student.tc_issue_date.isoformat() if student.tc_issue_date else "",
        "bonafide_purpose": student.bonafide_purpose,
        "emergency_contact_number": student.emergency_contact_number,
        "is_active": "Yes" if student.is_active else "No",
    }
    return [field_values[field] for field in selected_fields]


def parse_excel_date_value(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {text}. Use YYYY-MM-DD.")


def bool_from_excel(value):
    if value in (True, False):
        return bool(value)
    text = str(value or "Yes").strip().lower()
    return text not in {"no", "false", "0", "inactive"}


def match_student_import_headers(headers):
    expected = student_import_columns()
    submitted = headers[: len(expected)]
    if len(submitted) < len(expected):
        submitted.extend([""] * (len(expected) - len(submitted)))

    mismatched = [
        (index, actual, expected_header)
        for index, (actual, expected_header) in enumerate(zip(submitted, expected), start=1)
        if actual and actual != expected_header
    ]
    if mismatched:
        return False, []

    matching_count = sum(actual == expected_header for actual, expected_header in zip(submitted, expected))
    if matching_count < len(expected) - 2:
        return False, []

    missing_headers = [
        expected_header
        for actual, expected_header in zip(submitted, expected)
        if not actual
    ]
    return True, missing_headers


def validate_student_import_file(upload, institute, academic_year=None):
    academic_year = academic_year or get_or_create_academic_year(institute)
    try:
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook["Students"] if "Students" in workbook.sheetnames else workbook.active
    except Exception as exc:
        return {
            "valid": False,
            "valid_count": 0,
            "errors": [f"Could not read Excel file: {exc}"],
            "warnings": [],
        }

    expected = student_import_columns()
    headers = [str(cell.value or "").strip() for cell in sheet[3]]
    headers_match, missing_headers = match_student_import_headers(headers)
    if not headers_match:
        return {
            "valid": False,
            "valid_count": 0,
            "errors": ["Invalid template format. Download the latest template and try again."],
            "warnings": [],
        }

    errors = []
    warnings = []
    if missing_headers:
        warnings.append(
            "Blank column header(s) were recovered from their template positions: "
            + ", ".join(missing_headers)
            + "."
        )
    cell_errors = {}
    preview_rows = []
    valid_count = 0
    seen_emails = {}
    seen_phones = {}
    existing_phones = set(
        UserProfile.objects.filter(institute=institute)
        .exclude(phone="")
        .values_list("phone", flat=True)
    )

    for row_number in range(4, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_number, column=col).value for col in range(1, len(expected) + 1)]
        if not any(value not in (None, "") for value in row_values):
            continue

        row = dict(zip(expected, row_values))
        row_errors = []
        row_cell_errors = {}
        first_name = str(row["First Name *"] or "").strip()
        email = str(row["Email"] or "").strip().lower()
        phone = str(row["Phone"] or "").strip()

        if not first_name:
            error = "First Name is required."
            row_errors.append(error)
            row_cell_errors["First Name *"] = error
        if phone:
            if phone in existing_phones:
                error = f"Mobile number already exists: {phone}."
                row_errors.append(error)
                row_cell_errors["Phone"] = error
            if phone in seen_phones:
                error = f"Duplicate mobile number in file. Also used on row {seen_phones[phone]}."
                row_errors.append(error)
                row_cell_errors["Phone"] = error
            seen_phones[phone] = row_number
        else:
            error = "Phone is required and must be unique."
            row_errors.append(error)
            row_cell_errors["Phone"] = error
        if email:
            if email in seen_emails:
                warnings.append(f"Row {row_number}: Email also appears on row {seen_emails[email]} ({email}).")
            seen_emails[email] = row_number
        for date_column in ("Date of Birth", "Joined On"):
            try:
                parse_excel_date_value(row[date_column])
            except ValueError as exc:
                row_errors.append(str(exc))
                row_cell_errors[date_column] = str(exc)
        if not row["Password"]:
            warnings.append(f"Row {row_number}: Password is blank. Default Student@123 will be used.")

        if row_errors:
            errors.extend(f"Row {row_number}: {error}" for error in row_errors)
        else:
            valid_count += 1
        if row_cell_errors:
            cell_errors[str(row_number)] = row_cell_errors
        preview_rows.append(
            {
                "row_number": row_number,
                "values": ["" if value is None else str(value) for value in row_values],
                "errors": row_cell_errors,
            }
        )

    if valid_count == 0 and not errors:
        errors.append("No student rows found in the file.")

    return {
        "valid": not errors and valid_count > 0,
        "valid_count": valid_count,
        "errors": errors,
        "warnings": warnings,
        "prefix": f"{get_student_admission_prefix(institute, academic_year)}0001",
        "headers": expected,
        "rows": preview_rows[:100],
        "cell_errors": cell_errors,
    }


@institute_admin_required
def student_import_template(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    columns = student_import_columns()
    border = style_student_workbook_header(sheet, "Student Bulk Import Template", len(columns))
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sheet.cell(
        row=2,
        column=1,
        value=f"Admission number will be generated automatically like {get_student_admission_prefix(institute, academic_year)}0001",
    )
    sheet.cell(row=2, column=1).font = Font(italic=True, color="64748B")
    for col, header in enumerate(columns, start=1):
        sheet.cell(row=3, column=col, value=header)
        sheet.column_dimensions[get_column_letter(col)].width = 22
    examples = [
        "Rohan",
        "Sharma",
        "",
        "rohan@example.com",
        "9876543210",
        "2010-05-12",
        timezone.localdate().isoformat(),
        "Student address",
        "Saint Monica International School",
        "School address",
        "Previous School",
        "10th",
        "Mahesh Sharma",
        "Father",
        "9876543210",
        "guardian@example.com",
        "Yes",
    ]
    for col, value in enumerate(examples, start=1):
        cell = sheet.cell(row=4, column=col, value=value)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
    sheet.freeze_panes = "A4"

    info = workbook.create_sheet("Instructions")
    info["A1"] = "Instructions"
    info["A1"].font = Font(size=16, bold=True, color="0F766E")
    info["A3"] = "1. Do not add Admission Number. It is generated automatically."
    info["A4"] = "2. First Name is required."
    info["A5"] = "3. Username is generated automatically and is the same as the admission number."
    info["A6"] = "4. Password is optional. If blank, Student@123 will be used."
    info["A7"] = "5. Dates can be YYYY-MM-DD or DD-MM-YYYY."
    info.column_dimensions["A"].width = 90

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
    return response


@institute_admin_required
def student_export(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    sessions = StudentAcademicSession.objects.select_related(
        "student",
        "student__institute",
        "student__user",
        "student__user__profile",
        "academic_year",
    )
    if institute:
        sessions = sessions.filter(institute=institute)
    if academic_year:
        sessions = sessions.filter(academic_year=academic_year)
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    batch_queryset = Batch.objects.filter(institute=institute, is_active=True) if institute else Batch.objects.none()
    if academic_year:
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
    selected_batch = batch_queryset.filter(pk=batch_filter).first() if batch_filter else None
    if batch_filter:
        if selected_batch:
            sessions = sessions.filter(enrollments__batch=selected_batch).distinct()
        else:
            sessions = sessions.none()
    if search_query:
        sessions = sessions.filter(
            Q(admission_number__icontains=search_query)
            | Q(student__user__first_name__icontains=search_query)
            | Q(student__user__last_name__icontains=search_query)
            | Q(student__user__username__icontains=search_query)
            | Q(student__user__email__icontains=search_query)
            | Q(student__user__profile__phone__icontains=search_query)
            | Q(student__guardians__name__icontains=search_query)
            | Q(student__guardians__phone__icontains=search_query)
        ).distinct()
    if status_filter == "active":
        sessions = sessions.filter(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True)
    elif status_filter == "inactive":
        sessions = sessions.exclude(status=StudentAcademicSession.Status.ACTIVE, student__is_active=True)

    selected_fields = get_student_export_field_keys(request)
    field_labels = dict(STUDENT_EXPORT_FIELD_OPTIONS)
    columns = [field_labels[field] for field in selected_fields]
    session_count = sessions.count()
    fee_summaries = get_student_export_fee_summaries(sessions)
    guardian_summaries, batch_summaries = get_student_export_related_summaries(sessions)
    sessions = sessions.order_by("admission_number")

    export_format = request.GET.get("format", "").strip().lower()
    stream_threshold = int(getattr(settings, "STUDENT_EXPORT_CSV_THRESHOLD", 2000))
    if export_format == "csv" or session_count >= stream_threshold:
        class CsvEcho:
            def write(self, value):
                return value

        writer = csv.writer(CsvEcho())

        def stream_rows():
            yield "\ufeff"
            yield writer.writerow(columns)
            for session in sessions.iterator(chunk_size=500):
                yield writer.writerow(
                    student_export_row(
                        session,
                        selected_fields,
                        fee_summaries,
                        guardian_summaries,
                        batch_summaries,
                    )
                )

        response = StreamingHttpResponse(
            stream_rows(),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = 'attachment; filename="student_export.csv"'
        response["X-Export-Mode"] = "streamed"
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    border = style_student_workbook_header(sheet, "Student Export", len(columns))
    for col, header in enumerate(columns, start=1):
        sheet.cell(row=3, column=col, value=header)
        sheet.column_dimensions[get_column_letter(col)].width = 22
    for row, session in enumerate(sessions, start=4):
        values = student_export_row(
            session,
            selected_fields,
            fee_summaries,
            guardian_summaries,
            batch_summaries,
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(columns))}{max(4, session_count + 3)}"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_export.xlsx"'
    return response


@institute_admin_required
def student_bulk_import_validate(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if request.method != "POST":
        return JsonResponse({"valid": False, "errors": ["Invalid request."], "warnings": [], "valid_count": 0}, status=405)
    upload = request.FILES.get("student_file")
    if not upload:
        return JsonResponse(
            {"valid": False, "errors": ["Select an Excel file before validation."], "warnings": [], "valid_count": 0},
            status=400,
        )
    result = validate_student_import_file(upload, institute, academic_year)
    return JsonResponse(result)


@institute_admin_required
def student_bulk_import(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if request.method != "POST":
        return redirect("institute_admin:student_list")
    upload = request.FILES.get("student_file")
    if not upload:
        messages.error(request, "Select an Excel file before importing students.")
        return redirect("institute_admin:student_list")

    validation = validate_student_import_file(upload, institute, academic_year)
    if hasattr(upload, "seek"):
        upload.seek(0)
    if validation["errors"]:
        messages.error(request, "Fix validation errors before importing: " + " | ".join(validation["errors"][:5]))
        return redirect("institute_admin:student_list")

    background_threshold = int(
        getattr(settings, "STUDENT_IMPORT_BACKGROUND_THRESHOLD", 500)
    )
    if validation["valid_count"] >= background_threshold:
        if hasattr(upload, "seek"):
            upload.seek(0)
        job = enqueue_background_job(
            BackgroundJob.JobType.STUDENT_IMPORT,
            institute=institute,
            academic_year=academic_year,
            created_by=request.user,
            payload={"row_count": validation["valid_count"]},
            input_file=upload,
        )
        messages.success(
            request,
            f"Import of {validation['valid_count']} students queued as job #{job.pk}. "
            "It will run in the background.",
        )
        return redirect("institute_admin:student_list")

    try:
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook["Students"] if "Students" in workbook.sheetnames else workbook.active
    except Exception as exc:
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect("institute_admin:student_list")

    headers = [str(cell.value or "").strip() for cell in sheet[3]]
    expected = student_import_columns()
    headers_match, _missing_headers = match_student_import_headers(headers)
    if not headers_match:
        messages.error(request, "Invalid template format. Download the latest template and try again.")
        return redirect("institute_admin:student_list")

    rows_to_import = []
    errors = []
    for row_number in range(4, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_number, column=col).value for col in range(1, len(expected) + 1)]
        if not any(value not in (None, "") for value in row_values):
            continue
        row = dict(zip(expected, row_values))
        try:
            first_name = str(row["First Name *"] or "").strip()
            if not first_name:
                raise ValidationError("First Name is required.")
            rows_to_import.append(
                {
                    "row_number": row_number,
                    "first_name": first_name,
                    "last_name": str(row["Last Name"] or "").strip(),
                    "password": str(row["Password"] or "Student@123"),
                    "email": str(row["Email"] or "").strip(),
                    "phone": str(row["Phone"] or "").strip(),
                    "date_of_birth": parse_excel_date_value(row["Date of Birth"]),
                    "joined_on": parse_excel_date_value(row["Joined On"]),
                    "address": str(row["Address"] or "").strip(),
                    "current_school_name": str(row["Current School / College"] or "").strip(),
                    "current_school_address": str(row["Current School Address"] or "").strip(),
                    "previous_school_name": str(row["Previous School / College"] or "").strip(),
                    "previous_class": str(row["Previous Class"] or "").strip(),
                    "guardian_name": str(row["Guardian Name"] or "").strip(),
                    "guardian_relation": str(row["Guardian Relation"] or "").strip(),
                    "guardian_phone": str(row["Guardian Phone"] or row["Phone"] or "").strip(),
                    "guardian_email": str(row["Guardian Email"] or "").strip(),
                    "is_active": bool_from_excel(row["Active"]),
                }
            )
        except Exception as exc:
            errors.append(f"Row {row_number}: {exc}")

    if errors:
        messages.error(request, "Import failed: " + " | ".join(errors[:5]))
        return redirect("institute_admin:student_list")

    prefix = get_student_admission_prefix(institute, academic_year)
    sequence = get_last_student_admission_sequence(institute, academic_year) + 1
    username_prefix = build_student_username(institute, prefix)
    reserved_usernames = set(
        User.objects.filter(username__startswith=username_prefix).values_list("username", flat=True)
    )

    password_hashes = {}
    users = []
    for row in rows_to_import:
        while True:
            admission_number = f"{prefix}{sequence:04d}"
            username = build_student_username(institute, admission_number)
            sequence += 1
            if username not in reserved_usernames:
                reserved_usernames.add(username)
                break
        row["admission_number"] = admission_number
        row["username"] = username
        password = row["password"]
        if password not in password_hashes:
            password_hashes[password] = make_password(password)
        users.append(
            User(
                username=row["username"],
                first_name=row["first_name"],
                last_name=row["last_name"],
                email=row["email"],
                password=password_hashes[password],
                is_active=row["is_active"],
            )
        )

    created_count = 0
    try:
        with transaction.atomic():
            User.objects.bulk_create(users, batch_size=500)
            users_by_username = User.objects.in_bulk(
                [row["username"] for row in rows_to_import],
                field_name="username",
            )

            user_profiles = []
            students = []
            for row in rows_to_import:
                user = users_by_username[row["username"]]
                user_profiles.append(
                    UserProfile(
                        user=user,
                        institute=institute,
                        role=UserProfile.Role.STUDENT_PARENT,
                        phone=row["phone"],
                    )
                )
                students.append(
                    StudentProfile(
                        institute=institute,
                        academic_year=academic_year,
                        user=user,
                        admission_number=row["admission_number"],
                        date_of_birth=row["date_of_birth"],
                        joined_on=row["joined_on"],
                        address=row["address"],
                        current_school_name=row["current_school_name"],
                        current_school_address=row["current_school_address"],
                        previous_school_name=row["previous_school_name"],
                        previous_class=row["previous_class"],
                        is_active=row["is_active"],
                    )
                )

            UserProfile.objects.bulk_create(user_profiles, batch_size=500)
            StudentProfile.objects.bulk_create(students, batch_size=500)
            students_by_user_id = StudentProfile.objects.in_bulk(
                [user.pk for user in users_by_username.values()],
                field_name="user_id",
            )

            sessions = []
            guardians = []
            for row in rows_to_import:
                user = users_by_username[row["username"]]
                student = students_by_user_id[user.pk]
                sessions.append(
                    StudentAcademicSession(
                        institute=institute,
                        student=student,
                        academic_year=academic_year,
                        admission_number=row["admission_number"],
                        joined_on=row["joined_on"],
                        status=(
                            StudentAcademicSession.Status.ACTIVE
                            if row["is_active"]
                            else StudentAcademicSession.Status.LEFT
                        ),
                        current_school_name=row["current_school_name"],
                        current_school_address=row["current_school_address"],
                        previous_school_name=row["previous_school_name"],
                        previous_class=row["previous_class"],
                    )
                )
                if row["guardian_name"] or row["guardian_phone"]:
                    guardians.append(
                        GuardianProfile(
                            student=student,
                            name=row["guardian_name"] or "Primary Guardian",
                            relation=row["guardian_relation"],
                            phone=row["guardian_phone"],
                            email=row["guardian_email"],
                            is_primary=True,
                        )
                    )
            StudentAcademicSession.objects.bulk_create(sessions, batch_size=500)
            if guardians:
                GuardianProfile.objects.bulk_create(guardians, batch_size=500)
            created_count = len(rows_to_import)
            student_credentials = [
                (
                    students_by_user_id[users_by_username[row["username"]].pk].pk,
                    row["password"],
                )
                for row in rows_to_import
            ]
            on_commit_email(send_bulk_student_welcomes, student_credentials)
    except Exception as exc:
        messages.error(request, f"Import failed. No students were created: {exc}")
        return redirect("institute_admin:student_list")

    if created_count:
        invalidate_dashboard_summary(institute.pk, academic_year.pk)
        messages.success(request, f"{created_count} student(s) imported successfully.")
    return redirect("institute_admin:student_list")


@institute_admin_required
def student_dashboard(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    session_queryset = StudentAcademicSession.objects.select_related(
        "institute",
        "academic_year",
        "student",
        "student__institute",
        "student__user",
        "student__user__profile",
    ).prefetch_related(
        "student__guardians",
        "student__documents",
        "student__enrollments__batch",
        "student__enrollments__courses",
    )
    if institute:
        session_queryset = session_queryset.filter(institute=institute)
    if academic_year:
        session_queryset = session_queryset.filter(academic_year=academic_year)
    student_session = get_object_or_404(session_queryset, student_id=pk)
    student = student_session.student

    enrollments = student_session.enrollments.select_related("batch").prefetch_related("courses").order_by("-enrolled_on", "-pk")
    fee_enrollments = enrollments.exclude(status=StudentEnrollment.Status.CANCELLED)
    invoices = (
        FeeInvoice.objects.filter(student=student, academic_session=student_session)
        .select_related("category", "batch", "enrollment")
        .prefetch_related("payments")
        .order_by("-created_at", "-pk")
    )
    payments = (
        Payment.objects.filter(invoice__student=student, invoice__academic_session=student_session)
        .select_related("invoice", "received_by")
        .order_by("-created_at", "-pk")
    )
    transfer_certificates = (
        student.transfer_certificates.select_related("generated_by", "cancelled_by", "academic_session")
        .filter(academic_session=student_session)
        .order_by("-generated_at", "-pk")
    )
    bonafide_certificates = (
        student.bonafide_certificates.select_related("generated_by", "cancelled_by", "academic_session")
        .filter(academic_session=student_session)
        .order_by("-generated_at", "-pk")
    )
    payment_activities = (
        PaymentActivity.objects.filter(payment__invoice__student=student, payment__invoice__academic_session=student_session)
        .select_related(
            "payment",
            "payment__invoice",
            "performed_by",
        )
        .order_by("-performed_at", "-pk")[:20]
    )
    attendance_date_from = request.GET.get("attendance_date_from", "").strip()
    attendance_date_to = request.GET.get("attendance_date_to", "").strip()
    attendance_status_filter = request.GET.get("attendance_status", "").strip()
    attendance_batch_filter = request.GET.get("attendance_batch", "").strip()
    attendance_filter_active = any(
        [
            attendance_date_from,
            attendance_date_to,
            attendance_status_filter,
            attendance_batch_filter,
        ]
    )
    attendance_records_queryset = Attendance.objects.filter(
        student=student,
        academic_session=student_session,
    ).select_related("batch", "marked_by")
    date_from = parse_date(attendance_date_from)
    date_to = parse_date(attendance_date_to)
    if date_from:
        attendance_records_queryset = attendance_records_queryset.filter(date__gte=date_from)
    if date_to:
        attendance_records_queryset = attendance_records_queryset.filter(date__lte=date_to)
    if attendance_status_filter in Attendance.Status.values:
        attendance_records_queryset = attendance_records_queryset.filter(status=attendance_status_filter)
    else:
        attendance_status_filter = ""
    if attendance_batch_filter:
        attendance_records_queryset = attendance_records_queryset.filter(batch_id=attendance_batch_filter)
    attendance_records_queryset = attendance_records_queryset.order_by("-date", "-pk")
    attendance_records = list(
        attendance_records_queryset if attendance_filter_active else attendance_records_queryset[:10]
    )
    attendance_batch_options = (
        Batch.objects.filter(
            pk__in=Attendance.objects.filter(
                student=student,
                academic_session=student_session,
            ).values("batch_id")
        )
        .order_by("name")
        .distinct()
    )

    invoice_rows = []
    invoiced_amount = Decimal("0.00")
    additional_fee_amount = Decimal("0.00")
    raw_paid_amount = Decimal("0.00")
    service_rows = []
    excess_credit = Decimal("0.00")
    enrollment_invoice_ids = set()
    additional_groups = {}

    for invoice in invoices:
        paid_amount = sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
        due_amount = invoice.amount - paid_amount
        if due_amount < 0:
            due_amount = Decimal("0.00")
        if invoice.status != FeeInvoice.Status.CANCELLED:
            invoiced_amount += invoice.amount
            if not invoice.enrollment_id:
                additional_fee_amount += invoice.amount
            raw_paid_amount += paid_amount
            if invoice.enrollment_id:
                enrollment_invoice_ids.add(invoice.pk)
            else:
                group_key = f"category-{invoice.category_id}" if invoice.category_id else f"invoice-{invoice.pk}"
                if group_key not in additional_groups:
                    additional_groups[group_key] = {
                        "title": invoice.category.name if invoice.category_id else invoice.title,
                        "category": invoice.category.name if invoice.category_id else "General",
                        "amount": Decimal("0.00"),
                        "actual_paid": Decimal("0.00"),
                        "due_date": invoice.due_date,
                        "adjusted_credit": Decimal("0.00"),
                    }
                additional_groups[group_key]["amount"] += invoice.amount
                additional_groups[group_key]["actual_paid"] += paid_amount
                if invoice.due_date and invoice.due_date > additional_groups[group_key]["due_date"]:
                    additional_groups[group_key]["due_date"] = invoice.due_date

    for enrollment in fee_enrollments:
        enrollment_paid = (
            Payment.objects.filter(
                invoice__student=student,
                invoice__academic_session=student_session,
                invoice__enrollment=enrollment,
                status=Payment.Status.ACTIVE,
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )
        display_paid = min(enrollment_paid, enrollment.total_course_fee)
        if enrollment_paid > enrollment.total_course_fee:
            excess_credit += enrollment_paid - enrollment.total_course_fee
        service_rows.append(
            {
                "title": f"{enrollment.batch.name} Fee",
                "category": "Enrollment",
                "amount": enrollment.total_course_fee,
                "paid_amount": display_paid,
                "actual_paid": enrollment_paid,
                "due_amount": enrollment.total_course_fee - display_paid,
                "due_date": enrollment.enrolled_on,
                "status": "PAID" if display_paid >= enrollment.total_course_fee else "PARTIAL" if display_paid > 0 else "UNPAID",
                "adjusted_credit": Decimal("0.00"),
            }
        )

    for group in additional_groups.values():
        display_paid = min(group["actual_paid"], group["amount"])
        if group["actual_paid"] > group["amount"]:
            excess_credit += group["actual_paid"] - group["amount"]
        service_rows.append(
            {
                "title": group["title"],
                "category": group["category"],
                "amount": group["amount"],
                "paid_amount": display_paid,
                "actual_paid": group["actual_paid"],
                "due_amount": group["amount"] - display_paid,
                "due_date": group["due_date"],
                "status": "PAID" if display_paid >= group["amount"] else "PARTIAL" if display_paid > 0 else "UNPAID",
                "adjusted_credit": Decimal("0.00"),
            }
        )

    for row in service_rows:
        if excess_credit <= 0:
            break
        if row["due_amount"] <= 0:
            continue
        credit = min(row["due_amount"], excess_credit)
        row["paid_amount"] += credit
        row["due_amount"] -= credit
        row["adjusted_credit"] += credit
        excess_credit -= credit

    for row in service_rows:
        if row["due_amount"] <= 0:
            row["due_amount"] = Decimal("0.00")
            row["status"] = "PAID"
        elif row["paid_amount"] > 0:
            row["status"] = "PARTIAL"
        else:
            row["status"] = "UNPAID"
        invoice_rows.append(row)

    enrollment_fee_amount = sum(enrollment.total_course_fee for enrollment in fee_enrollments)
    total_fee_amount = enrollment_fee_amount + additional_fee_amount if enrollment_fee_amount > 0 else invoiced_amount
    total_paid_amount = min(raw_paid_amount, total_fee_amount)
    overpaid_amount = raw_paid_amount - total_fee_amount
    if overpaid_amount < 0:
        overpaid_amount = Decimal("0.00")
    total_due_amount = total_fee_amount - total_paid_amount
    if total_due_amount < 0:
        total_due_amount = Decimal("0.00")

    attendance_queryset = Attendance.objects.filter(student=student, academic_session=student_session)
    attendance_total = attendance_queryset.count()
    present_count = attendance_queryset.filter(status=Attendance.Status.PRESENT).count()
    absent_count = attendance_queryset.filter(status=Attendance.Status.ABSENT).count()
    late_count = attendance_queryset.filter(status=Attendance.Status.LATE).count()
    attendance_percentage = round((present_count / attendance_total) * 100, 2) if attendance_total else 0
    documents = list(student.documents.all())
    document_rows = build_student_document_rows(documents)

    context = {
        "student": student,
        "student_session": student_session,
        "is_school_institute": institute.institute_type == Institute.InstituteType.SCHOOL,
        "primary_guardian": student.guardians.filter(is_primary=True).first() or student.guardians.first(),
        "enrollments": enrollments,
        "invoice_rows": invoice_rows,
        "payments": payments,
        "transfer_certificates": transfer_certificates,
        "bonafide_certificates": bonafide_certificates,
        "active_payment_count": payments.filter(status=Payment.Status.ACTIVE).count(),
        "all_payment_count": payments.count(),
        "payment_activities": payment_activities,
        "documents": documents,
        "document_rows": document_rows,
        "attendance_records": attendance_records,
        "attendance_status_choices": Attendance.Status.choices,
        "attendance_batch_options": attendance_batch_options,
        "attendance_filter_active": attendance_filter_active,
        "attendance_filtered_count": attendance_records_queryset.count(),
        "attendance_date_from": attendance_date_from,
        "attendance_date_to": attendance_date_to,
        "attendance_status_filter": attendance_status_filter,
        "attendance_batch_filter": attendance_batch_filter,
        "total_fee_amount": total_fee_amount,
        "total_paid_amount": total_paid_amount,
        "raw_paid_amount": raw_paid_amount,
        "overpaid_amount": overpaid_amount,
        "total_due_amount": total_due_amount,
        "invoiced_amount": invoiced_amount,
        "enrollment_fee_amount": enrollment_fee_amount,
        "additional_fee_amount": additional_fee_amount,
        "attendance_total": attendance_total,
        "present_count": present_count,
        "absent_count": absent_count,
        "late_count": late_count,
        "attendance_percentage": attendance_percentage,
    }
    return render(request, "institute_admin/student_dashboard.html", context)


def build_student_document_rows(documents):
    covered_document_ids = set()
    document_rows = []
    type_labels = dict(StudentDocument.DocumentType.choices)
    for field_name, (document_type, document_label) in StudentForm.DOCUMENT_UPLOAD_FIELDS.items():
        matched_document = next(
            (
                document
                for document in documents
                if document.pk not in covered_document_ids and document.document_type == document_type
            ),
            None,
        )
        if matched_document:
            covered_document_ids.add(matched_document.pk)
        document_rows.append(
            {
                "field_name": field_name,
                "label": document_label,
                "document_type": document_type,
                "document": matched_document,
                "type_display": (
                    matched_document.get_document_type_display()
                    if matched_document
                    else type_labels.get(document_type, document_label)
                ),
                "status": "Uploaded" if matched_document else "Pending",
            }
        )
    for document in documents:
        if document.pk in covered_document_ids:
            continue
        document_rows.append(
            {
                "field_name": "additional_document",
                "label": document.title or document.get_document_type_display(),
                "document_type": document.document_type,
                "document": document,
                "type_display": document.get_document_type_display(),
                "status": "Uploaded",
            }
        )
    return document_rows


@institute_admin_required
def student_admission_form(request, pk):
    student_session = get_current_session_or_404(request, pk, prefetch_documents=True)
    student = student_session.student
    enrollments = (
        student_session.enrollments.select_related("batch")
        .prefetch_related("courses")
        .order_by("-enrolled_on", "-pk")
    )
    documents = list(student.documents.all())

    context = {
        "student": student,
        "student_session": student_session,
        "primary_guardian": student.guardians.filter(is_primary=True).first()
        or student.guardians.first(),
        "enrollments": enrollments,
        "document_rows": build_student_document_rows(documents),
        "generated_at": timezone.localtime(),
    }
    return render_print_document(
        request,
        PrintDocumentType.ADMISSION_FORM,
        "institute_admin/student_admission_form.html",
        context,
    )


@institute_admin_required
def student_tc_generate(request, pk):
    student_session = get_current_session_or_404(request, pk)
    student = student_session.student

    if request.method == "POST":
        form = StudentTransferCertificateForm(
            request.POST,
            student=student,
            academic_session=student_session,
            generated_by=request.user,
        )
        if form.is_valid():
            tc_record = form.save()
            messages.success(request, "Transfer Certificate generated successfully.")
            receipt_url = reverse("institute_admin:student_tc_print", args=[tc_record.pk])
            return close_popup_response(receipt_url)
    else:
        form = StudentTransferCertificateForm(
            student=student,
            academic_session=student_session,
            generated_by=request.user,
        )

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "student": student,
            "title": "Generate Transfer Certificate",
            "subtitle": "Only TC-specific details are required; student details are pulled from the profile.",
            "button_text": "Generate TC",
            "icon": "bi-file-earmark-text",
        },
    )


@institute_admin_required
def student_tc_print(request, pk):
    institute = get_current_institute(request)
    queryset = StudentTransferCertificate.objects.select_related(
        "student",
        "student__institute",
        "student__user",
        "academic_session",
        "academic_session__academic_year",
        "generated_by",
    )
    if institute:
        queryset = queryset.filter(institute=institute)
    tc_record = get_object_or_404(queryset, pk=pk)
    context = {
        "tc": tc_record,
        "student": tc_record.student,
        "student_session": tc_record.academic_session,
        "snapshot": tc_record.student_snapshot or {},
        "generated_at": timezone.localtime(),
    }
    return render_print_document(
        request,
        PrintDocumentType.TRANSFER_CERTIFICATE,
        "institute_admin/student_transfer_certificate.html",
        context,
    )


@require_POST
@institute_admin_required
def student_tc_cancel(request, pk):
    institute = get_current_institute(request)
    queryset = StudentTransferCertificate.objects.select_related("student", "institute")
    if institute:
        queryset = queryset.filter(institute=institute)
    tc_record = get_object_or_404(queryset, pk=pk)
    if tc_record.status == StudentTransferCertificate.Status.CANCELLED:
        messages.info(request, "This TC record is already cancelled.")
    else:
        tc_record.status = StudentTransferCertificate.Status.CANCELLED
        tc_record.cancelled_by = request.user
        tc_record.cancelled_at = timezone.now()
        tc_record.cancel_reason = request.POST.get("cancel_reason", "").strip()
        tc_record.save(update_fields=["status", "cancelled_by", "cancelled_at", "cancel_reason", "updated_at"])
        messages.success(request, "Transfer Certificate cancelled successfully.")
    return redirect("institute_admin:student_dashboard", pk=tc_record.student_id)


@institute_admin_required
def student_bonafide_update(request, pk):
    student_session = get_current_session_or_404(request, pk)
    student = student_session.student
    if request.method == "POST":
        form = StudentBonafideCertificateForm(
            request.POST,
            student=student,
            academic_session=student_session,
            generated_by=request.user,
        )
        if form.is_valid():
            bonafide_record = form.save()
            messages.success(request, "Bonafide Certificate generated successfully.")
            receipt_url = reverse("institute_admin:student_bonafide_print", args=[bonafide_record.pk])
            return close_popup_response(receipt_url)
    else:
        form = StudentBonafideCertificateForm(
            student=student,
            academic_session=student_session,
            generated_by=request.user,
        )

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "student": student,
            "title": "Generate Bonafide Certificate",
            "subtitle": "Only Bonafide-specific values are required; student details are pulled from the profile.",
            "button_text": "Generate Bonafide",
            "icon": "bi-file-earmark-person",
        },
    )


@institute_admin_required
def student_bonafide_print(request, pk):
    institute = get_current_institute(request)
    queryset = StudentBonafideCertificate.objects.select_related(
        "student",
        "student__institute",
        "student__user",
        "academic_session",
        "academic_session__academic_year",
        "generated_by",
    )
    if institute:
        queryset = queryset.filter(institute=institute)
    bonafide_record = get_object_or_404(queryset, pk=pk)
    context = {
        "bonafide": bonafide_record,
        "student": bonafide_record.student,
        "student_session": bonafide_record.academic_session,
        "snapshot": bonafide_record.student_snapshot or {},
        "generated_at": timezone.localtime(),
    }
    return render_print_document(
        request,
        PrintDocumentType.BONAFIDE_CERTIFICATE,
        "institute_admin/student_bonafide_certificate.html",
        context,
    )


@require_POST
@institute_admin_required
def student_bonafide_cancel(request, pk):
    institute = get_current_institute(request)
    queryset = StudentBonafideCertificate.objects.select_related("student", "institute")
    if institute:
        queryset = queryset.filter(institute=institute)
    bonafide_record = get_object_or_404(queryset, pk=pk)
    if bonafide_record.status == StudentBonafideCertificate.Status.CANCELLED:
        messages.info(request, "This Bonafide Certificate record is already cancelled.")
    else:
        bonafide_record.status = StudentBonafideCertificate.Status.CANCELLED
        bonafide_record.cancelled_by = request.user
        bonafide_record.cancelled_at = timezone.now()
        bonafide_record.cancel_reason = request.POST.get("cancel_reason", "").strip()
        bonafide_record.save(update_fields=["status", "cancelled_by", "cancelled_at", "cancel_reason", "updated_at"])
        messages.success(request, "Bonafide Certificate cancelled successfully.")
    return redirect("institute_admin:student_dashboard", pk=bonafide_record.student_id)


@institute_admin_required
def student_id_card_update(request, pk):
    student = get_current_session_student_or_404(request, pk)
    if request.method == "POST":
        form = StudentIdCardForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "ID card details updated successfully.")
            return close_popup_response()
    else:
        form = StudentIdCardForm(instance=student)

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "student": student,
            "title": "Update ID Card Details",
            "subtitle": "Manage only ID card and emergency contact values for this student.",
            "button_text": "Update ID Card",
            "icon": "bi-person-badge",
        },
    )


def get_current_session_student_or_404(request, pk, *, prefetch_documents=False):
    return get_current_session_or_404(request, pk, prefetch_documents=prefetch_documents).student


def get_current_session_or_404(request, pk, *, prefetch_documents=False):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = StudentAcademicSession.objects.select_related(
        "student",
        "student__institute",
        "student__user",
        "student__user__profile",
    )
    if prefetch_documents:
        queryset = queryset.prefetch_related("student__documents")
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    return get_object_or_404(queryset, student_id=pk)


@institute_admin_required
def student_add_fee(request, pk):
    student_session = get_current_session_or_404(request, pk)
    student = student_session.student

    category_data = {
        str(category.pk): {
            "name": category.name,
            "default_amount": str(category.default_amount),
        }
        for category in FeeCategory.objects.filter(institute=student.institute, is_active=True)
    }

    if request.method == "POST":
        form = AddStudentFeeForm(request.POST, institute=student.institute)
        if form.is_valid():
            category = form.cleaned_data["category"]
            pending_invoice = (
                FeeInvoice.objects.filter(
                    student=student,
                    academic_session=student_session,
                    category=category,
                    enrollment__isnull=True,
                    status__in=[FeeInvoice.Status.UNPAID, FeeInvoice.Status.PARTIAL],
                )
                .order_by("-created_at", "-pk")
                .first()
            )
            if pending_invoice:
                pending_invoice.amount += form.cleaned_data["amount"]
                pending_invoice.title = form.cleaned_data["title"]
                pending_invoice.due_date = form.cleaned_data["due_date"]
                pending_invoice.save(update_fields=["amount", "title", "due_date"])
                refresh_invoice_status(pending_invoice)
            else:
                FeeInvoice.objects.create(
                    institute=student.institute,
                    student=student,
                    academic_session=student_session,
                    category=category,
                    title=form.cleaned_data["title"],
                    amount=form.cleaned_data["amount"],
                    due_date=form.cleaned_data["due_date"],
                    status=FeeInvoice.Status.UNPAID,
                )
            messages.success(request, "Fee added successfully.")
            return close_popup_response()
    else:
        form = AddStudentFeeForm(
            institute=student.institute,
            initial={
                "due_date": date.today(),
            },
        )

    return render(
        request,
        "institute_admin/add_student_fee_form.html",
        {
            "form": form,
            "student": student,
            "title": "Add Fee",
            "subtitle": "Add an extra charge to this student's total fees.",
            "button_text": "Add Fee",
            "category_data": category_data,
        },
    )


@institute_admin_required
def student_receive_fee(request, pk):
    student_session = get_current_session_or_404(request, pk)
    student = student_session.student

    if request.method == "POST":
        form = ReceiveFeeForm(request.POST, institute=student.institute, student=student, academic_session=student_session)
        if form.is_valid():
            with transaction.atomic():
                invoice = form.cleaned_data["existing_invoice"]
                enrollment = form.cleaned_data["enrollment"]

                if not invoice:
                    invoice = FeeInvoice.objects.create(
                        institute=student.institute,
                        student=student,
                        academic_session=student_session,
                        enrollment=enrollment,
                        batch=enrollment.batch if enrollment else None,
                        category=form.cleaned_data["category"],
                        title=form.cleaned_data["title"],
                        amount=form.cleaned_data["invoice_amount"],
                        due_date=form.cleaned_data["paid_on"],
                        status=FeeInvoice.Status.UNPAID,
                    )

                invoice_due_amount = form.get_invoice_due_amount(invoice)
                collectable_amount = invoice_due_amount
                if invoice.enrollment_id:
                    enrollment_due_amount = form.get_enrollment_due_amount(invoice.enrollment)
                    collectable_amount = min(invoice_due_amount, enrollment_due_amount)
                if form.cleaned_data["payment_amount"] > collectable_amount:
                    form.add_error("payment_amount", "Payment amount cannot be greater than remaining due amount.")
                else:
                    payment = Payment.objects.create(
                        invoice=invoice,
                        amount=form.cleaned_data["payment_amount"],
                        paid_on=form.cleaned_data["paid_on"],
                        method=form.cleaned_data["method"],
                        received_by=request.user,
                        receipt_number=form.cleaned_data["receipt_number"],
                        note=form.cleaned_data["note"],
                    )
                    if not payment.receipt_number:
                        payment.receipt_number = f"RCP-{payment.paid_on:%Y%m%d}-{payment.pk:05d}"
                        payment.save(update_fields=["receipt_number"])
                    PaymentActivity.objects.create(
                        payment=payment,
                        action=PaymentActivity.Action.CREATED,
                        performed_by=request.user,
                        new_amount=payment.amount,
                        new_method=payment.method,
                        new_receipt_number=payment.receipt_number,
                        note="Payment received.",
                    )
                    refresh_invoice_status(invoice)
                    transaction.on_commit(lambda payment_id=payment.pk: enqueue_fee_paid_notification(payment_id))
                    on_commit_email(send_payment_confirmation, payment.pk)
                    messages.success(request, "Fee payment received successfully.")
                    receipt_url = reverse("institute_admin:payment_receipt", args=[payment.pk])
                    return close_popup_response(receipt_url)
    else:
        initial = {}
        first_enrollment = student_session.enrollments.exclude(status=StudentEnrollment.Status.CANCELLED).select_related("batch").first()
        initial["paid_on"] = date.today()
        if first_enrollment:
            paid_amount = Payment.objects.filter(
                invoice__student=student,
                invoice__academic_session=student_session,
                invoice__enrollment=first_enrollment,
                status=Payment.Status.ACTIVE,
            ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
            remaining_amount = first_enrollment.total_course_fee - paid_amount
            if remaining_amount < 0:
                remaining_amount = Decimal("0.00")
            initial.update(
                {
                    "enrollment": first_enrollment,
                    "title": f"{first_enrollment.batch.name} Fee",
                    "invoice_amount": remaining_amount,
                    "payment_amount": remaining_amount,
                }
            )
        form = ReceiveFeeForm(institute=student.institute, student=student, academic_session=student_session, initial=initial)

    return render(
        request,
        "institute_admin/receive_fee_form.html",
        {
            "form": form,
            "student": student,
            "title": "Receive Fee",
            "subtitle": "Create or collect fee payment for this student.",
            "button_text": "Receive Payment",
            "enrollment_due_data": get_student_enrollment_due_data(student_session),
            "category_due_data": get_student_category_due_data(student_session),
        },
    )


@institute_admin_required
def payment_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Payment.objects.select_related("invoice", "invoice__student", "invoice__student__institute")
    if institute:
        queryset = queryset.filter(invoice__student__institute=institute)
    if academic_year:
        queryset = queryset.filter(invoice__academic_session__academic_year=academic_year)
    payment = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = PaymentUpdateForm(request.POST, payment=payment)
        if form.is_valid():
            old_amount = payment.amount
            old_method = payment.method
            old_receipt_number = payment.receipt_number

            payment.amount = form.cleaned_data["amount"]
            payment.paid_on = form.cleaned_data["paid_on"]
            payment.method = form.cleaned_data["method"]
            payment.receipt_number = form.cleaned_data["receipt_number"]
            payment.note = form.cleaned_data["note"]
            payment.save()

            PaymentActivity.objects.create(
                payment=payment,
                action=PaymentActivity.Action.UPDATED,
                performed_by=request.user,
                old_amount=old_amount,
                new_amount=payment.amount,
                old_method=old_method,
                new_method=payment.method,
                old_receipt_number=old_receipt_number,
                new_receipt_number=payment.receipt_number,
                note=form.cleaned_data["correction_reason"],
            )
            refresh_invoice_status(payment.invoice)
            on_commit_email(send_payment_update, payment.pk)
            messages.success(request, "Payment corrected successfully.")
            return close_popup_response()
    else:
        form = PaymentUpdateForm(payment=payment)

    return render(
        request,
        "institute_admin/payment_update_form.html",
        {
            "form": form,
            "payment": payment,
            "title": "Correct Payment",
            "subtitle": "Update payment details with a required audit reason.",
            "button_text": "Save Correction",
        },
    )


@institute_admin_required
def payment_receipt(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Payment.objects.select_related(
        "invoice",
        "invoice__academic_session",
        "invoice__student",
        "invoice__student__user",
        "invoice__student__institute",
        "invoice__category",
        "invoice__enrollment",
        "invoice__batch",
        "received_by",
    )
    if institute:
        queryset = queryset.filter(invoice__student__institute=institute)
    if academic_year:
        queryset = queryset.filter(invoice__academic_session__academic_year=academic_year)
    payment = get_object_or_404(queryset, pk=pk)
    invoice = payment.invoice
    student = invoice.student
    student_session = invoice.academic_session
    primary_guardian = student.guardians.filter(is_primary=True).first() or student.guardians.first()
    invoice_paid_amount = (
        invoice.payments.filter(status=Payment.Status.ACTIVE).aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )
    invoice_due_amount = invoice.amount - invoice_paid_amount
    if invoice_due_amount < 0:
        invoice_due_amount = Decimal("0.00")

    return render(
        request,
        "institute_admin/payment_receipt.html",
        {
            "payment": payment,
            "invoice": invoice,
            "student": student,
            "student_session": student_session,
            "guardian": primary_guardian,
            "institute": student.institute,
            "invoice_paid_amount": invoice_paid_amount,
            "invoice_due_amount": invoice_due_amount,
        },
    )


@institute_admin_required
def payment_void(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Payment.objects.select_related("invoice", "invoice__student", "invoice__student__institute")
    if institute:
        queryset = queryset.filter(invoice__student__institute=institute)
    if academic_year:
        queryset = queryset.filter(invoice__academic_session__academic_year=academic_year)
    payment = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = PaymentVoidForm(request.POST, payment=payment)
        if form.is_valid():
            old_amount = payment.amount
            payment.void(request.user, form.cleaned_data["void_reason"])
            PaymentActivity.objects.create(
                payment=payment,
                action=PaymentActivity.Action.VOIDED,
                performed_by=request.user,
                old_amount=old_amount,
                new_amount=Decimal("0.00"),
                old_method=payment.method,
                old_receipt_number=payment.receipt_number,
                note=form.cleaned_data["void_reason"],
            )
            refresh_invoice_status(payment.invoice)
            messages.success(request, "Payment voided successfully.")
            return close_popup_response()
    else:
        form = PaymentVoidForm(payment=payment)

    return render(
        request,
        "institute_admin/payment_void_form.html",
        {
            "form": form,
            "payment": payment,
            "title": "Void Payment",
            "subtitle": "Void this payment with a required reason.",
            "button_text": "Void Payment",
        },
    )


@institute_admin_required
def student_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    active_section = "basic"
    if not institute:
        messages.error(request, "Select an institute before creating a student.")
        return redirect("institute_admin:student_list")

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, institute=institute, academic_year=academic_year)
        if form.is_valid():
            student = form.save()
            temporary_password = form.cleaned_data.get("password") or "Student@123"
            on_commit_email(send_student_welcome, student.pk, temporary_password)
            messages.success(request, "Student created successfully.")
            return close_popup_response()
    else:
        form = StudentForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/student_form.html",
        {
            "form": form,
            "title": "Create Student",
            "subtitle": "Add student login, admission and parent details.",
            "button_text": "Save Student",
            "course_batch_data": get_course_batch_data(institute, academic_year),
            "active_section": active_section,
        },
    )


@require_POST
@institute_admin_required
def student_dummy_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute or not academic_year:
        messages.error(request, "Select an institute and academic year before creating dummy students.")
        return redirect("institute_admin:student_list")

    form = DummyStudentCreateForm(request.POST)
    if not form.is_valid():
        error = form.errors.get("count", ["Enter a valid record count from 1 to 5000."])[0]
        messages.error(request, str(error))
        return redirect("institute_admin:student_list")

    batches = list(
        Batch.objects.filter(
            institute=institute,
            academic_year=academic_year,
            is_active=True,
        ).prefetch_related("courses")
    )
    if not batches:
        messages.error(
            request,
            "Create at least one active batch in the selected academic year before generating dummy students.",
        )
        return redirect("institute_admin:student_list")

    count = form.cleaned_data["count"]
    rng = random.SystemRandom()
    first_names = [
        "Aarav", "Aditi", "Advait", "Ananya", "Arjun", "Avni", "Diya", "Ishaan",
        "Kabir", "Kavya", "Krisha", "Meera", "Neel", "Nisha", "Pranav", "Riya",
        "Rohan", "Saanvi", "Sara", "Vihaan",
    ]
    last_names = [
        "Bhosale", "Chavan", "Deshmukh", "Gupta", "Iyer", "Jadhav", "Joshi",
        "Kapoor", "Kulkarni", "Mehta", "Mishra", "Nair", "Patel", "Rao",
        "Shah", "Sharma", "Singh", "Verma",
    ]
    relations = ["Father", "Mother", "Guardian"]
    previous_classes = ["5th", "6th", "7th", "8th", "9th", "10th", "11th", "12th"]
    localities = [
        "Aundh", "Baner", "Hadapsar", "Hinjewadi", "Kharadi", "Kothrud",
        "Pimpri", "Shivajinagar", "Viman Nagar", "Wakad",
    ]
    schools = [
        "Bright Future School", "City International School", "Green Valley Academy",
        "New Horizon School", "Scholars Public School", "Sunrise English School",
    ]
    batch_courses = {
        batch.pk: list(batch.courses.all())
        for batch in batches
    }
    existing_phones = set(
        UserProfile.objects.exclude(phone="").values_list("phone", flat=True)
    )
    existing_phones.update(
        GuardianProfile.objects.exclude(phone="").values_list("phone", flat=True)
    )

    def reserve_phone():
        while True:
            phone = f"{rng.choice('6789')}{rng.randrange(10**9):09d}"
            if phone not in existing_phones:
                existing_phones.add(phone)
                return phone

    try:
        with transaction.atomic():
            academic_year = AcademicYear.objects.select_for_update().get(
                pk=academic_year.pk,
                institute=institute,
            )
            prefix = get_student_admission_prefix(institute, academic_year)
            sequence = get_last_student_admission_sequence(institute, academic_year) + 1
            username_prefix = build_student_username(institute, prefix)
            reserved_usernames = set(
                User.objects.filter(username__startswith=username_prefix).values_list("username", flat=True)
            )
            default_password = make_password("Student@123")
            rows = []
            users = []
            for index in range(count):
                while True:
                    admission_number = f"{prefix}{sequence:04d}"
                    username = build_student_username(institute, admission_number)
                    sequence += 1
                    if username not in reserved_usernames:
                        reserved_usernames.add(username)
                        break

                first_name = rng.choice(first_names)
                last_name = rng.choice(last_names)
                locality = rng.choice(localities)
                school = rng.choice(schools)
                joined_span = max((academic_year.end_date - academic_year.start_date).days, 0)
                joined_on = academic_year.start_date + timedelta(days=rng.randint(0, joined_span))
                age = rng.randint(6, 18)
                date_of_birth = joined_on - timedelta(days=(age * 365) + rng.randint(0, 364))
                phone = reserve_phone()
                guardian_phone = reserve_phone()
                batch = rng.choice(batches)
                row = {
                    "admission_number": admission_number,
                    "username": username,
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": f"{username.lower()}@dummy.ultracoachmatrix.test",
                    "phone": phone,
                    "date_of_birth": date_of_birth,
                    "joined_on": joined_on,
                    "address": f"House {rng.randint(1, 999)}, {locality}, Pune, Maharashtra",
                    "current_school_name": institute.name,
                    "current_school_address": institute.address or f"Main Road, {locality}, Pune, Maharashtra",
                    "previous_school_name": school,
                    "previous_class": rng.choice(previous_classes),
                    "guardian_name": f"{rng.choice(first_names)} {last_name}",
                    "guardian_relation": rng.choice(relations),
                    "guardian_phone": guardian_phone,
                    "guardian_email": f"guardian.{username.lower()}@dummy.ultracoachmatrix.test",
                    "batch": batch,
                    "courses": batch_courses[batch.pk],
                }
                rows.append(row)
                users.append(
                    User(
                        username=username,
                        first_name=first_name,
                        last_name=last_name,
                        email=row["email"],
                        password=default_password,
                        is_active=True,
                    )
                )

            User.objects.bulk_create(users, batch_size=500)
            users_by_username = User.objects.in_bulk(
                [row["username"] for row in rows],
                field_name="username",
            )
            UserProfile.objects.bulk_create(
                [
                    UserProfile(
                        user=users_by_username[row["username"]],
                        institute=institute,
                        role=UserProfile.Role.STUDENT_PARENT,
                        phone=row["phone"],
                    )
                    for row in rows
                ],
                batch_size=500,
            )
            StudentProfile.objects.bulk_create(
                [
                    StudentProfile(
                        institute=institute,
                        academic_year=academic_year,
                        user=users_by_username[row["username"]],
                        admission_number=row["admission_number"],
                        date_of_birth=row["date_of_birth"],
                        joined_on=row["joined_on"],
                        address=row["address"],
                        current_school_name=row["current_school_name"],
                        current_school_address=row["current_school_address"],
                        previous_school_name=row["previous_school_name"],
                        previous_class=row["previous_class"],
                        is_active=True,
                    )
                    for row in rows
                ],
                batch_size=500,
            )
            students_by_user_id = StudentProfile.objects.in_bulk(
                [user.pk for user in users_by_username.values()],
                field_name="user_id",
            )
            sessions = [
                StudentAcademicSession(
                    institute=institute,
                    student=students_by_user_id[users_by_username[row["username"]].pk],
                    academic_year=academic_year,
                    admission_number=row["admission_number"],
                    joined_on=row["joined_on"],
                    status=StudentAcademicSession.Status.ACTIVE,
                    current_school_name=row["current_school_name"],
                    current_school_address=row["current_school_address"],
                    previous_school_name=row["previous_school_name"],
                    previous_class=row["previous_class"],
                )
                for row in rows
            ]
            StudentAcademicSession.objects.bulk_create(sessions, batch_size=500)
            sessions_by_student_id = {
                session.student_id: session
                for session in StudentAcademicSession.objects.filter(
                    student_id__in=[student.pk for student in students_by_user_id.values()],
                    academic_year=academic_year,
                )
            }
            GuardianProfile.objects.bulk_create(
                [
                    GuardianProfile(
                        student=students_by_user_id[users_by_username[row["username"]].pk],
                        name=row["guardian_name"],
                        relation=row["guardian_relation"],
                        phone=row["guardian_phone"],
                        email=row["guardian_email"],
                        is_primary=True,
                    )
                    for row in rows
                ],
                batch_size=500,
            )
            enrollments = []
            for row in rows:
                student = students_by_user_id[users_by_username[row["username"]].pk]
                courses = row["courses"]
                enrollments.append(
                    StudentEnrollment(
                        student=student,
                        academic_session=sessions_by_student_id[student.pk],
                        batch=row["batch"],
                        enrolled_on=row["joined_on"],
                        status=StudentEnrollment.Status.ACTIVE,
                        custom_fee_amount=sum(
                            (course.fee_amount for course in courses),
                            Decimal("0.00"),
                        ),
                    )
                )
            StudentEnrollment.objects.bulk_create(enrollments, batch_size=500)
            enrollment_courses = []
            for enrollment, row in zip(enrollments, rows):
                enrollment_courses.extend(
                    StudentEnrollment.courses.through(
                        studentenrollment_id=enrollment.pk,
                        course_id=course.pk,
                    )
                    for course in row["courses"]
                )
            if enrollment_courses:
                StudentEnrollment.courses.through.objects.bulk_create(
                    enrollment_courses,
                    batch_size=1000,
                )
    except Exception as exc:
        messages.error(request, f"Dummy student creation failed. No records were created: {exc}")
        return redirect("institute_admin:student_list")

    invalidate_dashboard_summary(institute.pk, academic_year.pk)
    messages.success(
        request,
        f"{count} complete dummy student record(s) created for {academic_year.name}. "
        "Default password: Student@123",
    )
    return redirect("institute_admin:student_list")


@institute_admin_required
def student_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    student = get_current_session_student_or_404(request, pk)
    valid_sections = {"basic", "login", "parent", "address", "academic", "documents"}
    active_section = request.GET.get("section", "basic")
    if active_section not in valid_sections:
        active_section = "basic"

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, institute=student.institute, student=student, academic_year=academic_year)
        active_section = request.POST.get("active_section") or active_section
        if active_section not in valid_sections:
            active_section = "basic"
        if form.is_valid():
            form.save()
            messages.success(request, "Student updated successfully.")
            return close_popup_response()
    else:
        form = StudentForm(institute=student.institute, student=student, academic_year=academic_year)

    return render(
        request,
        "institute_admin/student_form.html",
        {
            "form": form,
            "title": "Edit Student",
            "subtitle": "Update login, admission and parent details.",
            "button_text": "Update Student",
            "course_batch_data": get_course_batch_data(student.institute, academic_year),
            "active_section": active_section,
        },
    )


@institute_admin_required
def student_basic_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    student = get_current_session_student_or_404(request, pk)

    if request.method == "POST":
        form = StudentBasicForm(request.POST, request.FILES, institute=student.institute, student=student, academic_year=academic_year)
        if form.is_valid():
            form.save()
            messages.success(request, "Student basic information updated successfully.")
            return close_popup_response()
    else:
        form = StudentBasicForm(institute=student.institute, student=student, academic_year=academic_year)

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Edit Student Basic Info",
            "subtitle": "Update login, admission, contact and profile image.",
            "button_text": "Save Basic Info",
            "icon": "bi-person-vcard",
        },
    )


@institute_admin_required
def student_education_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    session_queryset = StudentAcademicSession.objects.select_related("student", "student__institute")
    if institute:
        session_queryset = session_queryset.filter(institute=institute)
    if academic_year:
        session_queryset = session_queryset.filter(academic_year=academic_year)
    student_session = get_object_or_404(session_queryset, student_id=pk)

    if request.method == "POST":
        form = StudentEducationForm(request.POST, instance=student_session)
        if form.is_valid():
            form.save()
            messages.success(request, "Education details updated successfully.")
            return close_popup_response()
    else:
        form = StudentEducationForm(instance=student_session)

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Edit Education Details",
            "subtitle": "Update school or college information for this student.",
            "button_text": "Save Education",
            "icon": "bi-building",
        },
    )


@institute_admin_required
def student_guardian_update(request, pk):
    student = get_current_session_student_or_404(request, pk)

    if request.method == "POST":
        form = StudentGuardianForm(request.POST, student=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Guardian details updated successfully.")
            return close_popup_response()
    else:
        form = StudentGuardianForm(student=student)

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Edit Parent / Guardian",
            "subtitle": "Update the primary guardian contact for this student.",
            "button_text": "Save Guardian",
            "icon": "bi-person-hearts",
        },
    )


@institute_admin_required
def student_document_upload(request, pk):
    student = get_current_session_student_or_404(request, pk)
    initial = {}
    requested_document_type = request.GET.get("document_type", "").strip()
    if requested_document_type in StudentDocument.DocumentType.values:
        initial["document_type"] = requested_document_type
    requested_document_title = request.GET.get("document_title", "").strip()
    if requested_document_title:
        initial["document_title"] = requested_document_title

    if request.method == "POST":
        form = StudentDocumentUploadForm(request.POST, request.FILES, student=student)
        if form.is_valid():
            documents = form.save()
            messages.success(request, f"{len(documents)} document(s) uploaded successfully.")
            return close_popup_response()
    else:
        form = StudentDocumentUploadForm(student=student, initial=initial)

    return render(
        request,
        "institute_admin/student_section_form.html",
        {
            "form": form,
            "title": "Upload Student Documents",
            "subtitle": "Select one or more files to upload for this student.",
            "button_text": "Upload Documents",
            "icon": "bi-file-earmark-arrow-up",
        },
    )


@institute_admin_required
def student_document_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = StudentDocument.objects.select_related("student", "student__institute")
    if institute:
        queryset = queryset.filter(student__institute=institute)
    if academic_year:
        queryset = queryset.filter(student__academic_sessions__academic_year=academic_year)
    document = get_object_or_404(queryset, pk=pk)
    student = document.student

    if request.method == "POST":
        if document.file:
            document.file.delete(save=False)
        document.delete()
        messages.success(request, "Student document deleted successfully.")

    return redirect("institute_admin:student_dashboard", pk=student.pk)


@institute_admin_required
def student_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = StudentAcademicSession.objects.select_related("student", "student__user")
    if institute:
        queryset = queryset.filter(institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)
    current_session_id = request.POST.get("current_session_id")
    if current_session_id and current_session_id.isdigit():
        queryset = StudentAcademicSession.objects.select_related("student", "student__user").filter(
            institute=institute,
            pk=current_session_id,
            student_id=pk,
        )
    student_session = get_object_or_404(queryset, student_id=pk)
    student = student_session.student

    if request.method == "POST":
        delete_scope = request.POST.get("delete_scope", "current_session")
        if delete_scope == "all_sessions":
            student.user.delete()
            messages.success(request, "Student and all academic sessions deleted successfully.")
        else:
            with transaction.atomic():
                student_session.delete()
                remaining_sessions = StudentAcademicSession.objects.filter(
                    institute=institute,
                    student=student,
                ).exists()
                if not remaining_sessions:
                    student.user.delete()
                    messages.success(request, "Student deleted successfully because no other sessions were available.")
                else:
                    messages.success(request, "Student removed from the current academic session successfully.")

    return redirect("institute_admin:student_list")


@institute_admin_required
def student_delete_details(request, pk):
    institute = get_current_institute(request)
    current_session_id = request.GET.get("current_session_id", "").strip()
    student_queryset = StudentProfile.objects.select_related("user", "user__profile").filter(institute=institute)
    student = get_object_or_404(student_queryset, pk=pk)

    sessions = (
        StudentAcademicSession.objects.filter(institute=institute, student=student)
        .select_related("academic_year")
        .prefetch_related("enrollments__batch", "enrollments__courses")
        .order_by("-academic_year__start_date", "-pk")
    )
    current_session = sessions.filter(pk=current_session_id).first() if current_session_id.isdigit() else None

    pending_invoices = (
        FeeInvoice.objects.filter(
            institute=institute,
            student=student,
            status__in=[FeeInvoice.Status.UNPAID, FeeInvoice.Status.PARTIAL],
        )
        .select_related("academic_session", "academic_session__academic_year", "batch", "course", "category")
        .prefetch_related("payments")
        .order_by("due_date", "pk")
    )

    session_rows = []
    for session in sessions:
        enrollments = list(session.enrollments.all())
        batch_names = []
        course_names = []
        for enrollment in enrollments:
            if enrollment.batch:
                batch_names.append(enrollment.batch.name)
            course_names.extend(course.name for course in enrollment.courses.all())
        session_rows.append(
            {
                "id": session.pk,
                "academic_year": session.academic_year.name,
                "admission_number": session.admission_number,
                "joined_on": session.joined_on.strftime("%Y-%m-%d") if session.joined_on else "-",
                "status": session.get_status_display(),
                "batches": ", ".join(dict.fromkeys(batch_names)) or "-",
                "courses": ", ".join(dict.fromkeys(course_names)) or "-",
                "is_current": bool(current_session and session.pk == current_session.pk),
            }
        )

    invoice_rows = []
    total_pending = Decimal("0.00")
    for invoice in pending_invoices:
        paid_amount = sum(payment.amount for payment in invoice.payments.filter(status=Payment.Status.ACTIVE))
        due_amount = invoice.amount - paid_amount
        if due_amount <= 0:
            continue
        total_pending += due_amount
        invoice_rows.append(
            {
                "id": invoice.pk,
                "academic_year": invoice.academic_session.academic_year.name if invoice.academic_session_id else "-",
                "title": invoice.title,
                "category": invoice.category.name if invoice.category else "-",
                "course": invoice.course.name if invoice.course else "-",
                "batch": invoice.batch.name if invoice.batch else "-",
                "amount": str(invoice.amount),
                "paid": str(paid_amount),
                "due": str(due_amount),
                "due_date": invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "-",
                "status": invoice.get_status_display(),
            }
        )

    other_session_count = sum(1 for row in session_rows if not row["is_current"])
    profile_phone = student.user.profile.phone if hasattr(student.user, "profile") else ""
    payload = {
        "student": {
            "id": student.pk,
            "name": student.user.get_full_name() or student.user.username,
            "username": student.user.username,
            "email": student.user.email or "-",
            "phone": profile_phone or "-",
            "admission_number": student.admission_number,
        },
        "summary": {
            "session_count": len(session_rows),
            "other_session_count": other_session_count,
            "pending_invoice_count": len(invoice_rows),
            "pending_amount": str(total_pending),
            "current_session": current_session.academic_year.name if current_session else "-",
        },
        "sessions": session_rows,
        "pending_invoices": invoice_rows,
    }
    return JsonResponse(payload)


@institute_admin_required
def enrollment_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    enrollments = StudentEnrollment.objects.select_related(
        "academic_session", "student", "student__user", "batch"
    ).prefetch_related("courses")
    if institute:
        enrollments = enrollments.filter(student__institute=institute)
    if academic_year:
        enrollments = enrollments.filter(academic_session__academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        enrollments = enrollments.filter(
            Q(academic_session__admission_number__icontains=search_query)
            | Q(student__user__first_name__icontains=search_query)
            | Q(student__user__last_name__icontains=search_query)
            | Q(batch__name__icontains=search_query)
            | Q(courses__name__icontains=search_query)
        ).distinct()

    if batch_filter:
        enrollments = enrollments.filter(batch_id=batch_filter)

    if status_filter:
        enrollments = enrollments.filter(status=status_filter)

    batch_queryset = Batch.objects.filter(institute=institute) if institute else Batch.objects.all()
    base_queryset = StudentEnrollment.objects.filter(student__institute=institute) if institute else StudentEnrollment.objects.all()
    if academic_year:
        base_queryset = base_queryset.filter(academic_session__academic_year=academic_year)

    page_obj, paginator, pagination_query = paginate_queryset(request, enrollments)
    context = {
        "enrollments": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "enrollments",
        "batches": batch_queryset,
        "status_choices": StudentEnrollment.Status.choices,
        "search_query": search_query,
        "batch_filter": batch_filter,
        "status_filter": status_filter,
        "total_enrollments": base_queryset.count(),
        "active_enrollments": base_queryset.filter(status=StudentEnrollment.Status.ACTIVE).count(),
        "completed_enrollments": base_queryset.filter(status=StudentEnrollment.Status.COMPLETED).count(),
        "cancelled_enrollments": base_queryset.filter(status=StudentEnrollment.Status.CANCELLED).count(),
    }
    return render(request, "institute_admin/enrollment_list.html", context)


@institute_admin_required
def enrollment_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating an enrollment.")
        return redirect("institute_admin:enrollment_list")

    initial = {}
    student_id = request.GET.get("student")
    if student_id:
        initial["student"] = student_id

    if request.method == "POST":
        form = StudentEnrollmentForm(request.POST, institute=institute, academic_year=academic_year)
        if form.is_valid():
            enrollment = form.save(commit=False)
            enrollment.academic_session = get_object_or_404(
                StudentAcademicSession,
                institute=institute,
                academic_year=academic_year,
                student=enrollment.student,
            )
            enrollment.save()
            form.save_m2m()
            messages.success(request, "Enrollment created successfully.")
            return close_popup_response()
    else:
        form = StudentEnrollmentForm(institute=institute, academic_year=academic_year, initial=initial)

    return render(
        request,
        "institute_admin/enrollment_form.html",
        {
            "form": form,
            "title": "Create Enrollment",
            "subtitle": "Assign student to a batch and select courses.",
            "button_text": "Save Enrollment",
            "batch_course_data": get_batch_course_data(institute, academic_year),
        },
    )


@institute_admin_required
def enrollment_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = StudentEnrollment.objects.select_related(
        "academic_session",
        "student",
        "student__user",
        "student__institute",
        "batch",
    )
    if institute:
        queryset = queryset.filter(student__institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_session__academic_year=academic_year)
    enrollment = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = StudentEnrollmentForm(
            request.POST,
            institute=institute,
            academic_year=academic_year,
            instance=enrollment,
        )
        if form.is_valid():
            updated_enrollment = form.save(commit=False)
            updated_enrollment.academic_session = get_object_or_404(
                StudentAcademicSession,
                institute=enrollment.student.institute,
                academic_year=academic_year,
                student=updated_enrollment.student,
            )
            updated_enrollment.save()
            form.save_m2m()
            messages.success(request, "Enrollment updated successfully.")
            return close_popup_response()
    else:
        form = StudentEnrollmentForm(
            institute=institute,
            academic_year=academic_year,
            instance=enrollment,
        )

    return render(
        request,
        "institute_admin/enrollment_form.html",
        {
            "form": form,
            "title": "Edit Enrollment",
            "subtitle": "Update batch, courses, status or custom fee.",
            "button_text": "Update Enrollment",
            "batch_course_data": get_batch_course_data(institute, academic_year),
        },
    )


@institute_admin_required
def enrollment_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = StudentEnrollment.objects.select_related("academic_session", "student")
    if institute:
        queryset = queryset.filter(student__institute=institute)
    if academic_year:
        queryset = queryset.filter(academic_session__academic_year=academic_year)
    enrollment = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if enrollment.fee_invoices.exists():
            messages.error(request, "This enrollment has fee invoices. Cancel it instead of deleting.")
        else:
            enrollment.delete()
            messages.success(request, "Enrollment deleted successfully.")

    return redirect("institute_admin:enrollment_list")


@institute_admin_required
def homework_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    homework = Homework.objects.select_related("batch", "subject", "course", "created_by").prefetch_related("attachments")
    if institute:
        homework = homework.filter(batch__institute=institute)
    if academic_year:
        homework = homework.filter(batch__academic_year=academic_year)

    search_query = request.GET.get("search", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    subject_filter = request.GET.get("subject", "").strip()
    course_filter = request.GET.get("course", "").strip()

    if search_query:
        homework = homework.filter(
            Q(title__icontains=search_query)
            | Q(instructions__icontains=search_query)
            | Q(batch__name__icontains=search_query)
            | Q(subject__name__icontains=search_query)
            | Q(course__name__icontains=search_query)
        )
    if batch_filter:
        homework = homework.filter(batch_id=batch_filter)
    if subject_filter:
        homework = homework.filter(subject_id=subject_filter)
    if course_filter:
        homework = homework.filter(course_id=course_filter)

    base_queryset = Homework.objects.filter(batch__institute=institute) if institute else Homework.objects.all()
    batch_queryset = Batch.objects.filter(institute=institute, is_active=True) if institute else Batch.objects.none()
    subject_queryset = Subject.objects.filter(institute=institute, is_active=True) if institute else Subject.objects.none()
    course_queryset = Course.objects.filter(institute=institute, is_active=True) if institute else Course.objects.none()
    if academic_year:
        base_queryset = base_queryset.filter(batch__academic_year=academic_year)
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
        subject_queryset = subject_queryset.filter(academic_year=academic_year)
        course_queryset = course_queryset.filter(academic_year=academic_year)

    page_obj, paginator, pagination_query = paginate_queryset(request, homework)
    context = {
        "homework_list": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "homework",
        "batches": batch_queryset,
        "subjects": subject_queryset,
        "courses": course_queryset,
        "search_query": search_query,
        "batch_filter": batch_filter,
        "subject_filter": subject_filter,
        "course_filter": course_filter,
        "today": date.today(),
        "total_homework": base_queryset.count(),
        "due_homework": base_queryset.filter(due_date__gte=date.today()).count(),
        "expired_homework": base_queryset.filter(due_date__lt=date.today()).count(),
        "subject_count": subject_queryset.count(),
    }
    return render(request, "institute_admin/homework_list.html", context)


@institute_admin_required
def homework_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating homework.")
        return redirect("institute_admin:homework_list")

    if request.method == "POST":
        form = HomeworkForm(request.POST, request.FILES, institute=institute, academic_year=academic_year)
        if form.is_valid():
            homework = form.save(commit=False)
            homework.created_by = request.user
            homework.save()
            form.save_attachments(homework)
            messages.success(request, "Homework created successfully.")
            return close_popup_response()
    else:
        form = HomeworkForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/homework_form.html",
        {
            "form": form,
            "title": "Create Homework",
            "subtitle": "Assign work to a batch with separate subject and course details.",
            "button_text": "Save Homework",
            "batch_course_data": get_batch_course_data(institute, academic_year),
            "attachments": [],
        },
    )


@institute_admin_required
def homework_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Homework.objects.select_related("batch", "subject", "course")
    if institute:
        queryset = queryset.filter(batch__institute=institute)
    if academic_year:
        queryset = queryset.filter(batch__academic_year=academic_year)
    homework = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = HomeworkForm(
            request.POST,
            request.FILES,
            institute=homework.batch.institute,
            academic_year=homework.batch.academic_year,
            instance=homework,
        )
        if form.is_valid():
            homework = form.save()
            form.save_attachments(homework)
            messages.success(request, "Homework updated successfully.")
            return close_popup_response()
    else:
        form = HomeworkForm(institute=homework.batch.institute, academic_year=homework.batch.academic_year, instance=homework)

    return render(
        request,
        "institute_admin/homework_form.html",
        {
            "form": form,
            "title": "Edit Homework",
            "subtitle": "Update batch, subject, course, instructions and due date.",
            "button_text": "Update Homework",
            "batch_course_data": get_batch_course_data(homework.batch.institute, homework.batch.academic_year),
            "attachments": homework.attachments.all(),
        },
    )


@institute_admin_required
def homework_delete(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Homework.objects.select_related("batch")
    if institute:
        queryset = queryset.filter(batch__institute=institute)
    if academic_year:
        queryset = queryset.filter(batch__academic_year=academic_year)
    homework = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        homework.delete()
        messages.success(request, "Homework deleted successfully.")

    return redirect("institute_admin:homework_list")


@institute_admin_required
def notice_list(request):
    institute = get_current_institute(request)
    notices = Notice.objects.select_related("created_by").prefetch_related(
        "target_batches",
        "target_courses",
        "target_students",
        "read_receipts",
    )
    if institute:
        notices = notices.filter(institute=institute)

    search_query = request.GET.get("search", "").strip()
    audience_filter = request.GET.get("audience", "").strip()
    category_filter = request.GET.get("category", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        notices = notices.filter(Q(title__icontains=search_query) | Q(message__icontains=search_query))
    if audience_filter:
        notices = notices.filter(audience=audience_filter)
    if category_filter:
        notices = notices.filter(category=category_filter)

    now = timezone.now()
    if status_filter == "published":
        notices = notices.filter(is_published=True).filter(Q(publish_at__isnull=True) | Q(publish_at__lte=now)).filter(Q(expires_at__isnull=True) | Q(expires_at__gte=now))
    elif status_filter == "scheduled":
        notices = notices.filter(is_published=True, publish_at__gt=now)
    elif status_filter == "expired":
        notices = notices.filter(expires_at__lt=now)
    elif status_filter == "draft":
        notices = notices.filter(is_published=False)

    base_queryset = Notice.objects.filter(institute=institute) if institute else Notice.objects.all()
    paginator = Paginator(notices, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "notice_list": page_obj,
        "page_obj": page_obj,
        "search_query": search_query,
        "audience_filter": audience_filter,
        "category_filter": category_filter,
        "status_filter": status_filter,
        "audience_choices": Notice.Audience.choices,
        "category_choices": Notice.Category.choices,
        "now": now,
        "total_notices": base_queryset.count(),
        "active_notices": base_queryset.filter(is_published=True).filter(Q(publish_at__isnull=True) | Q(publish_at__lte=now)).filter(Q(expires_at__isnull=True) | Q(expires_at__gte=now)).count(),
        "scheduled_notices": base_queryset.filter(is_published=True, publish_at__gt=now).count(),
        "app_notices": base_queryset.filter(push_to_app=True).count(),
    }
    return render(request, "institute_admin/notice_list.html", context)


@institute_admin_required
def notice_create(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    if not institute:
        messages.error(request, "Select an institute before creating a notice.")
        return redirect("institute_admin:notice_list")

    if request.method == "POST":
        form = NoticeForm(request.POST, institute=institute, academic_year=academic_year)
        if form.is_valid():
            notice = form.save(commit=False)
            notice.institute = institute
            notice.created_by = request.user
            notice.save()
            form.save_m2m()
            transaction.on_commit(lambda notice_id=notice.pk: enqueue_notice_published_notification(notice_id))
            messages.success(request, "Notice created successfully.")
            return close_popup_response()
    else:
        form = NoticeForm(institute=institute, academic_year=academic_year)

    return render(
        request,
        "institute_admin/notice_form.html",
        {
            "form": form,
            "title": "Create Notice",
            "subtitle": "Publish announcements for web records and the future student/parent app.",
            "button_text": "Publish Notice",
        },
    )


@institute_admin_required
def notice_update(request, pk):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    queryset = Notice.objects.prefetch_related("target_batches", "target_courses", "target_students")
    if institute:
        queryset = queryset.filter(institute=institute)
    notice = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = NoticeForm(request.POST, institute=notice.institute, academic_year=academic_year, instance=notice)
        if form.is_valid():
            notification_fields = {
                "title",
                "message",
                "audience",
                "category",
                "priority",
                "target_batches",
                "target_courses",
                "target_students",
                "publish_at",
                "expires_at",
                "is_published",
                "push_to_app",
            }
            should_notify = bool(notification_fields.intersection(form.changed_data))
            with transaction.atomic():
                notice = form.save()
                if should_notify:
                    Notice.objects.filter(pk=notice.pk).update(
                        push_notification_queued_at=None,
                        push_notification_version=F("push_notification_version") + 1,
                    )
                    transaction.on_commit(
                        lambda notice_id=notice.pk: enqueue_notice_published_notification(
                            notice_id
                        )
                    )
            messages.success(request, "Notice updated successfully.")
            return close_popup_response()
    else:
        form = NoticeForm(institute=notice.institute, academic_year=academic_year, instance=notice)

    return render(
        request,
        "institute_admin/notice_form.html",
        {
            "form": form,
            "title": "Edit Notice",
            "subtitle": "Adjust targeting, schedule, priority and app visibility.",
            "button_text": "Update Notice",
        },
    )


@institute_admin_required
def notice_delete(request, pk):
    institute = get_current_institute(request)
    queryset = Notice.objects.all()
    if institute:
        queryset = queryset.filter(institute=institute)
    notice = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        notice.delete()
        messages.success(request, "Notice deleted successfully.")

    return redirect("institute_admin:notice_list")


@institute_admin_required
def attendance_list(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    selected_date = parse_date(request.GET.get("date", "")) or date.today()
    batch_id = request.GET.get("batch", "").strip()

    batch_queryset = Batch.objects.filter(is_active=True).prefetch_related("courses")
    if institute:
        batch_queryset = batch_queryset.filter(institute=institute)
    if academic_year:
        batch_queryset = batch_queryset.filter(academic_year=academic_year)
    selected_batch = batch_queryset.filter(pk=batch_id).first() if batch_id else None

    student_sessions = StudentAcademicSession.objects.none()
    attendance_map = {}
    attendance_page = None
    display_student_sessions = StudentAcademicSession.objects.none()
    if selected_batch:
        student_sessions = (
            StudentAcademicSession.objects.filter(
                enrollments__batch=selected_batch,
                enrollments__status=StudentEnrollment.Status.ACTIVE,
                student__is_active=True,
                academic_year=academic_year,
            )
            .select_related("student", "student__user")
            .order_by("admission_number", "student__user__first_name", "student__user__username")
            .distinct()
        )

    if request.method == "POST" and selected_batch:
        saved_count = bulk_save_attendance(
            student_sessions=student_sessions,
            posted_student_ids=request.POST.getlist("student_ids"),
            form_data=request.POST,
            batch=selected_batch,
            attendance_date=selected_date,
            marked_by=request.user,
        )
        messages.success(request, f"Attendance saved for {saved_count} student(s).")
        return redirect(f"{reverse('institute_admin:attendance_list')}?batch={selected_batch.pk}&date={selected_date.isoformat()}")

    if selected_batch:
        attendance_page = Paginator(student_sessions, 100).get_page(request.GET.get("page"))
        display_student_sessions = attendance_page.object_list
        existing_attendance = Attendance.objects.filter(
            batch=selected_batch,
            date=selected_date,
            academic_session__in=display_student_sessions,
        ).select_related("academic_session", "student", "marked_by")
        attendance_map = {record.academic_session_id: record for record in existing_attendance}

    selected_date_records = Attendance.objects.filter(date=selected_date)
    all_records = Attendance.objects.all()
    if institute:
        selected_date_records = selected_date_records.filter(batch__institute=institute)
        all_records = all_records.filter(batch__institute=institute)
    if academic_year:
        selected_date_records = selected_date_records.filter(academic_session__academic_year=academic_year)
        all_records = all_records.filter(academic_session__academic_year=academic_year)
        all_records = all_records.filter(batch__academic_year=academic_year)
    if selected_batch:
        selected_date_records = selected_date_records.filter(batch=selected_batch)

    attendance_counts = selected_date_records.aggregate(
        total=Count("pk"),
        present=Count("pk", filter=Q(status=Attendance.Status.PRESENT)),
        absent=Count("pk", filter=Q(status=Attendance.Status.ABSENT)),
        late=Count("pk", filter=Q(status=Attendance.Status.LATE)),
    )
    total_today = attendance_counts["total"]
    present_today = attendance_counts["present"]
    absent_today = attendance_counts["absent"]
    late_today = attendance_counts["late"]
    rate_today = round((present_today / total_today) * 100, 1) if total_today else 0

    rows = []
    for student_session in display_student_sessions:
        record = attendance_map.get(student_session.pk)
        rows.append(
            {
                "student": student_session.student,
                "student_session": student_session,
                "record": record,
                "status": record.status if record else Attendance.Status.PRESENT,
                "note": record.note if record else "",
            }
        )

    context = {
        "batches": batch_queryset,
        "selected_batch": selected_batch,
        "selected_date": selected_date,
        "attendance_rows": rows,
        "attendance_page": attendance_page,
        "status_choices": Attendance.Status.choices,
        "present_value": Attendance.Status.PRESENT,
        "absent_value": Attendance.Status.ABSENT,
        "late_value": Attendance.Status.LATE,
        "total_students": student_sessions.count() if selected_batch else 0,
        "marked_count": len(attendance_map),
        "total_today": total_today,
        "present_today": present_today,
        "absent_today": absent_today,
        "late_today": late_today,
        "rate_today": rate_today,
        "recent_attendance": all_records.select_related("student", "student__user", "batch", "marked_by")[:8],
    }
    return render(request, "institute_admin/attendance_list.html", context)


def get_attendance_export_queryset(request, institute, academic_year=None):
    records = Attendance.objects.select_related(
        "academic_session", "student", "student__user", "batch", "marked_by"
    )
    if institute:
        records = records.filter(batch__institute=institute)
    if academic_year:
        records = records.filter(academic_session__academic_year=academic_year)

    batch_id = request.GET.get("batch", "").strip()
    student_id = request.GET.get("student", "").strip()
    status = request.GET.get("status", "").strip()
    date_from = parse_date(request.GET.get("date_from", ""))
    date_to = parse_date(request.GET.get("date_to", ""))

    if batch_id:
        records = records.filter(batch_id=batch_id)
    if student_id:
        records = records.filter(student_id=student_id)
    if status in Attendance.Status.values:
        records = records.filter(status=status)
    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)

    return records.order_by("-date", "batch__name", "academic_session__admission_number")


def attendance_export_filename(file_format):
    stamp = timezone.now().strftime("%Y%m%d_%H%M")
    extension = "xlsx" if file_format == "excel" else "pdf"
    return f"attendance_report_{stamp}.{extension}"


def attendance_status_counts(records):
    total = len(records)
    present = sum(1 for record in records if record.status == Attendance.Status.PRESENT)
    absent = sum(1 for record in records if record.status == Attendance.Status.ABSENT)
    late = sum(1 for record in records if record.status == Attendance.Status.LATE)
    rate = round((present / total) * 100, 1) if total else 0
    return total, present, absent, late, rate


def build_attendance_excel(records, institute, filters, include_notes=True):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    primary = "2563EB"
    dark = "0F172A"
    muted = "64748B"
    light_blue = "DBEAFE"
    light_green = "DCFCE7"
    light_red = "FEE2E2"
    light_amber = "FEF3C7"
    border_color = "CBD5E1"

    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    columns = ["Date", "Admission No", "Student Name", "Batch", "Status", "Marked By"]
    if include_notes:
        columns.append("Note")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1, value="Attendance Report")
    title_cell.font = Font(size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=primary)
    title_cell.alignment = center
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    institute_name = getattr(institute, "name", "All Institutes") if institute else "All Institutes"
    sheet.cell(row=2, column=1, value=f"{institute_name} | Generated {timezone.now().strftime('%d-%m-%Y %I:%M %p')}").font = Font(size=10, color=muted)
    sheet.cell(row=2, column=1).alignment = center

    total, present, absent, late, rate = attendance_status_counts(records)
    summary = [
        ("Total", total, light_blue),
        ("Present", present, light_green),
        ("Absent", absent, light_red),
        ("Late", late, light_amber),
        ("Rate", f"{rate}%", light_blue),
    ]
    col = 1
    for label, value, fill in summary:
        sheet.cell(row=4, column=col, value=label).font = Font(bold=True, color=dark)
        sheet.cell(row=4, column=col).fill = PatternFill("solid", fgColor=fill)
        sheet.cell(row=4, column=col).alignment = center
        sheet.cell(row=4, column=col).border = border
        sheet.cell(row=5, column=col, value=value).font = Font(size=14, bold=True, color=dark)
        sheet.cell(row=5, column=col).alignment = center
        sheet.cell(row=5, column=col).border = border
        col += 1

    filter_text = " | ".join(part for part in filters if part)
    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(columns))
    sheet.cell(row=7, column=1, value=f"Filters: {filter_text or 'All records'}").font = Font(italic=True, color=muted)

    header_row = 9
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border

    status_fills = {
        Attendance.Status.PRESENT: light_green,
        Attendance.Status.ABSENT: light_red,
        Attendance.Status.LATE: light_amber,
    }
    for row_index, record in enumerate(records, start=header_row + 1):
        values = [
            record.date.strftime("%d-%m-%Y"),
            record.academic_session.admission_number,
            record.student.user.get_full_name() or record.student.user.username,
            record.batch.name,
            record.get_status_display(),
            record.marked_by.get_full_name() if record.marked_by else "Not set",
        ]
        if include_notes:
            values.append(record.note or "")
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = left if col_index in (3, 4, len(values)) else center
            if col_index == 5:
                cell.fill = PatternFill("solid", fgColor=status_fills.get(record.status, "F8FAFC"))
                cell.font = Font(bold=True, color=dark)

    widths = [14, 16, 28, 24, 14, 24, 34]
    for index, width in enumerate(widths[: len(columns)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:{get_column_letter(len(columns))}{max(10, header_row + len(records))}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def pdf_escape(value):
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(x, y, text, size=9, font="F1", color="0 0 0"):
    return f"{color} rg BT /{font} {size} Tf {x} {y} Td ({pdf_escape(text)}) Tj ET\n"


def pdf_rect(x, y, w, h, color):
    return f"{color} rg {x} {y} {w} {h} re f\n"


def build_pdf_document(page_streams, width=842, height=595):
    objects = ["<< /Type /Catalog /Pages 2 0 R >>"]
    kids = []
    current_id = 3
    for stream in page_streams:
        page_id = current_id
        content_id = current_id + 1
        kids.append(f"{page_id} 0 R")
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> /F2 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >> >> >> /Contents {content_id} 0 R >>")
        encoded = stream.encode("latin-1", errors="replace")
        objects.append(f"<< /Length {len(encoded)} >>\nstream\n{stream}endstream")
        current_id += 2
    objects.insert(1, f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>")

    output = "%PDF-1.4\n"
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output.encode("latin-1", errors="replace")))
        output += f"{index} 0 obj\n{obj}\nendobj\n"
    xref_offset = len(output.encode("latin-1", errors="replace"))
    output += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for offset in offsets[1:]:
        output += f"{offset:010d} 00000 n \n"
    output += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
    return output.encode("latin-1", errors="replace")


def build_attendance_pdf(records, institute, filters, include_notes=True, title="Attendance Report"):
    width, height = 842, 595
    rows_per_page = 18 if include_notes else 22
    pages = []
    total, present, absent, late, rate = attendance_status_counts(records)
    institute_name = getattr(institute, "name", "All Institutes") if institute else "All Institutes"
    chunks = [records[index:index + rows_per_page] for index in range(0, len(records), rows_per_page)] or [[]]

    for page_number, chunk in enumerate(chunks, start=1):
        stream = ""
        stream += pdf_rect(0, 545, width, 50, "0.145 0.388 0.922")
        stream += pdf_text(34, 570, title, 18, "F2", "1 1 1")
        stream += pdf_text(34, 552, f"{institute_name} | Generated {timezone.now().strftime('%d-%m-%Y %I:%M %p')}", 8, "F1", "1 1 1")
        stream += pdf_text(700, 552, f"Page {page_number} of {len(chunks)}", 8, "F1", "1 1 1")

        summary = [("Total", total), ("Present", present), ("Absent", absent), ("Late", late), ("Rate", f"{rate}%")]
        x = 34
        for label, value in summary:
            stream += pdf_rect(x, 500, 128, 32, "0.93 0.96 1")
            stream += pdf_text(x + 8, 519, label, 7, "F1", "0.39 0.45 0.55")
            stream += pdf_text(x + 8, 506, value, 12, "F2", "0.06 0.09 0.16")
            x += 138

        filter_text = " | ".join(part for part in filters if part) or "All records"
        stream += pdf_text(34, 480, f"Filters: {filter_text[:135]}", 8, "F1", "0.39 0.45 0.55")

        headers = ["Date", "Adm No", "Student", "Batch", "Status", "Marked By"]
        widths = [68, 76, 168, 142, 76, 126]
        if include_notes:
            headers.append("Note")
            widths.append(118)
        start_x = 34
        y = 452
        stream += pdf_rect(start_x, y - 6, sum(widths), 22, "0.06 0.09 0.16")
        x = start_x
        for header, col_width in zip(headers, widths):
            stream += pdf_text(x + 4, y, header, 8, "F2", "1 1 1")
            x += col_width
        y -= 24

        for record in chunk:
            status_color = "0.86 0.99 0.91" if record.status == Attendance.Status.PRESENT else "1 0.89 0.89" if record.status == Attendance.Status.ABSENT else "1 0.95 0.78"
            stream += pdf_rect(start_x, y - 5, sum(widths), 20, "0.98 0.99 1")
            row_values = [
                record.date.strftime("%d-%m-%Y"),
                record.academic_session.admission_number,
                (record.student.user.get_full_name() or record.student.user.username)[:28],
                record.batch.name[:22],
                record.get_status_display(),
                (record.marked_by.get_full_name() if record.marked_by else "Not set")[:20],
            ]
            if include_notes:
                row_values.append((record.note or "")[:24])
            x = start_x
            for index, (value, col_width) in enumerate(zip(row_values, widths)):
                if index == 4:
                    stream += pdf_rect(x + 2, y - 3, col_width - 4, 15, status_color)
                    stream += pdf_text(x + 5, y, value, 7, "F2", "0.06 0.09 0.16")
                else:
                    stream += pdf_text(x + 4, y, value, 7, "F1", "0.06 0.09 0.16")
                x += col_width
            y -= 21

        stream += pdf_text(34, 28, "UltraCoachMatrix attendance export", 8, "F1", "0.39 0.45 0.55")
        pages.append(stream)

    return build_pdf_document(pages, width=width, height=height)


@institute_admin_required
def attendance_export(request):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    file_format = request.GET.get("format", "excel").strip()
    include_notes = request.GET.get("include_notes", "1") == "1"
    records = list(get_attendance_export_queryset(request, institute, academic_year))

    batch_label = "All Batches"
    batch_id = request.GET.get("batch", "").strip()
    if batch_id:
        batch_queryset = Batch.objects.filter(pk=batch_id)
        if institute:
            batch_queryset = batch_queryset.filter(institute=institute)
        if academic_year:
            batch_queryset = batch_queryset.filter(academic_year=academic_year)
        batch = batch_queryset.first()
        batch_label = batch.name if batch else batch_label
    student_label = "All Students"
    student_id = request.GET.get("student", "").strip()
    if student_id:
        session_queryset = StudentAcademicSession.objects.select_related("student", "student__user")
        if institute:
            session_queryset = session_queryset.filter(institute=institute)
        if academic_year:
            session_queryset = session_queryset.filter(academic_year=academic_year)
        student_session = session_queryset.filter(student_id=student_id).first()
        if student_session:
            student_label = str(student_session)
    status_value = request.GET.get("status", "").strip()
    status_label = dict(Attendance.Status.choices).get(status_value, "All Status")
    filters = [
        f"Batch: {batch_label}",
        f"Student: {student_label}",
        f"Status: {status_label}",
        f"From: {request.GET.get('date_from') or 'Any'}",
        f"To: {request.GET.get('date_to') or 'Any'}",
    ]

    if file_format == "pdf":
        content = build_attendance_pdf(records, institute, filters, include_notes=include_notes)
        response = HttpResponse(content, content_type="application/pdf")
    else:
        content = build_attendance_excel(records, institute, filters, include_notes=include_notes)
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file_format = "excel"
    response["Content-Disposition"] = f'attachment; filename="{attendance_export_filename(file_format)}"'
    return response


@institute_admin_required
def teacher_list(request):
    institute = get_current_institute(request)
    teachers = TeacherProfile.objects.select_related("institute", "user", "user__profile")
    if institute:
        teachers = teachers.filter(institute=institute)
    teachers = teachers.filter(user__profile__role=UserProfile.Role.TEACHER)

    search_query = request.GET.get("search", "").strip()
    teacher_type_filter = request.GET.get("teacher_type", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if search_query:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(employee_id__icontains=search_query)
            | Q(specialization__icontains=search_query)
            | Q(user__profile__phone__icontains=search_query)
        )

    if teacher_type_filter:
        teachers = teachers.filter(teacher_type=teacher_type_filter)

    if status_filter == "active":
        teachers = teachers.filter(is_active=True)
    elif status_filter == "inactive":
        teachers = teachers.filter(is_active=False)

    base_queryset = TeacherProfile.objects.filter(institute=institute) if institute else TeacherProfile.objects.all()
    page_obj, paginator, pagination_query = paginate_queryset(request, teachers)
    context = {
        "teachers": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "teachers",
        "search_query": search_query,
        "teacher_type_filter": teacher_type_filter,
        "status_filter": status_filter,
        "total_teachers": base_queryset.count(),
        "active_teachers": base_queryset.filter(is_active=True).count(),
        "full_time_teachers": base_queryset.filter(teacher_type=TeacherProfile.TeacherType.FULL_TIME).count(),
        "part_time_teachers": base_queryset.filter(teacher_type=TeacherProfile.TeacherType.PART_TIME).count(),
    }
    return render(request, "institute_admin/teacher_list.html", context)


@institute_admin_required
def teacher_create(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "Select an institute before creating a teacher.")
        return redirect("institute_admin:teacher_list")

    if request.method == "POST":
        form = TeacherForm(request.POST, institute=institute)
        if form.is_valid():
            teacher = form.save()
            on_commit_email(
                send_teacher_welcome,
                teacher.pk,
                form.cleaned_data["password"],
            )
            messages.success(request, "Teacher created successfully.")
            return close_popup_response()
    else:
        form = TeacherForm(institute=institute)

    return render(
        request,
        "institute_admin/teacher_form.html",
        {
            "form": form,
            "title": "Create Teacher",
            "subtitle": "Add teacher identity, contact and workload details.",
            "button_text": "Save Teacher",
        },
    )


@institute_admin_required
def teacher_update(request, pk):
    institute = get_current_institute(request)
    queryset = TeacherProfile.objects.select_related("user", "institute")
    if institute:
        queryset = queryset.filter(institute=institute)
    teacher = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        form = TeacherForm(request.POST, institute=teacher.institute, teacher=teacher)
        if form.is_valid():
            form.save()
            messages.success(request, "Teacher updated successfully.")
            return close_popup_response()
    else:
        form = TeacherForm(institute=teacher.institute, teacher=teacher)

    return render(
        request,
        "institute_admin/teacher_form.html",
        {
            "form": form,
            "title": "Edit Teacher",
            "subtitle": "Update teacher identity, contact and workload details.",
            "button_text": "Update Teacher",
        },
    )


@institute_admin_required
def teacher_delete(request, pk):
    institute = get_current_institute(request)
    queryset = TeacherProfile.objects.select_related("user")
    if institute:
        queryset = queryset.filter(institute=institute)
    teacher = get_object_or_404(queryset, pk=pk)

    if request.method == "POST":
        if teacher.user.assigned_batches.exists():
            messages.error(request, "This teacher is assigned to batches. Remove them from batches before deleting.")
        else:
            teacher.user.delete()
            messages.success(request, "Teacher deleted successfully.")

    return redirect("institute_admin:teacher_list")
















BULK_QUESTION_HEADERS = [
    "Question Text",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Correct Answer",
    "Marks",
]
BULK_ANSWER_TO_ORDER = {"A": 1, "B": 2, "C": 3, "D": 4}


def close_institute_exam_popup_response(fallback_url="/institute/exams/"):
    return HttpResponse(
        f"""
        <script>
            if (window.opener) {{
                window.opener.location.reload();
                window.close();
            }} else {{
                window.location.href = "{fallback_url}";
            }}
        </script>
        """
    )


def institute_exam_batches(request):
    institute = get_current_institute(request)
    selected_year = get_current_academic_year(request, institute)
    if not institute or not selected_year:
        return Batch.objects.none()
    return Batch.objects.filter(institute=institute, academic_year=selected_year)


def institute_exam_queryset(request):
    return Exam.objects.filter(batch__in=institute_exam_batches(request)).select_related("batch", "academic_year", "subject")



@institute_admin_required
def exams(request):
    batches = institute_exam_batches(request)
    selected_year = get_current_academic_year(request)
    exam_qs = (
        Exam.objects.filter(batch__in=batches)
        .select_related("batch", "academic_year", "subject")
        .annotate(question_count=Count("questions", distinct=True), marks_from_questions=Sum("questions__marks"))
        .order_by("-exam_date", "title", "pk")
    )
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    batch_filter = request.GET.get("batch", "").strip()
    if search_query:
        exam_qs = exam_qs.filter(Q(title__icontains=search_query) | Q(batch__name__icontains=search_query) | Q(subject__name__icontains=search_query))
    if status_filter == "published":
        exam_qs = exam_qs.filter(is_published=True)
    elif status_filter == "draft":
        exam_qs = exam_qs.filter(is_published=False)
    if batch_filter:
        exam_qs = exam_qs.filter(batch_id=batch_filter)
    base_qs = Exam.objects.filter(batch__in=batches)
    page_obj, paginator, pagination_query = paginate_queryset(request, exam_qs)
    return render(
        request,
        "exam/institute_exams.html",
        {
            "exams": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "pagination_query": pagination_query,
            "pagination_label": "exams",
            "batches": batches,
            "selected_academic_year": selected_year,
            "search_query": search_query,
            "status_filter": status_filter,
            "batch_filter": batch_filter,
            "total_exams": base_qs.count(),
            "published_exams": base_qs.filter(is_published=True).count(),
            "draft_exams": base_qs.filter(is_published=False).count(),
            "total_attempts": ExamAttempt.objects.filter(exam__in=base_qs, submitted_at__isnull=False).count(),
        },
    )


@institute_admin_required
def exam_create(request):
    batches = institute_exam_batches(request)
    form = TeacherExamForm(request.POST or None, batches=batches)
    if request.method == "POST" and form.is_valid():
        exam = form.save(commit=False)
        exam.academic_year = exam.batch.academic_year
        exam.created_by = request.user
        exam.total_marks = 0
        exam.save()
        messages.success(request, "Exam created successfully.")
        return close_institute_exam_popup_response()
    return render(
        request,
        "exam/institute_exam_form.html",
        {
            "form": form,
            "title": "Create Exam",
            "subtitle": "Set exam details for the selected academic year.",
            "button_text": "Save Exam",
        },
    )


@institute_admin_required
def exam_update(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    form = TeacherExamForm(request.POST or None, instance=exam, batches=institute_exam_batches(request))
    if request.method == "POST" and form.is_valid():
        exam = form.save(commit=False)
        exam.academic_year = exam.batch.academic_year
        exam.save()
        messages.success(request, "Exam updated successfully.")
        return close_institute_exam_popup_response()
    return render(
        request,
        "exam/institute_exam_form.html",
        {
            "form": form,
            "title": "Edit Exam",
            "subtitle": "Update exam schedule, instructions and publish settings.",
            "button_text": "Update Exam",
        },
    )

def sync_exam_total_marks(exam):
    total = exam.questions.aggregate(total=Sum("marks")).get("total") or 0
    Exam.objects.filter(pk=exam.pk).update(total_marks=total)
    exam.total_marks = total


def recalculate_exam_attempt(attempt):
    questions = list(attempt.exam.questions.prefetch_related("options"))
    answer_map = {
        row.question_id: row
        for row in attempt.question_attempts.select_related("selected_option")
    }
    score = 0
    correct_count = 0
    wrong_count = 0
    unattempted_count = 0
    total_marks = sum(question.marks for question in questions)
    for question in questions:
        answer = answer_map.get(question.pk)
        selected_option = answer.selected_option if answer else None
        if selected_option and selected_option.question_id != question.pk:
            selected_option = None
        is_correct = bool(selected_option and selected_option.is_correct)
        marks_awarded = question.marks if is_correct else 0
        if selected_option is None:
            unattempted_count += 1
        elif is_correct:
            correct_count += 1
            score += question.marks
        else:
            wrong_count += 1
        ExamQuestionAttempt.objects.update_or_create(
            attempt=attempt,
            question=question,
            defaults={
                "selected_option": selected_option,
                "is_correct": is_correct,
                "marks_awarded": marks_awarded,
            },
        )
    attempt.score = score
    attempt.total_marks = total_marks
    attempt.correct_count = correct_count
    attempt.wrong_count = wrong_count
    attempt.unattempted_count = unattempted_count
    attempt.save(
        update_fields=[
            "score",
            "total_marks",
            "correct_count",
            "wrong_count",
            "unattempted_count",
        ]
    )
    if attempt.is_submitted:
        ExamResult.objects.update_or_create(
            exam=attempt.exam,
            student=attempt.student,
            defaults={
                "marks_obtained": score,
                "remark": "Updated by teacher from exam submission management.",
            },
        )
    return attempt


def option_formset_has_one_correct(formset):
    option_rows = [
        data
        for data in formset.cleaned_data
        if data and not data.get("DELETE") and data.get("text")
    ]
    return len(option_rows) == 4 and sum(1 for data in option_rows if data.get("is_correct")) == 1


@institute_admin_required
def exam_questions(request, pk):
    exam = get_object_or_404(
        institute_exam_queryset(request).prefetch_related("questions", "questions__options"),
        pk=pk,
    )
    return render(request, "exam/institute_exam_questions.html", {"exam": exam})


@institute_admin_required
def exam_question_create(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    question = ExamQuestion(exam=exam, order=exam.questions.count() + 1)
    form = ExamQuestionForm(request.POST or None, request.FILES or None, instance=question)
    formset = ExamQuestionOptionFormSet(request.POST or None, instance=question, prefix="options")
    if request.method == "POST" and form.is_valid() and formset.is_valid():
        if not option_formset_has_one_correct(formset):
            messages.error(request, "Add exactly four options and select exactly one correct option.")
        else:
            question = form.save(commit=False)
            question.exam = exam
            question.order = exam.questions.count() + 1
            question.save()
            formset.instance = question
            options = formset.save(commit=False)
            for index, option in enumerate(options, start=1):
                option.question = question
                option.order = index
                option.save()
            for deleted in formset.deleted_objects:
                deleted.delete()
            sync_exam_total_marks(exam)
            messages.success(request, "Question added successfully.")
            return close_institute_exam_popup_response(reverse("institute_admin:institute_exam_questions", args=[exam.pk]))
    return render(
        request,
        "exam/institute_exam_question_form.html",
        {"exam": exam, "form": form, "formset": formset, "title": "Add Question"},
    )


@institute_admin_required
def exam_question_update(request, exam_pk, question_pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=exam_pk)
    question = get_object_or_404(ExamQuestion.objects.filter(exam=exam), pk=question_pk)
    form = ExamQuestionForm(request.POST or None, request.FILES or None, instance=question)
    formset = ExamQuestionOptionFormSet(request.POST or None, instance=question, prefix="options")
    if request.method == "POST" and form.is_valid() and formset.is_valid():
        if not option_formset_has_one_correct(formset):
            messages.error(request, "Add exactly four options and select exactly one correct option.")
        else:
            question = form.save()
            options = formset.save(commit=False)
            for deleted in formset.deleted_objects:
                deleted.delete()
            for index, option in enumerate(options, start=1):
                option.question = question
                option.order = index
                option.save()
            sync_exam_total_marks(exam)
            messages.success(request, "Question updated successfully.")
            return close_institute_exam_popup_response(reverse("institute_admin:institute_exam_questions", args=[exam.pk]))
    return render(
        request,
        "exam/institute_exam_question_form.html",
        {"exam": exam, "form": form, "formset": formset, "title": "Edit Question"},
    )


def normalize_bulk_cell(value):
    if value is None:
        return ""
    return str(value).strip()


@institute_admin_required
def exam_question_import_template(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Questions"
    worksheet.append(BULK_QUESTION_HEADERS)

    sample_rows = [
        [
            "A body is moving with uniform velocity. Its acceleration is?",
            "Zero",
            "Positive",
            "Negative",
            "Variable",
            "A",
            1,
        ],
        [
            "If x + 5 = 12, then x equals?",
            "5",
            "6",
            "7",
            "8",
            "C",
            1,
        ],
        [
            "Which option best represents the SI unit of force?",
            "Joule",
            "Newton",
            "Watt",
            "Pascal",
            "B",
            1,
        ],
    ]
    for row in sample_rows:
        worksheet.append(row)

    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_index, width in enumerate([48, 28, 28, 28, 28, 18, 12], start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    answer_validation = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
    answer_validation.error = "Select only A, B, C, or D as the correct answer."
    answer_validation.errorTitle = "Invalid answer"
    answer_validation.prompt = "Choose the correct option."
    answer_validation.promptTitle = "Correct Answer"
    worksheet.add_data_validation(answer_validation)
    answer_validation.add("F2:F500")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"bulk-question-template-{exam.pk}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@institute_admin_required
def exam_question_bulk_import(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    if request.method != "POST":
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)

    upload = request.FILES.get("question_file")
    if not upload:
        messages.error(request, "Please select the completed question template.")
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)
    if not upload.name.lower().endswith(".xlsx"):
        messages.error(request, "Please upload the Excel template in .xlsx format.")
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "Unable to read the uploaded Excel file. Please download the template and try again.")
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)

    worksheet = workbook.active
    imported_rows = []
    errors = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [normalize_bulk_cell(value) for value in row[:7]]
        values += [""] * (7 - len(values))
        question_text, option_a, option_b, option_c, option_d, correct_answer, marks_value = values

        if not any(values):
            continue

        row_errors = []
        if not question_text:
            row_errors.append("question text is required")
        options = [option_a, option_b, option_c, option_d]
        if any(not option for option in options):
            row_errors.append("all four options are required")
        correct_answer = correct_answer.upper()
        if correct_answer not in BULK_ANSWER_TO_ORDER:
            row_errors.append("correct answer must be A, B, C, or D")
        try:
            marks = int(float(marks_value or 1))
        except (TypeError, ValueError):
            marks = 0
        if marks < 1:
            row_errors.append("marks must be 1 or higher")

        if row_errors:
            errors.append(f"Row {row_number}: {', '.join(row_errors)}.")
            continue

        imported_rows.append(
            {
                "text": question_text,
                "options": options,
                "correct_order": BULK_ANSWER_TO_ORDER[correct_answer],
                "marks": marks,
            }
        )

    if errors:
        messages.error(request, "Bulk import failed. Fix these rows and upload again: " + " ".join(errors[:8]))
        if len(errors) > 8:
            messages.error(request, f"{len(errors) - 8} more row(s) also need correction.")
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)
    if not imported_rows:
        messages.error(request, "No question rows found in the uploaded template.")
        return redirect("institute_admin:institute_exam_questions", pk=exam.pk)

    start_order = exam.questions.count() + 1
    with transaction.atomic():
        for offset, row in enumerate(imported_rows):
            question = ExamQuestion.objects.create(
                exam=exam,
                text=row["text"],
                marks=row["marks"],
                order=start_order + offset,
            )
            ExamQuestionOption.objects.bulk_create(
                [
                    ExamQuestionOption(
                        question=question,
                        text=option_text,
                        order=option_order,
                        is_correct=option_order == row["correct_order"],
                    )
                    for option_order, option_text in enumerate(row["options"], start=1)
                ]
            )
        sync_exam_total_marks(exam)

    messages.success(request, f"{len(imported_rows)} question(s) imported successfully.")
    return redirect("institute_admin:institute_exam_questions", pk=exam.pk)


@institute_admin_required
def exam_submissions(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    enrollments = (
        StudentEnrollment.objects.filter(
            batch=exam.batch,
            academic_session__academic_year=exam.academic_year,
            academic_session__status=StudentAcademicSession.Status.ACTIVE,
            status=StudentEnrollment.Status.ACTIVE,
            student__is_active=True,
        )
        .select_related("academic_session", "student", "student__user")
        .order_by("academic_session__admission_number", "student__user__first_name", "student__user__username")
    )
    attempts = list(
        ExamAttempt.objects.filter(exam=exam)
        .select_related("student", "student__user", "academic_session")
        .prefetch_related("uploads")
    )
    attempts_by_session = {attempt.academic_session_id: attempt for attempt in attempts}
    rows = []
    seen_session_ids = set()

    for enrollment in enrollments:
        attempt = attempts_by_session.get(enrollment.academic_session_id)
        seen_session_ids.add(enrollment.academic_session_id)
        rows.append(
            {
                "academic_session": enrollment.academic_session,
                "student": enrollment.student,
                "attempt": attempt,
                "upload_count": len(attempt.uploads.all()) if attempt else 0,
                "status": "Attempted" if attempt and attempt.is_submitted else "In Progress" if attempt else "Not Attempted",
            }
        )

    for attempt in attempts:
        if attempt.academic_session_id in seen_session_ids:
            continue
        rows.append(
            {
                "academic_session": attempt.academic_session,
                "student": attempt.student,
                "attempt": attempt,
                "upload_count": len(attempt.uploads.all()),
                "status": "Attempted" if attempt.is_submitted else "In Progress",
            }
        )

    attempted_count = sum(1 for row in rows if row["attempt"] and row["attempt"].is_submitted)
    in_progress_count = sum(1 for row in rows if row["attempt"] and not row["attempt"].is_submitted)
    not_attempted_count = sum(1 for row in rows if not row["attempt"])
    return render(
        request,
        "exam/institute_exam_submissions.html",
        {
            "exam": exam,
            "rows": rows,
            "total_students": len(rows),
            "attempted_count": attempted_count,
            "in_progress_count": in_progress_count,
            "not_attempted_count": not_attempted_count,
        },
    )


@institute_admin_required
def exam_attempt_manage(request, exam_pk, attempt_pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=exam_pk)
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related("student", "student__user", "academic_session", "exam")
        .prefetch_related("question_attempts", "uploads", "activities"),
        pk=attempt_pk,
        exam=exam,
    )
    questions = list(exam.questions.prefetch_related("options"))

    if request.method == "POST":
        action = request.POST.get("action", "save_answers")
        with transaction.atomic():
            for question in questions:
                selected_option = None
                option_id = request.POST.get(f"answer_{question.pk}")
                if option_id:
                    selected_option = question.options.filter(pk=option_id).first()
                ExamQuestionAttempt.objects.update_or_create(
                    attempt=attempt,
                    question=question,
                    defaults={"selected_option": selected_option},
                )
            if action == "force_submit" and not attempt.is_submitted:
                attempt.submitted_at = timezone.now()
                attempt.save(update_fields=["submitted_at"])
            recalculate_exam_attempt(attempt)
        if action == "force_submit":
            messages.success(request, "Attempt answers saved and marked as submitted.")
        else:
            messages.success(request, "Student answers updated and score recalculated.")
        return redirect("institute_admin:institute_exam_attempt_manage", exam_pk=exam.pk, attempt_pk=attempt.pk)

    recalculate_exam_attempt(attempt)
    attempts_by_question = {
        answer.question_id: answer
        for answer in attempt.question_attempts.select_related("selected_option")
    }
    question_rows = [
        {
            "question": question,
            "answer": attempts_by_question.get(question.pk),
            "uploads": [upload for upload in attempt.uploads.all() if upload.question_id == question.pk],
        }
        for question in questions
    ]
    unlinked_uploads = [upload for upload in attempt.uploads.all() if not upload.question_id]
    return render(
        request,
        "exam/institute_exam_attempt_manage.html",
        {
            "exam": exam,
            "attempt": attempt,
            "question_rows": question_rows,
            "unlinked_uploads": unlinked_uploads,
        },
    )


@institute_admin_required
@require_POST
def exam_attempt_reset(request, exam_pk, attempt_pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=exam_pk)
    attempt = get_object_or_404(ExamAttempt.objects.select_related("student"), pk=attempt_pk, exam=exam)
    student_name = attempt.student.user.get_full_name() or attempt.student.user.username
    ExamResult.objects.filter(exam=exam, student=attempt.student).delete()
    attempt.delete()
    messages.success(request, f"Attempt reset for {student_name}. The student can attend this exam again.")
    return redirect("institute_admin:institute_exam_submissions", pk=exam.pk)


@institute_admin_required
@require_POST
def exam_publish(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    if exam.is_published:
        messages.info(request, "Exam is already published to students.")
    else:
        exam.is_published = True
        exam.save(update_fields=["is_published"])
        messages.success(request, "Exam published successfully. Students can now view and attempt this exam.")
    return redirect("institute_admin:institute_exam_submissions", pk=exam.pk)


@institute_admin_required
@require_POST
def exam_toggle_result_publish(request, pk):
    exam = get_object_or_404(institute_exam_queryset(request), pk=pk)
    action = request.POST.get("action")
    if action == "publish":
        if exam.show_result_after_submit:
            messages.info(request, "Exam results are already published to students.")
        else:
            exam.show_result_after_submit = True
            exam.save(update_fields=["show_result_after_submit"])
            transaction.on_commit(
                lambda exam_id=exam.pk: notify_exam_results_declared(exam_id)
            )
            messages.success(request, "Exam results published successfully. Students can now view their scores.")
    elif action == "hide":
        exam.show_result_after_submit = False
        exam.save(update_fields=["show_result_after_submit"])
        messages.success(request, "Exam results hidden successfully. Students will not see scores until you publish again.")
    else:
        messages.error(request, "Invalid result visibility action.")
    return redirect("institute_admin:institute_exam_submissions", pk=exam.pk)


@institute_admin_required
def results(request):
    report = get_exam_result_report(request, paginate=True)
    return render(request, "exam/institute_results.html", report)


def exam_attempt_percentage(attempt):
    if not attempt.total_marks:
        return Decimal("0.00")
    return (attempt.score / attempt.total_marks * Decimal("100")).quantize(Decimal("0.01"))


def ranked_exam_attempts(attempts):
    ranks = {}
    attempts_by_exam = {}
    for attempt in attempts:
        attempts_by_exam.setdefault(attempt.exam_id, []).append(attempt)

    for exam_attempts in attempts_by_exam.values():
        previous_score = None
        current_rank = 0
        for position, attempt in enumerate(
            sorted(exam_attempts, key=lambda item: (-item.score, item.submitted_at, item.pk)),
            start=1,
        ):
            if previous_score is None or attempt.score != previous_score:
                current_rank = position
                previous_score = attempt.score
            ranks[attempt.pk] = current_rank
    return ranks


def get_exam_attempt_percentage_expression():
    return Case(
        When(
            total_marks__gt=0,
            then=ExpressionWrapper(
                F("score") * Value(Decimal("100.00")) / F("total_marks"),
                output_field=DecimalField(max_digits=7, decimal_places=2),
            ),
        ),
        default=Value(Decimal("0.00")),
        output_field=DecimalField(max_digits=7, decimal_places=2),
    )


def get_ranked_exam_attempt_queryset(exams):
    higher_score_count = (
        ExamAttempt.objects.filter(
            exam_id=OuterRef("exam_id"),
            submitted_at__isnull=False,
            score__gt=OuterRef("score"),
        )
        .values("exam_id")
        .annotate(total=Count("pk"))
        .values("total")
    )
    rank_expression = ExpressionWrapper(
        Coalesce(Subquery(higher_score_count, output_field=IntegerField()), Value(0)) + Value(1),
        output_field=IntegerField(),
    )
    return (
        ExamAttempt.objects.filter(exam__in=exams, submitted_at__isnull=False)
        .select_related(
            "exam",
            "exam__batch",
            "exam__course",
            "exam__subject",
            "academic_session",
            "student",
            "student__user",
        )
        .annotate(
            percentage=get_exam_attempt_percentage_expression(),
            rank=rank_expression,
        )
    )


def build_exam_result_rows(attempts):
    rows = []
    for attempt in attempts:
        percentage = getattr(attempt, "percentage", None)
        if percentage is None:
            percentage = exam_attempt_percentage(attempt)
        else:
            percentage = percentage.quantize(Decimal("0.01"))
        student_name = attempt.student.user.get_full_name() or attempt.student.user.username
        rows.append(
            {
                "attempt": attempt,
                "student_name": student_name,
                "percentage": percentage,
                "rank": getattr(attempt, "rank", None),
                "performance": "Passed" if percentage >= Decimal("40") else "Needs improvement",
            }
        )
    return rows


def get_exam_result_report(request, paginate=False):
    institute = get_current_institute(request)
    academic_year = get_current_academic_year(request, institute)
    batches = institute_exam_batches(request).order_by("name")
    exams = Exam.objects.filter(batch__in=batches).select_related(
        "batch",
        "course",
        "subject",
        "academic_year",
    )

    batch_id = request.GET.get("batch", "").strip()
    course_id = request.GET.get("course", "").strip()
    subject_id = request.GET.get("subject", "").strip()
    exam_id = request.GET.get("exam", "").strip()
    student_query = request.GET.get("student", "").strip()
    date_from = parse_date(request.GET.get("date_from", "").strip())
    date_to = parse_date(request.GET.get("date_to", "").strip())
    performance = request.GET.get("performance", "").strip()
    min_percentage = request.GET.get("min_percentage", "").strip()
    max_percentage = request.GET.get("max_percentage", "").strip()

    try:
        minimum = Decimal(min_percentage) if min_percentage else None
    except (ValueError, ArithmeticError):
        minimum = None
    try:
        maximum = Decimal(max_percentage) if max_percentage else None
    except (ValueError, ArithmeticError):
        maximum = None

    attempts = get_ranked_exam_attempt_queryset(exams)
    if batch_id:
        attempts = attempts.filter(exam__batch_id=batch_id)
    if course_id:
        attempts = attempts.filter(exam__course_id=course_id)
    if subject_id:
        attempts = attempts.filter(exam__subject_id=subject_id)
    if exam_id:
        attempts = attempts.filter(exam_id=exam_id)
    if student_query:
        attempts = attempts.filter(
            Q(student__user__first_name__icontains=student_query)
            | Q(student__user__last_name__icontains=student_query)
            | Q(student__user__username__icontains=student_query)
            | Q(academic_session__admission_number__icontains=student_query)
        )
    if date_from:
        attempts = attempts.filter(submitted_at__date__gte=date_from)
    if date_to:
        attempts = attempts.filter(submitted_at__date__lte=date_to)
    if minimum is not None:
        attempts = attempts.filter(percentage__gte=minimum)
    if maximum is not None:
        attempts = attempts.filter(percentage__lte=maximum)
    if performance == "passed":
        attempts = attempts.filter(percentage__gte=Decimal("40"))
    elif performance == "failed":
        attempts = attempts.filter(percentage__lt=Decimal("40"))

    attempts = attempts.order_by("exam__title", "rank", "student__user__first_name", "student__user__username", "pk")
    result_count = attempts.count()
    passed_count = attempts.filter(percentage__gte=Decimal("40")).count()
    aggregates = attempts.aggregate(
        average_percentage=Avg("percentage"),
        highest_percentage=Max("percentage"),
    )

    page_obj = paginator = pagination_query = None
    rows_queryset = attempts
    if paginate:
        page_obj, paginator, pagination_query = paginate_queryset(request, attempts)
        rows_queryset = page_obj.object_list
    rows = build_exam_result_rows(rows_queryset)

    courses = Course.objects.filter(institute=institute, academic_year=academic_year, exams__in=exams).distinct().order_by("name")
    subjects = Subject.objects.filter(institute=institute, academic_year=academic_year, exams__in=exams).distinct().order_by("name")
    filters = {
        "batch": batch_id,
        "course": course_id,
        "subject": subject_id,
        "exam": exam_id,
        "student": student_query,
        "date_from": request.GET.get("date_from", "").strip(),
        "date_to": request.GET.get("date_to", "").strip(),
        "performance": performance,
        "min_percentage": min_percentage,
        "max_percentage": max_percentage,
    }
    filter_labels = exam_result_filter_labels(filters, batches, courses, subjects, exams)

    return {
        "institute": institute,
        "academic_year": academic_year,
        "rows": rows,
        "batches": batches,
        "courses": courses,
        "subjects": subjects,
        "exams": exams.order_by("-exam_date", "title"),
        "filters": filters,
        "filter_labels": filter_labels,
        "result_count": result_count,
        "passed_count": passed_count,
        "needs_improvement_count": result_count - passed_count,
        "average_percentage": (aggregates["average_percentage"] or Decimal("0.00")).quantize(Decimal("0.01")),
        "highest_percentage": (aggregates["highest_percentage"] or Decimal("0.00")).quantize(Decimal("0.01")),
        "page_obj": page_obj,
        "paginator": paginator,
        "pagination_query": pagination_query,
        "pagination_label": "results",
    }


def exam_result_filter_labels(filters, batches, courses, subjects, exams):
    def selected_label(queryset, selected_id, fallback):
        if not selected_id:
            return fallback
        item = queryset.filter(pk=selected_id).first()
        return str(item) if item else fallback

    performance_labels = {
        "passed": "Passed (40% and above)",
        "failed": "Needs improvement (below 40%)",
    }
    return [
        f"Batch: {selected_label(batches, filters['batch'], 'All')}",
        f"Course: {selected_label(courses, filters['course'], 'All')}",
        f"Subject: {selected_label(subjects, filters['subject'], 'All')}",
        f"Exam: {selected_label(exams, filters['exam'], 'All')}",
        f"Student: {filters['student'] or 'All'}",
        f"Submitted: {filters['date_from'] or 'Any'} to {filters['date_to'] or 'Any'}",
        f"Percentage: {filters['min_percentage'] or '0'} to {filters['max_percentage'] or '100'}",
        f"Performance: {performance_labels.get(filters['performance'], 'All')}",
    ]


def build_exam_results_excel(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exam Results"
    columns = [
        "Rank", "Exam", "Exam Date", "Student", "Admission No.", "Batch",
        "Course", "Subject", "Score", "Total Marks", "Percentage",
        "Correct", "Wrong", "Unattempted", "Performance", "Submitted At",
    ]
    primary, dark, muted, border_color = "7C3AED", "0F172A", "64748B", "CBD5E1"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title = sheet.cell(row=1, column=1, value="Exam Results Report")
    title.font = Font(size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=primary)
    title.alignment = center
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sheet.cell(
        row=2,
        column=1,
        value=(
            f"{report['institute'].name} | Academic Year: {report['academic_year']} | "
            f"Generated {timezone.localtime().strftime('%d-%m-%Y %I:%M %p')}"
        ),
    ).alignment = center

    summary = [
        ("Results", report["result_count"]),
        ("Passed", report["passed_count"]),
        ("Needs Improvement", report["needs_improvement_count"]),
        ("Average", f"{report['average_percentage']}%"),
        ("Highest", f"{report['highest_percentage']}%"),
    ]
    for index, (label, value) in enumerate(summary, start=1):
        sheet.cell(row=4, column=index, value=label).font = Font(bold=True, color=dark)
        sheet.cell(row=4, column=index).fill = PatternFill("solid", fgColor="EDE9FE")
        sheet.cell(row=4, column=index).alignment = center
        sheet.cell(row=5, column=index, value=value).font = Font(size=13, bold=True, color=dark)
        sheet.cell(row=5, column=index).alignment = center

    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(columns))
    sheet.cell(row=7, column=1, value="Filters: " + " | ".join(report["filter_labels"])).font = Font(italic=True, color=muted)

    header_row = 9
    for index, label in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=label)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(report["rows"], start=header_row + 1):
        attempt = row["attempt"]
        values = [
            row["rank"],
            attempt.exam.title,
            attempt.exam.exam_date.strftime("%d-%m-%Y"),
            row["student_name"],
            attempt.academic_session.admission_number,
            attempt.exam.batch.name,
            attempt.exam.course.name if attempt.exam.course else "",
            attempt.exam.subject.name if attempt.exam.subject else "",
            float(attempt.score),
            float(attempt.total_marks),
            float(row["percentage"]),
            attempt.correct_count,
            attempt.wrong_count,
            attempt.unattempted_count,
            row["performance"],
            timezone.localtime(attempt.submitted_at).strftime("%d-%m-%Y %I:%M %p"),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = center
            if col_index == 15:
                cell.fill = PatternFill(
                    "solid",
                    fgColor="DCFCE7" if row["performance"] == "Passed" else "FEE2E2",
                )
                cell.font = Font(bold=True, color=dark)

    widths = [8, 28, 13, 25, 18, 20, 20, 20, 11, 12, 12, 10, 10, 13, 20, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:P{max(10, header_row + len(report['rows']))}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_exam_results_pdf(report):
    width, height = 842, 595
    rows_per_page = 16
    chunks = [
        report["rows"][index:index + rows_per_page]
        for index in range(0, len(report["rows"]), rows_per_page)
    ] or [[]]
    pages = []
    headers = ["Rank", "Exam", "Student", "Adm No", "Batch", "Score", "%", "C", "W", "U", "Status"]
    widths = [34, 112, 132, 86, 92, 60, 44, 28, 28, 28, 92]

    for page_number, chunk in enumerate(chunks, start=1):
        stream = pdf_rect(0, 545, width, 50, "0.486 0.227 0.929")
        stream += pdf_text(30, 570, "Exam Results Report", 18, "F2", "1 1 1")
        stream += pdf_text(
            30,
            552,
            f"{report['institute'].name} | Academic Year: {report['academic_year']} | Generated {timezone.localtime().strftime('%d-%m-%Y %I:%M %p')}",
            8,
            "F1",
            "1 1 1",
        )
        stream += pdf_text(730, 552, f"Page {page_number}/{len(chunks)}", 8, "F1", "1 1 1")

        summary = [
            ("Results", report["result_count"]),
            ("Passed", report["passed_count"]),
            ("Needs Improvement", report["needs_improvement_count"]),
            ("Average", f"{report['average_percentage']}%"),
            ("Highest", f"{report['highest_percentage']}%"),
        ]
        x = 30
        for label, value in summary:
            stream += pdf_rect(x, 502, 145, 31, "0.95 0.94 1")
            stream += pdf_text(x + 7, 520, label, 7, "F1", "0.39 0.45 0.55")
            stream += pdf_text(x + 7, 507, value, 11, "F2", "0.06 0.09 0.16")
            x += 155

        filter_text = " | ".join(report["filter_labels"])
        stream += pdf_text(30, 481, f"Filters: {filter_text[:145]}", 7, "F1", "0.39 0.45 0.55")
        start_x, y = 30, 452
        stream += pdf_rect(start_x, y - 6, sum(widths), 22, "0.06 0.09 0.16")
        x = start_x
        for header, col_width in zip(headers, widths):
            stream += pdf_text(x + 3, y, header, 7, "F2", "1 1 1")
            x += col_width
        y -= 24

        for row in chunk:
            attempt = row["attempt"]
            stream += pdf_rect(start_x, y - 5, sum(widths), 20, "0.98 0.99 1")
            values = [
                row["rank"],
                attempt.exam.title[:17],
                row["student_name"][:20],
                attempt.academic_session.admission_number[:13],
                attempt.exam.batch.name[:14],
                f"{attempt.score}/{attempt.total_marks}",
                row["percentage"],
                attempt.correct_count,
                attempt.wrong_count,
                attempt.unattempted_count,
                row["performance"][:15],
            ]
            x = start_x
            for value, col_width in zip(values, widths):
                stream += pdf_text(x + 3, y, value, 6.5, "F1", "0.06 0.09 0.16")
                x += col_width
            y -= 21

        stream += pdf_text(30, 28, "C = Correct | W = Wrong | U = Unattempted | Rank is within each exam", 8, "F1", "0.39 0.45 0.55")
        pages.append(stream)
    return build_pdf_document(pages, width=width, height=height)


@institute_admin_required
def results_export(request):
    report = get_exam_result_report(request)
    file_format = request.GET.get("format", "excel").strip().lower()
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    if file_format == "pdf":
        response = HttpResponse(build_exam_results_pdf(report), content_type="application/pdf")
        extension = "pdf"
    else:
        response = HttpResponse(
            build_exam_results_excel(report),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        extension = "xlsx"
    response["Content-Disposition"] = f'attachment; filename="exam_results_report_{stamp}.{extension}"'
    return response
