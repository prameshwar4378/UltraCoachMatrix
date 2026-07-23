from datetime import date
from decimal import Decimal
from io import BytesIO
import json

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIClient

from institute_admin.models import AcademicYear, Batch, Course, Subject
from report_card.forms import (
    ReportCardAssessmentForm,
    ReportCardAssessmentSubjectForm,
    ReportCardGradeRuleForm,
    ReportCardTeacherSubjectAllocationForm,
)
from report_card.exports import import_marks_workbook, report_card_pdf_response
from report_card.models import (
    ReportCardAssessment,
    ReportCardAssessmentSubject,
    ReportCardAssessmentSubjectComponent,
    ReportCardAuditLog,
    ReportCardComponentMarkEntry,
    ReportCardGradeRule,
    ReportCardMarkEntry,
    ReportCardSubjectResult,
    ReportCardStudentResult,
    ReportCardTeacherSubjectAllocation,
)
from report_card.permissions import (
    student_can_view_result,
    teacher_can_edit_assessment,
    teacher_has_subject_allocation as permission_teacher_has_subject_allocation,
)
from report_card.selectors import (
    get_assessments_for_teacher,
    get_deleted_assessments_for_admin,
    get_completion_summary,
    get_published_results_for_student,
    get_teacher_accessible_assessment_subjects,
    get_teacher_accessible_assessments,
    get_teacher_allocated_batches,
    get_teacher_allocated_subjects,
    get_teacher_report_card_allocations,
    teacher_has_subject_allocation,
)
from report_card.services import (
    add_assessment_subject,
    add_assessment_subject_component,
    bulk_save_subject_marks,
    create_assessment,
    generate_assessment_results,
    get_assessment_delete_impact,
    lock_assessment,
    open_marks_entry,
    permanent_delete_assessment,
    publish_assessment_results,
    reopen_marks_entry,
    restore_deleted_assessment,
    soft_delete_assessment,
    sync_assessment_subject_components,
    update_assessment,
)
from student_parent.models import StudentAcademicSession, StudentEnrollment, StudentProfile
from super_admin.models import Institute, UserProfile
from teacher.models import Exam, ExamResult, TeacherProfile


class ReportCardFeatureTests(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(name="Report Institute", code="report")
        self.other_institute = Institute.objects.create(name="Other Institute", code="other-report")
        self.year = AcademicYear.objects.create(
            institute=self.institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.other_year = AcademicYear.objects.create(
            institute=self.other_institute,
            name="2026-27",
            start_date=date(2026, 4, 1),
            end_date=date(2027, 3, 31),
        )
        self.next_year = AcademicYear.objects.create(
            institute=self.institute,
            name="2027-28",
            start_date=date(2027, 4, 1),
            end_date=date(2028, 3, 31),
        )

        self.teacher = User.objects.create_user(username="report-teacher", password="pass")
        UserProfile.objects.create(
            user=self.teacher,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )
        self.other_teacher = User.objects.create_user(username="other-report-teacher", password="pass")
        UserProfile.objects.create(
            user=self.other_teacher,
            institute=self.institute,
            role=UserProfile.Role.TEACHER,
        )

        self.course = Course.objects.create(institute=self.institute, academic_year=self.year, name="Class 8")
        self.batch = Batch.objects.create(institute=self.institute, academic_year=self.year, name="8-A")
        self.batch.courses.add(self.course)
        self.batch.teachers.add(self.teacher)
        self.unassigned_batch = Batch.objects.create(
            institute=self.institute,
            academic_year=self.year,
            name="8-B",
        )
        self.unassigned_batch.courses.add(self.course)
        self.unassigned_batch.teachers.add(self.other_teacher)

        self.math = Subject.objects.create(institute=self.institute, academic_year=self.year, name="Mathematics")
        self.science = Subject.objects.create(institute=self.institute, academic_year=self.year, name="Science")
        self.art = Subject.objects.create(institute=self.institute, academic_year=self.year, name="Art")
        self.other_subject = Subject.objects.create(
            institute=self.other_institute,
            academic_year=self.other_year,
            name="Other Science",
        )
        self.next_year_subject = Subject.objects.create(
            institute=self.institute,
            academic_year=self.next_year,
            name="Future Math",
        )

        self.session_1 = self._create_student("student-one", "A001", "Ada", "Lovelace", "1")
        self.session_2 = self._create_student("student-two", "A002", "Grace", "Hopper", "2")
        self.session_3 = self._create_student("student-three", "A003", "Katherine", "Johnson", "3")

    def _create_student(self, username, admission_number, first_name, last_name, roll_number):
        user = User.objects.create_user(
            username=username,
            password="pass",
            first_name=first_name,
            last_name=last_name,
        )
        UserProfile.objects.create(
            user=user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )
        student = StudentProfile.objects.create(
            institute=self.institute,
            academic_year=self.year,
            user=user,
            admission_number=admission_number,
            roll_number=roll_number,
        )
        session = StudentAcademicSession.objects.create(
            institute=self.institute,
            student=student,
            academic_year=self.year,
            admission_number=admission_number,
        )
        enrollment = StudentEnrollment.objects.create(
            student=student,
            academic_session=session,
            batch=self.batch,
        )
        enrollment.courses.add(self.course)
        return session

    def _assessment(self, title="Mid Term"):
        return create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            title=title,
            created_by=self.teacher,
        )

    def _subject(self, assessment, subject=None, **kwargs):
        defaults = {
            "subject": subject or self.math,
            "max_marks": Decimal("100"),
            "passing_marks": Decimal("35"),
            "weightage": Decimal("100"),
            "display_order": assessment.assessment_subjects.count() + 1,
            "actor": self.teacher,
        }
        defaults.update(kwargs)
        return add_assessment_subject(assessment, **defaults)

    def _open_and_save(self, assessment_subject, rows):
        assessment = assessment_subject.assessment
        if assessment.status != ReportCardAssessment.Status.MARKS_ENTRY_OPEN:
            open_marks_entry(assessment, actor=self.teacher)
            assessment.refresh_from_db()
        self._ensure_allocation(assessment_subject)
        return bulk_save_subject_marks(assessment_subject, rows, actor=self.teacher)

    def _ensure_allocation(self, assessment_subject, teacher=None):
        return ReportCardTeacherSubjectAllocation.objects.get_or_create(
            institute=assessment_subject.assessment.institute,
            academic_year=assessment_subject.assessment.academic_year,
            batch=assessment_subject.assessment.batch,
            subject=assessment_subject.subject,
            teacher=teacher or self.teacher,
            defaults={"created_by": self.teacher},
        )[0]

    def _allocation(self, teacher=None, batch=None, subject=None, academic_year=None, is_active=True):
        return ReportCardTeacherSubjectAllocation.objects.create(
            institute=self.institute,
            academic_year=academic_year or self.year,
            batch=batch or self.batch,
            subject=subject or self.math,
            teacher=teacher or self.teacher,
            is_active=is_active,
            created_by=self.teacher,
        )

    def _admin_user(self, username="report-admin"):
        user = User.objects.create_user(username=username, password="pass")
        UserProfile.objects.create(
            user=user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        return user

    def test_teacher_subject_allocation_can_be_created_for_same_institute_master_data(self):
        allocation = ReportCardTeacherSubjectAllocation.objects.create(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            subject=self.math,
            teacher=self.teacher,
            created_by=self.teacher,
        )

        self.assertTrue(allocation.is_active)
        self.assertIn("report-teacher", str(allocation))
        self.assertIn("8-A", str(allocation))
        self.assertIn("Mathematics", str(allocation))

    def test_teacher_subject_allocation_validates_master_data_scope(self):
        with self.assertRaises(ValidationError):
            ReportCardTeacherSubjectAllocation.objects.create(
                institute=self.institute,
                academic_year=self.year,
                batch=self.batch,
                subject=self.other_subject,
                teacher=self.teacher,
            )
        with self.assertRaises(ValidationError):
            ReportCardTeacherSubjectAllocation.objects.create(
                institute=self.institute,
                academic_year=self.year,
                batch=self.batch,
                subject=self.next_year_subject,
                teacher=self.teacher,
            )

    def test_teacher_subject_allocation_requires_same_institute_teacher(self):
        outside_teacher = User.objects.create_user(username="outside-teacher", password="pass")
        UserProfile.objects.create(
            user=outside_teacher,
            institute=self.other_institute,
            role=UserProfile.Role.TEACHER,
        )
        student_user = User.objects.create_user(username="allocation-student", password="pass")
        UserProfile.objects.create(
            user=student_user,
            institute=self.institute,
            role=UserProfile.Role.STUDENT_PARENT,
        )

        with self.assertRaises(ValidationError):
            ReportCardTeacherSubjectAllocation.objects.create(
                institute=self.institute,
                academic_year=self.year,
                batch=self.batch,
                subject=self.math,
                teacher=outside_teacher,
            )
        with self.assertRaises(ValidationError):
            ReportCardTeacherSubjectAllocation.objects.create(
                institute=self.institute,
                academic_year=self.year,
                batch=self.batch,
                subject=self.math,
                teacher=student_user,
            )

    def test_teacher_subject_allocation_is_unique_per_year_batch_subject_teacher(self):
        ReportCardTeacherSubjectAllocation.objects.create(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            subject=self.math,
            teacher=self.teacher,
        )

        with self.assertRaises(ValidationError):
            ReportCardTeacherSubjectAllocation.objects.create(
                institute=self.institute,
                academic_year=self.year,
                batch=self.batch,
                subject=self.math,
                teacher=self.teacher,
            )

    def test_teacher_allocation_selectors_are_institute_and_year_safe(self):
        self._allocation(subject=self.math)
        self._allocation(subject=self.science)
        self._allocation(subject=self.art, academic_year=self.year, is_active=False)
        ReportCardTeacherSubjectAllocation.objects.create(
            institute=self.institute,
            academic_year=self.year,
            batch=self.unassigned_batch,
            subject=self.science,
            teacher=self.other_teacher,
        )

        allocations = list(get_teacher_report_card_allocations(self.teacher, academic_year=self.year))
        batches = list(get_teacher_allocated_batches(self.teacher, academic_year=self.year))
        subjects = list(get_teacher_allocated_subjects(self.teacher, batch=self.batch, academic_year=self.year))

        self.assertEqual({allocation.subject for allocation in allocations}, {self.math, self.science})
        self.assertEqual(batches, [self.batch])
        self.assertEqual(set(subjects), {self.math, self.science})
        self.assertEqual(list(get_teacher_report_card_allocations(self.teacher, academic_year=self.next_year)), [])

    def test_teacher_accessible_assessment_subjects_use_exact_allocated_subjects(self):
        self._allocation(subject=self.math)
        assessment = self._assessment()
        math = self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)

        accessible_subjects = list(get_teacher_accessible_assessment_subjects(self.teacher, assessment))

        self.assertEqual(accessible_subjects, [math])
        self.assertTrue(teacher_has_subject_allocation(self.teacher, math))
        self.assertTrue(permission_teacher_has_subject_allocation(self.teacher, math))
        self.assertFalse(teacher_has_subject_allocation(self.teacher, science))
        self.assertFalse(permission_teacher_has_subject_allocation(self.teacher, science))

    def test_teacher_accessible_assessments_use_allocated_batch_and_subject(self):
        self._allocation(subject=self.math)
        visible_assessment = self._assessment("Visible Allocation")
        self._subject(visible_assessment, self.math)
        hidden_subject_assessment = self._assessment("Hidden Subject")
        self._subject(hidden_subject_assessment, self.science)
        hidden_batch_assessment = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.unassigned_batch,
            title="Hidden Batch",
            created_by=self.teacher,
        )
        add_assessment_subject(
            hidden_batch_assessment,
            subject=self.math,
            max_marks=Decimal("100"),
            passing_marks=Decimal("35"),
            weightage=Decimal("100"),
            display_order=1,
            actor=self.teacher,
        )

        assessments = list(get_teacher_accessible_assessments(self.teacher, academic_year=self.year))

        self.assertEqual(assessments, [visible_assessment])
        self.assertFalse(teacher_can_edit_assessment(self.teacher, visible_assessment))
        self.assertFalse(teacher_can_edit_assessment(self.teacher, hidden_subject_assessment))
        self.assertFalse(teacher_can_edit_assessment(self.teacher, hidden_batch_assessment))

    def test_teacher_accessible_assessments_do_not_mix_allocation_tuples(self):
        self._allocation(batch=self.batch, subject=self.math)
        self._allocation(batch=self.unassigned_batch, subject=self.science)
        visible_assessment = self._assessment("Visible Exact Tuple")
        self._subject(visible_assessment, self.math)
        mixed_tuple_assessment = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.unassigned_batch,
            title="Hidden Mixed Tuple",
            created_by=self.teacher,
        )
        add_assessment_subject(
            mixed_tuple_assessment,
            subject=self.math,
            max_marks=Decimal("100"),
            passing_marks=Decimal("35"),
            weightage=Decimal("100"),
            display_order=1,
            actor=self.teacher,
        )

        assessments = list(get_teacher_accessible_assessments(self.teacher, academic_year=self.year))

        self.assertIn(visible_assessment, assessments)
        self.assertNotIn(mixed_tuple_assessment, assessments)

    def test_allocation_form_prevents_duplicates_gracefully(self):
        self._allocation(subject=self.math)
        form = ReportCardTeacherSubjectAllocationForm(
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "subject": self.math.pk,
                "teacher": self.teacher.pk,
                "is_active": "on",
            },
            institute=self.institute,
            created_by=self.teacher,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("teacher", form.errors)

    def test_institute_admin_can_create_teacher_subject_allocation_from_ui(self):
        admin_user = self._admin_user("allocation-admin")
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("report_card_admin:allocation_create"),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "subject": self.math.pk,
                "teacher": self.teacher.pk,
                "is_active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        allocation = ReportCardTeacherSubjectAllocation.objects.get(
            academic_year=self.year,
            batch=self.batch,
            subject=self.math,
            teacher=self.teacher,
        )
        self.assertEqual(allocation.institute, self.institute)
        self.assertEqual(allocation.created_by, admin_user)

    def test_institute_admin_can_bulk_save_class_subject_allocations_from_ui(self):
        admin_user = self._admin_user("allocation-bulk-admin")
        self.client.force_login(admin_user)
        session = self.client.session
        session["academic_year_id"] = self.year.pk
        session.save()

        response = self.client.post(
            reverse("report_card_admin:allocation_create"),
            data={
                "batch": self.batch.pk,
                "allocation_rows_json": json.dumps(
                    [
                        {"subject_id": self.math.pk, "teacher_id": self.teacher.pk},
                        {"subject_id": self.science.pk, "teacher_id": self.teacher.pk},
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        allocations = ReportCardTeacherSubjectAllocation.objects.filter(
            academic_year=self.year,
            batch=self.batch,
            teacher=self.teacher,
        )
        self.assertEqual(allocations.count(), 2)
        self.assertSetEqual(set(allocations.values_list("subject_id", flat=True)), {self.math.pk, self.science.pk})
        self.assertTrue(allocations.filter(created_by=admin_user).exists())

    def test_institute_admin_can_filter_update_and_delete_allocations_from_ui(self):
        admin_user = self._admin_user("allocation-admin-two")
        allocation = self._allocation(subject=self.math)
        self.client.force_login(admin_user)

        list_response = self.client.get(
            reverse("report_card_admin:allocation_list"),
            data={"academic_year": self.year.pk, "teacher": self.teacher.pk, "q": "Mathematics"},
        )
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Mathematics")

        update_response = self.client.post(
            reverse("report_card_admin:allocation_update", args=[allocation.pk]),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "subject": self.science.pk,
                "teacher": self.teacher.pk,
                "is_active": "",
            },
        )
        self.assertEqual(update_response.status_code, 302)
        allocation.refresh_from_db()
        self.assertEqual(allocation.subject, self.science)
        self.assertFalse(allocation.is_active)

        delete_response = self.client.post(reverse("report_card_admin:allocation_delete", args=[allocation.pk]))
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(ReportCardTeacherSubjectAllocation.objects.filter(pk=allocation.pk).exists())

    def test_institute_admin_sidebar_shows_report_card_management_links(self):
        admin_user = self._admin_user("navigation-admin")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:assessment_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tests / Exams")
        self.assertContains(response, "Report Card Generator")
        self.assertContains(response, 'data-bs-target="#report-card-nav"')
        self.assertContains(response, 'id="report-card-nav"')
        self.assertContains(response, "Report Cards")
        self.assertContains(response, "Recycle Bin")
        self.assertContains(response, "Teacher Subject Allocation")
        self.assertContains(response, "Grade Rules")
        self.assertContains(response, reverse("report_card_admin:assessment_list"))
        self.assertContains(response, reverse("report_card_admin:assessment_bin"))
        self.assertContains(response, reverse("report_card_admin:allocation_list"))
        self.assertContains(response, reverse("report_card_admin:grade_rule_list"))

    def test_institute_admin_cannot_edit_other_institute_allocation(self):
        admin_user = self._admin_user("allocation-admin-three")
        outside_teacher = User.objects.create_user(username="outside-allocation-teacher", password="pass")
        UserProfile.objects.create(
            user=outside_teacher,
            institute=self.other_institute,
            role=UserProfile.Role.TEACHER,
        )
        other_batch = Batch.objects.create(institute=self.other_institute, academic_year=self.other_year, name="Other Batch")
        other_allocation = ReportCardTeacherSubjectAllocation.objects.create(
            institute=self.other_institute,
            academic_year=self.other_year,
            batch=other_batch,
            subject=self.other_subject,
            teacher=outside_teacher,
        )
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:allocation_update", args=[other_allocation.pk]))

        self.assertEqual(response.status_code, 404)

    def test_admin_assessment_create_form_scopes_batches_to_selected_session(self):
        admin_user = self._admin_user("assessment-scope-admin")
        next_course = Course.objects.create(institute=self.institute, academic_year=self.next_year, name="Future Class")
        next_batch = Batch.objects.create(institute=self.institute, academic_year=self.next_year, name="Future 8-A")
        next_batch.courses.add(next_course)
        self.client.force_login(admin_user)
        session = self.client.session
        session["academic_year_id"] = self.year.pk
        session.save()

        response = self.client.get(reverse("report_card_admin:assessment_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.batch.name)
        self.assertNotContains(response, next_batch.name)

    def test_institute_admin_can_create_and_edit_report_card_assessment_from_ui(self):
        admin_user = self._admin_user("assessment-admin")
        online_exam_count = Exam.objects.count()
        self.client.force_login(admin_user)

        create_response = self.client.post(
            reverse("report_card_admin:assessment_create"),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "title": "Admin Mid Term",
                "assessment_date": "2026-07-19",
                "result_date": "2026-07-24",
            },
        )

        self.assertEqual(create_response.status_code, 302)
        assessment = ReportCardAssessment.objects.get(title="Admin Mid Term")
        self.assertEqual(assessment.institute, self.institute)
        self.assertEqual(assessment.created_by, admin_user)
        self.assertEqual(Exam.objects.count(), online_exam_count)

        multi_response = self.client.post(
            reverse("report_card_admin:assessment_create"),
            data={
                "academic_year": self.year.pk,
                "batches": [self.batch.pk, self.unassigned_batch.pk],
                "title": "Admin Multi Batch Test",
                "assessment_date": "2026-07-20",
                "result_date": "2026-07-25",
            },
        )

        self.assertEqual(multi_response.status_code, 302)
        multi_assessments = ReportCardAssessment.objects.filter(title="Admin Multi Batch Test")
        self.assertEqual(multi_assessments.count(), 2)
        self.assertSetEqual(set(multi_assessments.values_list("batch_id", flat=True)), {self.batch.pk, self.unassigned_batch.pk})
        self.assertEqual(Exam.objects.count(), online_exam_count)

        class_list_response = self.client.get(
            reverse("report_card_admin:assessment_classes", args=[multi_assessments.first().pk])
        )
        self.assertEqual(class_list_response.status_code, 200)
        self.assertContains(class_list_response, self.batch.name)
        self.assertContains(class_list_response, self.unassigned_batch.name)
        self.assertContains(class_list_response, reverse("report_card_admin:assessment_detail", args=[assessment.pk]), count=0)

        update_response = self.client.post(
            reverse("report_card_admin:assessment_update", args=[assessment.pk]),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "title": "Admin Final Term",
                "assessment_date": "2026-08-19",
                "result_date": "2026-08-24",
            },
        )

        self.assertEqual(update_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.title, "Admin Final Term")
        self.assertEqual(assessment.assessment_date, date(2026, 8, 19))
        self.assertEqual(Exam.objects.count(), online_exam_count)

    def test_soft_deleted_assessment_is_hidden_from_admin_pages(self):
        admin_user = self._admin_user("soft-delete-admin")
        assessment = self._assessment("Soft Deleted Admin Assessment")
        assessment.is_deleted = True
        assessment.deleted_at = timezone.now()
        assessment.deleted_by = admin_user
        assessment.delete_reason = "Created by mistake."
        assessment.save()
        self.client.force_login(admin_user)

        list_response = self.client.get(reverse("report_card_admin:assessment_list"))
        detail_response = self.client.get(reverse("report_card_admin:assessment_detail", args=[assessment.pk]))

        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, "Soft Deleted Admin Assessment")
        self.assertEqual(detail_response.status_code, 404)

    def test_soft_deleted_assessment_is_hidden_from_teacher_access(self):
        assessment = self._assessment("Soft Deleted Teacher Assessment")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        assessment.is_deleted = True
        assessment.deleted_at = timezone.now()
        assessment.deleted_by = self._admin_user("soft-delete-teacher-admin")
        assessment.delete_reason = "Duplicate assessment."
        assessment.save()
        self.client.force_login(self.teacher)

        selector_assessments = list(get_teacher_accessible_assessments(self.teacher, academic_year=self.year))
        list_response = self.client.get(reverse("report_card:assessment_list"))
        detail_response = self.client.get(reverse("report_card:assessment_detail", args=[assessment.pk]))

        self.assertNotIn(assessment, selector_assessments)
        self.assertNotIn(assessment, list(get_assessments_for_teacher(self.teacher, academic_year=self.year)))
        self.assertEqual(list(get_teacher_accessible_assessment_subjects(self.teacher, assessment)), [])
        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, "Soft Deleted Teacher Assessment")
        self.assertEqual(detail_response.status_code, 404)

    def test_recycle_bin_selector_shows_only_deleted_assessments_for_scope(self):
        admin_user = self._admin_user("recycle-bin-admin")
        active_assessment = self._assessment("Active Assessment Outside Bin")
        deleted_assessment = self._assessment("Deleted Assessment In Bin")
        next_year_assessment = create_assessment(
            institute=self.institute,
            academic_year=self.next_year,
            batch=Batch.objects.create(
                institute=self.institute,
                academic_year=self.next_year,
                name="Future 8-C",
            ),
            title="Future Deleted Assessment",
            created_by=admin_user,
        )
        other_batch = Batch.objects.create(
            institute=self.other_institute,
            academic_year=self.other_year,
            name="Other Deleted Batch",
        )
        other_deleted = create_assessment(
            institute=self.other_institute,
            academic_year=self.other_year,
            batch=other_batch,
            title="Other Institute Deleted Assessment",
            created_by=admin_user,
        )
        for assessment in [deleted_assessment, next_year_assessment, other_deleted]:
            assessment.is_deleted = True
            assessment.deleted_at = timezone.now()
            assessment.deleted_by = admin_user
            assessment.delete_reason = "Recycle bin scope test."
            assessment.save()

        current_year_bin = list(get_deleted_assessments_for_admin(self.institute, academic_year=self.year))
        all_institute_bin = list(get_deleted_assessments_for_admin(self.institute))
        other_institute_bin = list(get_deleted_assessments_for_admin(self.other_institute, academic_year=self.other_year))

        self.assertIn(deleted_assessment, current_year_bin)
        self.assertNotIn(active_assessment, current_year_bin)
        self.assertNotIn(next_year_assessment, current_year_bin)
        self.assertNotIn(other_deleted, current_year_bin)
        self.assertIn(next_year_assessment, all_institute_bin)
        self.assertEqual(other_institute_bin, [other_deleted])

    def test_soft_deleted_assessment_results_are_hidden_from_student_parent(self):
        admin_user = self._admin_user("soft-delete-result-admin")
        assessment = self._assessment("Soft Deleted Published Assessment")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "75"},
                {"academic_session": self.session_3, "marks_obtained": "70"},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)
        publish_assessment_results(assessment, actor=admin_user)
        result = ReportCardStudentResult.objects.get(assessment=assessment, academic_session=self.session_1)

        self.assertEqual(list(get_published_results_for_student(self.session_1.student)), [result])
        self.assertTrue(student_can_view_result(self.session_1.student.user, result))

        assessment.is_deleted = True
        assessment.deleted_at = timezone.now()
        assessment.deleted_by = admin_user
        assessment.delete_reason = "Confidential duplicate deletion."
        assessment.save()
        result.refresh_from_db()

        self.assertEqual(list(get_published_results_for_student(self.session_1.student)), [])
        self.assertFalse(student_can_view_result(self.session_1.student.user, result))

    def test_soft_deleted_assessment_title_can_be_reused_for_same_class(self):
        admin_user = self._admin_user("soft-delete-reuse-admin")
        assessment = self._assessment("Reusable Assessment Title")
        assessment.is_deleted = True
        assessment.deleted_at = timezone.now()
        assessment.deleted_by = admin_user
        assessment.delete_reason = "Recreate with corrected setup."
        assessment.save()

        replacement = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            title="Reusable Assessment Title",
            created_by=admin_user,
        )

        self.assertNotEqual(assessment.pk, replacement.pk)
        self.assertFalse(replacement.is_deleted)

    def test_assessment_delete_impact_counts_dependent_report_card_data(self):
        assessment = self._assessment("Delete Impact")
        subject = self._subject(assessment, self.math)
        add_assessment_subject_component(
            subject,
            name="Theory Exam",
            max_marks=Decimal("100"),
            weightage=Decimal("100"),
            include_in_total=True,
            display_order=1,
            actor=self.teacher,
        )
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [{"academic_session": self.session_1, "component_marks": {subject.components.first().pk: "82"}}],
        )
        generate_assessment_results(assessment, actor=self._admin_user("impact-admin"), require_complete=False)

        impact = get_assessment_delete_impact(assessment)

        self.assertEqual(impact["subject_count"], 1)
        self.assertEqual(impact["component_count"], 1)
        self.assertEqual(impact["mark_entry_count"], 1)
        self.assertEqual(impact["component_mark_entry_count"], 1)
        self.assertEqual(impact["generated_result_count"], 3)
        self.assertEqual(impact["student_count"], 3)
        self.assertTrue(impact["has_data"])

    def test_soft_delete_requires_reason_when_assessment_has_data(self):
        admin_user = self._admin_user("soft-delete-reason-admin")
        assessment = self._assessment("Needs Delete Reason")
        self._subject(assessment, self.math)

        with self.assertRaisesMessage(ValidationError, "delete reason is required"):
            soft_delete_assessment(assessment, actor=admin_user, reason="")

        assessment.refresh_from_db()
        self.assertFalse(assessment.is_deleted)

    def test_soft_delete_allows_empty_reason_when_assessment_has_no_data(self):
        admin_user = self._admin_user("simple-soft-delete-admin")
        assessment = self._assessment("Empty Draft Delete")

        soft_delete_assessment(assessment, actor=admin_user, reason="")

        assessment.refresh_from_db()
        self.assertTrue(assessment.is_deleted)
        self.assertEqual(assessment.deleted_by, admin_user)
        self.assertTrue(
            ReportCardAuditLog.objects.filter(
                assessment=assessment,
                action=ReportCardAuditLog.Action.ASSESSMENT_DELETED,
            ).exists()
        )

    def test_soft_delete_published_assessment_requires_detailed_reason(self):
        admin_user = self._admin_user("published-soft-delete-admin")
        assessment = self._assessment("Published Delete Guard")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "75"},
                {"academic_session": self.session_3, "marks_obtained": "70"},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)
        publish_assessment_results(assessment, actor=admin_user)

        with self.assertRaisesMessage(ValidationError, "detailed delete reason"):
            soft_delete_assessment(assessment, actor=admin_user, reason="Wrong")

        soft_delete_assessment(
            assessment,
            actor=admin_user,
            reason="Published duplicate assessment created for the wrong class.",
        )
        assessment.refresh_from_db()
        self.assertTrue(assessment.is_deleted)

    def test_published_delete_confirmation_page_requires_exact_title_in_ui(self):
        admin_user = self._admin_user("published-delete-ui-admin")
        assessment = self._assessment("Published Delete UI Guard")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "75"},
                {"academic_session": self.session_3, "marks_obtained": "70"},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)
        publish_assessment_results(assessment, actor=admin_user)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:assessment_delete", args=[assessment.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Critical:")
        self.assertContains(response, "Type exact assessment title")
        self.assertContains(response, assessment.title)

    def test_soft_delete_assessment_with_marks_and_results_requires_reason(self):
        admin_user = self._admin_user("marks-results-delete-admin")
        assessment = self._assessment("Marks Results Delete Guard")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "60"},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)

        with self.assertRaisesMessage(ValidationError, "delete reason is required"):
            soft_delete_assessment(assessment, actor=admin_user, reason="")

        soft_delete_assessment(
            assessment,
            actor=admin_user,
            reason="Generated results exist and this duplicate assessment must be removed.",
        )
        assessment.refresh_from_db()
        self.assertTrue(assessment.is_deleted)

    def test_restore_deleted_assessment_blocks_active_duplicate(self):
        admin_user = self._admin_user("restore-duplicate-admin")
        assessment = self._assessment("Restore Duplicate Guard")
        soft_delete_assessment(assessment, actor=admin_user, reason="")
        create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            title="Restore Duplicate Guard",
            created_by=admin_user,
        )

        with self.assertRaisesMessage(ValidationError, "same name already exists"):
            restore_deleted_assessment(assessment, actor=admin_user)

    def test_restore_deleted_assessment_clears_delete_fields_and_writes_audit(self):
        admin_user = self._admin_user("restore-soft-delete-admin")
        assessment = self._assessment("Restore Soft Delete")
        soft_delete_assessment(assessment, actor=admin_user, reason="")

        restore_deleted_assessment(assessment, actor=admin_user)

        assessment.refresh_from_db()
        self.assertFalse(assessment.is_deleted)
        self.assertIsNone(assessment.deleted_at)
        self.assertIsNone(assessment.deleted_by)
        self.assertEqual(assessment.delete_reason, "")
        self.assertTrue(
            ReportCardAuditLog.objects.filter(
                assessment=assessment,
                action=ReportCardAuditLog.Action.ASSESSMENT_RESTORED,
            ).exists()
        )

    def test_permanent_delete_requires_recycle_bin_exact_title_and_reason(self):
        admin_user = self._admin_user("permanent-delete-admin")
        assessment = self._assessment("Permanent Delete Guard")

        with self.assertRaisesMessage(ValidationError, "only from the recycle bin"):
            permanent_delete_assessment(assessment, actor=admin_user, confirmation_text=assessment.title, reason="Cleanup")

        soft_delete_assessment(assessment, actor=admin_user, reason="")

        with self.assertRaisesMessage(ValidationError, "exact assessment name"):
            permanent_delete_assessment(assessment, actor=admin_user, confirmation_text="Wrong", reason="Cleanup")

        with self.assertRaisesMessage(ValidationError, "permanent delete reason is required"):
            permanent_delete_assessment(assessment, actor=admin_user, confirmation_text=assessment.title, reason="")

    def test_permanent_delete_removes_assessment_and_preserves_audit_log(self):
        admin_user = self._admin_user("permanent-delete-audit-admin")
        assessment = self._assessment("Permanent Delete Audit")
        assessment_id = assessment.pk
        soft_delete_assessment(assessment, actor=admin_user, reason="")

        impact = permanent_delete_assessment(
            assessment,
            actor=admin_user,
            confirmation_text="Permanent Delete Audit",
            reason="Cleanup empty duplicate assessment.",
        )

        self.assertFalse(ReportCardAssessment.objects.filter(pk=assessment_id).exists())
        self.assertEqual(impact["subject_count"], 0)
        audit = ReportCardAuditLog.objects.get(
            action=ReportCardAuditLog.Action.ASSESSMENT_PERMANENTLY_DELETED,
            metadata__assessment_id=assessment_id,
        )
        self.assertIsNone(audit.assessment)
        self.assertEqual(audit.metadata["assessment_title"], "Permanent Delete Audit")

    def test_permanent_delete_removes_dependent_report_card_data_only(self):
        admin_user = self._admin_user("permanent-delete-dependent-admin")
        exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.math,
            title="Online Exam Untouched By Report Card Delete",
            exam_date=date(2026, 9, 1),
            total_marks=50,
            created_by=self.teacher,
        )
        ExamResult.objects.create(
            exam=exam,
            student=self.session_1.student,
            marks_obtained=Decimal("42"),
            remark="Keep online result.",
        )
        before_exams = list(Exam.objects.values("id", "title", "total_marks", "is_published"))
        before_exam_results = list(ExamResult.objects.values("id", "exam_id", "student_id", "marks_obtained", "remark"))

        assessment = self._assessment("Permanent Delete Dependencies")
        subject = self._subject(assessment, self.math)
        component = add_assessment_subject_component(
            subject,
            name="Theory Exam",
            max_marks=Decimal("100"),
            weightage=Decimal("100"),
            include_in_total=True,
            display_order=1,
            actor=admin_user,
        )
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "component_marks": {component.pk: "80"}},
                {"academic_session": self.session_2, "component_marks": {component.pk: "70"}},
                {"academic_session": self.session_3, "component_marks": {component.pk: "60"}},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)
        assessment_id = assessment.pk
        subject_id = subject.pk
        component_id = component.pk

        self.assertTrue(ReportCardMarkEntry.objects.filter(assessment_subject_id=subject_id).exists())
        self.assertTrue(ReportCardComponentMarkEntry.objects.filter(component_id=component_id).exists())
        self.assertTrue(ReportCardStudentResult.objects.filter(assessment_id=assessment_id).exists())
        self.assertTrue(ReportCardSubjectResult.objects.filter(assessment_subject_id=subject_id).exists())

        soft_delete_assessment(
            assessment,
            actor=admin_user,
            reason="Removing duplicate assessment with dependent report-card records.",
        )
        permanent_delete_assessment(
            assessment,
            actor=admin_user,
            confirmation_text=assessment.title,
            reason="Permanent cleanup after confirming duplicate report-card assessment.",
        )

        self.assertFalse(ReportCardAssessment.objects.filter(pk=assessment_id).exists())
        self.assertFalse(ReportCardAssessmentSubject.objects.filter(pk=subject_id).exists())
        self.assertFalse(ReportCardAssessmentSubjectComponent.objects.filter(pk=component_id).exists())
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject_id=subject_id).exists())
        self.assertFalse(ReportCardComponentMarkEntry.objects.filter(component_id=component_id).exists())
        self.assertFalse(ReportCardStudentResult.objects.filter(assessment_id=assessment_id).exists())
        self.assertFalse(ReportCardSubjectResult.objects.filter(assessment_subject_id=subject_id).exists())
        self.assertEqual(list(Exam.objects.values("id", "title", "total_marks", "is_published")), before_exams)
        self.assertEqual(
            list(ExamResult.objects.values("id", "exam_id", "student_id", "marks_obtained", "remark")),
            before_exam_results,
        )

    def test_delete_restore_and_permanent_delete_audit_logs_store_scope_reason_and_impact(self):
        admin_user = self._admin_user("delete-audit-metadata-admin")
        assessment = self._assessment("Delete Audit Metadata")
        self._subject(assessment, self.math)
        soft_reason = "Duplicate report-card assessment for the same class."

        soft_delete_assessment(assessment, actor=admin_user, reason=soft_reason)
        soft_audit = ReportCardAuditLog.objects.get(
            assessment=assessment,
            action=ReportCardAuditLog.Action.ASSESSMENT_DELETED,
        )

        self.assertEqual(soft_audit.actor, admin_user)
        self.assertEqual(soft_audit.metadata["reason"], soft_reason)
        self.assertEqual(soft_audit.metadata["assessment_id"], assessment.pk)
        self.assertEqual(soft_audit.metadata["assessment_title"], assessment.title)
        self.assertEqual(soft_audit.metadata["institute_id"], self.institute.pk)
        self.assertEqual(soft_audit.metadata["academic_year_id"], self.year.pk)
        self.assertEqual(soft_audit.metadata["batch_id"], self.batch.pk)
        self.assertEqual(soft_audit.metadata["impact"]["subject_count"], 1)

        restore_deleted_assessment(assessment, actor=admin_user)
        restore_audit = ReportCardAuditLog.objects.get(
            assessment=assessment,
            action=ReportCardAuditLog.Action.ASSESSMENT_RESTORED,
        )
        self.assertEqual(restore_audit.actor, admin_user)
        self.assertEqual(restore_audit.metadata["reason"], soft_reason)
        self.assertEqual(restore_audit.metadata["previous_delete_reason"], soft_reason)
        self.assertEqual(restore_audit.metadata["impact"]["subject_count"], 1)
        self.assertEqual(restore_audit.metadata["institute_id"], self.institute.pk)

        permanent_assessment = self._assessment("Permanent Audit Metadata")
        permanent_reason = "Permanent cleanup after duplicate assessment review."
        soft_delete_assessment(permanent_assessment, actor=admin_user, reason="")
        permanent_assessment_id = permanent_assessment.pk
        permanent_delete_assessment(
            permanent_assessment,
            actor=admin_user,
            confirmation_text=permanent_assessment.title,
            reason=permanent_reason,
        )
        permanent_audit = ReportCardAuditLog.objects.get(
            action=ReportCardAuditLog.Action.ASSESSMENT_PERMANENTLY_DELETED,
            metadata__assessment_id=permanent_assessment_id,
        )
        self.assertIsNone(permanent_audit.assessment)
        self.assertEqual(permanent_audit.actor, admin_user)
        self.assertEqual(permanent_audit.metadata["reason"], permanent_reason)
        self.assertEqual(permanent_audit.metadata["assessment_title"], "Permanent Audit Metadata")
        self.assertEqual(permanent_audit.metadata["institute_id"], self.institute.pk)
        self.assertEqual(permanent_audit.metadata["impact"]["generated_result_count"], 0)

    def test_institute_admin_can_soft_delete_assessment_from_view(self):
        admin_user = self._admin_user("delete-view-admin")
        assessment = self._assessment("Delete View Assessment")
        self._subject(assessment, self.math)
        self.client.force_login(admin_user)

        get_response = self.client.get(reverse("report_card_admin:assessment_delete", args=[assessment.pk]))
        post_response = self.client.post(
            reverse("report_card_admin:assessment_delete", args=[assessment.pk]),
            data={"delete_reason": "Created for wrong class by mistake."},
        )

        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, "Delete Report Card Assessment")
        self.assertEqual(post_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertTrue(assessment.is_deleted)
        self.assertEqual(assessment.deleted_by, admin_user)

    def test_assessment_bin_shows_deleted_assessments_only(self):
        admin_user = self._admin_user("bin-view-admin")
        active_assessment = self._assessment("Active Bin Hidden")
        deleted_assessment = self._assessment("Deleted Bin Visible")
        soft_delete_assessment(deleted_assessment, actor=admin_user, reason="")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:assessment_bin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Deleted Bin Visible")
        self.assertNotContains(response, "Active Bin Hidden")

    def test_assessment_bin_filters_by_batch_status_deleted_by_and_search(self):
        admin_user = self._admin_user("bin-filter-admin")
        other_admin = self._admin_user("bin-filter-other-admin")
        draft_deleted = self._assessment("Draft Deleted Searchable")
        open_deleted = self._assessment("Open Deleted Filter Target")
        open_deleted.status = ReportCardAssessment.Status.MARKS_ENTRY_OPEN
        open_deleted.save()
        other_batch_deleted = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.unassigned_batch,
            title="Other Batch Deleted",
            created_by=admin_user,
        )
        other_admin_deleted = self._assessment("Other Admin Deleted")
        soft_delete_assessment(draft_deleted, actor=admin_user, reason="Searchable reason")
        soft_delete_assessment(open_deleted, actor=admin_user, reason="Status filter reason")
        soft_delete_assessment(other_batch_deleted, actor=admin_user, reason="Other batch reason")
        soft_delete_assessment(other_admin_deleted, actor=other_admin, reason="Other user reason")
        self.client.force_login(admin_user)

        batch_response = self.client.get(reverse("report_card_admin:assessment_bin"), {"batch": self.unassigned_batch.pk})
        status_response = self.client.get(
            reverse("report_card_admin:assessment_bin"),
            {"status": ReportCardAssessment.Status.MARKS_ENTRY_OPEN},
        )
        deleted_by_response = self.client.get(reverse("report_card_admin:assessment_bin"), {"deleted_by": other_admin.pk})
        search_response = self.client.get(reverse("report_card_admin:assessment_bin"), {"q": "Searchable"})

        self.assertContains(batch_response, "Other Batch Deleted")
        self.assertNotContains(batch_response, "Draft Deleted Searchable")
        self.assertContains(status_response, "Open Deleted Filter Target")
        self.assertNotContains(status_response, "Draft Deleted Searchable")
        self.assertContains(deleted_by_response, "Other Admin Deleted")
        self.assertNotContains(deleted_by_response, "Open Deleted Filter Target")
        self.assertContains(search_response, "Draft Deleted Searchable")
        self.assertNotContains(search_response, "Open Deleted Filter Target")

    def test_restore_assessment_view_requires_post_and_recycle_bin(self):
        admin_user = self._admin_user("restore-view-admin")
        assessment = self._assessment("Restore View Assessment")
        self.client.force_login(admin_user)

        active_post = self.client.post(reverse("report_card_admin:assessment_restore", args=[assessment.pk]))
        soft_delete_assessment(assessment, actor=admin_user, reason="")
        get_response = self.client.get(reverse("report_card_admin:assessment_restore", args=[assessment.pk]))
        restore_response = self.client.post(reverse("report_card_admin:assessment_restore", args=[assessment.pk]))

        self.assertEqual(active_post.status_code, 404)
        self.assertEqual(get_response.status_code, 405)
        self.assertEqual(restore_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertFalse(assessment.is_deleted)

    def test_restore_brings_assessment_back_to_normal_admin_and_teacher_lists(self):
        admin_user = self._admin_user("restore-normal-list-admin")
        assessment = self._assessment("Restore Normal Lists")
        subject = self._subject(assessment, self.math)
        self._ensure_allocation(subject)
        soft_delete_assessment(assessment, actor=admin_user, reason="Temporary bin test.")
        self.client.force_login(admin_user)

        deleted_admin_list = self.client.get(reverse("report_card_admin:assessment_list"))
        self.assertNotContains(deleted_admin_list, "Restore Normal Lists")

        restore_deleted_assessment(assessment, actor=admin_user)

        restored_admin_list = self.client.get(reverse("report_card_admin:assessment_list"))
        self.assertContains(restored_admin_list, "Restore Normal Lists")
        self.assertIn(assessment, list(get_teacher_accessible_assessments(self.teacher, academic_year=self.year)))

    def test_permanent_delete_view_requires_post_bin_exact_title_and_reason(self):
        admin_user = self._admin_user("permanent-delete-view-admin")
        assessment = self._assessment("Permanent Delete View")
        assessment_id = assessment.pk
        soft_delete_assessment(assessment, actor=admin_user, reason="")
        self.client.force_login(admin_user)

        get_response = self.client.get(reverse("report_card_admin:assessment_permanent_delete", args=[assessment.pk]))
        bad_response = self.client.post(
            reverse("report_card_admin:assessment_permanent_delete", args=[assessment.pk]),
            data={"confirmation_text": "Wrong", "delete_reason": "Cleanup duplicate."},
        )
        assessment.refresh_from_db()
        self.assertTrue(ReportCardAssessment.objects.filter(pk=assessment_id).exists())
        good_response = self.client.post(
            reverse("report_card_admin:assessment_permanent_delete", args=[assessment.pk]),
            data={
                "confirmation_text": "Permanent Delete View",
                "delete_reason": "Cleanup duplicate empty assessment.",
            },
        )

        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, "Permanent Delete Assessment")
        self.assertContains(get_response, "I understand this action is permanent and cannot be restored.")
        self.assertEqual(bad_response.status_code, 302)
        self.assertEqual(good_response.status_code, 302)
        self.assertFalse(ReportCardAssessment.objects.filter(pk=assessment_id).exists())

    def test_teacher_cannot_access_admin_assessment_delete_routes(self):
        assessment = self._assessment("Teacher Delete Route Blocked")
        soft_delete_assessment(assessment, actor=self._admin_user("teacher-delete-route-admin"), reason="")
        self.client.force_login(self.teacher)

        delete_response = self.client.get(reverse("report_card_admin:assessment_delete", args=[assessment.pk]))
        bin_response = self.client.get(reverse("report_card_admin:assessment_bin"))
        restore_response = self.client.post(reverse("report_card_admin:assessment_restore", args=[assessment.pk]))

        self.assertNotEqual(delete_response.status_code, 200)
        self.assertNotEqual(bin_response.status_code, 200)
        self.assertNotEqual(restore_response.status_code, 200)

    def test_teacher_cannot_create_or_edit_report_card_assessment_from_ui(self):
        assessment = self._assessment("Teacher Blocked Edit")
        self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        online_exam_count = Exam.objects.count()
        self.client.force_login(self.teacher)

        create_response = self.client.post(
            reverse("report_card:assessment_create"),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "title": "Teacher Created Assessment",
                "assessment_date": "2026-07-19",
            },
        )

        self.assertEqual(create_response.status_code, 302)
        self.assertFalse(ReportCardAssessment.objects.filter(title="Teacher Created Assessment").exists())

        update_response = self.client.post(
            reverse("report_card:assessment_update", args=[assessment.pk]),
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "title": "Teacher Edited Assessment",
                "assessment_date": "2026-07-20",
            },
        )

        self.assertEqual(update_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.title, "Teacher Blocked Edit")
        self.assertEqual(Exam.objects.count(), online_exam_count)

    def test_institute_admin_can_create_update_and_delete_assessment_subject_structure_from_ui(self):
        admin_user = self._admin_user("structure-admin")
        assessment = self._assessment("Admin Structure")
        self.client.force_login(admin_user)
        components = [
            {
                "id": None,
                "name": "Theory Exam",
                "max_marks": "70",
                "passing_marks": "0",
                "weightage": "70",
                "display_order": 1,
                "include_in_total": True,
                "is_primary": True,
            },
            {
                "id": None,
                "name": "Practical",
                "max_marks": "30",
                "passing_marks": "0",
                "weightage": "30",
                "display_order": 2,
                "include_in_total": True,
                "is_primary": False,
            },
        ]

        create_response = self.client.post(
            reverse("report_card_admin:assessment_subject_create", args=[assessment.pk]),
            data={
                "subject": self.math.pk,
                "max_marks": "",
                "passing_marks": "35",
                "weightage": "",
                "display_order": "",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(components),
            },
        )

        self.assertEqual(create_response.status_code, 302)
        assessment_subject = ReportCardAssessmentSubject.objects.get(assessment=assessment, subject=self.math)
        self.assertEqual(assessment_subject.max_marks, Decimal("100.00"))
        self.assertEqual(
            list(assessment_subject.components.order_by("display_order").values_list("name_snapshot", "max_marks")),
            [("Theory Exam", Decimal("70.00")), ("Practical", Decimal("30.00"))],
        )

        existing_component_id = assessment_subject.components.order_by("display_order").first().pk
        updated_components = [
            {
                "id": existing_component_id,
                "name": "Theory Exam",
                "max_marks": "80",
                "passing_marks": "0",
                "weightage": "80",
                "display_order": 1,
                "include_in_total": True,
                "is_primary": True,
            },
            {
                "id": None,
                "name": "Notebook",
                "max_marks": "20",
                "passing_marks": "0",
                "weightage": "20",
                "display_order": 2,
                "include_in_total": True,
                "is_primary": False,
            },
        ]
        update_response = self.client.post(
            reverse("report_card_admin:assessment_subject_update", args=[assessment.pk, assessment_subject.pk]),
            data={
                "subject": self.math.pk,
                "max_marks": "",
                "passing_marks": "40",
                "weightage": "",
                "display_order": "",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(updated_components),
            },
        )

        self.assertEqual(update_response.status_code, 302)
        assessment_subject.refresh_from_db()
        self.assertEqual(assessment_subject.max_marks, Decimal("100.00"))
        self.assertTrue(assessment_subject.components.filter(pk=existing_component_id, max_marks=Decimal("80.00")).exists())
        self.assertTrue(assessment_subject.components.filter(name_snapshot="Notebook", max_marks=Decimal("20.00")).exists())

        delete_response = self.client.post(
            reverse("report_card_admin:assessment_subject_delete", args=[assessment.pk, assessment_subject.pk])
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(ReportCardAssessmentSubject.objects.filter(pk=assessment_subject.pk).exists())

    def test_teacher_cannot_create_update_delete_assessment_subject_or_components_from_ui(self):
        assessment = self._assessment("Teacher Structure Blocked")
        assessment_subject = self._subject(assessment, self.math)
        component = assessment_subject.components.create(
            name="Theory Exam",
            max_marks=Decimal("100"),
            passing_marks=Decimal("0"),
            weightage=Decimal("100"),
            display_order=1,
            include_in_total=True,
        )
        self._allocation(subject=self.math)
        self.client.force_login(self.teacher)

        create_response = self.client.post(
            reverse("report_card:assessment_subject_create", args=[assessment.pk]),
            data={
                "subject": self.science.pk,
                "passing_marks": "35",
                "include_in_total": "on",
                "components_json": json.dumps([{"name": "Theory Exam", "max_marks": "100", "include_in_total": True}]),
            },
        )
        self.assertEqual(create_response.status_code, 302)
        self.assertFalse(ReportCardAssessmentSubject.objects.filter(assessment=assessment, subject=self.science).exists())

        update_response = self.client.post(
            reverse("report_card:assessment_subject_update", args=[assessment.pk, assessment_subject.pk]),
            data={
                "subject": self.math.pk,
                "passing_marks": "10",
                "include_in_total": "on",
                "components_json": json.dumps([{"id": component.pk, "name": "Changed", "max_marks": "50", "include_in_total": True}]),
            },
        )
        self.assertEqual(update_response.status_code, 302)
        assessment_subject.refresh_from_db()
        component.refresh_from_db()
        self.assertEqual(assessment_subject.passing_marks, Decimal("35.00"))
        self.assertEqual(component.name_snapshot, "Theory Exam")

        component_update_response = self.client.post(
            reverse("report_card:assessment_subject_component_update", args=[assessment.pk, assessment_subject.pk, component.pk]),
            data={
                "name": "Oral",
                "max_marks": "20",
                "passing_marks": "0",
                "weightage": "20",
                "display_order": "1",
                "include_in_total": "on",
            },
        )
        self.assertEqual(component_update_response.status_code, 302)
        component.refresh_from_db()
        self.assertEqual(component.name_snapshot, "Theory Exam")

        delete_response = self.client.post(
            reverse("report_card:assessment_subject_delete", args=[assessment.pk, assessment_subject.pk])
        )
        component_delete_response = self.client.post(
            reverse("report_card:assessment_subject_component_delete", args=[assessment.pk, assessment_subject.pk, component.pk])
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertEqual(component_delete_response.status_code, 302)
        self.assertTrue(ReportCardAssessmentSubject.objects.filter(pk=assessment_subject.pk).exists())
        self.assertTrue(ReportCardAssessmentSubjectComponent.objects.filter(pk=component.pk).exists())

    def test_teacher_dashboard_shows_only_allocated_subjects_and_own_completion(self):
        assessment = self._assessment("Allocated Subject Dashboard")
        math = self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("open-admin"))
        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_1, "marks_obtained": "45"}],
            actor=self.teacher,
        )
        self.client.force_login(self.teacher)

        detail_response = self.client.get(reverse("report_card:assessment_detail", args=[assessment.pk]))
        structure_response = self.client.get(reverse("report_card:assessment_structure", args=[assessment.pk]))
        completion_response = self.client.get(reverse("report_card:completion_summary", args=[assessment.pk]))

        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Mathematics")
        self.assertNotContains(detail_response, "Science")
        self.assertEqual(list(detail_response.context["subjects"]), [math])
        self.assertEqual(list(structure_response.context["subjects"]), [math])
        self.assertEqual(completion_response.context["summary"]["subject_count"], 1)
        self.assertEqual(
            completion_response.context["summary"]["subjects"][0]["assessment_subject"],
            math,
        )
        self.assertNotIn(science, [item["assessment_subject"] for item in completion_response.context["summary"]["subjects"]])

    def test_teacher_sidebar_keeps_report_cards_separate_from_online_exams(self):
        self.client.force_login(self.teacher)

        response = self.client.get(reverse("report_card:assessment_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exams")
        self.assertContains(response, "Report Cards")
        self.assertContains(response, reverse("report_card:assessment_list"))
        self.assertNotContains(response, "Recycle Bin")
        self.assertNotContains(response, "assessment_delete")
        self.assertNotContains(response, "Teacher Subject Allocation")
        self.assertNotContains(response, "Grade Rules")

    def test_teacher_marks_entry_and_template_are_limited_to_allocated_subject(self):
        assessment = self._assessment("Allocated Marks Only")
        math = self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("marks-open-admin"))
        self.client.force_login(self.teacher)

        allowed_response = self.client.post(
            reverse("report_card:marks_entry", args=[assessment.pk, math.pk]),
            data={
                f"{self.session_1.pk}-academic_session_id": self.session_1.pk,
                f"{self.session_1.pk}-marks_obtained": "40",
                f"{self.session_1.pk}-is_absent": "",
                f"{self.session_1.pk}-remark": "",
                f"{self.session_2.pk}-academic_session_id": self.session_2.pk,
                f"{self.session_2.pk}-marks_obtained": "",
                f"{self.session_2.pk}-is_absent": "",
                f"{self.session_2.pk}-remark": "",
                f"{self.session_3.pk}-academic_session_id": self.session_3.pk,
                f"{self.session_3.pk}-marks_obtained": "",
                f"{self.session_3.pk}-is_absent": "",
                f"{self.session_3.pk}-remark": "",
            },
        )
        self.assertEqual(allowed_response.status_code, 302)
        self.assertTrue(
            ReportCardMarkEntry.objects.filter(
                assessment_subject=math,
                academic_session=self.session_1,
                marks_obtained=Decimal("40.00"),
            ).exists()
        )

        blocked_response = self.client.post(
            reverse("report_card:marks_entry", args=[assessment.pk, science.pk]),
            data={
                f"{self.session_1.pk}-academic_session_id": self.session_1.pk,
                f"{self.session_1.pk}-marks_obtained": "50",
            },
        )
        self.assertEqual(blocked_response.status_code, 302)
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

        template_response = self.client.get(reverse("report_card:marks_entry_template", args=[assessment.pk, math.pk]))
        blocked_template_response = self.client.get(reverse("report_card:marks_entry_template", args=[assessment.pk, science.pk]))
        self.assertEqual(template_response.status_code, 200)
        self.assertEqual(blocked_template_response.status_code, 302)

    def test_bulk_save_subject_marks_rejects_teacher_without_allocation(self):
        assessment = self._assessment("Service Allocation Guard")
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("service-open-admin"))

        with self.assertRaisesMessage(ValidationError, "allocated report-card subject"):
            bulk_save_subject_marks(
                science,
                [{"academic_session": self.session_1, "marks_obtained": "50"}],
                actor=self.teacher,
            )

        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

    def test_bulk_save_subject_marks_allows_same_institute_admin(self):
        assessment = self._assessment("Admin Marks Save")
        science = self._subject(assessment, self.science)
        admin_user = self._admin_user("service-marks-admin")
        open_marks_entry(assessment, actor=admin_user)

        saved = bulk_save_subject_marks(
            science,
            [{"academic_session": self.session_1, "marks_obtained": "50"}],
            actor=admin_user,
        )

        self.assertEqual(len(saved), 1)
        self.assertTrue(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

    def test_teacher_url_tampering_cannot_save_marks_for_unallocated_subject(self):
        assessment = self._assessment("URL Tampering Guard")
        self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("url-tamper-admin"))
        self.client.force_login(self.teacher)

        response = self.client.post(
            reverse("report_card:marks_entry", args=[assessment.pk, science.pk]),
            data={
                f"{self.session_1.pk}-academic_session_id": self.session_1.pk,
                f"{self.session_1.pk}-marks_obtained": "70",
                f"{self.session_1.pk}-is_absent": "",
                f"{self.session_1.pk}-remark": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

    def test_excel_import_cannot_save_unallocated_subject_marks(self):
        assessment = self._assessment("Excel Allocation Guard")
        self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("excel-tamper-admin"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Academic Session ID", "Admission Number", "Student Name", "Marks", "Absent", "Remark"])
        sheet.append([self.session_1.pk, self.session_1.admission_number, "Ada Lovelace", 55, "", ""])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)

        result = import_marks_workbook(science, upload, actor=self.teacher)

        self.assertEqual(result["saved_count"], 0)
        self.assertEqual(len(result["errors"]), 1)
        self.assertIn("allocated report-card subject", result["errors"][0]["errors"][0])
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

    def test_api_marks_save_cannot_save_unallocated_subject_marks(self):
        assessment = self._assessment("API Allocation Guard")
        self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("api-tamper-admin"))
        api_client = APIClient()
        api_client.force_authenticate(user=self.teacher)

        response = api_client.post(
            f"/api/mobile/report-cards/teacher/assessments/{assessment.pk}/subjects/{science.pk}/marks/",
            data={
                "rows": [
                    {
                        "academic_session_id": self.session_1.pk,
                        "marks_obtained": "65",
                        "is_absent": False,
                        "remark": "",
                    }
                ]
            },
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=science).exists())

    def test_teacher_api_returns_only_allocated_assessments_and_subjects(self):
        visible_assessment = self._assessment("API Visible Assessment")
        math = self._subject(visible_assessment, self.math)
        science = self._subject(visible_assessment, self.science)
        hidden_assessment = self._assessment("API Hidden Assessment")
        self._subject(hidden_assessment, self.art)
        self._allocation(subject=self.math)
        api_client = APIClient()
        api_client.force_authenticate(user=self.teacher)

        list_response = api_client.get("/api/mobile/report-cards/teacher/assessments/")
        detail_response = api_client.get(f"/api/mobile/report-cards/teacher/assessments/{visible_assessment.pk}/")
        subjects_response = api_client.get(f"/api/mobile/report-cards/teacher/assessments/{visible_assessment.pk}/subjects/")
        hidden_response = api_client.get(f"/api/mobile/report-cards/teacher/assessments/{hidden_assessment.pk}/")

        self.assertEqual(list_response.status_code, 200)
        self.assertEqual([item["id"] for item in list_response.data["results"]], [visible_assessment.pk])
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual([item["id"] for item in detail_response.data["subjects"]], [math.pk])
        self.assertNotIn(science.pk, [item["id"] for item in detail_response.data["subjects"]])
        self.assertEqual(subjects_response.status_code, 200)
        self.assertEqual([item["id"] for item in subjects_response.data["results"]], [math.pk])
        self.assertEqual(hidden_response.status_code, 404)

    def test_teacher_api_saves_and_returns_component_marks_for_allocated_subject(self):
        assessment = self._assessment("API Component Marks")
        math = self._subject(assessment, self.math, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        notebook = add_assessment_subject_component(
            math,
            name="Notebook",
            max_marks=Decimal("10"),
            weightage=Decimal("10"),
            display_order=2,
            actor=self.teacher,
        )
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("api-component-admin"))
        api_client = APIClient()
        api_client.force_authenticate(user=self.teacher)

        save_response = api_client.post(
            f"/api/mobile/report-cards/teacher/assessments/{assessment.pk}/subjects/{math.pk}/marks/",
            data={
                "rows": [
                    {
                        "academic_session_id": self.session_1.pk,
                        "component_marks": {
                            str(theory.pk): "42",
                            str(notebook.pk): "8",
                        },
                        "is_absent": False,
                        "remark": "API save",
                    }
                ]
            },
            format="json",
        )
        grid_response = api_client.get(
            f"/api/mobile/report-cards/teacher/assessments/{assessment.pk}/subjects/{math.pk}/marks/"
        )

        self.assertEqual(save_response.status_code, 200)
        subject_entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_1)
        self.assertEqual(subject_entry.marks_obtained, Decimal("50.00"))
        self.assertEqual(grid_response.status_code, 200)
        self.assertEqual([component["id"] for component in grid_response.data["subject"]["components"]], [theory.pk, notebook.pk])
        first_row = next(row for row in grid_response.data["rows"] if row["academic_session_id"] == self.session_1.pk)
        self.assertEqual(first_row["mark_entry"]["marks_obtained"], "50.00")
        self.assertEqual(first_row["mark_entry"]["component_marks"][str(theory.pk)]["marks_obtained"], "42.00")
        self.assertEqual(first_row["mark_entry"]["component_marks"][str(notebook.pk)]["marks_obtained"], "8.00")

    def test_teacher_api_allows_blank_component_marks_for_partial_save(self):
        assessment = self._assessment("API Partial Component Marks")
        math = self._subject(assessment, self.math, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        notebook = add_assessment_subject_component(
            math,
            name="Notebook",
            max_marks=Decimal("10"),
            weightage=Decimal("10"),
            display_order=2,
            actor=self.teacher,
        )
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("api-partial-admin"))
        api_client = APIClient()
        api_client.force_authenticate(user=self.teacher)

        save_response = api_client.post(
            f"/api/mobile/report-cards/teacher/assessments/{assessment.pk}/subjects/{math.pk}/marks/",
            data={
                "rows": [
                    {
                        "academic_session_id": self.session_1.pk,
                        "component_marks": {
                            str(theory.pk): "42",
                            str(notebook.pk): "",
                        },
                        "is_absent": False,
                        "remark": "",
                    }
                ]
            },
            format="json",
        )

        self.assertEqual(save_response.status_code, 200)
        subject_entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_1)
        self.assertIsNone(subject_entry.marks_obtained)
        theory_entry = ReportCardComponentMarkEntry.objects.get(component=theory, academic_session=self.session_1)
        self.assertEqual(theory_entry.marks_obtained, Decimal("42.00"))
        notebook_entry = ReportCardComponentMarkEntry.objects.get(component=notebook, academic_session=self.session_1)
        self.assertIsNone(notebook_entry.marks_obtained)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)

    def test_teacher_cannot_run_report_card_admin_workflow_actions(self):
        assessment = self._assessment("Teacher Admin Actions Blocked")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        self.client.force_login(self.teacher)

        blocked_posts = [
            reverse("report_card:open_marks_entry", args=[assessment.pk]),
            reverse("report_card:generate_results", args=[assessment.pk]),
            reverse("report_card:publish_results", args=[assessment.pk]),
            reverse("report_card:lock_assessment", args=[assessment.pk]),
        ]
        for url in blocked_posts:
            response = self.client.post(url)
            self.assertEqual(response.status_code, 302)

        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.STRUCTURE_READY)

        results_response = self.client.get(reverse("report_card:results_preview", args=[assessment.pk]))
        export_response = self.client.get(reverse("report_card:results_export", args=[assessment.pk]))
        self.assertEqual(results_response.status_code, 302)
        self.assertEqual(export_response.status_code, 302)

    def test_admin_structure_shows_warnings_before_opening_marks_entry(self):
        admin_user = self._admin_user("open-warning-admin")
        no_subject_assessment = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.batch,
            title="No Subject Warning",
            created_by=admin_user,
        )
        assessment = self._assessment("Missing Allocation Warning")
        self._subject(assessment, self.math)
        self.client.force_login(admin_user)

        no_subject_response = self.client.get(
            reverse("report_card_admin:assessment_structure", args=[no_subject_assessment.pk])
        )
        missing_allocation_response = self.client.get(
            reverse("report_card_admin:assessment_structure", args=[assessment.pk])
        )

        self.assertContains(no_subject_response, "Add at least one subject before opening marks entry.")
        self.assertContains(missing_allocation_response, "Teacher allocation is missing for: Mathematics")

        blocked_response = self.client.post(reverse("report_card_admin:open_marks_entry", args=[assessment.pk]))
        self.assertEqual(blocked_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.STRUCTURE_READY)

    def test_admin_can_open_marks_entry_after_structure_and_allocations_are_ready(self):
        admin_user = self._admin_user("open-marks-admin")
        assessment = self._assessment("Ready For Marks")
        self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        self.client.force_login(admin_user)

        response = self.client.post(reverse("report_card_admin:open_marks_entry", args=[assessment.pk]))

        self.assertEqual(response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)

    def test_admin_can_reopen_completed_generated_published_and_locked_marks_entry(self):
        admin_user = self._admin_user("reopen-service-admin")
        for status in [
            ReportCardAssessment.Status.MARKS_ENTRY_COMPLETED,
            ReportCardAssessment.Status.GENERATED,
            ReportCardAssessment.Status.PUBLISHED,
            ReportCardAssessment.Status.LOCKED,
        ]:
            assessment = self._assessment(f"Reopen {status}")
            self._subject(assessment)
            assessment.status = status
            if status in {ReportCardAssessment.Status.PUBLISHED, ReportCardAssessment.Status.LOCKED}:
                assessment.published_at = timezone.now()
                assessment.published_by = admin_user
            if status == ReportCardAssessment.Status.LOCKED:
                assessment.locked_at = timezone.now()
                assessment.locked_by = admin_user
            assessment.save()

            reopen_marks_entry(assessment, actor=admin_user, reason="Allow teachers to correct entered marks.")

            assessment.refresh_from_db()
            self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)
            self.assertIsNone(assessment.published_at)
            self.assertIsNone(assessment.published_by)
            self.assertIsNone(assessment.locked_at)
            self.assertIsNone(assessment.locked_by)

    def test_reopen_marks_entry_marks_generated_results_stale(self):
        admin_user = self._admin_user("reopen-stale-admin")
        assessment = self._assessment("Reopen Stale Results")
        subject = self._subject(assessment)
        self._ensure_allocation(subject)
        self._open_and_save(
            subject,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "60"},
            ],
        )
        generate_assessment_results(assessment, actor=admin_user)

        reopen_marks_entry(assessment, actor=admin_user, reason="Marks correction required.")

        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)
        self.assertEqual(ReportCardStudentResult.objects.filter(assessment=assessment, is_stale=True).count(), 3)

    def test_admin_reopen_marks_entry_view_and_subject_route(self):
        admin_user = self._admin_user("reopen-view-admin")
        assessment = self._assessment("Reopen View")
        subject = self._subject(assessment)
        assessment.status = ReportCardAssessment.Status.MARKS_ENTRY_COMPLETED
        assessment.save()
        self.client.force_login(admin_user)

        detail_response = self.client.get(reverse("report_card_admin:assessment_detail", args=[assessment.pk]))
        subject_response = self.client.post(
            reverse("report_card_admin:reopen_subject_marks_entry", args=[assessment.pk, subject.pk]),
            data={"reason": "Subject marks need correction."},
        )

        self.assertContains(detail_response, "Reopen Marks")
        self.assertEqual(subject_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)
        self.assertTrue(
            ReportCardAuditLog.objects.filter(
                assessment=assessment,
                action=ReportCardAuditLog.Action.MARKS_ENTRY_OPENED,
                metadata__assessment_subject_id=subject.pk,
            ).exists()
        )

    def test_teacher_cannot_reopen_marks_entry_from_admin_routes(self):
        assessment = self._assessment("Teacher Reopen Blocked")
        subject = self._subject(assessment)
        assessment.status = ReportCardAssessment.Status.MARKS_ENTRY_COMPLETED
        assessment.save()
        self.client.force_login(self.teacher)

        whole_response = self.client.post(reverse("report_card_admin:reopen_marks_entry", args=[assessment.pk]))
        subject_response = self.client.post(
            reverse("report_card_admin:reopen_subject_marks_entry", args=[assessment.pk, subject.pk])
        )

        self.assertNotEqual(whole_response.status_code, 200)
        self.assertNotEqual(subject_response.status_code, 200)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_COMPLETED)

    def test_teacher_can_enter_marks_only_after_admin_opens_marks_entry(self):
        admin_user = self._admin_user("teacher-open-admin")
        assessment = self._assessment("Teacher Waits For Admin")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        self.client.force_login(self.teacher)

        before_open_response = self.client.post(
            reverse("report_card:marks_entry", args=[assessment.pk, math.pk]),
            data={
                f"{self.session_1.pk}-academic_session_id": self.session_1.pk,
                f"{self.session_1.pk}-marks_obtained": "40",
            },
        )
        self.assertEqual(before_open_response.status_code, 302)
        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=math).exists())

        self.client.force_login(admin_user)
        self.client.post(reverse("report_card_admin:open_marks_entry", args=[assessment.pk]))
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)

        self.client.force_login(self.teacher)
        after_open_response = self.client.post(
            reverse("report_card:marks_entry", args=[assessment.pk, math.pk]),
            data={
                f"{self.session_1.pk}-academic_session_id": self.session_1.pk,
                f"{self.session_1.pk}-marks_obtained": "40",
                f"{self.session_1.pk}-is_absent": "",
                f"{self.session_1.pk}-remark": "",
                f"{self.session_2.pk}-academic_session_id": self.session_2.pk,
                f"{self.session_2.pk}-marks_obtained": "",
                f"{self.session_2.pk}-is_absent": "",
                f"{self.session_2.pk}-remark": "",
                f"{self.session_3.pk}-academic_session_id": self.session_3.pk,
                f"{self.session_3.pk}-marks_obtained": "",
                f"{self.session_3.pk}-is_absent": "",
                f"{self.session_3.pk}-remark": "",
            },
        )

        self.assertEqual(after_open_response.status_code, 302)
        self.assertTrue(
            ReportCardMarkEntry.objects.filter(
                assessment_subject=math,
                academic_session=self.session_1,
                marks_obtained=Decimal("40.00"),
            ).exists()
        )

    def test_institute_admin_completion_dashboard_shows_subject_and_teacher_pending_work(self):
        admin_user = self._admin_user("completion-admin")
        assessment = self._assessment("Completion Dashboard")
        math = self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math, teacher=self.teacher)
        self._allocation(subject=self.science, teacher=self.other_teacher)
        open_marks_entry(assessment, actor=admin_user)
        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_1, "marks_obtained": "45"}],
            actor=self.teacher,
        )
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:completion_dashboard", args=[assessment.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["summary"]["student_count"], 3)
        self.assertEqual(response.context["summary"]["expected_mark_count"], 6)
        self.assertEqual(response.context["summary"]["entered_mark_count"], 1)
        self.assertEqual(response.context["summary"]["missing_mark_count"], 5)
        subject_rows = {row["assessment_subject"]: row for row in response.context["rows"]}
        self.assertEqual(subject_rows[math]["missing_mark_count"], 2)
        self.assertEqual(subject_rows[science]["missing_mark_count"], 3)
        self.assertEqual(len(response.context["teacher_rows"]), 2)
        self.assertContains(response, "report-teacher")
        self.assertContains(response, "other-report-teacher")
        self.assertContains(response, "Grid")

    def test_institute_admin_can_inspect_subject_marks_grid_read_only(self):
        admin_user = self._admin_user("grid-admin")
        assessment = self._assessment("Admin Grid")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=admin_user)
        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_1, "marks_obtained": "45"}],
            actor=self.teacher,
        )
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:marks_grid", args=[assessment.pk, math.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mathematics Marks Grid")
        self.assertContains(response, "A001")
        self.assertContains(response, "45.00")

    def test_teacher_completion_page_remains_limited_to_assigned_subjects(self):
        assessment = self._assessment("Teacher Completion Scope")
        math = self._subject(assessment, self.math)
        science = self._subject(assessment, self.science)
        self._allocation(subject=self.math, teacher=self.teacher)
        self._allocation(subject=self.science, teacher=self.other_teacher)
        open_marks_entry(assessment, actor=self._admin_user("scope-admin"))
        self.client.force_login(self.teacher)

        response = self.client.get(reverse("report_card:completion_summary", args=[assessment.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["summary"]["subject_count"], 1)
        self.assertEqual(response.context["summary"]["subjects"][0]["assessment_subject"], math)
        self.assertNotIn(science, [item["assessment_subject"] for item in response.context["summary"]["subjects"]])

    def test_admin_results_preview_shows_generation_warnings(self):
        admin_user = self._admin_user("result-warning-admin")
        assessment = self._assessment("Result Warning")
        self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=admin_user)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:results_preview", args=[assessment.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "required mark field(s) are still missing")
        self.assertContains(response, "No active grade rules are configured")

    def test_admin_can_generate_publish_and_lock_report_card_results(self):
        admin_user = self._admin_user("result-admin")
        assessment = self._assessment("Admin Result Lifecycle")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=admin_user)
        bulk_save_subject_marks(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "90"},
                {"academic_session": self.session_2, "marks_obtained": "80"},
                {"academic_session": self.session_3, "marks_obtained": "70"},
            ],
            actor=self.teacher,
        )
        self.client.force_login(admin_user)

        generate_response = self.client.post(reverse("report_card_admin:generate_results", args=[assessment.pk]))
        self.assertEqual(generate_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.GENERATED)
        self.assertEqual(ReportCardStudentResult.objects.filter(assessment=assessment, is_stale=False).count(), 3)

        publish_response = self.client.post(reverse("report_card_admin:publish_results", args=[assessment.pk]))
        self.assertEqual(publish_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.PUBLISHED)

        lock_response = self.client.post(reverse("report_card_admin:lock_assessment", args=[assessment.pk]))
        self.assertEqual(lock_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.LOCKED)

    def test_admin_publish_blocks_stale_results_until_regenerated(self):
        admin_user = self._admin_user("stale-result-admin")
        assessment = self._assessment("Stale Result Safety")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=admin_user)
        bulk_save_subject_marks(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "90"},
                {"academic_session": self.session_2, "marks_obtained": "80"},
                {"academic_session": self.session_3, "marks_obtained": "70"},
            ],
            actor=self.teacher,
        )
        generate_assessment_results(assessment, actor=admin_user)
        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_1, "marks_obtained": "91"}],
            actor=self.teacher,
        )
        self.client.force_login(admin_user)

        preview_response = self.client.get(reverse("report_card_admin:results_preview", args=[assessment.pk]))
        publish_response = self.client.post(reverse("report_card_admin:publish_results", args=[assessment.pk]))

        self.assertContains(preview_response, "generated result(s) are stale")
        self.assertEqual(publish_response.status_code, 302)
        assessment.refresh_from_db()
        self.assertNotEqual(assessment.status, ReportCardAssessment.Status.PUBLISHED)

    def test_teacher_cannot_generate_publish_or_lock_after_admin_move(self):
        assessment = self._assessment("Teacher Result Lifecycle Blocked")
        math = self._subject(assessment, self.math)
        self._allocation(subject=self.math)
        open_marks_entry(assessment, actor=self._admin_user("teacher-block-admin"))
        self.client.force_login(self.teacher)

        for url in [
            reverse("report_card:generate_results", args=[assessment.pk]),
            reverse("report_card:publish_results", args=[assessment.pk]),
            reverse("report_card:lock_assessment", args=[assessment.pk]),
        ]:
            response = self.client.post(url)
            self.assertEqual(response.status_code, 302)

        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)
        self.assertFalse(ReportCardStudentResult.objects.filter(assessment=assessment).exists())

    def test_report_card_assessment_creation_does_not_create_or_modify_teacher_exam(self):
        exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.math,
            title="Online Unit Test",
            exam_date=date(2026, 7, 1),
            total_marks=20,
            created_by=self.teacher,
        )
        before = list(Exam.objects.values("id", "title", "total_marks", "is_published"))

        assessment = self._assessment()

        self.assertEqual(Exam.objects.count(), 1)
        self.assertEqual(list(Exam.objects.values("id", "title", "total_marks", "is_published")), before)
        self.assertNotEqual(assessment.title, exam.title)

    def test_report_card_result_generation_does_not_create_or_modify_teacher_exam_result(self):
        exam = Exam.objects.create(
            academic_year=self.year,
            batch=self.batch,
            course=self.course,
            subject=self.math,
            title="Online Final",
            exam_date=date(2026, 8, 1),
            total_marks=50,
            created_by=self.teacher,
        )
        ExamResult.objects.create(
            exam=exam,
            student=self.session_1.student,
            marks_obtained=Decimal("44"),
            remark="Online result remains separate.",
        )
        before = list(ExamResult.objects.values("id", "exam_id", "student_id", "marks_obtained", "remark"))
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "60"},
            ],
        )

        generate_assessment_results(assessment, actor=self.teacher)

        self.assertEqual(ExamResult.objects.count(), 1)
        self.assertEqual(list(ExamResult.objects.values("id", "exam_id", "student_id", "marks_obtained", "remark")), before)

    def test_teacher_sees_only_assigned_batch_assessments(self):
        assigned = self._assessment("Assigned Assessment")
        unassigned = create_assessment(
            institute=self.institute,
            academic_year=self.year,
            batch=self.unassigned_batch,
            title="Unassigned Assessment",
            created_by=self.other_teacher,
        )

        assessments = list(get_assessments_for_teacher(self.teacher, academic_year=self.year))

        self.assertIn(assigned, assessments)
        self.assertNotIn(unassigned, assessments)

    def test_teacher_profile_institute_is_used_for_report_card_scoping(self):
        self._allocation(subject=self.math)
        self.teacher.profile.institute = self.other_institute
        self.teacher.profile.save(update_fields=["institute"])
        TeacherProfile.objects.create(user=self.teacher, institute=self.institute)
        assessment = self._assessment("Teacher Profile Scoped")
        self._subject(assessment, self.math)

        assessments = list(get_assessments_for_teacher(self.teacher, academic_year=self.year))

        self.assertIn(assessment, assessments)
        self.assertFalse(teacher_can_edit_assessment(self.teacher, assessment))

    def test_assigned_batch_institute_is_used_when_user_profile_institute_is_stale(self):
        self.teacher.profile.institute = self.other_institute
        self.teacher.profile.save(update_fields=["institute"])
        assessment = self._assessment("Assigned Batch Scoped")
        form = ReportCardAssessmentForm(
            data={
                "academic_year": self.year.pk,
                "batch": self.batch.pk,
                "title": "Mid Term",
                "assessment_date": "2026-07-19",
                "result_date": "2026-07-21",
            },
            user=self.teacher,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.effective_institute, self.institute)
        self.assertIn(assessment, list(get_assessments_for_teacher(self.teacher, academic_year=self.year)))

    def test_batch_and_subject_fk_validation_works(self):
        with self.assertRaises(ValidationError):
            create_assessment(
                institute=self.institute,
                academic_year=self.year,
                batch=Batch.objects.create(
                    institute=self.institute,
                    academic_year=self.next_year,
                    name="Future Batch",
                ),
                title="Wrong Batch Year",
                created_by=self.teacher,
            )

        assessment = self._assessment()
        with self.assertRaises(ValidationError):
            self._subject(assessment, subject=self.other_subject)
        with self.assertRaises(ValidationError):
            self._subject(assessment, subject=self.next_year_subject)

    def test_snapshot_fields_are_copied_correctly(self):
        assessment = self._assessment()
        subject = self._subject(assessment, self.math)
        self._open_and_save(subject, [{"academic_session": self.session_1, "marks_obtained": "78"}])
        mark = ReportCardMarkEntry.objects.get(assessment_subject=subject, academic_session=self.session_1)

        self.assertEqual(assessment.institute_name_snapshot, self.institute.name)
        self.assertEqual(assessment.academic_year_name_snapshot, self.year.name)
        self.assertEqual(assessment.batch_name_snapshot, self.batch.name)
        self.assertEqual(subject.subject_name_snapshot, self.math.name)
        self.assertEqual(mark.student_name_snapshot, "Ada Lovelace")
        self.assertEqual(mark.admission_number_snapshot, "A001")

    def test_different_subjects_per_assessment_work(self):
        first = self._assessment("Term One")
        second = self._assessment("Term Two")
        first_math = self._subject(first, self.math)
        first_science = self._subject(first, self.science)
        second_art = self._subject(second, self.art)

        self.assertEqual(set(first.assessment_subjects.values_list("subject_id", flat=True)), {self.math.pk, self.science.pk})
        self.assertEqual(list(second.assessment_subjects.values_list("subject_id", flat=True)), [self.art.pk])
        self.assertNotEqual(first_math.pk, second_art.pk)
        self.assertNotEqual(first_science.subject_id, second_art.subject_id)

    def test_different_max_marks_and_weightage_work(self):
        assessment = self._assessment()
        math = self._subject(
            assessment,
            self.math,
            max_marks=Decimal("50"),
            passing_marks=Decimal("18"),
            weightage=Decimal("50"),
        )
        science = self._subject(
            assessment,
            self.science,
            max_marks=Decimal("100"),
            passing_marks=Decimal("35"),
            weightage=Decimal("150"),
        )
        open_marks_entry(assessment, actor=self.teacher)
        self._ensure_allocation(math)
        self._ensure_allocation(science)
        bulk_save_subject_marks(math, [{"academic_session": self.session_1, "marks_obtained": "45"}], actor=self.teacher)
        bulk_save_subject_marks(science, [{"academic_session": self.session_1, "marks_obtained": "60"}], actor=self.teacher)

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.total_obtained, Decimal("105.00"))
        self.assertEqual(result.total_max_marks, Decimal("150.00"))
        self.assertEqual(result.total_weightage, Decimal("200.00"))
        self.assertEqual(result.percentage, Decimal("67.50"))

    def test_bulk_marks_save_creates_and_updates_marks_correctly(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "76", "remark": "Initial"},
                {"academic_session": self.session_2, "marks_obtained": "66"},
            ],
        )
        self.assertEqual(ReportCardMarkEntry.objects.filter(assessment_subject=math).count(), 2)

        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_1, "marks_obtained": "88", "remark": "Updated"}],
            actor=self.teacher,
        )
        entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_1)

        self.assertEqual(ReportCardMarkEntry.objects.filter(assessment_subject=math).count(), 2)
        self.assertEqual(entry.marks_obtained, Decimal("88.00"))
        self.assertEqual(entry.remark, "Updated")

    def test_bulk_marks_save_allows_partial_single_column_marks(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "76"},
                {"academic_session": self.session_2, "marks_obtained": ""},
            ],
        )

        blank_entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_2)
        summary = get_completion_summary(assessment)

        self.assertIsNone(blank_entry.marks_obtained)
        self.assertEqual(summary["entered_mark_count"], 1)
        self.assertEqual(summary["missing_mark_count"], 2)
        self.assertFalse(summary["is_complete"])

    def test_incomplete_resave_moves_completed_assessment_back_to_open(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "76"},
                {"academic_session": self.session_2, "marks_obtained": "82"},
                {"academic_session": self.session_3, "marks_obtained": "68"},
            ],
        )
        assessment.refresh_from_db()
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_COMPLETED)

        bulk_save_subject_marks(
            math,
            [{"academic_session": self.session_2, "marks_obtained": ""}],
            actor=self.teacher,
        )

        assessment.refresh_from_db()
        summary = get_completion_summary(assessment)
        self.assertEqual(assessment.status, ReportCardAssessment.Status.MARKS_ENTRY_OPEN)
        self.assertFalse(summary["is_complete"])
        self.assertEqual(summary["missing_mark_count"], 1)

    def test_bulk_marks_save_allows_partial_component_marks(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        viva = add_assessment_subject_component(
            math,
            name="Viva",
            max_marks=Decimal("10"),
            weightage=Decimal("10"),
            display_order=2,
            actor=self.teacher,
        )
        self._open_and_save(
            math,
            [
                {
                    "academic_session": self.session_1,
                    "component_marks": {theory.pk: "40", viva.pk: ""},
                }
            ],
        )

        subject_entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_1)
        theory_entry = ReportCardComponentMarkEntry.objects.get(component=theory, academic_session=self.session_1)
        viva_entry = ReportCardComponentMarkEntry.objects.get(component=viva, academic_session=self.session_1)
        summary = get_completion_summary(assessment)

        self.assertIsNone(subject_entry.marks_obtained)
        self.assertEqual(theory_entry.marks_obtained, Decimal("40.00"))
        self.assertIsNone(viva_entry.marks_obtained)
        self.assertEqual(summary["entered_mark_count"], 1)
        self.assertEqual(summary["missing_mark_count"], 5)
        self.assertFalse(summary["is_complete"])

    def test_excel_import_allows_partial_component_marks(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        notebook = add_assessment_subject_component(
            math,
            name="Notebook",
            max_marks=Decimal("10"),
            weightage=Decimal("10"),
            display_order=2,
            actor=self.teacher,
        )
        open_marks_entry(assessment, actor=self.teacher)
        self._ensure_allocation(math)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Academic Session ID", "Admission Number", "Student Name", "Theory Exam", "Notebook", "Absent", "Remark"])
        sheet.append([self.session_1.pk, self.session_1.admission_number, "Ada Lovelace", 42, "", "", "Notebook later"])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)

        result = import_marks_workbook(math, upload, actor=self.teacher)

        self.assertEqual(result["errors"], [])
        self.assertEqual(result["saved_count"], 1)
        self.assertEqual(len(result["warnings"]), 1)
        self.assertIn("Notebook", result["warnings"][0]["missing_fields"])
        subject_entry = ReportCardMarkEntry.objects.get(assessment_subject=math, academic_session=self.session_1)
        theory_entry = ReportCardComponentMarkEntry.objects.get(component=theory, academic_session=self.session_1)
        notebook_entry = ReportCardComponentMarkEntry.objects.get(component=notebook, academic_session=self.session_1)
        self.assertIsNone(subject_entry.marks_obtained)
        self.assertEqual(theory_entry.marks_obtained, Decimal("42.00"))
        self.assertIsNone(notebook_entry.marks_obtained)

    def test_excel_import_rejects_invalid_filled_component_marks(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("50"), passing_marks=Decimal("18"))
        add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        open_marks_entry(assessment, actor=self.teacher)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Academic Session ID", "Admission Number", "Student Name", "Theory Exam", "Absent", "Remark"])
        sheet.append([self.session_1.pk, self.session_1.admission_number, "Ada Lovelace", 55, "", ""])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)

        result = import_marks_workbook(math, upload, actor=self.teacher)

        self.assertEqual(result["saved_count"], 0)
        self.assertEqual(len(result["errors"]), 1)
        self.assertIn("cannot exceed", result["errors"][0]["errors"][0])

    def test_structure_edit_is_blocked_after_marks_entry_opens(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        open_marks_entry(assessment, actor=self.teacher)

        with self.assertRaisesMessage(ValidationError, "before marks entry opens"):
            sync_assessment_subject_components(
                math,
                [
                    {
                        "id": theory.pk,
                        "name": "Theory Exam",
                        "max_marks": Decimal("60"),
                        "display_order": 1,
                        "include_in_total": True,
                    }
                ],
                actor=self.teacher,
            )

        theory.refresh_from_db()
        self.assertEqual(theory.max_marks, Decimal("50.00"))

    def test_component_sync_updates_existing_components_without_recreating_them(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("60"), passing_marks=Decimal("21"))
        theory = add_assessment_subject_component(
            math,
            name="Theory Exam",
            max_marks=Decimal("50"),
            weightage=Decimal("50"),
            display_order=1,
            actor=self.teacher,
        )
        notebook = add_assessment_subject_component(
            math,
            name="Notebook",
            max_marks=Decimal("10"),
            weightage=Decimal("10"),
            display_order=2,
            actor=self.teacher,
        )
        mark_entry = ReportCardComponentMarkEntry.objects.create(
            component=theory,
            academic_session=self.session_1,
            marks_obtained=Decimal("40"),
            entered_by=self.teacher,
            updated_by=self.teacher,
        )

        sync_assessment_subject_components(
            math,
            [
                {
                    "id": notebook.pk,
                    "name": "Notebook",
                    "max_marks": Decimal("10"),
                    "display_order": 1,
                    "include_in_total": True,
                },
                {
                    "id": theory.pk,
                    "name": "Theory Exam",
                    "max_marks": Decimal("50"),
                    "display_order": 2,
                    "include_in_total": True,
                },
            ],
            actor=self.teacher,
        )

        theory.refresh_from_db()
        notebook.refresh_from_db()
        mark_entry.refresh_from_db()
        self.assertEqual(theory.display_order, 2)
        self.assertEqual(notebook.display_order, 1)
        self.assertEqual(mark_entry.component_id, theory.pk)
        self.assertEqual(mark_entry.marks_obtained, Decimal("40.00"))

    def test_marks_above_max_and_negative_marks_are_rejected(self):
        assessment = self._assessment()
        math = self._subject(assessment, max_marks=Decimal("50"), passing_marks=Decimal("18"))
        open_marks_entry(assessment, actor=self.teacher)

        with self.assertRaises(ValidationError):
            bulk_save_subject_marks(
                math,
                [{"academic_session": self.session_1, "marks_obtained": "51"}],
                actor=self.teacher,
            )
        with self.assertRaises(ValidationError):
            bulk_save_subject_marks(
                math,
                [{"academic_session": self.session_1, "marks_obtained": "-1"}],
                actor=self.teacher,
            )

        self.assertFalse(ReportCardMarkEntry.objects.filter(assessment_subject=math).exists())

    def test_absent_students_are_handled_correctly(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(math, [{"academic_session": self.session_1, "is_absent": True}])

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.result_status, ReportCardStudentResult.ResultStatus.ABSENT)
        self.assertIsNone(result.percentage)

    def test_include_in_total_false_subjects_are_excluded_from_percentage(self):
        assessment = self._assessment()
        math = self._subject(assessment, self.math, max_marks=Decimal("50"), passing_marks=Decimal("18"))
        art = self._subject(
            assessment,
            self.art,
            max_marks=Decimal("100"),
            passing_marks=Decimal("0"),
            include_in_total=False,
        )
        open_marks_entry(assessment, actor=self.teacher)
        self._ensure_allocation(math)
        self._ensure_allocation(art)
        bulk_save_subject_marks(math, [{"academic_session": self.session_1, "marks_obtained": "50"}], actor=self.teacher)
        bulk_save_subject_marks(art, [{"academic_session": self.session_1, "marks_obtained": "10"}], actor=self.teacher)

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.total_obtained, Decimal("50.00"))
        self.assertEqual(result.total_max_marks, Decimal("50.00"))
        self.assertEqual(result.percentage, Decimal("100.00"))

    def test_optional_subjects_work_correctly(self):
        assessment = self._assessment()
        math = self._subject(assessment, self.math)
        self._subject(assessment, self.art, is_optional=True, display_order=2)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "75"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "65"},
            ],
        )

        results = generate_assessment_results(assessment, actor=self.teacher)

        self.assertEqual(len(results), 3)
        self.assertFalse(any(result.result_status == ReportCardStudentResult.ResultStatus.INCOMPLETE for result in results))

    def test_passing_marks_affect_pass_fail(self):
        assessment = self._assessment()
        math = self._subject(assessment, self.math, passing_marks=Decimal("35"))
        self._open_and_save(math, [{"academic_session": self.session_1, "marks_obtained": "34"}])

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.result_status, ReportCardStudentResult.ResultStatus.FAIL)
        self.assertIsNone(result.rank)

    def test_grade_rules_apply_correctly(self):
        ReportCardGradeRule.objects.create(
            institute=self.institute,
            academic_year=self.year,
            min_percentage=Decimal("80"),
            max_percentage=Decimal("100"),
            grade="A",
            remark="Excellent",
        )
        ReportCardGradeRule.objects.create(
            institute=self.institute,
            academic_year=self.year,
            min_percentage=Decimal("0"),
            max_percentage=Decimal("79.99"),
            grade="B",
        )
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(math, [{"academic_session": self.session_1, "marks_obtained": "88"}])

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.grade, "A")
        self.assertEqual(result.remark, "Excellent")

    def test_generation_without_grade_rules_keeps_grade_blank(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(math, [{"academic_session": self.session_1, "marks_obtained": "88"}])

        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        self.assertEqual(result.grade, "")
        self.assertEqual(result.remark, "")

    def test_institute_admin_can_create_default_grade_rules(self):
        admin_user = User.objects.create_user(username="grade-admin", password="pass")
        UserProfile.objects.create(
            user=admin_user,
            institute=self.institute,
            role=UserProfile.Role.INSTITUTE_ADMIN,
        )
        self.client.force_login(admin_user)

        response = self.client.post(reverse("report_card_admin:grade_rule_create_defaults"))

        self.assertEqual(response.status_code, 302)
        rules = ReportCardGradeRule.objects.filter(institute=self.institute, academic_year=self.year).order_by("display_order")
        self.assertEqual(rules.count(), 7)
        self.assertEqual(rules.first().grade, "A1")
        self.assertEqual(rules.last().grade, "F")
        self.assertFalse(ReportCardGradeRule.objects.filter(institute=self.institute, academic_year__isnull=True).exists())

    def test_default_grade_rule_creation_warns_when_active_defaults_exist(self):
        admin_user = self._admin_user("grade-existing-admin")
        ReportCardGradeRule.objects.create(
            institute=self.institute,
            academic_year=self.year,
            min_percentage=Decimal("90"),
            max_percentage=Decimal("100"),
            grade="A1",
            display_order=1,
            is_active=True,
        )
        self.client.force_login(admin_user)

        list_response = self.client.get(reverse("report_card_admin:grade_rule_list"))
        create_response = self.client.post(reverse("report_card_admin:grade_rule_create_defaults"), follow=True)

        self.assertContains(list_response, "Default Grade Rules")
        self.assertContains(create_response, f"Active grade rules already exist for {self.year.name}")
        self.assertEqual(ReportCardGradeRule.objects.filter(institute=self.institute, academic_year=self.year).count(), 1)

    def test_institute_admin_can_create_academic_year_default_grade_rules(self):
        admin_user = self._admin_user("grade-year-admin")
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("report_card_admin:grade_rule_create_defaults"),
            data={"scope": "academic_year", "academic_year_id": self.year.pk},
        )

        self.assertEqual(response.status_code, 302)
        rules = ReportCardGradeRule.objects.filter(institute=self.institute, academic_year=self.year).order_by("display_order")
        self.assertEqual(rules.count(), 7)
        self.assertEqual(rules.first().grade, "A1")
        self.assertFalse(ReportCardGradeRule.objects.filter(institute=self.institute, academic_year__isnull=True).exists())

    def test_grade_rule_list_shows_assessments_that_may_generate_blank_grades(self):
        admin_user = self._admin_user("grade-risk-admin")
        assessment = self._assessment("Blank Grade Risk")
        self._subject(assessment, self.math)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("report_card_admin:grade_rule_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assessments That May Generate Blank Grades")
        self.assertContains(response, "Blank Grade Risk")

    def test_teacher_cannot_manage_report_card_grade_rules(self):
        self.client.force_login(self.teacher)

        list_response = self.client.get(reverse("report_card_admin:grade_rule_list"))
        create_response = self.client.post(reverse("report_card_admin:grade_rule_create_defaults"))

        self.assertNotEqual(list_response.status_code, 200)
        self.assertNotEqual(create_response.status_code, 200)
        self.assertFalse(ReportCardGradeRule.objects.filter(institute=self.institute).exists())

    def test_grade_rule_form_accepts_matching_academic_year(self):
        form = ReportCardGradeRuleForm(
            data={
                "academic_year": self.year.pk,
                "min_percentage": "91",
                "max_percentage": "100",
                "grade": "A++",
                "remark": "A++",
                "display_order": "1",
                "is_active": "on",
            },
            institute=self.institute,
        )

        self.assertTrue(form.is_valid(), form.errors.as_text())
        rule = form.save()
        self.assertEqual(rule.institute, self.institute)
        self.assertEqual(rule.academic_year, self.year)

    def test_grade_rule_form_handles_missing_percentage_without_crashing(self):
        form = ReportCardGradeRuleForm(
            data={
                "academic_year": self.year.pk,
                "min_percentage": "91",
                "max_percentage": "",
                "grade": "A++",
                "remark": "A++",
                "display_order": "1",
                "is_active": "on",
            },
            institute=self.institute,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("max_percentage", form.errors)

    def test_custom_subject_components_generate_subject_grade(self):
        ReportCardGradeRule.objects.create(
            institute=self.institute,
            academic_year=self.year,
            min_percentage=Decimal("90"),
            max_percentage=Decimal("100"),
            grade="A1",
        )
        ReportCardGradeRule.objects.create(
            institute=self.institute,
            academic_year=self.year,
            min_percentage=Decimal("70"),
            max_percentage=Decimal("89.99"),
            grade="B1",
        )
        assessment = self._assessment()
        english = self._subject(assessment, self.math, max_marks=Decimal("100"), passing_marks=Decimal("0"))
        unit = add_assessment_subject_component(
            english,
            name="Unit Test",
            max_marks=Decimal("10"),
            passing_marks=Decimal("0"),
            display_order=1,
            actor=self.teacher,
        )
        notebook = add_assessment_subject_component(
            english,
            name="Notebook",
            max_marks=Decimal("10"),
            passing_marks=Decimal("0"),
            display_order=2,
            actor=self.teacher,
        )
        theory = add_assessment_subject_component(
            english,
            name="Theory",
            max_marks=Decimal("80"),
            passing_marks=Decimal("0"),
            display_order=3,
            actor=self.teacher,
        )
        open_marks_entry(assessment, actor=self.teacher)
        self._ensure_allocation(english)

        bulk_save_subject_marks(
            english,
            [
                {
                    "academic_session": self.session_1,
                    "component_marks": {
                        unit.pk: "10",
                        notebook.pk: "9",
                        theory.pk: "75",
                    },
                }
            ],
            actor=self.teacher,
        )
        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]
        subject_result = ReportCardSubjectResult.objects.get(result=result, assessment_subject=english)

        self.assertEqual(subject_result.obtained_marks, Decimal("94.00"))
        self.assertEqual(subject_result.percentage, Decimal("94.00"))
        self.assertEqual(subject_result.grade, "A1")
        self.assertEqual(result.total_obtained, Decimal("94.00"))
        self.assertEqual(ReportCardAssessmentSubjectComponent.objects.filter(assessment_subject=english).count(), 3)

    def test_subject_form_saves_local_additional_columns_on_submit(self):
        assessment = self._assessment()
        components = [
            {
                "id": None,
                "name": "Notebook",
                "max_marks": "10",
                "passing_marks": "999",
                "weightage": "999",
                "display_order": 1,
                "include_in_total": True,
            },
            {
                "id": None,
                "name": "Theory",
                "max_marks": "90",
                "passing_marks": "999",
                "weightage": "999",
                "display_order": 2,
                "include_in_total": True,
            },
        ]
        form = ReportCardAssessmentSubjectForm(
            data={
                "subject": self.math.pk,
                "max_marks": "100",
                "passing_marks": "0",
                "weightage": "100",
                "display_order": "1",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(components),
            },
            assessment=assessment,
        )

        self.assertTrue(form.is_valid(), form.errors)
        assessment_subject = add_assessment_subject(
            assessment,
            subject=form.cleaned_data["subject"],
            max_marks=form.cleaned_data["max_marks"],
            passing_marks=form.cleaned_data["passing_marks"],
            weightage=form.cleaned_data["weightage"],
            display_order=form.cleaned_data["display_order"],
            is_optional=form.cleaned_data["is_optional"],
            include_in_total=form.cleaned_data["include_in_total"],
            actor=self.teacher,
        )
        sync_assessment_subject_components(
            assessment_subject,
            form.cleaned_data["components"],
            actor=self.teacher,
        )

        self.assertEqual(
            list(assessment_subject.components.values_list("name_snapshot", "max_marks", "weightage")),
            [("Notebook", Decimal("10.00"), Decimal("10.00")), ("Theory", Decimal("90.00"), Decimal("90.00"))],
        )
        self.assertEqual(
            list(assessment_subject.components.values_list("passing_marks", flat=True)),
            [Decimal("0.00"), Decimal("0.00")],
        )

    def test_subject_display_order_is_assigned_automatically(self):
        assessment = self._assessment()
        first = self._subject(assessment, self.math)
        form = ReportCardAssessmentSubjectForm(
            data={
                "subject": self.science.pk,
                "passing_marks": "30",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(
                    [
                        {
                            "name": "Theory Exam",
                            "max_marks": "80",
                            "include_in_total": True,
                        }
                    ]
                ),
            },
            assessment=assessment,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["display_order"], first.display_order + 1)

    def test_subject_passing_marks_cannot_exceed_additional_column_total(self):
        assessment = self._assessment()
        form = ReportCardAssessmentSubjectForm(
            data={
                "subject": self.math.pk,
                "max_marks": "80",
                "passing_marks": "81",
                "weightage": "100",
                "display_order": "1",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(
                    [
                        {
                            "name": "Notebook",
                            "max_marks": "10",
                            "display_order": 1,
                            "include_in_total": True,
                        },
                        {
                            "name": "Theory",
                            "max_marks": "70",
                            "display_order": 2,
                            "include_in_total": True,
                        },
                    ]
                ),
            },
            assessment=assessment,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("passing_marks", form.errors)

    def test_subject_form_cannot_be_edited_after_marks_entry_opens(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        open_marks_entry(assessment, actor=self.teacher)
        assessment.refresh_from_db()

        form = ReportCardAssessmentSubjectForm(
            data={
                "subject": self.math.pk,
                "max_marks": "80",
                "passing_marks": "30",
                "weightage": "100",
                "display_order": "1",
                "is_optional": "",
                "include_in_total": "on",
                "components_json": json.dumps(
                    [
                        {
                            "name": "Theory Exam",
                            "max_marks": "80",
                            "display_order": 1,
                            "include_in_total": True,
                        }
                    ]
                ),
            },
            instance=math,
            assessment=assessment,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("before marks entry opens", form.errors["__all__"][0])

    def test_rank_generation_works(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "95"},
                {"academic_session": self.session_3, "marks_obtained": "80"},
            ],
        )

        generate_assessment_results(assessment, actor=self.teacher)
        ranks = {
            result.admission_number_snapshot: result.rank
            for result in ReportCardStudentResult.objects.filter(assessment=assessment)
        }

        self.assertEqual(ranks["A002"], 1)
        self.assertEqual(ranks["A001"], 2)
        self.assertEqual(ranks["A003"], 2)

    def test_report_card_pdf_supports_multiple_pages_without_truncation_notice(self):
        assessment = self._assessment()
        for index in range(24):
            subject = Subject.objects.create(
                institute=self.institute,
                academic_year=self.year,
                name=f"Long PDF Subject {index + 1}",
            )
            self._subject(
                assessment,
                subject=subject,
                max_marks=Decimal("100"),
                passing_marks=Decimal("35"),
                weightage=Decimal("100"),
            )
        result = generate_assessment_results(assessment, actor=self.teacher, require_complete=False)[0]

        response = report_card_pdf_response(result)
        pdf_content = bytes(response.content)

        self.assertIn(b"/Count 2", pdf_content)
        self.assertNotIn(b"Continued on next page is not supported", pdf_content)

    def test_published_results_are_visible_and_unpublished_results_are_hidden(self):
        assessment = self._assessment("Visible Term")
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "60"},
            ],
        )
        generated = generate_assessment_results(assessment, actor=self.teacher)
        unpublished_result = next(result for result in generated if result.academic_session_id == self.session_1.pk)

        self.assertFalse(student_can_view_result(self.session_1.student.user, unpublished_result))
        self.assertEqual(list(get_published_results_for_student(self.session_1.student)), [])

        publish_assessment_results(assessment, actor=self.teacher)
        unpublished_result.refresh_from_db()

        self.assertTrue(student_can_view_result(self.session_1.student.user, unpublished_result))
        self.assertEqual(list(get_published_results_for_student(self.session_1.student)), [unpublished_result])

    def test_student_api_shows_only_published_or_locked_report_cards(self):
        assessment = self._assessment("Student API Visibility")
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "88"},
                {"academic_session": self.session_2, "marks_obtained": "77"},
                {"academic_session": self.session_3, "marks_obtained": "66"},
            ],
        )
        generated = generate_assessment_results(assessment, actor=self.teacher)
        student_result = next(result for result in generated if result.academic_session_id == self.session_1.pk)
        api_client = APIClient()
        api_client.force_authenticate(user=self.session_1.student.user)

        unpublished_list = api_client.get("/api/mobile/report-cards/student/")
        unpublished_detail = api_client.get(f"/api/mobile/report-cards/student/{student_result.pk}/")
        publish_assessment_results(assessment, actor=self.teacher)
        published_list = api_client.get("/api/mobile/report-cards/student/")
        published_detail = api_client.get(f"/api/mobile/report-cards/student/{student_result.pk}/")
        lock_assessment(assessment, actor=self.teacher)
        locked_list = api_client.get("/api/mobile/report-cards/student/")

        self.assertEqual(unpublished_list.status_code, 200)
        self.assertEqual(unpublished_list.data["results"], [])
        self.assertEqual(unpublished_detail.status_code, 404)
        self.assertEqual([item["id"] for item in published_list.data["results"]], [student_result.pk])
        self.assertEqual(published_detail.status_code, 200)
        self.assertEqual(published_detail.data["result"]["id"], student_result.pk)
        self.assertEqual([item["id"] for item in locked_list.data["results"]], [student_result.pk])

    def test_locked_assessments_cannot_be_edited_by_teacher(self):
        assessment = self._assessment()
        math = self._subject(assessment)
        self._open_and_save(
            math,
            [
                {"academic_session": self.session_1, "marks_obtained": "80"},
                {"academic_session": self.session_2, "marks_obtained": "70"},
                {"academic_session": self.session_3, "marks_obtained": "60"},
            ],
        )
        generate_assessment_results(assessment, actor=self.teacher)
        publish_assessment_results(assessment, actor=self.teacher)
        lock_assessment(assessment, actor=self.teacher)
        assessment.refresh_from_db()

        self.assertFalse(teacher_can_edit_assessment(self.teacher, assessment))
        with self.assertRaises(ValidationError):
            update_assessment(assessment, actor=self.teacher, title="Changed after lock")
        with self.assertRaises(ValidationError):
            bulk_save_subject_marks(
                math,
                [{"academic_session": self.session_1, "marks_obtained": "90"}],
                actor=self.teacher,
            )
