from io import BytesIO
from textwrap import wrap
from xml.sax.saxutils import escape

from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from .models import ReportCardComponentMarkEntry, ReportCardMarkEntry, ReportCardStudentResult
from .selectors import (
    get_active_student_sessions_for_assessment,
    get_assessment_subjects,
    get_generated_results,
    get_marks_grid,
    get_result_subject_rows,
)
from .services import bulk_save_subject_marks


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_filename(value):
    return "".join(char if char.isalnum() or char in {"-", "_"} else "-" for char in str(value or "").strip()).strip("-") or "report-card"


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _style_header(sheet):
    fill = PatternFill("solid", fgColor="E0F2FE")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="075985")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 3, 14), 34)


def marks_entry_template_response(assessment_subject):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Marks Entry"
    components = list(assessment_subject.components.all())
    mark_headers = [component.name_snapshot for component in components] or ["Marks"]
    sheet.append(
        [
            "Academic Session ID",
            "Student ID",
            "Admission Number",
            "Student Name",
            "Roll Number",
            *mark_headers,
            "Absent",
            "Remark",
        ]
    )
    for row in get_marks_grid(assessment_subject):
        session = row["academic_session"]
        student = row["student"]
        entry = row["mark_entry"]
        if components:
            component_values = [
                "" if not row["component_entries"].get(component.pk) or row["component_entries"][component.pk].marks_obtained is None else row["component_entries"][component.pk].marks_obtained
                for component in components
            ]
        else:
            component_values = ["" if not entry or entry.marks_obtained is None else entry.marks_obtained]
        sheet.append(
            [
                session.pk,
                student.pk,
                session.admission_number,
                student.user.get_full_name() or student.user.username,
                student.roll_number,
                *component_values,
                "Yes" if entry and entry.is_absent else "No",
                entry.remark if entry else "",
            ]
        )
    _style_header(sheet)
    filename = f"marks-template-{_safe_filename(assessment_subject.assessment.title)}-{_safe_filename(assessment_subject.subject_name_snapshot)}.xlsx"
    return _workbook_response(workbook, filename)


def _header_map(sheet):
    headers = {}
    for index, cell in enumerate(sheet[1], start=1):
        key = str(cell.value or "").strip().lower()
        if key:
            headers[key] = index
    return headers


def _cell(row, headers, name):
    index = headers.get(name)
    if not index:
        return None
    return row[index - 1].value


def _truthy(value):
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "absent", "a"}


def import_marks_workbook(assessment_subject, upload, *, actor=None):
    workbook = load_workbook(upload, data_only=True)
    sheet = workbook.active
    headers = _header_map(sheet)
    components = list(assessment_subject.components.all())
    required_headers = {"academic session id", "admission number", "absent"}
    if components:
        required_headers.update((component.name_snapshot or component.name).strip().lower() for component in components)
    else:
        required_headers.add("marks")
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        return {
            "saved_count": 0,
            "errors": [{"row": 1, "errors": [f"Missing column: {header}"]} for header in missing_headers],
        }

    sessions = {
        session.pk: session
        for session in get_active_student_sessions_for_assessment(assessment_subject.assessment)
    }
    sessions_by_admission = {session.admission_number: session for session in sessions.values()}
    rows = []
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        raw_session_id = _cell(row, headers, "academic session id")
        raw_admission = str(_cell(row, headers, "admission number") or "").strip()
        raw_marks = _cell(row, headers, "marks") if not components else None
        component_marks = {}
        for component in components:
            component_marks[component.pk] = _cell(row, headers, (component.name_snapshot or component.name).strip().lower())
        is_absent = _truthy(_cell(row, headers, "absent"))
        remark = str(_cell(row, headers, "remark") or "").strip()
        row_errors = []

        if not raw_session_id and not raw_admission and raw_marks in (None, "") and not is_absent and not remark:
            continue

        session = None
        if raw_session_id:
            try:
                session = sessions.get(int(raw_session_id))
            except (TypeError, ValueError):
                row_errors.append("Academic Session ID must be a number.")
        if not session and raw_admission:
            session = sessions_by_admission.get(raw_admission)

        if not session:
            row_errors.append("Student session was not found in this assessment class.")
        elif raw_admission and raw_admission != session.admission_number:
            row_errors.append("Admission number does not match the academic session.")

        if not is_absent and not components:
            if raw_marks in (None, ""):
                row_errors.append("Marks are required unless Absent is Yes.")
            else:
                try:
                    if float(raw_marks) < 0:
                        row_errors.append("Marks cannot be negative.")
                    if float(raw_marks) > float(assessment_subject.max_marks):
                        row_errors.append("Marks cannot exceed subject maximum marks.")
                except (TypeError, ValueError):
                    row_errors.append("Marks must be a number.")
        if not is_absent and components:
            for component in components:
                raw_component_marks = component_marks.get(component.pk)
                if raw_component_marks in (None, ""):
                    row_errors.append(f"{component.name_snapshot} marks are required unless Absent is Yes.")
                    continue
                try:
                    if float(raw_component_marks) < 0:
                        row_errors.append(f"{component.name_snapshot} marks cannot be negative.")
                    if float(raw_component_marks) > float(component.max_marks):
                        row_errors.append(f"{component.name_snapshot} marks cannot exceed {component.max_marks}.")
                except (TypeError, ValueError):
                    row_errors.append(f"{component.name_snapshot} marks must be a number.")

        if row_errors:
            errors.append({"row": row_number, "errors": row_errors})
            continue

        rows.append(
            {
                "academic_session": session,
                "marks_obtained": None if raw_marks in (None, "") else raw_marks,
                "component_marks": component_marks,
                "is_absent": is_absent,
                "remark": remark,
            }
        )

    if errors:
        return {"saved_count": 0, "errors": errors}

    try:
        saved_entries = bulk_save_subject_marks(assessment_subject, rows, actor=actor)
    except ValidationError as error:
        return {"saved_count": 0, "errors": [{"row": None, "errors": getattr(error, "messages", [str(error)])}]}
    return {"saved_count": len(saved_entries), "errors": []}


def consolidated_results_response(assessment):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidated Results"
    subjects = list(get_assessment_subjects(assessment))
    headers = [
        "Rank",
        "Admission Number",
        "Student Name",
        "Total Obtained",
        "Total Max Marks",
        "Percentage",
        "Grade",
        "Result",
        "Remark",
    ] + [f"{subject.subject_name_snapshot} ({subject.max_marks})" for subject in subjects] + [
        f"{subject.subject_name_snapshot} Grade" for subject in subjects
    ]
    sheet.append(headers)

    for result in get_generated_results(assessment):
        subject_results = {subject_result.assessment_subject_id: subject_result for subject_result in result.subject_results.all()}
        subject_values = []
        subject_grades = []
        for subject in subjects:
            subject_result = subject_results.get(subject.pk)
            if not subject_result:
                subject_values.append("")
                subject_grades.append("")
            elif subject_result.is_absent:
                subject_values.append("Absent")
                subject_grades.append("AB")
            else:
                subject_values.append(subject_result.obtained_marks)
                subject_grades.append(subject_result.grade)
        sheet.append(
            [
                result.rank,
                result.admission_number_snapshot,
                result.student_name_snapshot,
                result.total_obtained,
                result.total_max_marks,
                result.percentage,
                result.grade,
                result.get_result_status_display(),
                result.remark,
                *subject_values,
                *subject_grades,
            ]
        )
    _style_header(sheet)
    filename = f"report-card-results-{_safe_filename(assessment.title)}.xlsx"
    return _workbook_response(workbook, filename)


def _pdf_escape(value):
    return str(value or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_text(x, y, text, size=10, bold=False, color="111827"):
    font = "F2" if bold else "F1"
    return f"{_hex_to_rgb(color)} rg BT /{font} {size} Tf {x} {y} Td ({_pdf_escape(text)}) Tj ET\n"


def _hex_to_rgb(value):
    value = (value or "111827").lstrip("#")
    return " ".join(f"{int(value[index:index + 2], 16) / 255:.3f}" for index in (0, 2, 4))


def _pdf_rect(x, y, width, height, *, stroke="D1D5DB", fill=None, line_width=1):
    commands = [f"{line_width} w"]
    if fill:
        commands.append(f"{_hex_to_rgb(fill)} rg")
    commands.append(f"{_hex_to_rgb(stroke)} RG")
    commands.append(f"{x} {y} {width} {height} re {'B' if fill else 'S'}")
    return "\n".join(commands) + "\n"


def _pdf_line(x1, y1, x2, y2, *, color="D1D5DB", line_width=1):
    return f"{line_width} w {_hex_to_rgb(color)} RG {x1} {y1} m {x2} {y2} l S\n"


def _pdf_wrapped_text(x, y, text, *, size=9, bold=False, color="111827", chars=80, line_gap=12):
    stream = ""
    current_y = y
    for line in wrap(str(text or ""), width=chars) or [""]:
        stream += _pdf_text(x, current_y, line, size=size, bold=bold, color=color)
        current_y -= line_gap
    return stream, current_y


def _pdf_centered_text(x, width, y, text, *, size=10, bold=False, color="111827"):
    text = str(text or "")
    estimated_width = len(text) * size * 0.52
    return _pdf_text(x + max((width - estimated_width) / 2, 0), y, text, size=size, bold=bold, color=color)


def _pdf_image(x, y, width, height, image_name="Im1"):
    return f"q {width} 0 0 {height} {x} {y} cm /{image_name} Do Q\n"


def _prepare_pdf_logo(institute, *, max_size=180):
    logo = getattr(institute, "logo", None)
    if not logo:
        return None
    try:
        with default_storage.open(logo.name, "rb") as logo_file:
            image = Image.open(logo_file)
            image.thumbnail((max_size, max_size))
            canvas = Image.new("RGB", image.size, "white")
            if image.mode in {"RGBA", "LA"}:
                canvas.paste(image.convert("RGBA"), mask=image.convert("RGBA").getchannel("A"))
            else:
                canvas.paste(image.convert("RGB"))
            buffer = BytesIO()
            canvas.save(buffer, format="JPEG", quality=92)
            return {
                "data": buffer.getvalue(),
                "width": canvas.width,
                "height": canvas.height,
            }
    except Exception:
        return None


def _build_pdf_stream(stream, *, width=595, height=842, images=None):
    content = stream.encode("latin-1", errors="replace")
    images = images or {}
    image_resource = ""
    if images:
        image_resource = " /XObject << " + " ".join(
            f"/{name} {7 + index} 0 R" for index, name in enumerate(images)
        ) + " >>"
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 4 0 R /F2 5 0 R >>{image_resource} >> /Contents 6 0 R >>".encode(),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
        b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n" + content + b"endstream",
    ]
    for image in images.values():
        objects.append(
            (
                f"<< /Type /XObject /Subtype /Image /Width {image['width']} /Height {image['height']} "
                f"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length {len(image['data'])} >>\n"
            ).encode()
            + b"stream\n"
            + image["data"]
            + b"\nendstream"
        )
    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf += f"{index} 0 obj\n".encode() + obj + b"\nendobj\n"
    xref_offset = len(pdf)
    pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode()
    for offset in offsets[1:]:
        pdf += f"{offset:010d} 00000 n \n".encode()
    pdf += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode()
    return pdf


def _display(value, fallback="-"):
    return fallback if value in (None, "") else str(value)


def _institute_initials(name):
    words = [word for word in str(name or "RC").split() if word]
    initials = "".join(word[0].upper() for word in words[:3])
    return initials or "RC"


def report_card_pdf_response(result, *, download=True):
    subject_rows = get_result_subject_rows(result)
    assessment = result.assessment
    institute = assessment.institute
    institute_name = assessment.institute_name_snapshot or getattr(institute, "name", "")
    institute_address = getattr(institute, "address", "") or "Address not configured"
    contact_parts = []
    if getattr(institute, "phone", ""):
        contact_parts.append(f"Phone: {institute.phone}")
    if getattr(institute, "email", ""):
        contact_parts.append(f"Email: {institute.email}")
    contact_line = " | ".join(contact_parts) or "Contact details not configured"
    logo_image = _prepare_pdf_logo(institute)
    pdf_images = {"Im1": logo_image} if logo_image else {}

    stream = ""
    width = 595
    height = 842
    margin = 32
    y = height - 32

    stream += _pdf_rect(margin, 30, width - (margin * 2), height - 60, stroke="111827", line_width=1.4)
    header_height = 132
    header_bottom = y - header_height
    logo_x = margin + 18
    logo_y = y - 92
    text_x = margin + 90
    text_width = width - margin - text_x - 16
    identity_width_chars = 52
    name_lines = wrap(institute_name, width=identity_width_chars) or [institute_name]
    name_size = 12.2 if len(institute_name) > 58 else 13.6
    name_line_gap = 13

    stream += _pdf_rect(margin, header_bottom, width - (margin * 2), header_height, stroke="CBD5E1", fill="F8FAFC")
    stream += _pdf_line(margin, y - 10, width - margin, y - 10, color="6042C8", line_width=2.4)
    stream += _pdf_line(margin, header_bottom + 22, width - margin, header_bottom + 22, color="CBD5E1", line_width=.8)
    if logo_image:
        logo_box_size = 58
        scale = min(logo_box_size / logo_image["width"], logo_box_size / logo_image["height"])
        draw_width = logo_image["width"] * scale
        draw_height = logo_image["height"] * scale
        stream += _pdf_image(
            logo_x + ((logo_box_size - draw_width) / 2),
            logo_y + ((logo_box_size - draw_height) / 2),
            draw_width,
            draw_height,
        )
    else:
        stream += _pdf_text(logo_x + 8, logo_y + 24, _institute_initials(institute_name), size=16, bold=True, color="442C9F")

    name_y = y - 26
    for name_line in name_lines[:3]:
        stream += _pdf_centered_text(text_x, text_width, name_y, name_line, size=name_size, bold=True, color="111827")
        name_y -= name_line_gap
    if len(name_lines) > 3:
        stream += _pdf_centered_text(text_x, text_width, name_y + 2, "...", size=10, bold=True, color="111827")

    address_lines = wrap(institute_address, width=68)[:2] or [institute_address]
    address_y = max(name_y - 2, header_bottom + 48)
    for address_line in address_lines:
        stream += _pdf_centered_text(text_x, text_width, address_y, address_line, size=7.6, color="475569")
        address_y -= 9

    for contact_line_part in wrap(contact_line, width=72)[:2]:
        stream += _pdf_centered_text(text_x, text_width, address_y - 1, contact_line_part, size=7.6, color="334155")
        address_y -= 9

    stream += _pdf_centered_text(text_x, text_width, header_bottom + 8, "OFFICIAL RESULT | REPORT CARD", size=8.6, bold=True, color="6042C8")

    y = header_bottom - 24
    stream += _pdf_text(margin, y, assessment.title.upper(), size=12.5, bold=True, color="111827")
    stream += _pdf_text(margin, y - 15, f"Academic Year: {assessment.academic_year_name_snapshot}    Class / Batch: {assessment.batch_name_snapshot}", size=8.5, color="475569")

    y -= 34
    stream += _pdf_rect(margin, y - 50, width - (margin * 2), 50, stroke="CBD5E1", fill="FFFFFF")
    stream += _pdf_text(margin + 14, y - 15, "Student Name", size=6.8, bold=True, color="64748B")
    stream += _pdf_text(margin + 14, y - 31, result.student_name_snapshot, size=9.3, bold=True)
    stream += _pdf_text(margin + 250, y - 15, "Admission Number", size=6.8, bold=True, color="64748B")
    stream += _pdf_text(margin + 250, y - 31, result.admission_number_snapshot, size=9.3, bold=True)
    stream += _pdf_text(margin + 410, y - 15, "Result Date", size=6.8, bold=True, color="64748B")
    stream += _pdf_text(margin + 410, y - 31, _display(assessment.result_date), size=9.3, bold=True)

    y -= 72
    stream += _pdf_text(margin, y, "SUBJECT-WISE PERFORMANCE", size=10, bold=True, color="111827")
    y -= 15
    table_x = margin
    table_w = width - (margin * 2)
    columns = [("Subject", 145), ("Components / Marks", 165), ("Total", 64), ("Max", 56), ("%", 47), ("Grade", 54)]
    row_h = 26
    stream += _pdf_rect(table_x, y - row_h, table_w, row_h, stroke="CBD5E1", fill="EEF2FF")
    x = table_x
    for title, col_w in columns:
        stream += _pdf_text(x + 8, y - 17, title, size=7.4, bold=True, color="334155")
        if x > table_x:
            stream += _pdf_line(x, y, x, y - row_h, color="CBD5E1")
        x += col_w
    y -= row_h

    for index, row in enumerate(subject_rows, start=1):
        subject = row["assessment_subject"]
        subject_result = row["subject_result"]
        entry = row["mark_entry"]
        component_bits = []
        for component_entry in row["component_entries"]:
            component_marks = "Absent" if component_entry.is_absent else component_entry.marks_obtained
            component_bits.append(f"{component_entry.component.name_snapshot}: {component_marks}/{component_entry.component.max_marks}")
        component_text = "; ".join(component_bits) if component_bits else "Single marks column"
        marks = "Absent" if subject_result and subject_result.is_absent else (
            subject_result.obtained_marks if subject_result else (entry.marks_obtained if entry else "-")
        )
        max_marks = subject_result.max_marks if subject_result else subject.max_marks
        percentage = subject_result.percentage if subject_result else "-"
        grade = subject_result.grade if subject_result else "-"

        wrapped_components = wrap(component_text, width=50) or [component_text]
        row_height = max(30, 16 + (min(len(wrapped_components), 3) * 8))
        if y - row_height < 170:
            stream += _pdf_text(margin, 58, "Continued on next page is not supported in this compact PDF. Use Excel export for full long-format records.", size=8, color="B45309")
            break
        stream += _pdf_rect(table_x, y - row_height, table_w, row_height, stroke="E5E7EB", fill="FFFFFF" if index % 2 else "F8FAFC")
        x = table_x
        for _, col_w in columns:
            if x > table_x:
                stream += _pdf_line(x, y, x, y - row_height, color="E5E7EB")
            x += col_w
        stream += _pdf_text(table_x + 8, y - 17, subject.subject_name_snapshot, size=8.3, bold=True)
        component_y = y - 16
        for component_line in wrapped_components[:3]:
            stream += _pdf_text(table_x + 153, component_y, component_line, size=6.4, color="475569")
            component_y -= 8
        stream += _pdf_text(table_x + 318, y - 17, _display(marks), size=8.2, bold=True)
        stream += _pdf_text(table_x + 382, y - 17, _display(max_marks), size=8.2)
        stream += _pdf_text(table_x + 438, y - 17, _display(percentage), size=8.2)
        stream += _pdf_text(table_x + 485, y - 17, _display(grade), size=8.2, bold=True, color="6042C8")
        y -= row_height

    y -= 18
    box_width = (width - (margin * 2) - 20) / 3
    summary_boxes = [
        ("Total Marks", f"{result.total_obtained} / {result.total_max_marks}", "6042C8"),
        ("Percentage", _display(result.percentage), "15803D"),
        ("Grade / Rank", f"{_display(result.grade)} / {_display(result.rank)}", "D97706"),
    ]
    for index, (label, value, color) in enumerate(summary_boxes):
        x = margin + index * (box_width + 10)
        stream += _pdf_rect(x, y - 42, box_width, 42, stroke="CBD5E1", fill="F8FAFC")
        stream += _pdf_text(x + 10, y - 15, label, size=6.8, bold=True, color="64748B")
        stream += _pdf_text(x + 10, y - 32, value, size=10.5, bold=True, color=color)

    y -= 58
    result_color = "15803D" if result.result_status == "PASS" else "B91C1C" if result.result_status == "FAIL" else "C2410C"
    stream += _pdf_rect(margin, y - 27, width - (margin * 2), 27, stroke="CBD5E1", fill="F8FAFC")
    stream += _pdf_text(margin + 12, y - 18, f"Final Result: {result.get_result_status_display()}", size=10, bold=True, color=result_color)
    if result.remark:
        stream += _pdf_text(margin + 238, y - 18, f"Remark: {result.remark}", size=8, color="475569")

    y -= 46
    if y > 90:
        stream += _pdf_line(margin, y, margin + 145, y, color="111827")
        stream += _pdf_line(width - margin - 145, y, width - margin, y, color="111827")
        stream += _pdf_text(margin + 28, y - 15, "Class Teacher", size=8, color="475569")
        stream += _pdf_text(width - margin - 112, y - 15, "Principal / Director", size=8, color="475569")
    stream += _pdf_text(margin, 48, "This is a computer-generated report card. Verify school seal/signature where required.", size=7.5, color="64748B")
    stream += _pdf_text(width - 174, 48, "Powered by UltraCoachMatrix", size=7.5, color="64748B")

    response = HttpResponse(_build_pdf_stream(stream, images=pdf_images), content_type="application/pdf")
    disposition = "attachment" if download else "inline"
    filename = f"report-card-{_safe_filename(result.admission_number_snapshot)}-{_safe_filename(assessment.title)}.pdf"
    response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
    return response
